import datetime
from http.client import NOT_FOUND
import locale
from threading import local

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMultiAlternatives
from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Group
from django.core.files.storage import default_storage

import openpyxl
from xlwt import Workbook

from receiving.models import (
    LineRequestFinishedHistory,
    LineRequestFinished,
    ICDR,
    ICDRBackup,
)
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseNotFound,
    Http404,
)
from shippers.models import Gateway, GatewayBackup
from .models import *
from qad_ee.models import WoMstr
from xlwt import Workbook

@login_required()
@user_passes_test(lambda u: u.groups.filter(name="crossdocking").exists())
def prodlineFilter(request):
    firstDayPrevMonth = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)
    ).replace(day=15)

    elements = ProdlineTableHistory.objects
    #elements = EmailDiarioSchedule.objects
    #print(elements.all())
    todayItems = ProdlineTable.objects
    return render(
        request,
        "crossdocking/prodlinesFilter.html",
        {"dadosDia": todayItems, "elements": elements},
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists())
def configurations(request):
    prodlines = Prodlines.objects
    users = User.objects
    return render(
        request,
        "crossdocking/configurationsWhProd.html",
        {"prodlines": prodlines, "users": users},
    )


def create(request):
    if request.method == "POST":
        try:
            nome = Prodlines.objects.get(nome=request.POST["nome"].replace(" ", ""))
        except ObjectDoesNotExist:
            prodline = Prodlines()
            prodline.nome = request.POST["nome"].replace(" ", "")
            prodline.save()

            message = "Adicionada Prodline " + prodline.nome
            subject, from_email, to = (
                "Alteração em Crossdocking - Configurations",
                "noreply@visteon.com",
                ["pmarti30@visteon.com", "npires2@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect("crossdocking:configurations")

    prodlines = Prodlines.objects
    return render(
        request,
        "crossdocking/configurationsWhProd.html",
        {"prodlines": prodlines, "erro": "Prodline already exists"},
    )


def delete(request):
    if request.method == "POST":
        nome = request.POST["nome2"]
        try:
            nome = Prodlines.objects.get(nome=nome)
        except ObjectDoesNotExist:
            prodlines = Prodlines.objects
            return render(
                request,
                "crossdocking/configurationsWhProd.html",
                {"prodlines": prodlines, "erro": "Prodline not exists"},
            )

        message = "Eliminada Prodline " + nome.nome
        subject, from_email, to = (
            "Alteração em Crossdocking - Configurations",
            "noreply@visteon.com",
            ["pmarti30@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("crossdocking:configurations")


def submitAll(request):
    if request.method == "POST":
        user = request.user
        try:
            todayItems = ProdlineTable.objects.all()
            if request.POST["select"] == "select":
                if request.POST["posicao"] == "receiving":
                    for element in todayItems:
                        element.allOkReceiving = True
                        element.save()
                    enviaEmail(request, user)
                elif request.POST["posicao"] == "shipping":
                    for element in todayItems:
                        element.allOkShipping = True
                        element.save()
                    enviaEmailShipping(request, user)
            elif request.POST["select"] == "deselect":
                if request.POST["posicao"] == "receiving":
                    for element in todayItems:
                        element.allOkReceiving = False
                        element.save()
                    enviaEmail(request, user)
                elif request.POST["posicao"] == "shipping":
                    for element in todayItems:
                        element.allOkShipping = False
                        element.save()
                    enviaEmailShipping(request, user)
        except IndexError:
            todayItems = ProdlineTable.objects
            return render(
                request, "crossdocking/prodlinesFilter.html", {"dadosDia": todayItems}
            )
    todayItems = ProdlineTable.objects
    return render(
        request, "crossdocking/prodlinesFilter.html", {"dadosDia": todayItems}
    )


def changeCheckbox(request):
    if request.method == "POST":
        id = request.POST["id"]

        if request.POST["shippingCheck"] == "true":
            element = ProdlineTable.objects.get(id=id)
            element.shipping = True
            element.save()
        elif request.POST["shippingCheck"] == "false":
            element = ProdlineTable.objects.get(id=id)
            element.shipping = False
            element.save()
        if request.POST["receivingCheck"] == "true":
            element = ProdlineTable.objects.get(id=id)
            element.receiving = True
            element.save()
        elif request.POST["receivingCheck"] == "false":
            element = ProdlineTable.objects.get(id=id)
            element.receiving = False
            element.save()
        todayItems = ProdlineTable.objects
        return render(
            request, "crossdocking/prodlinesFilter.html", {"dadosDia": todayItems}
        )


def updateComentReceiving(request):
    if request.method == "POST":
        id = request.POST["id"]
        comentario = request.POST["comentario"]

        element = ProdlineTable.objects.get(id=id)
        element.comentarioReceiving = comentario
        element.save()

        todayItems = ProdlineTable.objects
        return render(
            request, "crossdocking/prodlinesFilter.html", {"dadosDia": todayItems}
        )


def updateComentShipping(request):
    if request.method == "POST":
        id = request.POST["id"]
        comentario = request.POST["comentario"]

        element = ProdlineTable.objects.get(id=id)
        element.comentarioShipping = comentario
        element.save()

        todayItems = ProdlineTable.objects
        return render(
            request, "crossdocking/prodlinesFilter.html", {"dadosDia": todayItems}
        )


def schedulePlanning(request):
    if request.method == "POST":
        print("Request->", request.POST)
        prodlines = Prodlines.objects
        path = request.POST.get("textPath")
        path1 = request.POST.get("textPath1")
        
        list = []
        # BUG func que verifica se o file é aberto com pontos ou _
        workbook = EmailDiarioSchedule.objects.raw(f""" SELECT * FROM crossdocking_emaildiarioschedule WHERE due_date BETWEEN '{request.POST["requestDay"]} 00:00:00' AND '{request.POST["requestDay"]} 23:59:59' """)
        #print("querry", workbook)
        try:
            workbook = EmailDiarioSchedule.objects.raw(f""" SELECT * FROM crossdocking_emaildiarioschedule WHERE due_date BETWEEN '{request.POST["requestDay"]} 00:00:00' AND '{request.POST["requestDay"]} 23:59:59' """)
            print("Tamanho dos dados Da BD--->", len(workbook))
            if len(workbook) == 0:
                todayItems = ProdlineTable.objects
                elements = ProdlineTableHistory.objects
                return render(
                    request,
                    "crossdocking/prodlinesFilter.html",
                    {
                        "dadosDia": todayItems,
                        "erro5": "No Data found",
                        "elements": elements,
                    },
                )
            try:

                worksheet =  EmailDiarioSchedule.objects.raw(f""" SELECT * FROM crossdocking_emaildiarioschedule WHERE due_date BETWEEN '{request.POST["requestDay"]} 00:00:00' AND '{request.POST["requestDay"]} 23:59:59' """)
                for element in worksheet:
                    for prod in prodlines.all():
                        print("Tas ai",prod.nome ,"|", element.line )

                        if prod.nome == element.line:
                            print("dados->",element)
                            elem = {
                                "line": element.line,
                                "site": element.site,
                                "dueDate": str(element.due_date),
                                "itemNumber": element.item_number,
                                "description": element.description ,
                                "toComplete": str(element.to_complete),
                            }
                            list.append(elem)
                        print("tamanho da lista", len(list))
                todayItems = ProdlineTable.objects
                elements = ProdlineTableHistory.objects
                if len(list) == 0:
                     return render(
                        request,
                        "crossdocking/prodlinesFilter.html",
                        {
                            "dadosDia": todayItems,
                            "erro5": "No Data to present",
                            "elements": elements,
                        },
                    )
                else:
                    return render(
                        request,
                        "crossdocking/prodlinesFilter.html",
                        {"dadosDia": todayItems, "lista": list, "elements": elements},
                    )
            except:
                try:
                    worksheet = EmailDiarioSchedule.objects.raw(f""" SELECT * FROM crossdocking_emaildiarioschedule WHERE due_date BETWEEN '{request.POST["requestDay"]} 00:00:00' AND '{request.POST["requestDay"]} 23:59:59' """)
                    for element in worksheet:
                        for prod in prodlines.all():
                            if prod.nome == element[0].value:
                                elem = {
                                    "line": element[0].value,
                                    "site": element[1].value,
                                    "dueDate": str(element[2].value),
                                    "itemNumber": element[3].value,
                                    "description": element[4].value,
                                    "toComplete": str(element[5].value),
                                }
                                list.append(elem)
                    todayItems = ProdlineTable.objects
                    elements = ProdlineTableHistory.objects
                    return render(
                        request,
                        "crossdocking/prodlinesFilter.html",
                        {"dadosDia": todayItems, "lista": list, "elements": elements},
                    )
                except:
                    todayItems = ProdlineTable.objects
                    elements = ProdlineTableHistory.objects
                    return render(
                        request,
                        "crossdocking/prodlinesFilter.html",
                        {
                            "dadosDia": todayItems,
                            "erro5": "No Data found",
                            "elements": elements,
                        },
                    )

        except:
            todayItems = ProdlineTable.objects
            elements = ProdlineTableHistory.objects
            return render(
                request,
                "crossdocking/prodlinesFilter.html",
                {
                    "dadosDia": todayItems,
                    "erro5": "Data not exists",
                    "elements": elements,
                },
            )

    elements = ProdlineTableHistory.objects
    todayItems = ProdlineTable.objects
    return render(
        request,
        "crossdocking/prodlinesFilter.html",
        {"dadosDia": todayItems, "elements": elements},
    )


def changeUserGroups(request):
    if request.method == "POST":
        user = User.objects.get(username=request.POST["username"])
        grupo = request.POST["paginas"]
        if User.objects.filter(
            username=request.POST["username"], groups__name="crossdocking"
        ):
            my_group = Group.objects.using("default").get(name="crossdocking")
            my_group.user_set.remove(user)
        if grupo == "none":
            message = (
                "User " + user.username + " perdeu acesso à página de Crossdocking."
            )
            subject, from_email, to = (
                "Alteração em Crossdocking - Configurations",
                "noreply@visteon.com",
                ["pmarti30@visteon.com", "npires2@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect("crossdocking:configurations")
        if grupo == "crossdocking":
            my_group = Group.objects.using("default").get(name="crossdocking")
            my_group.user_set.add(user)
        message = (
            "User "
            + user.username
            + " com páginas acessiveis - "
            + grupo.replace("/", " , ")
        )
        subject, from_email, to = (
            "Alteração em Crossdocking - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("crossdocking:configurations")


def reportCrossdocking(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Crossdocking",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("crossdocking:prodlineFilter")


def enviaEmail(request, user):
    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime("%A") == "Saturday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime("%A") == "Sunday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime("%m")
    if mes == "01":
        mes = "Janeiro"
    if mes == "02":
        mes = "Fevereiro"
    if mes == "03":
        mes = "Março"
    if mes == "04":
        mes = "Abril"
    if mes == "05":
        mes = "Maio"
    if mes == "06":
        mes = "Junho"
    if mes == "07":
        mes = "Julho"
    if mes == "08":
        mes = "Agosto"
    if mes == "09":
        mes = "Setembro"
    if mes == "10":
        mes = "Outubro"
    if mes == "11":
        mes = "Novembro"
    if mes == "12":
        mes = "Dezembro"
    mensagem = ""
    #BUG FILESYS
    textPath = (
        "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
        + actualDay.strftime("%Y")
        + "/"
        + mes
        + " "
        + actualDay.strftime("%Y")
        + "/Daily_Schedule_"
        + actualDay.strftime("%d.%m.%Y")
        + ".xlsx"
    )

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = (
        "</br>User: <b>"
        + user.username
        + '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
    )

    valueSubmitReceiving = ""
    valueSubmitShipping = ""

    for elem in elementosDiaSeguinte.all():
        valueReceiving = ""
        valueShipping = ""

        if elem.comentarioShipping == None:
            elem.comentarioShipping = " "
        if elem.comentarioReceiving == None:
            elem.comentarioReceiving = " "
        if elem.shipping == True:
            valueShipping = "X"
        if elem.receiving == True:
            valueReceiving = "X"
        if elem.allOkShipping == True:
            valueSubmitShipping = "X"
        if elem.allOkReceiving == True:
            valueSubmitReceiving = "X"
        table += (
            '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
            + elem.line
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.site
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.due_date
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.item_number
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.description
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.to_complete
            + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueReceiving
            + '</td><td style="padding:0 15px 0 15px;">'
            + str(elem.comentarioReceiving)
            + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueShipping
            + '</td><td style="padding:0 15px 0 15px;">'
            + str(elem.comentarioShipping)
            + "</td></tr>"
        )

    table += (
        '<tr style="background-color: lightgray"><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
        + valueSubmitReceiving
        + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
        + valueSubmitShipping
        + "</td><td></td></tr>"
    )
    table += "</tbody></table>"
    table += "</br></br><b>Prodlines</b></br>"
    for prod in prodline.all():
        table += prod.nome + "</br>"

    table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
    subject, from_email, to = (
        "Wh Production " + actualDay.strftime("%d-%m-%Y"),
        "noreply@visteon.com",
        [
            "pmarti30@visteon.com",
        ],
    )

    msg = EmailMultiAlternatives(subject, table, from_email, to)
    msg.attach_alternative(table, "text/html")
    msg.send()


def enviaEmailShipping(request, user):
    dadosQAD = WoMstr.objects.values(
        "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
    )

    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime("%A") == "Saturday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime("%A") == "Sunday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime("%m")
    if mes == "01":
        mes = "Janeiro"
    if mes == "02":
        mes = "Fevereiro"
    if mes == "03":
        mes = "Março"
    if mes == "04":
        mes = "Abril"
    if mes == "05":
        mes = "Maio"
    if mes == "06":
        mes = "Junho"
    if mes == "07":
        mes = "Julho"
    if mes == "08":
        mes = "Agosto"
    if mes == "09":
        mes = "Setembro"
    if mes == "10":
        mes = "Outubro"
    if mes == "11":
        mes = "Novembro"
    if mes == "12":
        mes = "Dezembro"
    mensagem = ""
    #BUG FILESYS
    textPath = (
        "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
        + actualDay.strftime("%Y")
        + "/"
        + mes
        + " "
        + actualDay.strftime("%Y")
        + "/Daily_Schedule_"
        + actualDay.strftime("%d.%m.%Y")
        + ".xlsx"
    )

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = (
        "</br>User: <b>"
        + user.username
        + '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
    )

    valueSubmitReceiving = ""
    valueSubmitShipping = ""

    for elem in elementosDiaSeguinte.all():
        valueReceiving = ""
        valueShipping = ""

        if elem.comentarioShipping == None:
            elem.comentarioShipping = " "
        if elem.comentarioReceiving == None:
            elem.comentarioReceiving = " "
        if elem.shipping == True:
            valueShipping = "X"
        if elem.receiving == True:
            valueReceiving = "X"
        if elem.allOkShipping == True:
            valueSubmitShipping = "X"
        if elem.allOkReceiving == True:
            valueSubmitReceiving = "X"
        table += (
            '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
            + elem.line
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.site
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.due_date
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.item_number
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.description
            + '</td><td style="padding:0 15px 0 15px;">'
            + elem.to_complete
            + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueReceiving
            + '</td><td style="padding:0 15px 0 15px;">'
            + str(elem.comentarioReceiving)
            + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueShipping
            + '</td><td style="padding:0 15px 0 15px;">'
            + str(elem.comentarioShipping)
            + "</td></tr>"
        )

    table += (
        '<tr style="background-color: lightgray"><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
        + valueSubmitReceiving
        + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
        + valueSubmitShipping
        + "</td><td></td></tr>"
    )
    table += "</tbody></table>"

    table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

    for elem3 in elementosDiaSeguinte.all():
        elem2 = dadosQAD.filter(
            wo_due_date=elem3.due_date, wo_part=elem3.item_number, wo_site="3515"
        )
        for elem1 in elem2.all():
            if (
                str(elem1["wo_qty_exp_complete"])[:-11]
                != str(elem1["wo_qty_comp"])[:-11]
            ):
                table += (
                    '<tr style="background-color: red">'
                    '<td style="padding:0 15px 0 15px;">'
                    + elem3.line
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem3.site
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_due_date"])
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem1["wo_part"]
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_qty_exp_complete"])[:-11]
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_qty_comp"])[:-11]
                    + "</td></tr>"
                )
            else:
                table += (
                    '<tr style="background-color: whitesmoke">'
                    '<td style="padding:0 15px 0 15px;">'
                    + elem3.line
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem3.site
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_due_date"])
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem1["wo_part"]
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_qty_exp_complete"])[:-11]
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem1["wo_qty_comp"])[:-11]
                    + "</td></tr>"
                )

    table += "</tbody></table>"
    table += "</br></br><b>Prodlines</b></br>"
    for prod in prodline.all():
        table += prod.nome + "</br>"

    table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
    subject, from_email, to = (
        "Wh Production " + actualDay.strftime("%d-%m-%Y"),
        "noreply@visteon.com",
        ["pmarti30@visteon.com"],
    )
    msg = EmailMultiAlternatives(subject, table, from_email, to)
    msg.attach_alternative(table, "text/html")
    msg.send()



def enviarEmailSchedule():
    """  ["pmarti30@visteon.com",'npires2@visteon.com',
                     'abrandao@visteon.com','sanasta1@visteon.com',
                    'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                    'abilro1@visteon.com', 'jrodri80@visteon.com',
                    'evenanc1@visteon.com'] """
    import locale
    locale.setlocale(locale.LC_ALL, "pt-PT")
    timeNow = datetime.datetime.now()
    timeEnd = timeNow.replace(hour=8, minute=10)
    timeBefore = timeNow.replace(hour=7, minute=55)
    timeEnd2 = timeNow.replace(hour=16, minute=10)
    timeBefore2 = timeNow.replace(hour=15, minute=55)
    timeEnd3 = timeNow.replace(hour=22, minute=10)
    timeBefore3 = timeNow.replace(hour=21, minute=55)

    dadosQAD = WoMstr.objects.values(
        "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
    )

    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today() - datetime.timedelta(days=2)
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime("%A") == "Saturday" or nextDay.strftime("%A") == "sábado": #sábado
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime("%A") == "Sunday" or nextDay.strftime("%A") == "domingo": #domingo
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    month = nextDay.strftime("%B")
    mes = nextDay.strftime("%m")
    fmt_num = nextDay.strftime("%d.%m.%Y")
    mensagem = ""
    textPath = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num}.xlsx"  # + nextDay.strftime(
    #'%Y') + "/" + mes + " " + nextDay.strftime('%Y') + "/Daily_Schedule_" + nextDay.strftime(
    #'%d.%m.%Y') + ".xlsx"
    print(textPath + " :TEXTPATH")
    fmt_num1 = nextDay.strftime("%d_%m_%Y")
    fmt_num2 = nextDay.strftime("%d-%m-%Y")
    fmt_num3 = nextDay.strftime("%d/%m/%Y")
    textPath1 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num1}.xlsx"
    textPath2 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num2}.xlsx"
    textPath3 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num3}.xlsx"
    def defineBook():
        try:
            workbook = openpyxl.load_workbook(textPath)
            print("Fez o textPath")
        except:
            try: 
             workbook = openpyxl.load_workbook(textPath1)
             print("Fez o textPath 1")
            except:
                try:
                    workbook = openpyxl.load_workbook(textPath2)
                    print("Fez o textPath 2")
                except:
                    workbook = openpyxl.load_workbook(textPath3)
                    print("Fez o textPath 3")
        return workbook

    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/#Daily_Schedule_03.11.2021.xlsx'

    table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
    print("ENTROU")
    if (
        (timeNow > timeEnd and timeNow > timeBefore)
        and (actualDay.strftime("%A") != "Saturday" ) #sábado
        and (actualDay.strftime("%A") != "Sunday" ) #domingo
    ):
        table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
        try:
            workbook = defineBook()
            sheet_num = ""
            for sheet in workbook:
                # print(sheet.title)
                # print("CELL 1 --> ",sheet.cell(1,1).value)
                if sheet.cell(1, 1)._value == "Line":
                    sheet_num = sheet.title
                    print(sheet_num)
            try:
                worksheet = workbook["Schedule Data"]
                for element in worksheet:
                    for prod in prodline.all():
                        if prod.nome == element[0].value:
                            table += (
                                '<tr><td style="padding:0 15px 0 15px;">'
                                + element[0].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[1].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + str(element[2].value)
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[3].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[4].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + str(element[5].value)
                                + "</td></tr>"
                            )

                table += "</tbody></table>"
                table += "</br></br><b>Prodlines</b></br>"
                for prod in prodline.all():
                    table += prod.nome + "</br>"

                table += (
                    "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
                )

                subject, from_email, to = (
                    "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                    "noreply@visteon.com",
                    ["pmarti30@visteon.com"],
                    
                ) 
                """  """
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                print(" Mensagem a ser enviada no mail")
                msg.send()
            except:
                try:
                    worksheet = workbook[sheet_num]
                    for element in worksheet:
                        for prod in prodline.all():
                            if prod.nome == element[0].value:
                                table += (
                                    '<tr><td style="padding:0 15px 0 15px;">'
                                    + element[0].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[1].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + str(element[2].value)
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[3].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[4].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + str(element[5].value)
                                    + "</td></tr>"
                                )

                    table += "</tbody></table>"
                    table += "</br></br><b>Prodlines</b></br>"
                    for prod in prodline.all():
                        table += prod.nome + "</br>"

                    table += (
                        "</b></b></br><b>Ficheiro diário de Schedule em: </b>"
                        + textPath
                    )

                    subject, from_email, to = (
                        "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                        "noreply@visteon.com",
                       ["pmarti30@visteon.com"],
                    )
                    msg = EmailMultiAlternatives(subject, table, from_email, to)
                    msg.attach_alternative(table, "text/html")
                    print("vai mandar o email")
                    msg.send()
                except KeyError:
                    mensagem = (
                        "</br>Worksheet name not found. Should be: Schedule Data or Sheet1 </br>"
                        + textPath
                        + ""
                    )
                    subject, from_email, to = (
                        "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                        "noreply@visteon.com",
                        ["pmarti30@visteon.com"],
                    )
                    msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
                    msg.attach_alternative(mensagem, "text/html")
                    print("vai mandar o email")
                    msg.send()
        except FileNotFoundError:
            mensagem = "</br> File not found. </br></br>" + textPath + ""
            subject, from_email, to = (
                "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
            msg.attach_alternative(mensagem, "text/html")
            msg.send()

    elif (
        (timeNow < timeEnd2 and timeNow > timeBefore2)
        and (not ProdlineTable.objects.filter(allOkReceiving=True).exists())
        and (actualDay.strftime("%A") != "Saturday" and actualDay.strftime("%A") != "sábado" )
        and (actualDay.strftime("%A") != "Sunday" and actualDay.strftime("%A") != "domingo")
        and (elementosDiaSeguinte.count() > 0)
    ):

        valueSubmitReceiving = ""
        valueSubmitShipping = ""

        textPath = (
            "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
            + actualDay.strftime("%Y")
            + "/"
            + mes
            + " "
            + actualDay.strftime("%Y")
            + "/Daily_Schedule_"
            + actualDay.strftime("%d.%m.%Y")
            + ".xlsx"
        )

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ""
            valueShipping = ""

            if elem.comentarioShipping == None:
                elem.comentarioShipping = " "
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = " "
            if elem.shipping == True:
                valueShipping = "X"
            if elem.receiving == True:
                valueReceiving = "X"
            if elem.allOkShipping == True:
                valueSubmitShipping = "X"
            if elem.allOkReceiving == True:
                valueSubmitReceiving = "X"
            table += (
                '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
                + elem.line
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.site
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.item_number
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.description
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.to_complete
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueReceiving
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioReceiving)
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueShipping
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioShipping)
                + "</td></tr>"
            )

        table += (
            '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitReceiving
            + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitShipping
            + "</td><td></td></tr>"
        )
        table += "</tbody></table>"
        table += "</br></br><b>Prodlines</b></br>"
        for prod in prodline.all():
            table += prod.nome + "</br>"

        table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

        subject, from_email, to = (
            "Wh Production " + actualDay.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        msg.send()

    elif (
        (timeNow < timeEnd3 and timeNow > timeBefore3)
        and (
            not ProdlineTable.objects.filter(allOkReceiving=True).exists()
            or not ProdlineTable.objects.filter(allOkShipping=True).exists()
        )
        and (actualDay.strftime("%A") != "Saturday")
        and (actualDay.strftime("%A") != "Sunday")
        and (elementosDiaSeguinte.count() > 0)
    ):

        valueSubmitReceiving = ""
        valueSubmitShipping = ""

        textPath = (
            "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
            + actualDay.strftime("%Y")
            + "/"
            + mes
            + " "
            + actualDay.strftime("%Y")
            + "/Daily_Schedule_"
            + actualDay.strftime("%d.%m.%Y")
            + ".xlsx"
        )

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ""
            valueShipping = ""

            if elem.comentarioShipping == None:
                elem.comentarioShipping = " "
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = " "
            if elem.shipping == True:
                valueShipping = "X"
            if elem.receiving == True:
                valueReceiving = "X"
            if elem.allOkShipping == True:
                valueSubmitShipping = "X"
            if elem.allOkReceiving == True:
                valueSubmitReceiving = "X"
            table += (
                '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
                + elem.line
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.site
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.item_number
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.description
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.to_complete
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueReceiving
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioReceiving)
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueShipping
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioShipping)
                + "</td></tr>"
            )

        table += (
            '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitReceiving
            + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitShipping
            + "</td><td></td></tr>"
        )
        table += "</tbody></table>"

        table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

        for elem3 in elementosDiaSeguinte.all():
            elem2 = dadosQAD.filter(
                wo_due_date=elem3.due_date, wo_part=elem3.item_number
            )
            for elem1 in elem2.all():
                if (
                    str(elem1["wo_qty_exp_complete"])[:-11]
                    != str(elem1["wo_qty_comp"])[:-11]
                ):
                    table += (
                        '<tr style="background-color: red">'
                        '<td style="padding:0 15px 0 15px;">'
                        + elem3.line
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem3.site
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_due_date"])
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem1["wo_part"]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_exp_complete"])[:-11]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_comp"])[:-11]
                        + "</td></tr>"
                    )
                else:
                    table += (
                        '<tr style="background-color: whitesmoke">'
                        '<td style="padding:0 15px 0 15px;">'
                        + elem3.line
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem3.site
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_due_date"])
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem1["wo_part"]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_exp_complete"])[:-11]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_comp"])[:-11]
                        + "</td></tr>"
                    )

        table += "</tbody></table>"

        table += "</br></br><b>Prodlines</b></br>"
        for prod in prodline.all():
            table += prod.nome + "</br>"

        table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

        subject, from_email, to = (
            "Wh Production " + actualDay.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        msg.send()


# def enviarEmailSchedule():
#     """  ["pmarti30@visteon.com",'npires2@visteon.com',
#                      'abrandao@visteon.com','sanasta1@visteon.com',
#                     'rsalgue2@visteon.com', 'nlopes8@visteon.com',
#                     'abilro1@visteon.com', 'jrodri80@visteon.com',
#                     'evenanc1@visteon.com'] """
#     import locale
#     locale.setlocale(locale.LC_ALL, "pt-PT")
#     timeNow = datetime.datetime.now()
#     timeEnd = timeNow.replace(hour=8, minute=10)
#     timeBefore = timeNow.replace(hour=7, minute=55)
#     timeEnd2 = timeNow.replace(hour=16, minute=10)
#     timeBefore2 = timeNow.replace(hour=15, minute=55)
#     timeEnd3 = timeNow.replace(hour=22, minute=10)
#     timeBefore3 = timeNow.replace(hour=21, minute=55)

#     dadosQAD = WoMstr.objects.values(
#         "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
#     )



#     prodline = Prodlines.objects
#     elementosDiaSeguinte = ProdlineTable.objects
#     actualDay = datetime.datetime.today() - datetime.timedelta(days=2)
#     nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
#     if nextDay.strftime("%A") == "Saturday":
#         nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
#     if nextDay.strftime("%A") == "Sunday":
#         nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
#     month = nextDay.strftime("%B")
#     mes = nextDay.strftime("%m")
#     fmt_num = nextDay.strftime("%d.%m.%Y")
#     mensagem = ""
#     print("DADOS PARA EMAIL->", "/",actualDay, " 2022-12-06")
#     #BUG FILESYS
#     textPath = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num}.xlsx"  # + nextDay.strftime(
#     #'%Y') + "/" + mes + " " + nextDay.strftime('%Y') + "/Daily_Schedule_" + nextDay.strftime(
#     #'%d.%m.%Y') + ".xlsx"
#     #print(textPath + " :TEXTPATH")
#     fmt_num1 = nextDay.strftime("%d_%m_%Y")
#     fmt_num2 = nextDay.strftime("%d-%m-%Y")
#     fmt_num3 = nextDay.strftime("%d/%m/%Y")
#     #BUG FILESYS
#     textPath1 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num1}.xlsx"
#     textPath2 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num2}.xlsx"
#     textPath3 = f"//PAVPD002/E_Proj/sharedir/MP&L/Schedule/{nextDay.year}/{month} {nextDay.year}/Daily_Schedule_{fmt_num3}.xlsx"
#     a = 0
#     def defineBook():
#         try:
#             workbook = openpyxl.load_workbook(textPath) 
#             a = 1
#         except:
#             try: 
#              workbook = openpyxl.load_workbook(textPath1)
#              a = 2
#             except:
#                 try:
#                     workbook = openpyxl.load_workbook(textPath2)
#                     a = 3
#                 except:
#                     workbook = openpyxl.load_workbook(textPath3)
#                     a = 4
#         print("WorkBook-->", workbook, "A->", a)
#         return workbook

#     # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/#Daily_Schedule_03.11.2021.xlsx'

#     table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
#     print("ENTROU")
#     if (
#         (timeNow > timeEnd and timeNow > timeBefore)
#         and (actualDay.strftime("%A") != "Saturday")
#         and (actualDay.strftime("%A") != "Sunday")
#     ):
#         table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
#         try:
#             workbook = defineBook()
#             sheet_num = ""
#             for sheet in workbook:
#                 # print(sheet.title)
#                 # print("CELL 1 --> ",sheet.cell(1,1).value)
#                 if sheet.cell(1, 1)._value == "Line":
#                     sheet_num = sheet.title
#                     print(sheet_num)
#             try:
#                 worksheet = workbook["Schedule Data"]
#                 for element in worksheet:
#                     for prod in prodline.all():
#                         if prod.nome == element[0].value:
#                             table += (
#                                 '<tr><td style="padding:0 15px 0 15px;">'
#                                 + element[0].value
#                                 + '</td><td style="padding:0 15px 0 15px;">'
#                                 + element[1].value
#                                 + '</td><td style="padding:0 15px 0 15px;">'
#                                 + str(element[2].value)
#                                 + '</td><td style="padding:0 15px 0 15px;">'
#                                 + element[3].value
#                                 + '</td><td style="padding:0 15px 0 15px;">'
#                                 + element[4].value
#                                 + '</td><td style="padding:0 15px 0 15px;">'
#                                 + str(element[5].value)
#                                 + "</td></tr>"
#                             )

#                 table += "</tbody></table>"
#                 table += "</br></br><b>Prodlines</b></br>"
#                 for prod in prodline.all():
#                     table += prod.nome + "</br>"

#                 table += (
#                     "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
#                 )

#                 subject, from_email, to = (
#                     "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#                     "noreply@visteon.com",
#                     ["pmarti30@visteon.com"] ,
                    
#                 ) 
#                 """ ,'npires2@visteon.com',
#                      'abrandao@visteon.com','sanasta1@visteon.com',
#                     'rsalgue2@visteon.com', 'nlopes8@visteon.com',
#                     'abilro1@visteon.com', 'jrodri80@visteon.com',
#                     'evenanc1@visteon.com' """
#                 msg = EmailMultiAlternatives(subject, table, from_email, to)
#                 msg.attach_alternative(table, "text/html")
#                 print( " A enviar mail")
#                 msg.send()
#             except:
#                 try:
#                     worksheet = workbook[sheet_num]
#                     for element in worksheet:
#                         for prod in prodline.all():
#                             if prod.nome == element[0].value:
#                                 table += (
#                                     '<tr><td style="padding:0 15px 0 15px;">'
#                                     + element[0].value
#                                     + '</td><td style="padding:0 15px 0 15px;">'
#                                     + element[1].value
#                                     + '</td><td style="padding:0 15px 0 15px;">'
#                                     + str(element[2].value)
#                                     + '</td><td style="padding:0 15px 0 15px;">'
#                                     + element[3].value
#                                     + '</td><td style="padding:0 15px 0 15px;">'
#                                     + element[4].value
#                                     + '</td><td style="padding:0 15px 0 15px;">'
#                                     + str(element[5].value)
#                                     + "</td></tr>"
#                                 )

#                     table += "</tbody></table>"
#                     table += "</br></br><b>Prodlines</b></br>"
#                     for prod in prodline.all():
#                         table += prod.nome + "</br>"

#                     table += (
#                         "</b></b></br><b>Ficheiro diário de Schedule em: </b>"
#                         + textPath
#                     )

#                     subject, from_email, to = (
#                         "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#                         "noreply@visteon.com",
#                        ["pmarti30@visteon.com"] ,
#                     )
#                     msg = EmailMultiAlternatives(subject, table, from_email, to)
#                     msg.attach_alternative(table, "text/html")
#                     msg.send()
#                 except KeyError:
#                     mensagem = (
#                         "</br>Worksheet name not found. Should be: Schedule Data or Sheet1 </br>"
#                         + textPath
#                         + ""
#                     )
#                     subject, from_email, to = (
#                         "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#                         "noreply@visteon.com",
#                         ["pmarti30@visteon.com"],
#                     )
#                     msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
#                     msg.attach_alternative(mensagem, "text/html")
#                     msg.send()
#         except FileNotFoundError:
#             mensagem = "</br> File not found. </br></br>" + textPath + ""
#             subject, from_email, to = (
#                 "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#                 "noreply@visteon.com",
#                 ["pmarti30@visteon.com"],
#             )
#             msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
#             msg.attach_alternative(mensagem, "text/html")
#             msg.send()

#     elif (
#         (timeNow < timeEnd2 and timeNow > timeBefore2)
#         and (not ProdlineTable.objects.filter(allOkReceiving=True).exists())
#         and (actualDay.strftime("%A") != "Saturday")
#         and (actualDay.strftime("%A") != "Sunday")
#         and (elementosDiaSeguinte.count() > 0)
#     ):

#         valueSubmitReceiving = ""
#         valueSubmitShipping = ""

#         #BUG FILESYS
#         textPath = (
#             "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
#             + actualDay.strftime("%Y")
#             + "/"
#             + mes
#             + " "
#             + actualDay.strftime("%Y")
#             + "/Daily_Schedule_"
#             + actualDay.strftime("%d.%m.%Y")
#             + ".xlsx"
#         )

#         for elem in elementosDiaSeguinte.all():
#             valueReceiving = ""
#             valueShipping = ""

#             if elem.comentarioShipping == None:
#                 elem.comentarioShipping = " "
#             if elem.comentarioReceiving == None:
#                 elem.comentarioReceiving = " "
#             if elem.shipping == True:
#                 valueShipping = "X"
#             if elem.receiving == True:
#                 valueReceiving = "X"
#             if elem.allOkShipping == True:
#                 valueSubmitShipping = "X"
#             if elem.allOkReceiving == True:
#                 valueSubmitReceiving = "X"
#             table += (
#                 '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
#                 + elem.line
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.site
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.due_date
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.item_number
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.description
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.to_complete
#                 + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
#                 + valueReceiving
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + str(elem.comentarioReceiving)
#                 + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
#                 + valueShipping
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + str(elem.comentarioShipping)
#                 + "</td></tr>"
#             )

#         table += (
#             '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
#             + valueSubmitReceiving
#             + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
#             + valueSubmitShipping
#             + "</td><td></td></tr>"
#         )
#         table += "</tbody></table>"
#         table += "</br></br><b>Prodlines</b></br>"
#         for prod in prodline.all():
#             table += prod.nome + "</br>"

#         table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

#         subject, from_email, to = (
#             "Wh Production " + actualDay.strftime("%d-%m-%Y"),
#             "noreply@visteon.com",
#             ["pmarti30@visteon.com"],
#         )
#         msg = EmailMultiAlternatives(subject, table, from_email, to)
#         msg.attach_alternative(table, "text/html")
#         msg.send()

#     elif (
#         (timeNow < timeEnd3 and timeNow > timeBefore3)
#         and (
#             not ProdlineTable.objects.filter(allOkReceiving=True).exists()
#             or not ProdlineTable.objects.filter(allOkShipping=True).exists()
#         )
#         and (actualDay.strftime("%A") != "Saturday")
#         and (actualDay.strftime("%A") != "Sunday")
#         and (elementosDiaSeguinte.count() > 0)
#     ):

#         valueSubmitReceiving = ""
#         valueSubmitShipping = ""

#         #BUG FILESYS
#         textPath = (
#             "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
#             + actualDay.strftime("%Y")
#             + "/"
#             + mes
#             + " "
#             + actualDay.strftime("%Y")
#             + "/Daily_Schedule_"
#             + actualDay.strftime("%d.%m.%Y")
#             + ".xlsx"
#         )

#         for elem in elementosDiaSeguinte.all():
#             valueReceiving = ""
#             valueShipping = ""

#             if elem.comentarioShipping == None:
#                 elem.comentarioShipping = " "
#             if elem.comentarioReceiving == None:
#                 elem.comentarioReceiving = " "
#             if elem.shipping == True:
#                 valueShipping = "X"
#             if elem.receiving == True:
#                 valueReceiving = "X"
#             if elem.allOkShipping == True:
#                 valueSubmitShipping = "X"
#             if elem.allOkReceiving == True:
#                 valueSubmitReceiving = "X"
#             table += (
#                 '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
#                 + elem.line
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.site
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.due_date
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.item_number
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.description
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + elem.to_complete
#                 + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
#                 + valueReceiving
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + str(elem.comentarioReceiving)
#                 + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
#                 + valueShipping
#                 + '</td><td style="padding:0 15px 0 15px;">'
#                 + str(elem.comentarioShipping)
#                 + "</td></tr>"
#             )

#         table += (
#             '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
#             + valueSubmitReceiving
#             + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
#             + valueSubmitShipping
#             + "</td><td></td></tr>"
#         )
#         table += "</tbody></table>"

#         table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

#         for elem3 in elementosDiaSeguinte.all():
#             elem2 = dadosQAD.filter(
#                 wo_due_date=elem3.due_date, wo_part=elem3.item_number
#             )
#             for elem1 in elem2.all():
#                 if (
#                     str(elem1["wo_qty_exp_complete"])[:-11]
#                     != str(elem1["wo_qty_comp"])[:-11]
#                 ):
#                     table += (
#                         '<tr style="background-color: red">'
#                         '<td style="padding:0 15px 0 15px;">'
#                         + elem3.line
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + elem3.site
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_due_date"])
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + elem1["wo_part"]
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_qty_exp_complete"])[:-11]
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_qty_comp"])[:-11]
#                         + "</td></tr>"
#                     )
#                 else:
#                     table += (
#                         '<tr style="background-color: whitesmoke">'
#                         '<td style="padding:0 15px 0 15px;">'
#                         + elem3.line
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + elem3.site
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_due_date"])
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + elem1["wo_part"]
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_qty_exp_complete"])[:-11]
#                         + '</td><td style="padding:0 15px 0 15px;">'
#                         + str(elem1["wo_qty_comp"])[:-11]
#                         + "</td></tr>"
#                     )

#         table += "</tbody></table>"

#         table += "</br></br><b>Prodlines</b></br>"
#         for prod in prodline.all():
#             table += prod.nome + "</br>"

#         table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

#         subject, from_email, to = (
#             "Wh Production " + actualDay.strftime("%d-%m-%Y"),
#             "noreply@visteon.com",
#             ["pmarti30@visteon.com"],
#         )
#         msg = EmailMultiAlternatives(subject, table, from_email, to)
#         msg.attach_alternative(table, "text/html")
#         msg.send()


def enviarEmailSchedule1():
    print("ENTROU")
    timeNow = datetime.datetime.now()
    timeEnd = timeNow.replace(hour=8, minute=10)
    timeBefore = timeNow.replace(hour=7, minute=55)
    timeEnd2 = timeNow.replace(hour=16, minute=10)
    timeBefore2 = timeNow.replace(hour=15, minute=55)
    timeEnd3 = timeNow.replace(hour=22, minute=10)
    timeBefore3 = timeNow.replace(hour=21, minute=55)
    dadosQAD = WoMstr.objects.values(
        "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
    )

    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today()
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime("%A") == "Saturday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime("%A") == "Sunday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    mes = nextDay.strftime("%m")
    if mes == "01":
        mes = "Janeiro"
    if mes == "02":
        mes = "Fevereiro"
    if mes == "03":
        mes = "Março"
    if mes == "04":
        mes = "Abril"
    if mes == "05":
        mes = "Maio"
    if mes == "06":
        mes = "Junho"
    if mes == "07":
        mes = "Julho"
    if mes == "08":
        mes = "Agosto"
    if mes == "09":
        mes = "Setembro"
    if mes == "10":
        mes = "Outubro"
    if mes == "11":
        mes = "Novembro"
    if mes == "12":
        mes = "Dezembro"
    mensagem = ""
    #BUG FILESYS
    textPath = (
        "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
        + nextDay.strftime("%Y")
        + "/"
        + mes
        + " "
        + nextDay.strftime("%Y")
        + "/Daily_Schedule_"
        + nextDay.strftime("%d.%m.%Y")
        + ".xlsx"
    )
    # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

    table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'

    if (
        (timeNow < timeEnd and timeNow > timeBefore)
        and (actualDay.strftime("%A") != "Saturday")
        and (actualDay.strftime("%A") != "Sunday")
    ):
        table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
        try:
            workbook = openpyxl.load_workbook(textPath)
            try:
                worksheet = workbook["Schedule Data"]
                for element in worksheet:
                    for prod in prodline.all():
                        if prod.nome == element[0].value:
                            table += (
                                '<tr><td style="padding:0 15px 0 15px;">'
                                + element[0].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[1].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + str(element[2].value)
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[3].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + element[4].value
                                + '</td><td style="padding:0 15px 0 15px;">'
                                + str(element[5].value)
                                + "</td></tr>"
                            )

                table += "</tbody></table>"
                table += "</br></br><b>Prodlines</b></br>"
                for prod in prodline.all():
                    table += prod.nome + "</br>"

                table += (
                    "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
                )

                subject, from_email, to = (
                    "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                    "noreply@visteon.com",
                    ["pmarti30@visteon.com"],
                )
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                msg.send()
            except:
                try:
                    worksheet = workbook["Sheet1"]
                    for element in worksheet:
                        for prod in prodline.all():
                            if prod.nome == element[0].value:
                                table += (
                                    '<tr><td style="padding:0 15px 0 15px;">'
                                    + element[0].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[1].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + str(element[2].value)
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[3].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + element[4].value
                                    + '</td><td style="padding:0 15px 0 15px;">'
                                    + str(element[5].value)
                                    + "</td></tr>"
                                )

                    table += "</tbody></table>"
                    table += "</br></br><b>Prodlines</b></br>"
                    for prod in prodline.all():
                        table += prod.nome + "</br>"

                    table += (
                        "</b></b></br><b>Ficheiro diário de Schedule em: </b>"
                        + textPath
                    )

                    subject, from_email, to = (
                        "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                        "noreply@visteon.com",
                        ["pmarti30@visteon.com"],
                    )
                    msg = EmailMultiAlternatives(subject, table, from_email, to)
                    msg.attach_alternative(table, "text/html")
                    msg.send()
                except KeyError:
                    mensagem = (
                        "</br>Worksheet name not found. Should be: Schedule Data or Sheet1 </br>"
                        + textPath
                        + ""
                    )
                    subject, from_email, to = (
                        "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                        "noreply@visteon.com",
                        ["pmarti30@visteon.com"],
                    )
                    msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
                    msg.attach_alternative(mensagem, "text/html")
                    msg.send()
        except FileNotFoundError:
            mensagem = "</br> File not found. </br></br>" + textPath + ""
            subject, from_email, to = (
                "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
            msg.attach_alternative(mensagem, "text/html")
            msg.send()

    elif (
        (timeNow < timeEnd2 and timeNow > timeBefore2)
        and (not ProdlineTable.objects.filter(allOkReceiving=True).exists())
        and (actualDay.strftime("%A") != "Saturday")
        and (actualDay.strftime("%A") != "Sunday")
        and (elementosDiaSeguinte.count() > 0)
    ):
        valueSubmitReceiving = ""
        valueSubmitShipping = ""

        #BUG FILESYS
        textPath = (
            "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
            + actualDay.strftime("%Y")
            + "/"
            + mes
            + " "
            + actualDay.strftime("%Y")
            + "/Daily_Schedule_"
            + actualDay.strftime("%d.%m.%Y")
            + ".xlsx"
        )

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ""
            valueShipping = ""

            if elem.comentarioShipping == None:
                elem.comentarioShipping = " "
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = " "
            if elem.shipping == True:
                valueShipping = "X"
            if elem.receiving == True:
                valueReceiving = "X"
            if elem.allOkShipping == True:
                valueSubmitShipping = "X"
            if elem.allOkReceiving == True:
                valueSubmitReceiving = "X"
            table += (
                '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
                + elem.line
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.site
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.item_number
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.description
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.to_complete
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueReceiving
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioReceiving)
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueShipping
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioShipping)
                + "</td></tr>"
            )

        table += (
            '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitReceiving
            + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitShipping
            + "</td><td></td></tr>"
        )
        table += "</tbody></table>"
        table += "</br></br><b>Prodlines</b></br>"
        for prod in prodline.all():
            table += prod.nome + "</br>"

        table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

        subject, from_email, to = (
            "Wh Production " + actualDay.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        print("ENVIOU MAIL DAS 4.30 ")
        msg.send()

    elif (
        (timeNow < timeEnd3 and timeNow > timeBefore3)
        and (
            not ProdlineTable.objects.filter(allOkReceiving=True).exists()
            or not ProdlineTable.objects.filter(allOkShipping=True).exists()
        )
        and (actualDay.strftime("%A") != "Saturday")
        and (actualDay.strftime("%A") != "Sunday")
        and (elementosDiaSeguinte.count() > 0)
    ):

        valueSubmitReceiving = ""
        valueSubmitShipping = ""
        #BUG FILESYS
        textPath = (
            "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
            + actualDay.strftime("%Y")
            + "/"
            + mes
            + " "
            + actualDay.strftime("%Y")
            + "/Daily_Schedule_"
            + actualDay.strftime("%d.%m.%Y")
            + ".xlsx"
        )

        for elem in elementosDiaSeguinte.all():
            valueReceiving = ""
            valueShipping = ""

            if elem.comentarioShipping == None:
                elem.comentarioShipping = " "
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = " "
            if elem.shipping == True:
                valueShipping = "X"
            if elem.receiving == True:
                valueReceiving = "X"
            if elem.allOkShipping == True:
                valueSubmitShipping = "X"
            if elem.allOkReceiving == True:
                valueSubmitReceiving = "X"
            table += (
                '<tr style="background-color: red"><td style="padding:0 15px 0 15px; background-color: red">'
                + elem.line
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.site
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.item_number
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.description
                + '</td><td style="padding:0 15px 0 15px;">'
                + elem.to_complete
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueReceiving
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioReceiving)
                + '</td><td style="text-align: center; padding:0 15px 0 15px;">'
                + valueShipping
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(elem.comentarioShipping)
                + "</td></tr>"
            )

        table += (
            '<tr><td colspan="6" style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitReceiving
            + '</td><td style="text-align: right"><b>Submit</b></td><td style="text-align: center; padding:0 15px 0 15px;">'
            + valueSubmitShipping
            + "</td><td></td></tr>"
        )
        table += "</tbody></table>"

        table += '</b></br></br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Qty to Complete</th><th>Actual Qty Completed</th></tr></thead><tbody>'

        for elem3 in elementosDiaSeguinte.all():
            elem2 = dadosQAD.filter(
                wo_due_date=elem3.due_date, wo_part=elem3.item_number
            )
            for elem1 in elem2.all():
                if (
                    str(elem1["wo_qty_exp_complete"])[:-11]
                    != str(elem1["wo_qty_comp"])[:-11]
                ):
                    table += (
                        '<tr style="background-color: red">'
                        '<td style="padding:0 15px 0 15px;">'
                        + elem3.line
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem3.site
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_due_date"])
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem1["wo_part"]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_exp_complete"])[:-11]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_comp"])[:-11]
                        + "</td></tr>"
                    )
                else:
                    table += (
                        '<tr style="background-color: whitesmoke">'
                        '<td style="padding:0 15px 0 15px;">'
                        + elem3.line
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem3.site
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_due_date"])
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem1["wo_part"]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_exp_complete"])[:-11]
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem1["wo_qty_comp"])[:-11]
                        + "</td></tr>"
                    )

        table += "</tbody></table>"

        table += "</br></br><b>Prodlines</b></br>"
        for prod in prodline.all():
            table += prod.nome + "</br>"

        table += "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath

        subject, from_email, to = (
            "Wh Production " + actualDay.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )

        #'npires2@visteon.com', 'abrandao@visteon.com',
        # 'aroque1@visteon.com', 'sanasta1@visteon.com',
        # 'rsalgue2@visteon.com', 'nlopes8@visteon.com',
        # 'abilro1@visteon.com', 'jrodri80@visteon.com',
        # 'evenanc1@visteon.com'

        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        print("ENVIOU MAIL das 22.10")
        msg.send()


def updateSchedule():
    prodline = Prodlines.objects

    day = datetime.datetime.today()

    if day.strftime("%A") != "Saturday" and day.strftime("%A") != "Sunday":
        mes = day.strftime("%m")
        if mes == "01":
            mes = "Janeiro"
        if mes == "02":
            mes = "Fevereiro"
        if mes == "03":
            mes = "Março"
        if mes == "04":
            mes = "Abril"
        if mes == "05":
            mes = "Maio"
        if mes == "06":
            mes = "Junho"
        if mes == "07":
            mes = "Julho"
        if mes == "08":
            mes = "Agosto"
        if mes == "09":
            mes = "Setembro"
        if mes == "10":
            mes = "Outubro"
        if mes == "11":
            mes = "Novembro"
        if mes == "12":
            mes = "Dezembro"
        #BUG FILESYS
        textPath = (
            "//PAVPD002/E_Proj/sharedir/MP&L/Schedule/"
            + day.strftime("%Y")
            + "/"
            + mes
            + " "
            + day.strftime("%Y")
            + "/Daily_Schedule_"
            + day.strftime("%d.%m.%Y")
            + ".xlsx"
        )

        # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/Daily_Schedule_03.11.2021.xlsx'

        # guarda elementos do dia anterior na tabela histórico
        # BUG elem tá DAY/Month/year
        elementosDiaAnterior = ProdlineTable.objects
        for elem in elementosDiaAnterior.all():
            if elem.comentarioShipping == None:
                elem.comentarioShipping = ""
            if elem.comentarioReceiving == None:
                elem.comentarioReceiving = ""
            novoElem = ProdlineTableHistory(
                None,
                elem.day,
                elem.line,
                elem.site,
                elem.due_date,
                elem.item_number,
                elem.description,
                elem.to_complete,
                elem.receiving,
                elem.allOkReceiving,
                elem.comentarioReceiving,
                elem.shipping,
                elem.allOkShipping,
                elem.comentarioShipping,
            )
            novoElem.save()

        # limpa a tabela provisória

        ProdlineTable.objects.filter().delete()
        """ --------------------------///////////////----------------------------------------------------///////////////-------------------------- """
        #Cria um ficheiro Excell é preciso confirmar se os dados que vão para a bd são os mesmos que vão para o ficheiro
        # Esta func está a guardar em 2 bds, 
        """ --------------------------///////////////----------------------------------------------------///////////////-------------------------- """
        #penso que ele faz backup na bd e num ficheiro que se encontra no path
        workbook = openpyxl.load_workbook(textPath)
        try:
            worksheet = workbook["Schedule Data"]
            for element in worksheet:
                for prod in prodline.all():
                    if prod.nome == element[0].value:
                        novo = ProdlineTable(
                            None,
                            day.strftime("%d-%m-%Y"),
                            element[0].value,
                            element[1].value,
                            str(element[2].value),
                            element[3].value,
                            element[4].value,
                            str(element[5].value),
                            None,
                            None,
                            None,
                            None,
                            None,
                            None,
                        )
                        novo.save()
        except:
            worksheet = workbook["Sheet1"]
            for element in worksheet:
                for prod in prodline.all():
                    if prod.nome == element[0].value:
                        novo = ProdlineTable(
                            None,
                            day.strftime("%d-%m-%Y"),
                            element[0].value,
                            element[1].value,
                            str(element[2].value),
                            element[3].value,
                            element[4].value,
                            str(element[5].value),
                            None,
                            None,
                            None,
                            None,
                            None,
                            None,
                        )
                        novo.save()
    print("Confirmar update: ->  ", novo)


def updateLineRequestDia1():
    lineRequest = LineRequestFinished.objects

    firstDayPrevMonth = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)
    ).replace(day=1)
    prev2MonthName = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%B")
    prev2MonthYear = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%Y")



    row2 = 0
    col2 = 0


    row2 += 1
    # datetime3 = datetime.datetime.strptime(elem.horaPedido, '%Y-%m-%d %H:%M:%S')
    for elem in lineRequest.all():
        # DÁ ERRO AQUI pois o formato que ele pede (firstDayPrevMonth) tem o formato '%Y-%m-%d %H:%M:%S
        # o datetime2 vem com %D %M %Y %H %M %S falta fazer formatação
        # O ELEM È QUE ESTÀ A VIR COM A FORMATAÇÂO ERRADA

        try:
            datetime2 = datetime.datetime.strptime(elem.horaPedido, "%d-%m-%Y %H:%M:%S")
        except ValueError:

            continue

        if datetime2 < firstDayPrevMonth:
            novoRequest = LineRequestFinishedHistory(
                None,
                elem.horaPedido,
                elem.partNumber,
                elem.linha,
                elem.requisitante,
                elem.receiver,
                elem.justificacao,
                elem.comentario,
            )
            novoRequest.save()

     
            row2 += 1





def updatePortariaDia1():
    portaria = Gateway.objects

    fifteenDayPrevMonth = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)
    ).replace(day=15)
    prev2MonthName = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%B")
    prev2MonthYear = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%Y")

    wbPortaria = Workbook()
   

    row = 0
    col = 0


    row += 1
    # BUG formatação OK
    for elem in portaria.all():

        datetime2 = datetime.datetime.strptime(elem.dataHoraChegada, "%Y-%m-%d %H:%M")

        if datetime2 < fifteenDayPrevMonth:
            abandono = ""
            if elem.abandono == "true":
                abandono = "X"

            eleme = GatewayBackup(
                None,
                elem.dataHoraChegada,
                elem.condutor,
                elem.ident,
                elem.contacto,
                elem.empresa,
                elem.primeiraMatricula,
                elem.segundaMatricula,
                elem.cargaDescarga,
                elem.doca,
                elem.destinoCarga,
                elem.tipoViatura,
                elem.dataHoraEntrada,
                abandono,
                elem.comentEntrada,
                elem.dataHoraSaida,
                elem.comentSaida,
            )
            eleme.save()

           

def updatePortariaDia15():
    portaria = Gateway.objects
    print("PORTARIA -->  ", portaria.all())
    firstDayCurrMonth = datetime.datetime.today().replace(day=1)
    prev2MonthName = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%B")
    prev2MonthYear = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%Y")


    row = 0
    col = 0


    row += 1
    # BUG data desformatada
    for elem in portaria.all():
        datetime2 = datetime.datetime.strptime(elem.dataHoraChegada, "%Y-%m-%d %H:%M")
        if datetime2 < firstDayCurrMonth:
            abandono = ""
            if elem.abandono == "true":
                abandono = "X"

            eleme = GatewayBackup(
                None,
                elem.dataHoraChegada,
                elem.condutor,
                elem.ident,
                elem.contacto,
                elem.empresa,
                elem.primeiraMatricula,
                elem.segundaMatricula,
                elem.cargaDescarga,
                elem.doca,
                elem.destinoCarga,
                elem.tipoViatura,
                elem.dataHoraEntrada,
                abandono,
                elem.comentEntrada,
                elem.dataHoraSaida,
                elem.comentSaida,
            )
            eleme.save()


def updateProductionDia1():
    prodline = ProdlineTableHistory.objects

    firstDayPrevMonth = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=10)
    ).replace(day=1)
    prev2MonthName = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%B")
    prev2MonthYear = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%Y")


    row = 0
    col = 0



    row += 1
    for elem in prodline.all():
        # BUG
        try:
            datetime2 = datetime.datetime.strptime(elem.day, "%d-%m-%Y %H:%M:%S")
        except ValueError:
            # Data mal formatada
            continue

        # datetime2 = datetime.datetime.strptime(elem.day, '%d-%m-%Y')
        if datetime2 < firstDayPrevMonth:
            eleme = ProdlineHistoryAllTime(
                None,
                elem.day,
                elem.line,
                elem.site,
                elem.due_date,
                elem.item_number,
                elem.description,
                elem.to_complete,
                elem.receiving,
                elem.allOkReceiving,
                elem.comentarioReceiving,
                elem.shipping,
                elem.allOkShipping,
                elem.comentarioShipping,
            )
            eleme.save()




def updateICDRDia1():
    icdr = ICDR.objects

    firstDayPrevMonth = datetime.datetime.today()
    prev2MonthName = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%B")
    prev2MonthYear = (
        datetime.datetime.today().replace(day=1) - datetime.timedelta(days=32)
    ).strftime("%Y")

    row = 0
    col = 0

    row += 1
    # BUG confirmar formatação
    for elem in icdr.all():
        datetime2 = datetime.datetime.strptime(elem.aberturaICDR, "%Y-%m-%d %H:%M:%S")
        if datetime2 < firstDayPrevMonth:

            elemeBackup = ICDRBackup()
            elemeBackup.aberturaICDR = elem.aberturaICDR
            elemeBackup.ageing = elem.ageing
            elemeBackup.nAno = elem.nAno
            elemeBackup.fornecedor = elem.fornecedor
            elemeBackup.partnumber = elem.partnumber
            elemeBackup.quantidade = elem.quantidade
            elemeBackup.tipo = elem.tipo
            elemeBackup.simNao = elem.simNao
            elemeBackup.responsavel = elem.responsavel
            elemeBackup.departamento = elem.departamento
            elemeBackup.comentarioFecho = elem.comentarioFecho
            elemeBackup.unCost = elem.unCost
            elemeBackup.totalCost = elem.totalCost
            elemeBackup.rctUnpCheck = elem.rctUnpCheck
            elemeBackup.cycleCountCheck = elem.cycleCountCheck
            elemeBackup.consumption = elem.consumption
            elemeBackup.po = elem.po
            elemeBackup.comentarioFechoICDR = elem.comentarioFechoICDR
            elemeBackup.date = elem.date
            elemeBackup.cycleCount = elem.cycleCount
            elemeBackup.dataCycleCount = elem.dataCycleCount
            elemeBackup.auditCheck = elem.auditCheck
            elemeBackup.save()



def configurationsEmailDiario(request):
     return render(request, "crossdocking/configurations.html")






""" 
def def uploadFicheiroDiarioBD(request): 
    #tens de criar uma linha na tabela ("dadosEmailDiario") onde guardas pelo menos a data
    ...
 """


#def uploadFicheiroDiarioManual(request): #A ir buscar o ficheiro e a guardar na bd
#   from datetime import datetime
#   from datetime import date
#   import os
#   meses = [ "Janeiro" , "Fevereiro" , "Março" , "Abril" , "Maio" , "Junho" , "Julho" , "Agosto" , 
#   "Setembro" , "Outubro" , "Novembro" , "Dezembro"]
#   datetime_object = datetime.now()
#por o caminho da pasta de acordo com o mês e o ano (era fixe se desse para ficar uma tree como tens no sharedir, está organizado por ANO/Mes)
#   ficheiro = request.FILES.get("ficheiro") 
#   print("REQUEST--->",ficheiro, request.FILES["ficheiro"])
#   #workbook = openpyxl.load_workbook(ficheiro[0])
#   caminho = f"crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\"
#   print("LINK", ficheiro.name)
#   print("mes",meses[datetime_object.month -1 ])
#   if ficheiro is not None:
#default_storage.save(caminho+ficheiro.name, ficheiro)
#       print("LINK", ficheiro.name)
#       #Caso não exista a pasta com o ano atual nesse diretorio, ele cria um novo com o ano atual
#   if not os.path.exists(f"media\\crossdocking\\Schedule\\{datetime_object.year}\\"):
#       print("Não existe pasta para este ano")
#       os.makedirs(f"media\\crossdocking\\Schedule\\{datetime_object.year}\\")
#    
#   else :
#       print("Já existe esta pasta deste ano vai verificar a do mes")
#   if not os.path.exists(f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}"):
#       os.makedirs(f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}")
#       print("Não tinha diretoria mas já criou uma para este mês")
#       default_storage.save(caminho + ficheiro.name, ficheiro)
#   else:
#       default_storage.save(caminho + ficheiro.name , ficheiro)
#       print("Tinha diretoria mas já criou uma para este mês")
#   bd_ficheiro = EmailDiarioInformacao(ficheiro = ficheiro, data_guardado = date.today().strftime('%Y-%m-%d'))
#   #bd_ficheiro.save()
#   return redirect("crossdocking:configurationsEmailDiario")



def uploadFicheiroDiarioAutomatico(request): 
    from datetime import datetime
    from datetime import date
    import os
    from openpyxl import load_workbook

    datetime_object = datetime.now()
    timestampField = datetime_object.strftime("%Y-%m-%d %H:%M:%S")
    file = request.FILES.get("ficheiro") 

    workbook = load_workbook(file.open("rb"), data_only = True)
    
    for x in workbook.sheetnames:
        sheet = workbook.get_sheet_by_name(x)

        
        
        if len(list(sheet.columns)) != 7:
            print("não pode ser este sheet!",x)
            
        else:
            print("já encontrou a sheet")
            for index, i in enumerate(sheet):
                if index == 0: 
                    continue
                lineField, siteField,  dueDateField, itemNumberField, descriptionField, toCompleteField, qtyCompletedField = (
                    i[0].value,
                    i[1].value,
                    i[2].value,
                    i[3].value,
                    i[4].value,
                    i[5].value,
                    i[6].value,
                )
                db = EmailDiarioSchedule(
                    line=lineField,
                    site=siteField,
                    due_date=dueDateField,
                    item_number=itemNumberField,
                    to_complete=toCompleteField,
                    description=descriptionField,
                    qty_completed=qtyCompletedField,
                    timeStamp = timestampField,
                )
                db.save()
                print("guardou na bd")
                print("|DADOS->",lineField, siteField, dueDateField, itemNumberField,descriptionField, toCompleteField, qtyCompletedField, timestampField)
    
    return redirect("crossdocking:configurationsEmailDiario")





#em principio não vai ser preciso
#ef sendFicheiroDiarioManual(request): 
#   import datetime
#   from datetime import datetime as dt
#   import os
#   meses = [ "Janeiro" , "Fevereiro" , "Março" , "Abril" , "Maio" , "Junho" , "Julho" , "Agosto" , 
#   "Setembro" , "Outubro" , "Novembro" , "Dezembro"]
#   datetime_object = dt.now()
#   caminho ='media\crossdocking\Schedule\\' #por o caminho da pasta de acordo com o mês e o ano (era fixe se desse para ficar uma tree como tens no sharedir, está organizado por ANO/Mes)
#  
#   #se não conseguir ir buscar os dados à bd vai pelo ficheiro
#   
#   #ver como são os nomes dos ficheiros e pesquisar por data e não por nome e não precisas de ir á bd
#   local = f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{datetime_object.day}.{datetime_object.month}.{datetime_object.year}.xlsx"
#   
#   """  ["pmarti30@visteon.com",'npires2@visteon.com',
#                    'abrandao@visteon.com','sanasta1@visteon.com',
#                   'rsalgue2@visteon.com', 'nlopes8@visteon.com',
#                   'abilro1@visteon.com', 'jrodri80@visteon.com',
#                   'evenanc1@visteon.com'] """
#   """  ["pmarti30@visteon.com",'npires2@visteon.com',
#                    'abrandao@visteon.com','sanasta1@visteon.com',
#                   'rsalgue2@visteon.com', 'nlopes8@visteon.com',
#                   'abilro1@visteon.com', 'jrodri80@visteon.com',
#                   'evenanc1@visteon.com'] """
#   import locale
#   locale.setlocale(locale.LC_ALL, "pt-PT")
#   timeNow = datetime.datetime.now()
#   timeEnd = timeNow.replace(hour=8, minute=10)
#   timeBefore = timeNow.replace(hour=7, minute=55)
#   timeEnd2 = timeNow.replace(hour=16, minute=10)
#   timeBefore2 = timeNow.replace(hour=15, minute=55)
#   timeEnd3 = timeNow.replace(hour=22, minute=10)
#   timeBefore3 = timeNow.replace(hour=21, minute=55)
#
#   dadosQAD = WoMstr.objects.values(
#       "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
#   )
#
#   prodline = Prodlines.objects
#   elementosDiaSeguinte = ProdlineTable.objects
#   actualDay = datetime.datetime.today() - datetime.timedelta(days=2)
#   nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
#   if nextDay.strftime("%A") == "Saturday":
#       nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
#   if nextDay.strftime("%A") == "Sunday":
#       nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
#   month = nextDay.strftime("%B")
#   mes = nextDay.strftime("%m")
#   fmt_num = nextDay.strftime("%d.%m.%Y")
#   mensagem = ""
#  
#   #'%Y') + "/" + mes + " " + nextDay.strftime('%Y') + "/Daily_Schedule_" + nextDay.strftime(
#   #'%d.%m.%Y') + ".xlsx"
#   fmt_num1 = nextDay.strftime("%d_%m_%Y")
#   fmt_num2 = nextDay.strftime("%d-%m-%Y")
#   fmt_num3 = nextDay.strftime("%d/%m/%Y")
#   local = f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{datetime_object.day+1}.{datetime_object.month}.{datetime_object.year}.xlsx"
#   textPath = local
#   #BUG FILESYS
#
#
#   # textPath = '//PAVPD002/E_Proj/sharedir/MP&L/Schedule/2021/Novembro 2021/#Daily_Schedule_03.11.2021.xlsx'
#
#   table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
#   print("ENTROU")
#   
#   if (
#       (timeNow > timeEnd and timeNow > timeBefore)
#       and (actualDay.strftime("%A") != "Saturday")
#       and (actualDay.strftime("%A") != "Sunday")
#    
#   ):
#       table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
#       dados_bd_schedule = EmailDiarioSchedule_view.objects
#       print("VAR--->>", dados_bd_schedule.all(), type(dados_bd_schedule))
#       if dados_bd_schedule.all():
#           
#           for element in dados_bd_schedule.all(): #in View db
#               print("elemento --)",element.line)
#               for prod in prodline.all():
#                   if prod.nome == element.line:
#                       table += (
#                           '<tr><td style="padding:0 15px 0 15px;">'
#                           + element.line
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                           + element.site
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                           + str(element.due_date)
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                           + element.item_number
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                           + element.description
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                           + str(element.to_complete) 
#                           + '</td><td style="padding:0 15px 0 15px;">'
#                       
#                       )
#                       """ + str(element.qty_completed)
#                           + "</td></tr>" """
#           table += "</tbody></table>"
#           table += "</br></br><b>Prodlines</b></br>"
#           for prod in prodline.all():
#               table += prod.nome + "</br>"
#
#           table += (
#               "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + textPath
#           )
#
#           subject, from_email, to = (
#               "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#               "noreply@visteon.com",
#               ["pmarti30@visteon.com "],
#               
#           ) 
#           """ ,'npires2@visteon.com',
#                   'abrandao@visteon.com','sanasta1@visteon.com',
#               'rsalgue2@visteon.com', 'nlopes8@visteon.com',
#               'abilro1@visteon.com', 'jrodri80@visteon.com',
#               'evenanc1@visteon.com' """
#           msg = EmailMultiAlternatives(subject, table, from_email, to)
#           msg.attach_alternative(table, "text/html")
#           print(" Mensagem a ser enviada no mail")
#           msg.send()
#       else:
#           print("Estas no else")
#           mensagem = "</br> Data not found. </br></br>" + "W:\sharedir\MP&L\Schedule" + ""
#           subject, from_email, to = (
#               "Wh Production " + nextDay.strftime("%d-%m-%Y"),
#               "noreply@visteon.com",
#               ["pmarti30@visteon.com"],   
#           )
#           msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
#           msg.attach_alternative(mensagem, "text/html")
#           msg.send()
#
#   return redirect("crossdocking:configurationsEmailDiario")
#   



""" ------------------------///////////-----------------------------//////////////////////  """
""" ------------------------///////////-----------------------------//////////////////////   """
def sendFicheiroDiarioAutomatico(*request): 
    #Esta func é para correr todos os dias às 8 caso não encontre dados na bd manda um email a dizer que não deu para encontrar dados para o dia seguinte 
    #Adicionar a opção que caso não encontre dados da bd que os procure pelo file (SQUE NÂO É PRECISO).
    #Ver como seria feito nos fds ( com a opção de por ele a corre apenas em dias uteis, de segunda a sexta)
    #Ver como seria com os dados do QAD a virem corretamente
    
    import datetime
    from datetime import datetime as dt
    import os
    meses = [ "January" , "February " , "March " ,  "April " , "May " , "June " , "July " , "August " , 
    "September" , "October " , "November " , "December"]
    datetime_object = dt.now()
    caminho ='media\crossdocking\Schedule\\' #por o caminho da pasta de acordo com o mês e o ano (era fixe se desse para ficar uma tree como tens no sharedir, está organizado por ANO/Mes)
    
    local = f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{datetime_object.day}.{datetime_object.month}.{datetime_object.year}.xlsx"
    
    """  ["pmarti30@visteon.com",'npires2@visteon.com',
                     'abrandao@visteon.com','sanasta1@visteon.com',
                    'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                    'abilro1@visteon.com', 'jrodri80@visteon.com',
                    'evenanc1@visteon.com'] """
    """  ["pmarti30@visteon.com",'npires2@visteon.com',
                     'abrandao@visteon.com','sanasta1@visteon.com',
                    'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                    'abilro1@visteon.com', 'jrodri80@visteon.com',
                    'evenanc1@visteon.com'] """
    import locale
    timeNow = datetime.datetime.now()
    
    print("FUNC EMAIL AUTOMATICO")
    dadosQAD = WoMstr.objects.values(
        "wo_due_date", "wo_part", "wo_qty_exp_complete", "wo_qty_comp"
    )
    elementos_data_base = EmailDiarioSchedule_view.objects
    prodline = Prodlines.objects
    elementosDiaSeguinte = ProdlineTable.objects
    actualDay = datetime.datetime.today() - datetime.timedelta(days=2)
    nextDay = datetime.datetime.today() + datetime.timedelta(days=1)
    if nextDay.strftime("%A") == "Saturday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=3)
    if nextDay.strftime("%A") == "Sunday":
        nextDay = datetime.datetime.today() + datetime.timedelta(days=2)
    month = nextDay.strftime("%B")
    mes = nextDay.strftime("%m")
    fmt_num = nextDay.strftime("%d.%m.%Y")
    mensagem = ""
   
    #'%Y') + "/" + mes + " " + nextDay.strftime('%Y') + "/Daily_Schedule_" + nextDay.strftime(
    #'%d.%m.%Y') + ".xlsx"
    

     
    if nextDay.strftime("%A") == "Saturday" or nextDay.strftime("%A") != "Sunday":
        local = f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{datetime_object.day+1}.{datetime_object.month}.{datetime_object.year}.xlsx"
        textPath = local

        #table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th><th>Receiving</th><th>Comentário Receiving</th><th>Shipping</th><th>Comentário Shipping</th></tr></thead><tbody>'
        
        table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>Line</th><th>Site</th><th>Due Date</th><th>Item Number</th><th>Description</th><th>To Complete</th></tr></thead><tbody>'
                
        print("TEste", elementos_data_base.count())
        if elementos_data_base.count() !=0:
            for element in elementos_data_base.all():

                for prod in prodline.all():
                    if prod.nome == element.line:
                        table += (
                            '<tr><td style="padding:0 15px 0 15px;">'
                        + element.line # line
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + element.site # site
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(element.due_date) # due_date
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + element.item_number # item_number
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + element.description # to_complete
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(element.to_complete) # description
                        + "</td></tr>"
                        )

            table += "</tbody></table>"
            table += "</br></br><b>Prodlines</b></br>"
            for prod in prodline.all():
                table += prod.nome + "</br>"

            table += (
                "</b></b></br><b>Ficheiro diário de Schedule em: </b>" + f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{nextDay.strftime('%d')}.{datetime_object.month}.{datetime_object.year}.xlsx" + ""
            )

            subject, from_email, to = (
                "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                "noreply@visteon.com",
                ["pmarti30@visteon.com",'npires2@visteon.com',
                    'abrandao@visteon.com','sanasta1@visteon.com',
                'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                'abilro1@visteon.com', 'jrodri80@visteon.com',
                'evenanc1@visteon.com'],
                
            ) 
            """ ,'npires2@visteon.com',
                    'abrandao@visteon.com','sanasta1@visteon.com',
                'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                'abilro1@visteon.com', 'jrodri80@visteon.com',
                'evenanc1@visteon.com' """
            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            print("Mensagem a ser enviada no mail", nextDay.strftime("%d-%m-%Y"))
            msg.send()
            return redirect("crossdocking:configurationsEmailDiario")
        
        else: #A view não tem dados, quer dizer que ainda não foi importado o ficheiro para o dia seguinte
            
            mensagem = "</br> No data found in data base. </br></br>" + f"media\\crossdocking\\Schedule\\{datetime_object.year}\\{meses[datetime_object.month -1 ]} {datetime_object.year}\\" +  f"Daily_Schedule_{nextDay}.{datetime_object.month}.{datetime_object.year}.xlsx" + ""
            subject, from_email, to = (
                "Wh Production " + nextDay.strftime("%d-%m-%Y"),
                "noreply@visteon.com",
                 ["pmarti30@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, mensagem, from_email, to)
            msg.attach_alternative(mensagem, "text/html")
            #msg.send()
        print("Não é para madar email, é fim-de-seman")
        return redirect("crossdocking:configurationsEmailDiario")
   
   #function for login 


def render_karbox(request):
     return render(request,"crossdocking/karbox.html")

#FALTA ADICIONAR ao ficheiro URLS.py

def add_karbox(request):
    ref = Karbox()
    ref.pn
    return render(request,"crossdocking/karbox.html")

def delete_karbox(request):
    Karbox(pn = request.POST["pn"], descricao = request.POST["descricao"], qty = request.POST["qty"])
    return render(request,"crossdocking/karbox.html")

# SUB ITEMS OPERATIONS
def add_subItem(request):
    KarboxSubItem(pn = request.POST["pn"], descricao = request.POST["descricao"], qty = request.POST["qty"]).save()
    return render(request,"crossdocking/karbox.html")

def delete_subItem(request):
    Karbox(pn = request.POST["pn"], descricao = request.POST["descricao"], qty = request.POST["qty"])
    return render(request,"crossdocking/karbox.html")

def importKarbox(request):
    return render(request,"crossdocking/karbox.html")