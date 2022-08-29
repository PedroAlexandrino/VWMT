from pickle import NONE
from sys import getsizeof
from time import CLOCK_UPTIME
import tablib
import io
import math
import typing as t
import aspose.slides as slides
import win32com.client

from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMultiAlternatives
from django.http import JsonResponse, FileResponse
from django.shortcuts import render, redirect
from django.db.models import Q
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
from django.conf import settings
from django.db import models
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseNotFound,
    Http404,
)
from django.http import HttpResponseRedirect
from django.core.files.storage import FileSystemStorage
from django.core.files.storage import default_storage
from contextlib import contextmanager

from pManagement.models import *
from vware.models import (
    Produtos,
    StockPackage,
    SupplyPackage,
    TipoEmbalagem,
    ClientesOEM,
    ClienteProduto,
)

from .models import PackageType


# Mostra a tabela dos Requests
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="requests")).exists()
)
def requests(request):
    clientes = ClientesOEM.objects.all()
    produtos = Produtos.objects.all()
    tipoEmbalagem = TipoEmbalagem.objects.all()

    lista = []
    listaFinal = []
    produtoOld = ""

    for cliente in clientes:
        for produto in produtos:
            for embalagem in tipoEmbalagem:
                if (
                    cliente.oem == produto.cliente.oem
                    and embalagem.produto.nome == produto.nome
                    and embalagem.produto.cliente.oem == produto.cliente.oem
                ):
                    lista_objecto = {
                        "cliente": cliente.oem,
                        "produto": produto.nome,
                        "tipo": embalagem.nome,
                        "tempo": embalagem.tempoSupply,
                        "link": embalagem.link,
                        "quantidade": embalagem.quantidade,
                    }
                    lista.append(lista_objecto)
            produtoOld = produto

    for element in lista:
        if element in listaFinal:
            continue
        else:
            listaFinal.append(element)

    return render(
        request,
        "pManagement/requests.html",
        {
            "clientes": clientes,
            "produtos": produtos,
            "tipoEmbalagem": tipoEmbalagem,
            "data": listaFinal,
        },
    )


# Mostra a tabela dos Expendable
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(
        Q(name="pManagement") | Q(name="ExtendablePackage")
    ).exists()
)
def expendablePacking(request):
    partnumbers = PartNumbersExpendable.objects
    return render(
        request, "pManagement/expendablePackage.html", {"partnumbers": partnumbers}
    )


# Render de dados para a pag stpckPackage, PARA ADICIONAR AO SERVER
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def stockPackage(request):
    partnumbers = StockPackage.objects.all()
    #ver se dá para fazer a query e ordenar os dados por ordem decrescente
    # onde estás a limitar as querys

    """  p = Paginator(partnumbers, 20)
    page_num = request.GET.get("page", 1)
    page_content = p.page(page_num) """
    # return render(request, "pManagement/stockPackage.html",{'items':partnumbers})
    return render(
        request,
        "pManagement/stockPackage.html",
        {
            "items": partnumbers,
        },
    )

    return render(
        request,
        "pManagement/stockPackage1.html",
        {
            "items": page_content,
            "page_max": math.ceil(len(partnumbers) / 20),
            "current_page": page_num,
        },
    )


# Criação de uma linha na pag StockPackage
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def createStockPackage(request):
    if request.method == "POST":
        print("rESQUEST", request.POST)
        partNumber = request.POST.get("partNumberNovo")
        descricao = request.POST.get("descricaoNovo")
        link = request.FILES.get("linkNovo")
        comentario = request.POST.get("comentarioNovo")
        quantidade = request.POST.get("novoStock", 0) # BUG AQUI
        print("QTY",quantidade,type(quantidade))
        inventario = request.POST.get("quantidadeInventarioNovo", 0)
        tipo = PackageType(request.POST["tipo"])

        if not len(quantidade):
            quantidade=0
        if not len(inventario):
            inventario=0

        if not StockPackage.objects.filter(pn=partNumber).exists():
            newPartNumber = StockPackage()
            newPartNumber.pn = partNumber
            newPartNumber.descricao = descricao

            if link:
                # mudar para pasta do StockPAckage
                #C:\Users\PMARTI30\Desktop\visteon\media\StockPackage
                
                newPartNumber.link = default_storage.save(link.name, link)

            newPartNumber.comentario = comentario
            newPartNumber.quantidade = quantidade
            newPartNumber.inventario = inventario
            newPartNumber.tipo = tipo.value
            newPartNumber.save()

    return redirect("pManagement:stockPackage")


# Edita uma linha da table table(html) e faz edição na db
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def updateStockPackage(request):
    print("REquest do update StockP", request.POST)
    if request.method == "POST":
        rowId = request.POST.get("rowIdUpdate")
        edit_partNumber = request.POST.get("partNumberEdit")
        edit_description = request.POST.get("descricaoEdit")

        edit_link = request.FILES.get("linkEdit")
        edit_Links = request.POST.get("linkEdits")

        edit_comentario = request.POST.get("commentEdit")
        edit_quantidade = request.POST.get("stockEdit")
        edit_inv = request.POST.get("invEdit")
        #edit_tipo = request.POST.get("tipoEdit")
        edit_tipo = request.POST["tipoEdit"]

        print("MAMBOS->",request.POST)
        print("MAMBOS2->",type(edit_tipo))
        if rowId == "":
            return HttpResponseBadRequest(
            "Não recebe o ID "
            )
        
       
        ref = StockPackage(
            id=rowId,
            pn=edit_partNumber,
            comentario=edit_comentario,
            descricao=edit_description,
            quantidade=edit_quantidade,
            inventario = edit_inv,
           
        )

        
        if not edit_tipo =="undefined" :
            print("CHEGOU AO IF DO TIPO NOT NULOL")
            ref.tipo = edit_tipo
        else:
            print("CHEGOU AO IF DO TIPO")
            ref.tipo =StockPackage.objects.get(id=rowId).tipo
            #ref.save()

        if edit_link is not None:
            # Savar o link localmente na maquina
            ref.link = default_storage.save(edit_link.name, edit_link)
        else:
            ref.link = StockPackage.objects.get(id=rowId).link
            #ref.save()
        
        print("GUARDOU")
        ref.save()

        return HttpResponseRedirect("pManagement:stockPackage")

# Delete de uma linha da table(html) e faz edição na db
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def deleteLinhaStockPackage(request):
    # Acho que falta fazer as verificações de campos vazios
    stockPackagePack = StockPackage.objects
    if request.method == "POST":
        row_id = request.POST["rowIdDelete"]
        # Aqui ele vai buscar o partNumber id e nesse id vai eliminar a cell
        stockPackagePack.all().get(id=row_id).delete()
    return redirect("pManagement:stockPackage")


# view para o btn update,  iguala o stock e o inventario
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def updateInventory(request):
    ids = request.POST.getlist("ids[]")
    inventario = request.POST.getlist("inventorys[]")
    quantidade = request.POST.getlist("quantidades[]")
    partNumberPackage = StockPackage.objects

    for row_id, inv, quant in zip(ids, inventario, quantidade):
        data = partNumberPackage.get(id=row_id)
        data.quantidade = inv
        data.inventario = inv
        data.save()

    return redirect("pManagement:stockPackage")


# view para o btn update, guarda na bd o inventario
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def updateAllStock(request):
    ids = request.POST.getlist("ids[]")
    inventario = request.POST.getlist("inventorys[]")
    partNumberPackage = StockPackage.objects

    for row_id, inv in zip(ids, inventario):
        if not inv:
            continue
        data = partNumberPackage.get(id=row_id)
        data.inventario = inv
        data.save()

    return redirect("pManagement:stockPackage")


""" def uploadStockPackage(request):
    if request.method == 'POST':
        if request.FILES.get('myfile'):
            dataset = tablib.Databook()
            new_table = request.FILES['myfile']

            if not new_table.name.endswith('xlsx'):
                clientes = ClientesOEM.objects
                produtos = Produtos.objects
                stockPackage = StockPackage.objects
                return render(request, 'pManagement/stockPackage.html',
                              {'clientes': clientes, 'produtos': produtos, 'tipoEmbalagem': tipoEmbalagem,
                               "erro2": "WRONG FORMAT! Only xlsx files"})

            # elimina todos os elementos que estavam na base de dados
            ClientesOEM.objects.all().delete()
            Produtos.objects.all().delete()
            TipoEmbalagem.objects.all().delete()

            imported_data = dataset.load(new_table.read(), format='xlsx')

            for data in imported_data.sheets():
                if (data.title == 'Sheet3'):
                    for element in data:
                        if (element.__contains__('Cliente')):
                            continue

                        # caso nao sejam inseridas quantidades no ficheiro = 0
                        if (element[4] == None):
                            y = list(element)
                            y[4] = 0
                            element = tuple(y)

                        # caso o cliente no ficheiro nao exista, é criado Not needed
                        if (not ClientesOEM.objects.filter(oem=element[0]).exists()):
                            cliente = ClientesOEM(
                                None,
                                element[0]
                            )
                            cliente.save()

                        # caso nao exista um produto para aquele cliente e tipo embalagem, é criado not needed, i need something difrent
                        if not Produtos.objects.filter(nome=element[1], cliente__oem=element[
                            0]).exists() or TipoEmbalagem.objects.filter(produto__nome=element[1],
                                                                         produto__cliente__oem=element[0]).exists():
                            produto = Produtos(
                                None,
                                ClientesOEM.objects.get(oem=element[0]).id,
                                element[1]
                            )
                            produto.save()

                        tipoEmbalagem = TipoEmbalagem(
                            None,
                            Produtos.objects.get(nome=element[1], cliente__oem=element[0], tipoembalagem__nome=None).id,
                            element[2],
                            element[4],
                            element[5],
                            element[3]
                        )
                        tipoEmbalagem.save()
        return redirect('pManagement:stockPackage')
 """

# ver se é para por aqui decorarors
def downloadExcelStockPackage(request):
    if request.method == "GET":
        elementos = StockPackage.objects
        wbProduction = Workbook()
        file = io.BytesIO()
        # Não criar um sheet novo e ter um existente ;)
        sheetProduction = wbProduction.get_sheet_by_name("Sheet")

        sheetProduction.append(
            [
                "partNumber",
                "descricao",
                "comentario",
                "quantidade",
                "inventario",
                "link",
                "tipo",
            ]
        )

        for elem in elementos.all():
            sheetProduction.append(
                [
                    str(elem.pn),
                    str(elem.descricao),
                    str(elem.comentario),
                    str(elem.quantidade),
                    str(elem.inventario),
                    str(elem.link),
                    str(elem.tipo),
                ]
            )

        wbProduction.save(file)
        # Voltar a pocição inicial do IO file object
        file.seek(0)

        data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
        fresp = FileResponse(
            file, filename=f"StockPackage_{data_fmt}.xlsx", as_attachment=True
        )
        return fresp


# def uploadDataExcel(request):
#    print("CHEGOU AO EXCEL")
#    if request.method == 'POST':
#        excel_file = ['//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Shippers Tracking.xlsx']
#        workbook = openpyxl.load_workbook(excel_file[0])
#        for sheet in workbook:
#            if sheet.title == 'Pre-Shipper Browse':
#                # if PreShipperBrowse.objects.get(shipDate=sheet[1][6].value):
#                # return
#                # caso seja igual, ver as diferenças linha a linha e acrescentar
#                # else:
#                StockPackage.objects.all().delete()
#                for element in sheet:
#                    value2 = StockPackage(
#                        None,
#                        element[0].value,
#                        element[1].value,
#                        element[2].value,
#                        element[3].value,
#                        element[4].value,
#                        element[5].value,
#
#                    )
#                    value2.save()
#
#            if sheet.title == 'Pre-Shipper Detail Browse':
#                StockPackage.objects.all().delete()
#                for element in sheet:
#                    value2 = StockPackage(
#                        None,
#                        element[0].value,
#                        element[1].value,
#                        element[2].value,
#                        element[3].value,
#                        element[4].value,
#                        element[5].value,
#                      )
#                    value2.save()
#    return redirect('shippers:shippersTracking')
#

# Ver que decorator por aqui
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="SupplyPackage")).exists()
)
def uploadStockFileStock(request):
    if request.method == "POST":
        if not (file := request.FILES.get("myfile")):
            return HttpResponseBadRequest("Não recebi nenhum ficheiro :(")

        if not file.name.endswith("xlsx"):
            return HttpResponseBadRequest(
                "Apenas excel no formato .xlsx são permitidos."
            )

        # elimina todos os elementos que estavam na base de dados

        workbook = load_workbook(file.open("rb"))
        sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
        db = StockPackage.objects
        db.all().delete()

        if not sheet:
            return HttpResponseBadRequest(
                "Não foi encontrado nenhuma sheet no excel..."
            )

        if len(list(sheet.columns)) != 7:
            return HttpResponseBadRequest("Ficheiro mal formatado.")

        for index, i in enumerate(sheet):
            if index == 0:
                continue

            pn, desc,  com, link, stock, inv, tipo = (
                i[x].value if len(i) - 1 >= x else None for x in range(7)
            )

            db.create(
                pn=pn,
                descricao=desc,
                comentario=com,
                quantidade=stock,
                inventario=inv,
                link=link,
                tipo=tipo,
            )

        return redirect("pManagement:stockPackage")




#esta func dá para ir buscar À pen ou ao git
def uploadStockFile(request):
    import re

    if request.method == "POST":
        if not (file := request.FILES.get("myfile")):
            return HttpResponseBadRequest("Ficheiro não recebido :(")

        if not file.name.endswith("xlsx"):
            return HttpResponseBadRequest(
                "Apenas excel no formato .xlsx são permitidos."
            )

        # elimina todos os elementos que estavam na base de dados
       
        workbook = load_workbook(file.open("rb"))
        sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])

        SupplyPackage.objects.all().delete()
        ClientesOEM.objects.all().delete()
        Produtos.objects.all().delete()
        ClienteProduto.objects.all().delete()
        StockPackage.objects.all().delete()
        
        if not sheet:
            return HttpResponseBadRequest(
                "Não foi encontrado nenhum sheet no excel..."
            )

        if len(list(sheet.columns)) != 9:
            return HttpResponseBadRequest("Ficheiro com formatação errada.")

        for i in list(sheet)[1:]:
            pn, desc, com, cpkg, spkg, stime, stock, inv, link = (
                i[x].value if len(i) - 1 >= x else None for x in range(9)
            )

            cpkg = re.findall(r"(\S+)?-(\S+)?", cpkg) if cpkg else []
            spkg = spkg.split(";") if spkg else []

            ref = SupplyPackage.objects.create(
                part_number=pn,
                description=desc,
                comment=com,
                supply_time=stime,
                stock=stock,
                inventario=inv,
                link=link,
            )

            for pkg in cpkg:
                c, _ = ClientesOEM.objects.get_or_create(oem=pkg[0])
                p = Produtos(nome=pkg[1])
                p.save()

                c.suplyPackage.add(ref)

                cp = ClienteProduto(
                    cliente=c,
                    produto=p
                )
                cp.save()
                cp.supply_pkg.add(ref)

            for pkg in spkg:
                c = StockPackage(pn=pkg, tipo="Returnable")
                c.save()
                c.suplyPackage.add(ref)

        return redirect("pManagement:stockPackage")


# Mostra a tabela dos Expendable
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="requests")).exists()
)
def stock(request):
    stockElements = StockElements.objects
    return render(request, "pManagement/stock.html", {"stockElements": stockElements})


# Mostra a tabela dos Returnable
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(
        Q(name="pManagement") | Q(name="ReturnablePackage")
    ).exists()
)
def returnablePacking(request):
    partnumbers = PartNumbersReturnable.objects
    return render(
        request, "pManagement/returnablePackage.html", {"partnumbers": partnumbers}
    )


# Mostra a tabela das Supply Packages
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="Supply")).exists()
)
def get_json_client_data(request):
    qs_val = list(ClientesOEM.objects.values())
    return JsonResponse({"data": qs_val})


def get_json_product_data(request, *args, **kwargs):
    selected_client = kwargs.get("client").upper()
    obj_models = list(Produtos.objects.filter(cliente__oem=selected_client).values())
    return JsonResponse({"data": obj_models})


def get_json_type_data(request, *args, **kwargs):
    selected_product = kwargs.get("product").upper()
    obj_models = list(
        TipoEmbalagem.objects.filter(produto__nome=selected_product).values()
    )
    return JsonResponse({"data": obj_models})


def deleteLinhaStock(request):
    if request.method == "POST":
        id = request.POST["idDelete"]
        StockElements.objects.get(id=id).delete()
    return redirect("pManagement:stock")


# Mostra a página do supply
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="Supply")).exists()
)
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="SupplyPackage")).exists()
)
def supply(request):
    pedidosExtra = extraProductTable.objects
    pedidosHistorico = extraProductHistoric.objects
    historicoAutores = buttonAuthorHistory.objects
    return render(
        request,
        "pManagement/supply.html",
        {
            "pedidosExtra": pedidosExtra,
            "pedidosHistorico": pedidosHistorico,
            "historicoAutores": historicoAutores,
        },
    )


# Pedido extra para a tabela Requests
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="requests")).exists()
)
def pedidoExtra(request):
    clientes = ClientesOEM.objects
    return render(request, "pManagement/pedidoExtra.html", {"clientes": clientes})


# Acrescenta uma nova linha à tabela returnable package
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="SupplyPackage")).exists()
)
def create(request):
    if request.method == "POST":
        oem = request.POST["oem"].upper()
        produto = request.POST["produto"].upper()
        tipo = request.POST["embalagem"].upper()
        quantidade = request.POST["quantidade"]

        if oem == "SELECT OPTION":
            return render(
                request, "pManagement/pedidoExtra.html", {"erro": "Insert valid OEM"}
            )

        if (
            len(request.POST["produto"] or request.POST["embalagem"]) == 0
            or quantidade == ""
        ):
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Fill all fields"}
            )

        if quantidade <= "0":
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Invalid quantity"}
            )

        tabelaArmazem = ClientesOEM.objects.get(oem=oem)

        try:
            produtoF = list(Produtos.objects.filter(nome=produto))
        except ObjectDoesNotExist:
            product = Produtos()
            product.cliente = tabelaArmazem
            product.nome = produto
            product.save()
            novotipo = TipoEmbalagem()
            novotipo.produto = product
            novotipo.nome = tipo
            novotipo.quantidade = quantidade
            if request.FILES.get("link"):
                novotipo.link = request.FILES["link"]
            novotipo.save()
            return redirect("pManagement:requests")

        tipoF = list(TipoEmbalagem.objects.filter(nome=tipo))
        for p in produtoF:
            for t in tipoF:
                if p.cliente.oem == oem and t.produto.nome == produto:
                    return render(
                        request,
                        "pManagement/pedidoExtra.html",
                        {"erro2": "Client already have this product"},
                    )

        try:
            produtoBD = Produtos.objects.get(nome=produto, cliente__oem=oem)
        except ObjectDoesNotExist:
            product = Produtos()
            product.cliente = tabelaArmazem
            product.nome = produto
            product.save()
            novotipo = TipoEmbalagem()
            novotipo.produto = product
            novotipo.nome = tipo
            novotipo.quantidade = quantidade
            if request.FILES.get("link"):
                novotipo.link = request.FILES["link"]
            novotipo.save()
            return redirect("pManagement:requests")

        novotipo = TipoEmbalagem()
        novotipo.produto = produtoBD
        novotipo.nome = tipo
        novotipo.quantidade = quantidade
        if request.FILES.get("link"):
            novotipo.link = request.FILES["link"]

        novotipo.save()
        request.FILES.clear()
        return redirect("pManagement:requests")

    else:
        clientes = ClientesOEM.objects
        return render(request, "pManagement/pedidoExtra.html", {"clientes": clientes})


# retira produto com a quantidade definida da tabela returnable package e acrescenta à tabela extra orders
def takeQuantity(request):
    if request.method == "POST":

        novoProdutoExtra = extraProductTable()

        oem = request.POST["oem"]
        produto = request.POST["produto"]
        embalagem = request.POST["embalagem"]
        quantidadeExistente = request.POST["quantidadeExistente"]
        quantidade = request.POST["quantidade"]
        data = request.POST["data"]
        link = request.POST["link"]
        estado = request.POST["estado"]
        tempoSupply = request.POST["tempoSupply"]
        tempoSupplyRestante = request.POST["tempoSupply"]

        if int(quantidade) > int(quantidadeExistente) or int(quantidade) < 0:
            clientes = ClientesOEM.objects
            produtos = Produtos.objects
            tipoEmbalagem = TipoEmbalagem.objects
            return render(
                request,
                "pManagement/requests.html",
                {
                    "clientes": clientes,
                    "produtos": produtos,
                    "tipoEmbalagem": tipoEmbalagem,
                    "erro": "Insert valid quantity",
                },
            )
        elif quantidade == quantidadeExistente:
            novoProdutoExtra.cliente = oem
            novoProdutoExtra.produto = produto
            novoProdutoExtra.embalagem = embalagem
            novoProdutoExtra.quantidade = quantidade
            novoProdutoExtra.link = link
            novoProdutoExtra.dataPedido = data
            novoProdutoExtra.estado = estado
            novoProdutoExtra.tempoSupply = tempoSupply
            novoProdutoExtra.tempoSupplyRestante = tempoSupply
            novoProdutoExtra.save()
            tipoEmbalagem = TipoEmbalagem.objects.get(
                nome=embalagem, produto__nome=produto, produto__cliente__oem=oem
            )
            tipoEmbalagem.quantidade = 0
            tipoEmbalagem.save()
            return redirect("pManagement:supply")
        elif quantidade == "0":
            return redirect("pManagement:requests")
        else:
            tipoEmbalagem = TipoEmbalagem.objects.get(
                nome=embalagem, produto__nome=produto, produto__cliente__oem=oem
            )
            tipoEmbalagem.quantidade = int(quantidadeExistente) - int(quantidade)
            tipoEmbalagem.save()
            novoProdutoExtra.cliente = oem
            novoProdutoExtra.produto = produto
            novoProdutoExtra.embalagem = embalagem
            novoProdutoExtra.quantidade = quantidade
            novoProdutoExtra.dataPedido = data
            novoProdutoExtra.estado = estado
            novoProdutoExtra.tempoSupply = tempoSupply
            novoProdutoExtra.tempoSupplyRestante = tempoSupply
            novoProdutoExtra.link = link
            novoProdutoExtra.save()
            return redirect("pManagement:supply")


# retira produto da tabela dos pedidos e acrescenta à tabela historico
def addHistoric(request):
    if request.method == "POST":
        oem = request.POST["clienteForm"]
        produto = request.POST["produtoForm"]
        embalagem = request.POST["embalagemForm"]
        quantidade = request.POST["quantidadeForm"]
        dataTerminado = request.POST["dataTerminadoForm"]
        dataPedido = request.POST["dataPedidoForm"]
        comentario = request.POST["comentarioForm"]
        estado = request.POST["estadoForm"]

        if estado == "Rejected":
            tipoEmbalagem = TipoEmbalagem.objects.get(
                nome=embalagem, produto__nome=produto, produto__cliente__oem=oem
            )
            tipoEmbalagem.quantidade = tipoEmbalagem.quantidade + int(quantidade)
            tipoEmbalagem.save()

        extraProduct = extraProductTable.objects.get(dataPedido=dataPedido)
        extraProduct.delete()

        novoProdutoHistorico = extraProductHistoric()
        novoProdutoHistorico.cliente = oem
        novoProdutoHistorico.produto = produto
        novoProdutoHistorico.embalagem = embalagem
        novoProdutoHistorico.quantidade = quantidade
        novoProdutoHistorico.dataTerminado = dataTerminado
        novoProdutoHistorico.dataPedido = dataPedido
        novoProdutoHistorico.comentario = comentario
        novoProdutoHistorico.estado = estado
        novoProdutoHistorico.save()
        return redirect("pManagement:supply")


# acrescenta um produto expendable à tabela expendable packing
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(
        Q(name="pManagement") | Q(name="ExtendablePackage")
    ).exists()
)
def createExpendable(request):
    if request.method == "POST":
        partNumber = request.POST["partNumberNovo"]
        descricao = request.POST["descricaoNovo"]
        link = request.POST["linkNovo"]
        quantidade = request.POST["quantidadeNovo"]

        if (
            len(request.POST["partNumberNovo"] or request.POST["descricaoNovo"]) == 0
            or quantidade == ""
        ):
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Fill all fields"}
            )

        if quantidade <= "0":
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Invalid quantity"}
            )

        if not PartNumbersExpendable.objects.filter(pn=partNumber).exists():
            newPartNumber = PartNumbersExpendable()
            newPartNumber.pn = partNumber
            newPartNumber.descricao = descricao
            newPartNumber.link = link
            newPartNumber.quantidade = quantidade
            newPartNumber.save()
            return redirect("pManagement:expendablePacking")
        return redirect("pManagement:expendablePacking")


@login_required()
@user_passes_test(
    lambda u: u.groups.filter(
        Q(name="pManagement") | Q(name="ReturnablePackage")
    ).exists()
)
def createReturnable(request):
    if request.method == "POST":
        partNumber = request.POST["partNumberNovo"]
        descricao = request.POST["descricaoNovo"]
        link = request.POST["linkNovo"]
        quantidade = request.POST["quantidadeNovo"]

        if (
            len(request.POST["partNumberNovo"] or request.POST["descricaoNovo"]) == 0
            or quantidade == ""
        ):
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Fill all fields"}
            )

        if quantidade <= "0":
            return render(
                request, "pManagement/pedidoExtra.html", {"erro2": "Invalid quantity"}
            )

        if not PartNumbersReturnable.objects.filter(pn=partNumber).exists():
            newPartNumber = PartNumbersReturnable()
            newPartNumber.pn = partNumber
            newPartNumber.descricao = descricao
            newPartNumber.link = link
            newPartNumber.quantidade = quantidade
            newPartNumber.save()
            return redirect("pManagement:returnablePacking")
        return redirect("pManagement:returnablePacking")


# preenche a tabela de returnable package com os dados do ficheiro xlsx inserido
def upload(request):
    if request.method == "POST":
        if request.FILES.get("myfile"):
            dataset = tablib.Databook()
            new_table = request.FILES["myfile"]

            if not new_table.name.endswith("xlsx"):
                clientes = ClientesOEM.objects
                produtos = Produtos.objects
                tipoEmbalagem = TipoEmbalagem.objects
                return render(
                    request,
                    "pManagement/requests.html",
                    {
                        "clientes": clientes,
                        "produtos": produtos,
                        "tipoEmbalagem": tipoEmbalagem,
                        "erro2": "WRONG FORMAT! Only xlsx files",
                    },
                )

            # elimina todos os elementos que estavam na base de dados
            ClientesOEM.objects.all().delete()
            Produtos.objects.all().delete()
            TipoEmbalagem.objects.all().delete()

            imported_data = dataset.load(new_table.read(), format="xlsx")

            for data in imported_data.sheets():
                if data.title == "Sheet3":
                    for element in data:
                        if element.__contains__("Cliente"):
                            continue

                        # caso nao sejam inseridas quantidades no ficheiro = 0
                        if element[4] == None:
                            y = list(element)
                            y[4] = 0
                            element = tuple(y)

                        # caso o cliente no ficheiro nao exista, é criado
                        if not ClientesOEM.objects.filter(oem=element[0]).exists():
                            cliente = ClientesOEM(None, element[0])
                            cliente.save()

                        # caso nao exista um produto para aquele cliente e tipo embalagem, é criado
                        if (
                            not Produtos.objects.filter(
                                nome=element[1], cliente__oem=element[0]
                            ).exists()
                            or TipoEmbalagem.objects.filter(
                                produto__nome=element[1],
                                produto__cliente__oem=element[0],
                            ).exists()
                        ):
                            produto = Produtos(
                                None,
                                ClientesOEM.objects.get(oem=element[0]).id,
                                element[1],
                            )
                            produto.save()

                        tipoEmbalagem = TipoEmbalagem(
                            None,
                            Produtos.objects.get(
                                nome=element[1],
                                cliente__oem=element[0],
                                tipoembalagem__nome=None,
                            ).id,
                            element[2],
                            element[4],
                            element[5],
                            element[3],
                        )
                        tipoEmbalagem.save()
        return redirect("pManagement:requests")


def uploadReturnable(request):
    if request.method == "POST":
        if request.FILES.get("myfile"):
            dataset = tablib.Databook()
            new_table = request.FILES["myfile"]

            # apenas aceita ficheiros do tipo xlsx
            if not new_table.name.endswith("xlsx"):
                partnumbers = PartNumbersReturnable.objects
                return render(
                    request,
                    "pManagement/returnablePackage.html",
                    {
                        "partnumbers": partnumbers,
                        "erro2": "WRONG FORMAT! Only xlsx files",
                    },
                )

            # elimina todos os elementos que estavam na base de dados
            PartNumbersReturnable.objects.all().delete()

            imported_data = dataset.load(new_table.read(), format="xlsx")

            for data in imported_data.sheets():
                if data.title == "Sheet3":
                    for element in data:
                        if element.__contains__("Link"):
                            continue

                        if element[3] == None:
                            y = list(element)
                            y[3] = 0
                            element = tuple(y)

                        partNumber = PartNumbersReturnable(
                            None, element[0], element[1], element[2], element[3]
                        )
                        partNumber.save()
            return redirect("pManagement:returnablePacking")

        else:
            partNumber = PartNumbersReturnable.objects.get(id=request.POST["nome"])
            partNumber.quantidade = request.POST["quantidade"]
            partNumber.save()

            return redirect("pManagement:returnablePacking")


def updateStock(request):
    if request.method == "POST":
        id = request.POST["id"]
        quantidade = request.POST["quantidade"]

        elem = StockElements.objects.get(id=id)
        elem.quantidadeStock = quantidade
        elem.save()
        return redirect("pManagement:stock")


def criarElementoStock(request):
    if request.method == "POST":
        partNumber = request.POST["partNumberNovo"]
        descricao = request.POST["descricaoNovo"]
        link = request.POST["linkNovo"]
        expendable = "false"
        returnable = "false"
        if "expendableNovo" in request.POST:
            expendable = "true"
        if "returnableMovo" in request.POST:
            returnable = "true"
        quantidade = request.POST["quantidadeNovo"]

        if not StockElements.objects.filter(partNumber=partNumber).exists():
            novoElem = StockElements(
                None, partNumber, descricao, link, expendable, returnable, quantidade
            )
            novoElem.save()
            return redirect("pManagement:stock")
        return redirect("pManagement:stock")


# atualiza o estado do pedido especifico na tabela supply
def atualizaEstado(request):
    if request.method == "POST":
        estadoProcesso = request.POST["estadoProcesso"]
        cliente = request.POST["cliente1"]
        produto = request.POST["produto1"]
        embalagem = request.POST["embalagem1"]
        data = request.POST["data2"]
        autor = None
        if request.POST["autor"]:
            autor = request.POST["autor"]
        startTime = None
        if request.POST["startTime"]:
            startTime = request.POST["startTime"]
        endTime = None
        if request.POST["endTime"]:
            endTime = request.POST["endTime"]

        produtoExtra = extraProductTable.objects.get(
            dataPedido=data, cliente=cliente, produto=produto, embalagem=embalagem
        )
        if startTime:
            produtoExtra.startTime = startTime
        if endTime:
            produtoExtra.dataTerminado = endTime
        if autor:
            produtoExtra.autor = autor
        produtoExtra.estado = estadoProcesso
        produtoExtra.save()

        novaEntrada = buttonAuthorHistory()
        novaEntrada.embalagem = produtoExtra
        novaEntrada.estado = estadoProcesso
        if autor:
            novaEntrada.autor = autor
        novaEntrada.horaPedido = data
        novaEntrada.horaAlteracao = startTime
        novaEntrada.save()

        return redirect("pManagement:supply")


def deleteLinhaExpendable(request):
    if request.method == "POST":
        partNumber = request.POST["partNumberDelete"]
        descricao = request.POST["descricaoDelete"]
        link = request.POST["linkDelete"]
        quantidade = request.POST["quantidadeDelete"]

        tipoEmbalagem = PartNumbersExpendable.objects

        tipoEmbalagem.get(pn=partNumber, descricao=descricao).delete()
    return redirect("pManagement:expendablePacking")


def deleteLinhaReturnable(request):
    if request.method == "POST":
        print(
            "Post do Returnable->",
            request["partNumber", "descricao", "link" "quantidade"],
        )
        partNumber = request.POST["partNumber"]
        descricao = request.POST["descricao"]
        link = request.POST["link"]
        quantidade = request.POST["quantidade"]

        tipoEmbalagem = PartNumbersReturnable.objects

        tipoEmbalagem.get(pn=partNumber, descricao=descricao).delete()
    return redirect("pManagement:returnablePacking")


def deleteLinhaPacking(request):
    partNumber = request.POST["partNumber"]
    descricao = request.POST["descricao"]
    link = request.POST["link"]
    quantidade = request.POST["quantidade"]

    print("partNumber", partNumber)
    print("Link", link)
    print("Quantidade", quantidade)


def updateLinhaExpendable(request):
    if request.method == "POST":

        partNumber = request.POST["partNumberEdit"]
        description = request.POST["descricaoEdit"]
        link = request.POST["linkEdit"]

        novoPartNumber = request.POST["novoPartNumber"]
        novoDescription = request.POST["novoDescription"]
        novoLink = request.POST["novoLink"]

        partNumberExpendable = PartNumbersExpendable.objects

        linhaCompleta = partNumberExpendable.get(pn=partNumber, descricao=description)

        # supplyTimeFINAL
        if novoPartNumber != "":
            linhaCompleta.pn = novoPartNumber
            linhaCompleta.save()
        if novoDescription != "":
            linhaCompleta.descricao = novoDescription
            linhaCompleta.save()
        if novoLink != "":
            linhaCompleta.link = novoLink
            linhaCompleta.save()
    return redirect("pManagement:expendablePacking")


def updateLinhaReturnable(request):
    if request.method == "POST":

        partNumber = request.POST["partNumberEdit"]
        description = request.POST["descricaoEdit"]
        link = request.POST["linkEdit"]

        novoPartNumber = request.POST["novoPartNumber"]
        novoDescription = request.POST["novoDescription"]
        novoLink = request.POST["novoLink"]

        partNumberExpendable = PartNumbersReturnable.objects

        linhaCompleta = partNumberExpendable.get(pn=partNumber, descricao=description)

        # supplyTimeFINAL
        if novoPartNumber != "":
            linhaCompleta.pn = novoPartNumber
            linhaCompleta.save()
        if novoDescription != "":
            linhaCompleta.descricao = novoDescription
            linhaCompleta.save()
        if novoLink != "":
            linhaCompleta.link = novoLink
            linhaCompleta.save()
    return redirect("pManagement:returnablePacking")


# BUG
# Tens de criar mais duas funcs, para os Clientes e Produtos
def getPNExpendable(request):
    print("CLIENTPACKAGEs-> ", request.GET)
    print("entrou no dos clientes")
    # falta por ele a devolver os expendable do StockPackage
    if request.method == "GET":
        idPacote = int(request.GET["idPacote"])
        s = SupplyPackage.objects.get(id=idPacote)
        pacote = ClienteProduto.objects.filter(supply_pkg=s).values()
        print("PAck", pacote)

        return JsonResponse({"listaExpendable": list(pacote)}, safe=False)

def _model_to_dict(model) -> dict[str, t.Any]:
    ret = {}
    for k in model._meta.get_fields():
        if isinstance(k, (models.ManyToOneRel, models.ManyToManyField)):
            continue
        
        value = getattr(model, k.name, None)
        if isinstance(k, models.ForeignKey) and value is not None:
            ret[k.name] = _model_to_dict(value)
        elif isinstance(k, models.FileField):
            ret[k.name] = value.name
        else:
            ret[k.name] = value

    return ret

# NOT USED: Maybe can be useful later
def queryset_to_dict(queryset):
    ret = []
    for row in queryset:
        ret.append(_model_to_dict(row))
    return ret

def get_cliente_produto(request):
    # falta por ele a devolver os expendable do StockPackage
    if request.method == "GET":
        s = ClienteProduto.objects.all()
        return JsonResponse({"data": queryset_to_dict(s)}, safe=False)

def get_cliente_produto_rel(request):
    # falta por ele a devolver os expendable do StockPackage
    if request.method == "GET":
        SPId = request.GET["Id"]
        s = ClienteProduto.supply_pkg.through.objects.filter(supplypackage_id=SPId)
        return JsonResponse({"data": queryset_to_dict(s)}, safe=False)

def get_stock_package_rel(request):
    # falta por ele a devolver os expendable do StockPackage
    if request.method == "GET":
        SPId = request.GET["Id"]
        s = StockPackage.suplyPackage.through.objects.filter(supplypackage_id=SPId)
        return JsonResponse({"data": queryset_to_dict(s)}, safe=False)

def getPNReturnable(request):
    # falta por ele a devolver os returnable do StockPackage
    if request.method == "GET":
        print("entrou no dos stocks")
        idPacote = int(request.GET["idPacote"])
        s = SupplyPackage.objects.get(id=idPacote)
        pacote = StockPackage.objects.filter(suplyPackage=s).values()

        return JsonResponse({"listaReturnable": list(pacote)}, safe=False)


def getPNStockPackage(request):
    print("StokPackage Lista-> ", request.GET)
    # falta por ele a devolver os expendable do StockPackage
    if request.method == "GET":
        idPacote = int(request.GET["idPacote"], base=10)
        print("IDPACOTE", idPacote)
        pacote = StockPackage.objects.get(id=idPacote)

        return JsonResponse({"listaExpendable": pacote.partNumberExpendable})


def getPNClienteProduto(request):
    # falta por ele a devolver os returnable do StockPackage
    print("BOMB ClienteProduto->", request.GET)
    # REQUEST VEM {} empty
    if request.method == "GET":
        idPacote = int(request.GET["idPacote"], base=10)
        print("IDPACOTE", idPacote)
        pacote = Produtos.objects.get(cliente=idPacote)


def getProdSupplyPackage(request):
    # falta por ele a devolver os returnable do StockPackage
    print("BOMB ClienteProduto->", request.GET)
    # REQUEST VEM {} empty
    if request.method == "GET":
        idPacote = int(request.GET["idPacote"], base=10)
        print("IDPACOTE", idPacote)
        pacote = Produtos.objects.get(cliente=idPacote)

        return JsonResponse({"listaReturnable": pacote.partNumberReturnable})


def getClientSupplyPackage(request):
    # falta por ele a devolver os returnable do StockPackage
    print("BOMB ClienteProduto->", request.GET)
    # REQUEST VEM {} empty
    print("entrou no dos vlientes")
    if request.method == "GET":
        idPacote = int(request.GET["idPacote"], base=10)
        print("IDPACOTE", idPacote)
        pacote = ClienteProduto.objects.get(cliente=idPacote)

        return JsonResponse({"listaReturnable": pacote.partNumberReturnable})



def reportCustomerPacking(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Packing - Customer Packing",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("pManagement:requests")


def reportSupplyPackage(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Packing - Supply Package",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("pManagement:supply")


def reportReturnablePackage(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Packing - Returnable Package",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("pManagement:returnablePacking")


def reportExpendablePackage(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Packing - Expendable Package",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("pManagement:expendablePacking")


@login_required()
@user_passes_test(lambda u: u.groups.filter(Q(name="pManagement")).exists())
def changeUserGroups(request):
    print(
        "REQs -> ",
    )

    if request.method == "POST":
        user = User.objects.get(username=request.POST["username"])
        grupo = request.POST["paginas"]
        # print("GROUPS->",  Group.objects.all())

        if User.objects.filter(
            username=request.POST["username"], groups__name="requests"
        ):
            my_group = Group.objects.using("default").get_or_create(name="requests")
            my_group[0].user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="newItem"
        ):
            my_group = Group.objects.using("default").get_or_create(name="newItem")
            my_group[0].user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="Supply"
        ):
            my_group = Group.objects.using("default").get_or_create(name="Supply")
            my_group[0].user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="ReturnablePackage"
        ):
            my_group = Group.objects.using("default").get_or_create(
                name="ReturnablePackage"
            )
            my_group[0].user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="ExtendablePackage"
        ):
            my_group = Group.objects.using("default").get_or_create(
                name="ExtendablePackage"
            )
            my_group[0].user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="StockPackage"
        ):
            my_group = Group.objects.using("default").get_or_create(name="StockPackage")
            my_group[0].user_set.remove(user)

        if request.POST["paginas"] == "none":
            my_group = Group.objects.using("default").get_or_create(name="StockPackage")
            my_group[0].user_set.remove(user)
            my_group = Group.objects.using("default").get_or_create(name="requests")
            my_group[0].user_set.remove(user)
            my_group = Group.objects.using("default").get_or_create(name="newItem")
            my_group[0].user_set.remove(user)
            my_group = Group.objects.using("default").get_or_create(name="Supply")
            my_group[0].user_set.remove(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.remove(user)

        if grupo == "none":
            message = (
                "User " + user.username + " perdeu acesso às páginas de Receiving."
            )
            subject, from_email, to = (
                "Alteração em Packing  - Configurations",
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            # msg.send()
            return redirect("pManagement:configurations")

        if grupo == "requests":
            my_group = Group.objects.using("default").get_or_create(name="requests")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="newItem")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="Supply")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        if grupo == "newItem":
            my_group = Group.objects.using("default").get_or_create(name="newItem")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        if grupo == "Supply":
            my_group = Group.objects.using("default").get_or_create(name="Supply")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        if grupo == "ReturnablePackage":
            my_group = Group.objects.using("default").get_or_create(
                name="ReturnablePackage"
            )
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        if grupo == "ExtendablePackage":
            my_group = Group.objects.using("default").get_or_create(
                name="ExtendablePackage"
            )
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        if grupo == "StockPackage":
            my_group = Group.objects.using("default").get_or_create(name="StockPackage")
            my_group[0].user_set.add(user)
            my_group = Group.objects.using("default").get_or_create(name="pManagement")
            my_group[0].user_set.add(user)
        return redirect("pManagement:configurations")


# INICIO SUPLY PACKAGE


@login_required()
@user_passes_test(lambda u: u.groups.filter(Q(name="pManagement")).exists())
def configurations(request):

    users = User.objects.all()
    print("USERS: ", users)
    return render(request, "pManagement/configurations.html", {"users": users})


# --------------------//--------------//--------SUPLY--PACKAGE ----------------------- // ---------------------------------------//---------------------------------//-------------------------
# func que vai carregar a pagina do suply package


def supplyPackage(request):
    # pagina do supplyPackage
    # clientes = ClientesOEM.produto.through.objects.all()
    # produtos =  Produtos.objects.all()
    tipoEmbalagem = TipoEmbalagem.objects.all()
    stockPackage_items = StockPackage.objects.all()
    kits = SupplyPackage.objects.all()
    cliente_produto = ClienteProduto.objects.all()

    listaFinal = []

    """  for cliente in clientes:
        for produto in produtos:
            if cliente.id == produto.cliente:
                lista_objecto = {
                    "cliente": cliente.oem,
                
                    
                    
                }
                lista.append(lista_objecto)
    for element in lista:
        if element in listaFinal:
            continue
        else:
            listaFinal.append(element)
    """
    return render(
        request,
        "pManagement/supplyPackage.html",
        {
            "tipoEmbalagem": tipoEmbalagem,
            "kitSuplyPackage": kits,
            "data": listaFinal,
            "stock_package_parts": stockPackage_items,
            "Produtos": Produtos.objects,
            "cliente_produto": ClienteProduto.objects,
            "ClientesOEM": ClientesOEM.objects,
        },
    )


def downloadExcelSupplyPackage(request):
    # não tas a apanhar os valores para fazer as comparações para os pores no DataValidation
    if request.method == "GET":
        elementos = SupplyPackage.objects
        stockPackages = StockPackage.objects
        clientes = ClienteProduto.objects
        wbProduction = Workbook()
        file = io.BytesIO()
        print("Elementos a passar para o excel-> ", elementos.all())
        # Não criar um sheet novo e ter um existente ;)
        sheetProduction = wbProduction.get_sheet_by_name("Sheet")

        # Verificar se ele está a ir buscar os campos à view ou aos models
        sheetProduction.append(
            [
                "PartNumber",
                "Descricao",
                "Comentario",
                "Cliente Package",
                "Stock Package",
                "Supply Time",
                "Stock",
                "Inventario",
                "Link",
            ]
        )

        for elem in elementos.all():
            sheetProduction.append(
                [
                    str(elem.part_number),
                    str(elem.description),
                    str(elem.comment),
                    " ; ".join(x.cliente.oem +"-"+ x.produto.nome for x in clientes.filter(supply_pkg=elem).all()),
                    #"/".join(x.produto.nome for x in clientes.filter(supply_pkg=elem).all()),
                    " ; ".join(
                        x.pn for x in stockPackages.filter(suplyPackage=elem).all()
                    ),
                    str(elem.supply_time),
                    str(elem.stock),
                    str(elem.inventario),
                    str(elem.link),
                ]
            )

        wbProduction.save(file)
        # Voltar a pocição inicial do IO file object
        file.seek(0)

        data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
        fresp = FileResponse(
            file, filename=f"SupplyPackage{data_fmt}.xlsx", as_attachment=True
        )
        return fresp


# BUG SUPPLY PACKAGE
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="SupplyPackage")).exists()
)
def addRowSupplyPackage(request):
    
    print("REQUEST ADD->", request.POST)
    if request.method == "POST":
        pn = request.POST.get("pnAdd")
        kitDescription = request.POST.get("prodDescriptionAdd")
        kitComment = request.POST.get("prodCommentAdd")
        kitStock = request.POST.get("prodStockAdd")
        kitlink = request.FILES.get("novoLinkEdit")
        kit_supplyTime = request.POST.get("supplyTimeAdd", 60) or 60
        
        stockPackages = request.POST.getlist("stockPackages")
        clients = request.POST.getlist("clients")
        

        """ if not stockPackages:
            stockPackages = None
        if not clients:
            clients= None """
        print("NO ADD!")
        print("LINK-> ", str( request.FILES.get("novoLinkEdit")))
        print(f"PART-NUM{pn}, Desc{kitDescription}, Comm {kitComment}, Link{kitlink} StP{stockPackages} , ClienteProd{clients}" )
        print("CLDASD->", clients[0].split(","))
        ref = SupplyPackage(
            part_number=pn,
            description=kitDescription,
            comment=kitComment,
            supply_time=kit_supplyTime,
            stock=kitStock,
        )
        ref.save()
        print("GUARDOU!")
        # se vierem valores em branco para as listas, tem de dar para adicionar na mesma
        for stock in stockPackages[0].split(","):
            print("S StockP",stock)
            s = StockPackage.objects.get(id=stock)
         
            s.suplyPackage.add(ref)
            s.save()
        for client in clients[0].split(","):
            s = ClienteProduto.objects.get(id=client)
            print("S ClienteP",s)
            s.supply_pkg.add(ref)
            s.save()
            
       

       

    return redirect("pManagement:supplyPackage")

def display_pdf(fobj: io.FileIO, fname: str):
    response = HttpResponse(fobj.read(), content_type="application/pdf")
    # aqui talvez dê para ir buscar o nome do file ao titulo
    response["Content-Disposition"] = f'filename={fname}'
    return response

def pdf_view(request):
    import aspose.slides as slides
    print("CHEGASTE AO PDF")

    link = request.GET.get("link")
    filename = link.split("/")[-1]
    # é aqui que vais por o sistema de pastas, ele está aw ir ao defaultStorage buscar

    try:
        f = default_storage.open(filename)
    except FileNotFoundError:
        raise Http404("Ficheiro não existe.")

    fmt = link.split(".")[-1]
    if fmt == "pdf":
        response = display_pdf(f, filename)
    elif fmt == "pptx":
        ffile = io.BytesIO()
        slides.Presentation(f.name).save(ffile, slides.export.SaveFormat.PDF)
        ffile.seek(0)
        response = display_pdf(ffile, filename)
    else:
        response = FileResponse(f, filename=link, as_attachment=True)
    return response

    
def pdf_view1(request):
    print("CHEGASTE AO PDF1")
    link = request.GET.get("link")
    filename = link.split("/")[-1]
    # é aqui que vais por o sistema de pastas, ele está a ir ao defaultStorage buscar
    try:
        f = default_storage.open(filename)
    except FileNotFoundError:
        raise Http404("Ficheiro não existe.")

    if link.split(".")[-1] == "pdf":
        response = HttpResponse(f.read(), content_type="application/pdf")
        # aqui talvez dê para ir buscar o nome do file ao titulo
        response["Content-Disposition"] = f'filename={link.split("/")[-1]}'
    else:
        response = FileResponse(f, filename=link, as_attachment=True)
    return response


def updateSupplyPackage(request):
    print("REquest do update SUPPLY PACKAGE--> ", request.POST)

    if request.method == "POST":
        edit_partNumber = request.POST.get("novoPartNumberEdit")
        edit_description = request.POST.get("novoDescriptionEdit")
        edit_comment = request.POST.get("novoCommentEdit")
        edit_stock = request.POST.get("novoStockProdutoEdit")

        edit_supplyTime = request.POST.get("novoSupplyTimeEdit", 60) or 60
        edit_link = request.FILES.get("novolinkEdit")
        edit_stockPackages = request.POST.getlist("stockPackages")
        edit_clients = request.POST.getlist("clients")
        rowId = request.POST["rowIdUpdate"]
        print("edit_stockPackages->", edit_stockPackages)
        print("edit_clients->", edit_clients)

        ref = SupplyPackage(
            part_number=edit_partNumber,
            description=edit_description,
            comment=edit_comment,
            supply_time=edit_supplyTime,
            stock=edit_stock,
            id=rowId,
        )
        ref.save()
      
        if edit_link is not None:
            print("LINK", edit_link)
            # alterar o defaultStorage para a pasta desiganada (supplyPackage)
            ref.link = default_storage.save(edit_link.name, edit_link)

        StockPackage.suplyPackage.through.objects.filter(supplypackage_id=rowId).delete()
        ClienteProduto.supply_pkg.through.objects.filter(supplypackage_id=rowId).delete()
        print("Stock", edit_stockPackages)

        for stock in edit_stockPackages[0].split(","):
            s = StockPackage.objects.get(id=stock)
            s.suplyPackage.add(ref)
            s.save()
       

        for client in edit_clients[0].split(","):
            s = ClienteProduto.objects.get(id=client)
            s.supply_pkg.add(ref)
            s.save()
        ref.save()
    return redirect("pManagement:supplyPackage")


def deleteSuplyPackage(request):
    print("ESTA A APAGAR LINHA DA PAG SUPPLY PACKAGE")
    ProdutosObj = SupplyPackage.objects
    if request.method == "POST":
        row_id = request.POST["rowIdDelete"]
        ProdutosObj.all().get(id=row_id).delete()
    return redirect("pManagement:supplyPackage")


def updateInventorySupplyPackage(request):
    # print("REQ do btn Update--> ",request.POST)
    ids = request.POST.getlist("ids[]")
    inventario = request.POST.getlist("inventorys[]")
    quantidade = request.POST.getlist("quantidades[]")
    partNumberPackage = SupplyPackage.objects
    print("LISTA VALORES->", inventario)
    for row_id, inv, quant in zip(ids, inventario, quantidade):
        data = partNumberPackage.get(id=row_id)
        data.stock = inv
        data.inventario = inv
        data.save()

    return redirect("pManagement:supplyPackage")


def updateAllStockSupplyPackage(request):
    ids = request.POST.getlist("ids[]")
    inventario = request.POST.getlist("inventorys[]")
    partNumberPackage = SupplyPackage.objects

    for row_id, inv in zip(ids, inventario):
        if not inv:
            continue
        data = partNumberPackage.get(id=row_id)
        data.inventario = inv
        data.save()

    return redirect("pManagement:supplyPackage")


def uploadSupplyPackageFile(request):
    if request.method == "POST":

        def raise_err(message: str):
            clientes = produtos = StockElements.objects
            tipoEmbalagem = TipoEmbalagem.objects
            return render(
                request,
                "pManagement/stockPackage.html",
                {
                    "clientes": clientes,
                    "produtos": produtos,
                    "tipoEmbalagem": tipoEmbalagem,
                    "erro2": message,
                },
            )

        if file := request.FILES.get("myfile"):
            if not file.name.endswith("xlsx"):
                return raise_err("WRONG FORMAT! Only xlsx files.")

            # elimina todos os elementos que estavam na base de dados
            # FALTA ADICIONAR CAMPOS CERTOS; NO FOR JÀ ESTÂO ALGUNS MAS FALTA DEFINIR COMO GUARDAR AS LISTAS
            workbook = load_workbook(file.open("rb"))
            sheet = workbook.get_sheet_by_name(
                "Sheet1"
            )  # existe maneira dele procurar pelo sheet1 e pelo sheet
            db = StockPackage.objects
            if not sheet:
                return raise_err("Only accept 1st Sheet.")

            if len(list(sheet.columns)) != 7:
                return raise_err("Ficheiro mal formatado.")

            # Remover todos os dados para inserir novos...
            db.all().delete()

            for index, i in enumerate(sheet):
                if index == 0:
                    continue
                pn, desc, link, quant, inv, com, supTime, listClients, listStockPack = (
                    i[0].value,
                    i[1].value,
                    i[2].value,
                    i[3].value,
                    i[4].value,
                    i[5].value,
                    i[6].value,
                )
                db.create(
                    pn=pn,
                    descricao=desc,
                    link=link,
                    quantidade=quant,
                    inventario=inv,
                    comentario=com,
                    tipo=supTime,
                )

        return redirect("pManagement:supplyPackage")


def updateLinhaStock(request):
    print("Novo updateLinhaStock NO SUPPLY PACKAGE")
    if request.method == "POST":
        id = request.POST["idEdit"]
        # Falta fazer as verificações
        novoPartNumber = request.POST["novoPartNumber"]
        novoDescription = request.POST["novoDescription"]
        novoLink = request.POST["novoLink"]
        expendable = "false"
        returnable = "false"

        if "novoValorExpendable" in request.POST:
            expendable = "true"
        if "novoValorReturnable" in request.POST:
            returnable = "true"

        elemento = StockElements.objects.get(id=id)

        # supplyTimeFINAL
        if novoPartNumber != "":
            elemento.partNumber = novoPartNumber
        if novoDescription != "":
            elemento.descricao = novoDescription
        if novoLink != "":
            elemento.link = novoLink

        """  elemento.expendable = expendable
            elemento.returnable = returnable """
        elemento.save()
    return redirect("pManagement:stock")


# BUG SUPLYPACKAGE
def updateSupplyStock(request):
    print("Update Stock supplyPackage")
    if request.method == "POST":
        pack = TipoEmbalagem.objects.get(id=request.POST["nome"])
        pack.quantidade = request.POST["quantidade"]
        pack.save()

        return redirect("pManagement:supplyPackage")


# BUG SUPLYPACKAGE
def savePNExpendable(request):
    # Confirmar se esta func é para o popup, se sim, vai ter de ser alterado o nome
    print("Save Item from ClienttPackage")
    if request.method == "POST":
        pacote = int(request.POST["idExpendable"], base=10)
        listaPN = request.POST["listaPNExpendable"]

        embalagem = TipoEmbalagem.objects.get(id=pacote)
        embalagem.partNumberExpendable = listaPN
        embalagem.save()

        return redirect("pManagement:supplyPackage")


# BUG SUPLYPACKAGE
def savePNReturnable(request):
    print("Save Item from ClienttPackage")
    if request.method == "POST":
        pacote = int(request.POST["idReturnable"], base=10)
        listaPN = request.POST["listaPNReturnable"]

        embalagem = TipoEmbalagem.objects.get(id=pacote)
        embalagem.partNumberReturnable = listaPN
        embalagem.save()
        return redirect("pManagement:supplyPackage")


# BUG SUPLYPACKAGE
def updateLinhaSupply(request):
    # FAZER VERIFICAÇÔEs PARA ACEITARES CAMPOS EMPTY
    print("Update supplyPackage")
    if request.method == "POST":
        oem = request.POST["oemEdit"]
        client_oem = request.POST["oemEdit"]
        produto = request.POST["produtoEdit"]
        prod_description = request.POST["descriptionEdit"]
        prod_comment = request.POST["commentEdit"]
        prod_stock = request.POST["stockEdit"]
        prod_inventory = request.POST["inventoryEdit"]
        prod_link = request.POST["linkEdit"]

        novoClienteInput = request.POST["novoClienteInput"]
        novoProdutoInput = request.POST["novoProdutoInput"]
        novoDescricaoProduto = request.POST["novoDescriptionEdit"]
        novoCommentEdit = request.POST["novoCommentEdit"]
        novoSupplyTime = request.POST["novoSupplyTime"]
        novoStockProduto = request.POST["novoStockProduto"]
        novoInventarioProduto = request.POST["novoInventarioProduto"]
        # vars em baixo, em principio não serão usadas

        embalagem = request.POST["embalagemEdit"]
        supplyTime = request.POST["supplyTimeEdit"]

        novoCliente = request.POST["novoCliente"]
        novoProduto = request.POST["novoProduto"]
        novoPacote = request.POST["novoPacote"]
        novoSupplyTime = request.POST["novoSupplyTime"]

        novoPacoteInput = request.POST["novoPacoteInput"]

        clientes = ClientesOEM.objects
        produtos = Produtos.objects
        tipoEmbalagem = TipoEmbalagem.objects

        clienteFINAL = None
        produtoFINAL = None
        tipoEmbalagemFINAL = None
        supplyTimeFINAL = None

        embalagemCompleta = tipoEmbalagem.get(
            nome=embalagem,
            tempoSupply=supplyTime,
            produto__nome=produto,
            produto__cliente__oem=client_oem,
        )

        # supplyTimeFINAL
        if novoSupplyTime == "":
            supplyTimeFINAL = supplyTime
        elif novoSupplyTime != "":
            supplyTimeFINAL = novoSupplyTime
            embalagemCompleta.tempoSupply = supplyTimeFINAL

        # clienteFINAL
        if novoCliente != "":
            clienteFINAL = clientes.get(id=novoCliente)
            produtoLocal = produtos.get(id=embalagemCompleta.produto.id)
            produtoLocal.cliente = clienteFINAL
            produtoLocal.save()
            embalagemCompleta.produto.cliente = clienteFINAL
        elif novoClienteInput != "":
            clienteFINAL = novoClienteInput
            if not clientes.filter(oem=novoClienteInput).exists():
                cliente = ClientesOEM(None, novoClienteInput.upper())
                cliente.save()
                produtoLocal = produtos.get(id=embalagemCompleta.produto.id)
                produtoLocal.cliente = cliente
                produtoLocal.save()
                embalagemCompleta.produto.cliente = cliente
            else:
                produtoLocal = produtos.get(id=embalagemCompleta.produto.id)
                produtoLocal.cliente = clienteFINAL
                produtoLocal.save()
                embalagemCompleta.produto.cliente = clienteFINAL

        else:
            clienteFINAL = oem

        # produtoFINAL
        if novoProduto != "":
            produtoFINAL = produtos.get(id=novoProduto).nome
            novoProdutoLocal = Produtos(
                None, clientes.get(oem=clienteFINAL).id, produtoFINAL
            )
            novoProdutoLocal.save()
            produtos.get(id=embalagemCompleta.produto.id).delete()
            embalagemCompleta.produto = novoProdutoLocal

        elif novoProdutoInput != "":
            produtoFINAL = novoProdutoInput
            produtoLocal = Produtos(
                None, clientes.get(oem=clienteFINAL).id, produtoFINAL
            )
            produtoLocal.save()
            produtos.get(id=embalagemCompleta.produto.id).delete()
            embalagemCompleta.produto = produtoLocal
        else:
            produtoFINAL = produto

        # tipoEmbalagemFINAL
        if novoPacote != "" or novoPacoteInput != "" or supplyTimeFINAL != supplyTime:
            if novoPacote != "":
                tipoEmbalagemFINAL = tipoEmbalagem.get(id=novoPacote).nome
                if embalagem == "":
                    embalagem = tipoEmbalagem.get(
                        produto__nome=produtoFINAL,
                        produto__cliente__oem=clienteFINAL,
                        tempoSupply=supplyTimeFINAL,
                    )
                    embalagem.nome = tipoEmbalagemFINAL
                    embalagem.save()
                else:
                    embalagemCompleta.nome = tipoEmbalagemFINAL

            elif novoPacoteInput != "":
                tipoEmbalagemFINAL = novoPacoteInput
                if embalagem == "":
                    embalagem.nome = tipoEmbalagemFINAL
                    embalagem.save()
                else:
                    embalagemCompleta.nome = tipoEmbalagemFINAL

            elif supplyTimeFINAL != supplyTime:
                embalagemCompleta.tempoSupply = supplyTimeFINAL

        embalagemCompleta.save()
    return redirect("pManagement:supplyPackage")


# BUG SUPLYPACKAGE
def deleteLinhaSupply(request):

    print("TAS NO DELETE DO SUPLYPAKAGE")
    print("Delete supplyPackage", request.POST)
    produto = Produtos.objects

    if request.method == "POST":
        # oem = request.POST["oemDelete"]
        row_id = request.POST["rowIdDelete"]
        produto = request.POST["produtoDelete"]
        embalagem = request.POST["embalagemDelete"]
        supplyTime = request.POST["supplyTimeDelete"]

        tipoEmbalagem = TipoEmbalagem.objects

        produto.get(id=row_id).delete()
        """ 
        tipoEmbalagem.get(
            nome=embalagem,
            produto__nome=produto,
            
            tempoSupply=supplyTime,
        ).delete() """
    return redirect("pManagement:supplyPackage")


# BUG SUPLYPACKAGE
def atualizaTempo(request):
    if request.method == "POST":

        if extraProductTable.objects.filter(
            dataPedido=request.POST.get("dataPedido")
        ).exists():
            pedido = extraProductTable.objects.get(
                dataPedido=request.POST.get("dataPedido")
            )
            pedido.tempoSupplyRestante = request.POST["novoTempoSupply"]
            pedido.ultimosDez = request.POST["ultimosDez"]
            pedido.tempoLimite = request.POST["tempoLimite"]
            pedido.save()

    return redirect("pManagement:supplyPackage")


# --------------------//--------------//--------Cliente--Produto ----------------------- // ---------------------------------------//---------------------------------//-------------------------


def clienteProduto(request):
    return render(
        request,
        "pManagement/clienteProduto.html",
        {
            "ClienteProduto": ClienteProduto.objects,
            "Produtos": Produtos.objects,
            "ClientesOEM": ClientesOEM.objects,
        },
    )



def downloadExcelClienteProduto(request):
    #não tas a apanhar os valores para fazer as comparações para os pores no DataValidation
    if request.method == "GET":
        produtos = Produtos.objects
        #stockPackages = StockPackage.objects
        cliente_prod = ClienteProduto.objects
        wbProduction = Workbook()
        file = io.BytesIO()
        # Não criar um sheet novo e ter um existente ;)
        sheetProduction = wbProduction.get_sheet_by_name("Sheet")
       
        #Verificar se ele está a ir buscar os campos à view ou aos models
        sheetProduction.append(
            [
                "Cliente",
                "Produto",
                "Comentario",
            ]
        )
        #CONFIRAR COM O BOSS COMO ELE QUER AS TABLES pois é aqui que vai ser construido o excel
        for elem in cliente_prod.all():
                sheetProduction.append(
                    [
                        str(elem.cliente),
                        str(elem.produto),
                        str(elem.comment)
                    ]
                )

        wbProduction.save(file)
        # Voltar a pocição inicial do IO file object
        file.seek(0)

        data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
        fresp = FileResponse(
            file, filename=f"ClienteProduto{data_fmt}.xlsx", as_attachment=True
        )
        return fresp

def returnListaProdutos(request):
    if request.method != "GET":
        raise Http404

    targetId = request.GET.get("targetId")
    if targetId:
        prods = Produtos.objects.filter(cliente_id=targetId)
    else:
        prods = Produtos.objects.all()
        
    return JsonResponse({"listaProds": list(prods.values())})


def createClienteProduto(request):
    # falta receber id dos items dos popupsFF
    if request.method == "POST":
        # tem que vir um id/nome para os clientes e para os produtos
        print("REQUEST EDIT CLIENTEPROD",request.POST)
        cliente = request.POST["newClient"]
        prods = request.POST["newProd"]
        comment = request.POST["newComment"]
        print("CLIENTE->", cliente)
        print("PRODUTOS->", prods)
        
        if not ClientesOEM.objects.filter(oem=cliente).exists():
            ref = ClientesOEM()
            ref.oem = cliente
            ref.save()

        ref= ClientesOEM.objects.get(oem=cliente)
        
        if not Produtos.objects.filter(nome=prods):
            prod = Produtos()
            prod.nome = prods
            prod.save()

        prod = Produtos.objects.get(nome=prods)

        rel = ClienteProduto(
            cliente=ref,
            comment=comment,
            produto = prod,
        )
        rel.save()

      

    return redirect("pManagement:clienteProduto")





# Edita uma linha da table table(html) e faz edição na db
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def updateClienteProduto(request):
    print("REQUEST", request.POST)
    print("Update CLIENTE PRODUTO --------->")
    if request.method == "POST":

        rowId = request.POST["rowIdUpdate"]
        edit_client = request.POST.get("getCliente")
        edit_prod = request.POST.get("getProds", "")
        edit_comentario = request.POST.get("editComment")

        print("MAMBOS->",edit_comentario)
        client_ref = ClientesOEM.objects
        if not ClientesOEM.objects.filter(oem=edit_client).exists():
            client_ref.oem = edit_client

        prod_ref = Produtos.objects
        if not Produtos.objects.filter(nome=edit_prod).exists():
            prod_ref.nome = edit_prod
        

        ref_client = ClientesOEM.objects.get(oem=edit_client)
        ref_prod = Produtos.objects.get(nome=edit_prod)
        ref = ClienteProduto(
            id = rowId,
            comment=edit_comentario,
            produto = ref_prod,
            cliente=ref_client,
        )
        
        #Produtos.objects.filter(cliente_id=rowId).update(cliente_id=None)
        """ for prod in edit_prod:
            s = Produtos.objects.get(id=prod)
            s.cliente = ref 
            s.save()"""

        ref.save()


    return redirect("pManagement:clienteProduto")


# Delete de uma linha da table(html) e faz edição na db
@login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="pManagement") | Q(name="StockPackage")).exists()
)
def deleteLinhaClienteProduto(request):
    print("REQUEST", request.POST)
    if request.method == "POST":
        row_id = request.POST["rowIdDelete"]
        print("ID TO DELETE", row_id)
        # StockPackage.suplyPackage.through.objects.filter(supplypackage_id=rowId)
        ClienteProduto.objects.filter(id=row_id).delete()
    return redirect("pManagement:clienteProduto")
