import os
import time
import openpyxl
import typing as t


from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
from django.core.mail import EmailMultiAlternatives
from django.http import JsonResponse, FileResponse
from django.shortcuts import render, redirect
from django.db.models import Q
from django.core.paginator import Paginator
from qad_ee.models import *



# Create your views here.
from xlwt import Workbook

from shippers.tables import GatewayTable,GatewayRecorrenteTable
from shippers.models import *
from qad_ee.models import *
from contextlib import contextmanager



@contextmanager
def timer():
    try:
        # __enter__
        t = time.perf_counter()
        yield
    finally:
        # __exit__
        print(f"A função  demorou {time.perf_counter() -t}s")


""" @login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists()) """
def configurations(request):
    users = User.objects
    #numHoras = NumeroHorasEmailDiario.objects.first()
    #print("numHoras -->", int(numHoras.numHoras1), int(numHoras.numHoras2))

    return render(request, "shippers/configurations.html", {"users": users})


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="shippingTracking").exists())
def shippersTracking(request):
    with timer():
        # max 0.8 ms
        tabela = PreShipperBrowse.objects
        tabelaFilha = PreShipperDetailBrowse.objects
        return render(
            request,
            "shippers/shippers.html",
            {"tabela": tabela, "tabelaFilha": tabelaFilha},
        )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="shippingConfirmation").exists())
def shippersConfirmation(request):
    ficheiro_Shippers = finalFicheiroShippers.objects
    return render(
        request,
        "shippers/shippersConfirmation.html",
        {"ficheiro_Shippers": ficheiro_Shippers},
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="shippingSecurity").exists())
def security(request):
    return render(request, "shippers/security.html")

@login_required()
@user_passes_test(
    lambda u: u.groups.filter(
        Q(name="shippingPortaria") | Q(name="shippingPortariaReadOnly")
    ).exists()
)
def portaria(request):
    with timer():
        readOnly = request.user.groups.filter(name="shippingPortariaReadOnly").exists()
        entrada = request.session.get("entrada")
        saida = request.session.get("saida")
        empresas = GatewayEmpresa.objects
        condutores = GatewayCondutor.objects
        condutoresID = GatewayCondutorID.objects
        primeiraMatricula = GatewayPrimeiraMatricula.objects
        segundaMtricula = GatewaySegundaMatricula.objects
        cargaDescarga = GatewayCargaDescarga.objects
        docas = GatewayDoca.objects
        destinosCarga = GatewayDestinoCarga.objects
        elementosGateway = Gateway.objects.all() 
        viatura = GatewayTipoViatura.objects.first()
        infoCondutores = GatewayInfoCondutor.objects.raw("SELECT * FROM shippers_gatewayinfocondutor")
        id_last_update = LastUpdatePortaria.objects
        #elementosGateway = Teste_Historico.objects.all()  #teste_bd_view
        #mediaEmpresa = MediaMensalEmpresa.objects.raw("SELECT id,empresa,media FROM shippers_mediamensalempresa WHERE MONTH(data) =  MONTH(CURRENT_TIMESTAMP)")
        # = MediaTotalPortaria.objects.raw("SELECT * FROM shippers_mediatotalportaria WHERE MONTH(data) = MONTH(CURRENT_TIMESTAMP)")
        contador =0
        """ for i in elementosGateway:
            contador +=1
            print("elementos Gatway ->",contador) """
        """ mediaEmpresa_dict = {}
        for i in mediaEmpresa:
            mediaEmpresa_dict[i.empresa] = i.media
        print("Valores dentro do dict----->>", mediaEmpresa_dict )
        
        mediaTotalPortaria_var = 0
        for i in mediaTotalPortaria:
            mediaTotalPortaria_var = i.media
        id_to_update =0
        for i in id_last_update.all():
            id_to_update = i.id

        for i in teste_bd_view.all():
            print("VIEW -)", i) """
        id_to_update =0
        for i in id_last_update.all():
            id_to_update = i.id

        #elementosGatewayPage = Paginator(elementosGateway, per_page=10).page(request.GET.get("page", 1))

        return render(
            request,
            "shippers/portariaDT.html",
            {
                # "empresas": empresas,
                "table": GatewayTable.render_paginated_table(request),
                "tableRecorrente": GatewayRecorrenteTable.render_paginated_table(request),
                "condutores":condutores,
                "condutoresID": condutoresID,
                "primeiraMatricula": primeiraMatricula,
                "segundaMtricula": segundaMtricula,
                "cargaDescarga": cargaDescarga,
                "docas": docas,
                "destinosCarga": destinosCarga,
                #"elementosGatewayTableSize": elementosGatewayPage.num_pages,
                "viatura": viatura,
                "infoCondutores": infoCondutores,
                "readOnly": readOnly,
                "entrada": entrada,
                "saida": saida,
                "id_last_update": id_to_update,
                #"mediaEmpresa": mediaEmpresa_dict,
                #"mediaTotal" : mediaTotalPortaria_var,
            },
        )



def json_gateway(request):
    from django.http import JsonResponse

    #elementosGateway = Gateway.objects.all() 
    elementosGateway = Gateway_View.objects.all()
    dados_list = [item.to_dict() for item in elementosGateway]
    return JsonResponse(dados_list, safe=False)

def json_trackingShippers(request):
    from django.http import JsonResponse
    elementostrackingShippers = Tracking_ShippersPage.objects.filter(abs_id=request.POST["abs_id"]) #Tracking_ShippersPage.objects.all()
    dados_list = [item.to_dict() for item in elementostrackingShippers]
    print("POST---)",dados_list)

    return JsonResponse(dados_list, safe=False)

def json_trackingShippersUpdate(request):
    refresh_date = Tracking_ShippersPage_Refresh.objects.all().values('lastUpdate').first().get('lastUpdate')
    return JsonResponse({"resposta": refresh_date})


def getIdForChandedItemPortaria(*args):
    id_last_update = LastUpdatePortaria.objects 
    id_update =0
    for i in id_last_update.all():
        id_update = i.id
    print("Entrou",id_update)
    return JsonResponse({"resposta": id_update})



# elif request.user.groups.filter(name='shippingPortariaReadOnly').exists():
#     empresas = GatewayEmpresa.objects
#     condutores = GatewayCondutor.objects
#     condutoresID = GatewayCondutorID.objects
#     primeiraMatricula = GatewayPrimeiraMatricula.objects
#     segundaMtricula = GatewaySegundaMatricula.objects
#     cargaDescarga = GatewayCargaDescarga.objects
#     docas = GatewayDoca.objects
#     destinosCarga = GatewayDestinoCarga.objects
#     elementosGateway = Gateway.objects
#     viatura = GatewayTipoViatura.objects.first()
#     return render(request, 'shippers/portariaReadOnly.html',
#                   {'empresas': empresas, 'condutores': condutores, 'condutoresID': condutoresID,
#                    'primeiraMatricula': primeiraMatricula, 'segundaMtricula': segundaMtricula,
#                    'cargaDescarga': cargaDescarga, 'docas': docas, 'destinosCarga': destinosCarga,
#                    'elementosGateway': elementosGateway, 'viatura': viatura})


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="shippingTracking").exists())
def tracking(request):
    print("tracking")
    with timer():
        actualDay = datetime.today().strftime("%Y-%m-%d")
        dadosQAD = []
        gateway = Gateway.objects.filter(Q(estado="verde") | Q(estado="terminado"))
        gatewayValues = Gateway.objects.filter(
            Q(estado="verde") | Q(estado="terminado")
        ).values("primeiraMatricula")
        dadosQADabs = (
            AbsMstr.objects.filter(abs_shp_date=actualDay)
            .filter(Q(abs_id__startswith="s") | Q(abs_id__startswith="p"))
            .filter((Q(abs_type="s")) | (Q(abs_type="p")))
            .filter(abs_shipfrom="3515")
            .values(
                "abs_id",
                "abs_shp_date",
                "abs_shp_time",
                "abs_qad01",
                "oid_abs_mstr",
                "abs_status",
                "abs_shipto",
                "abs_item",
                "abs_domain",
            )
        )
        dadosQADad = (
            AdMstr.objects.filter(ad_addr__in=dadosQADabs.values_list("abs_shipto"))
            .values("ad_addr", "ad_city", "ad_country")
            .distinct()
        )

        for abs in dadosQADabs.all():
            abs["abs_shp_time"] = str(convert(abs["abs_shp_time"]))
            if abs["abs_status"][1:2] == "y":
                abs["abs_status"] = "Yes"
            else:
                abs["abs_status"] = "No"
            QADad = dadosQADad.filter(ad_addr=abs["abs_shipto"]).distinct()
            for ad in QADad.all():
                row = {
                    "id": abs["abs_id"],
                    "shipDate": str(abs["abs_shp_date"])[0:10],
                    "shipTime": abs["abs_shp_time"],
                    "city": ad["ad_city"],
                    "country": ad["ad_country"],
                    "carrier": "",
                    "fob": abs["abs_qad01"][20:40],
                    "modeOfTransport": abs["abs_qad01"][60:80],
                    "vehicleID": abs["abs_qad01"][80:100],
                    "totalMasterPacks": str(abs["oid_abs_mstr"]),
                    "confirmed": abs["abs_status"],
                    "itemNumber": "",
                    "description": "",
                    "qtyToShip": "",
                    "qtyShipped": "",
                }
                dadosQAD.append(row)

    return render(
        request, "shippers/shipping.html", {"dadosQAD": dadosQAD, "gateway": gateway}
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists())
def tracking2(request):
    with timer():
        actualDay = datetime.today().strftime("%Y-%m-%d")
        dadosQAD = []
        dadosQADabs = (
            AbsMstr.objects.filter(abs_shp_date=actualDay)
            .filter(Q(abs_id__startswith="s") | Q(abs_id__startswith="p"))
            .filter((Q(abs_type="s")) | (Q(abs_type="p")))
            .filter(abs_shipfrom="3515")
            .values(
                "abs_id",
                "abs_shp_date",
                "abs_shp_time",
                "abs_qad01",
                "oid_abs_mstr",
                "abs_status",
                "abs_shipto",
                "abs_item",
                "abs_domain",
            )
        )
        dadosQADad = (
            AdMstr.objects.filter(ad_addr__in=dadosQADabs.values_list("abs_shipto"))
            .values("ad_addr", "ad_city", "ad_country")
            .distinct()
        )
        dadosQADabsc = AbscDet.objects.filter(
            absc_abs_id__in=dadosQADabs.values_list("abs_id")
        ).values("absc_abs_id", "absc_carrier")
        dadosQADabsALTERADOS = AbsMstr.objects.filter(
            abs_par_id__in=dadosQADabs.values_list("abs_id")
        ).values("abs_par_id", "abs_item", "abs_qty", "abs_ship_qty")
        dadosQADpt = PtMstr.objects.filter(
            pt_part__in=dadosQADabsALTERADOS.values_list("abs_item")
        ).values("pt_desc1", "pt_desc2", "pt_part")
        AbsMstrPriv.objects.all().delete()
        AdPriv.objects.all().delete()
        AbscPriv.objects.all().delete()
        Abs2Priv.objects.all().delete()
        PtPriv.objects.all().delete()
        for abs in dadosQADabs.all():
            row = AbsMstrPriv(
                None,
                abs["abs_id"],
                str(abs["abs_shp_date"])[0:10],
                str(abs["abs_shp_time"]),
                abs["abs_qad01"],
                str(abs["oid_abs_mstr"]),
                abs["abs_status"],
                abs["abs_shipto"],
                abs["abs_item"],
                abs["abs_domain"],
            )
            row.save()

        for abs in AbsMstrPriv.objects.all():
            for ad in dadosQADad.all():
                if ad["ad_addr"] == abs.abs_shipto_2:
                    rowAD = AdPriv()
                    rowAD.ad_addr_2 = (abs,)
                    rowAD.ad_city_2 = (ad["ad_city"],)
                    rowAD.ad_country_2 = ad["ad_country"]
                    rowAD.save()
            for pt in dadosQADpt.all():
                if pt["pt_part"] == abs.abs_item:
                    rowAD = AdPriv(None, pt["pt_desc1"], pt["pt_desc2"], abs.abs_item)
                    rowAD.save()
            for absc in dadosQADabsc.all():
                if absc["absc_abs_id"] == abs.abs_id:
                    rowABSC = AbscPriv(None, abs.abs_id, absc["absc_carrier"])
                    rowABSC.save()
            for abs2 in dadosQADabsALTERADOS.all():
                if abs2["abs_par_id"] == abs.abs_id:
                    rowABS2 = Abs2Priv(
                        None,
                        abs.abs_id,
                        abs2["abs_item"],
                        abs2["abs_qty"],
                        abs2["abs_shp_qty"],
                    )
                    rowABS2.save()

            rowAbs = {
                "id": abs["abs_id"],
                "shipDate": str(abs["abs_shp_date"])[0:10],
                "shipTime": abs["abs_shp_time"],
                "fob": abs["abs_qad01"][20:40],
                "modeOfTransport": abs["abs_qad01"][60:80],
                "vehicleID": abs["abs_qad01"][80:100],
                "totalMasterPacks": str(abs["oid_abs_mstr"]),
                "confirmed": abs["abs_status"],
            }

            rowAd = {}
            if dadosQADad.filter(ad_addr=abs["abs_shipto"]).exists():
                ad = dadosQADad.filter(ad_addr=abs["abs_shipto"])[0]
                rowAd = {"city": ad["ad_city"], "country": ad["ad_country"]}

            rowAbsc = {}
            if dadosQADabsc.filter(absc_abs_id=abs["abs_id"]).exists():
                absc = dadosQADabsc.get(absc_abs_id=abs["abs_id"])
                rowAbsc = {
                    "carrier": absc["absc_carrier"],
                }

            if dadosQADabsALTERADOS.filter(abs_par_id=abs["abs_id"]).exists():
                for abs2 in dadosQADabsALTERADOS.filter(
                    abs_par_id=abs["abs_id"]
                ).distinct():
                    rowAbs2 = {
                        "itemNumber": abs2["abs_item"],
                        "qtyToShip": str(abs2["abs_qty"]),
                        "qtyShipped": str(abs2["abs_ship_qty"]),
                    }

                    rowPt = {}

                    if dadosQADpt.filter(pt_part=abs["abs_item"]).exists():
                        pt = dadosQADpt.get(pt_part=abs["abs_item"])
                        rowPt = {
                            "description": pt["pt_desc1"] + pt["pt_desc2"],
                        }
                    mergedRow = rowAbs | rowAd | rowAbsc | rowPt | rowAbs2
                    dadosQAD.append(mergedRow)
            else:
                mergedRow = rowAbs | rowAd | rowAbsc
                dadosQAD.append(mergedRow)

        return render(request, "shippers/shipping.html", {"dadosQAD": dadosQAD})

# utilizado em shippersTracking
def uploadFiles(request):
    if request.method == "POST":
        # if request.FILES.get('browseFile') and request.FILES.get('detailFile'):
        # browseDatabook = tablib.Databook()
        # detailDatabook = tablib.Databook()
        # browseFile = request.FILES['browseFile']
        # detailFile = request.FILES['detailFile']

        # if not browseFile.name.endswith('xlsx') or not detailFile.name.endswith('xlsx'):
        #    path = Path.objects
        #    return render(request, 'shippers/shippersTracking.html',
        #                  {"path": path, "erro2": "WRONG FORMAT! Only xlsx files"})
        #BUG FILESYS
        excel_files = [
            "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Teste_Browse.xlsx",
            "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Teste_Detail.xlsx",
        ]

        filteredTable.objects.all().delete()
        Teste_browse.objects.all().delete()
        Teste_detail.objects.all().delete()

        # imported_browse = browseDatabook.load(browseFile.read(), format='xlsx')
        # imported_detail = detailDatabook.load(detailFile.read(), format='xlsx')

        # for data in imported_browse.sheets():
        #BUG FILESYS Este ficheiro vai ter de ser substituido por uma tabela na bd
        for file in excel_files:
            if (
                file
                == "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Teste_Browse.xlsx"
            ):
                workbook = openpyxl.load_workbook(file)
                worksheet = workbook["Sheet1"]
                for element in worksheet:
                    value2 = Teste_browse(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                        element[16].value,
                        element[17].value,
                        element[18].value,
                        element[19].value,
                        element[20].value,
                        element[21].value,
                        element[22].value,
                    )
                    value2.save()
                    # cria elemento na filteredTable com dados apenas do ficheiro Teste Browse
                    tabelaFiltrada = filteredTable(
                        None,
                        element[2].value,
                        element[8].value,
                        element[9].value,
                        element[4].value,
                        element[5].value,
                        element[10].value,
                        element[13].value,
                        element[14].value,
                        None,
                        None,
                        None,
                        None,
                        element[19].value,
                        element[20].value,
                    )
                    tabelaFiltrada.save()

            #BUG FILESYS
            if (
                file
                == "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Teste_Detail.xlsx"
            ):
                workbook = openpyxl.load_workbook(file)
                worksheet = workbook["Sheet1"]
                # for data in imported_detail.sheets():
                for element in worksheet:
                    for elementoTabelaFiltrada in filteredTable.objects.all():
                        if elementoTabelaFiltrada.idComum == element[2].value:
                            # complementa os elementos que estao na filteredTable com dados do ficheiro Teste Detail
                            if (
                                elementoTabelaFiltrada.itemNumber == None
                                and elementoTabelaFiltrada.quantityShipped == None
                                and elementoTabelaFiltrada.description == None
                                and elementoTabelaFiltrada.quantityToShip == None
                            ):
                                elementoTabelaFiltrada.itemNumber = element[7].value
                                elementoTabelaFiltrada.description = element[8].value
                                elementoTabelaFiltrada.quantityToShip = element[9].value
                                elementoTabelaFiltrada.quantityShipped = element[
                                    10
                                ].value
                                elementoTabelaFiltrada.save()
                            # caso haja mais que um elemento com aquele idComum, cria novo
                            else:
                                novoElemento = filteredTable(
                                    None,
                                    elementoTabelaFiltrada.idComum,
                                    elementoTabelaFiltrada.shipDate,
                                    elementoTabelaFiltrada.shipTime,
                                    elementoTabelaFiltrada.name,
                                    elementoTabelaFiltrada.city,
                                    elementoTabelaFiltrada.carrier,
                                    elementoTabelaFiltrada.modeOfTransport,
                                    elementoTabelaFiltrada.vehicleID,
                                    element[7].value,
                                    element[8].value,
                                    element[9].value,
                                    element[10].value,
                                    elementoTabelaFiltrada.inProcess,
                                    elementoTabelaFiltrada.confirmed,
                                )
                                novoElemento.save()

                    value = Teste_detail(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                        element[16].value,
                        element[17].value,
                        element[18].value,
                        element[19].value,
                        element[20].value,
                        element[21].value,
                    )
                    value.save()

    return redirect("shippers:shippersTracking")


# filtra alguns campos de diferentes sheets e guarda na base de dados. utilizado em shippers
def uploadDataExcel(request):
    if request.method == "POST":
        #BUG FILESYS
        excel_file = [
            "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Shippers Tracking.xlsx"
        ]
        workbook = openpyxl.load_workbook(excel_file[0])
        for sheet in workbook:
            if sheet.title == "Pre-Shipper Browse":
                # if PreShipperBrowse.objects.get(shipDate=sheet[1][6].value):
                # return
                # caso seja igual, ver as diferenças linha a linha e acrescentar
                # else:
                PreShipperBrowse.objects.all().delete()
                for element in sheet:
                    value2 = PreShipperBrowse(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                        element[16].value,
                        element[17].value,
                        element[18].value,
                        element[19].value,
                        element[20].value,
                        element[21].value,
                    )
                    value2.save()

            if sheet.title == "Pre-Shipper Detail Browse":
                PreShipperDetailBrowse.objects.all().delete()
                for element in sheet:
                    value2 = PreShipperDetailBrowse(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                        element[16].value,
                        element[17].value,
                        element[18].value,
                        element[19].value,
                        element[20].value,
                        element[21].value,
                    )
                    value2.save()
    return redirect("shippers:shippersTracking")


def getChildValues(request):
    with timer():
        print("getChild Values")
        if request.method == "GET":
            shipperID = request.GET["shipperID"]
            listTableValues = list(
                PreShipperDetailBrowse.objects.filter(idShipper=shipperID).values()
            )
        return JsonResponse({"filteredTableValues": listTableValues})


def getDataFicheiro(request):
    if request.method == "GET":
        dataAtualFicheiro = request.GET["dataAtualFicheiro"]
        #BUG FILESYS
        excel_file = [
            "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Shipping/Shippers Tracking.xlsx"
        ]
        workbook = openpyxl.load_workbook(excel_file[0])
        data = ""
        for sheet in workbook:
            if sheet.title == "Pre-Shipper Browse":
                for element in sheet:
                    data = element[8].value
        return JsonResponse({"dataFicheiro": data})


def uploadShippersConfirmation(request):
    if request.method == "POST":
        if request.FILES.get("ficheiroConfirmacao"):

            ficheiroShippers.objects.all().delete()
            finalFicheiroShippers.objects.all().delete()

            ficheiroConfirmacao = [request.FILES["ficheiroConfirmacao"]]
            workbook = openpyxl.load_workbook(ficheiroConfirmacao[0])

            for sheet in workbook:
                if sheet.title == "extended packaging data browse":
                    for element in sheet:
                        novoElemento = ficheiroShippers(
                            None,
                            element[0].value,
                            element[2].value,
                            element[5].value,
                            element[10].value,
                        )
                        novoElemento.save()

            totalFicheiros = ficheiroShippers.objects.all()
            masterSerial = None
            preShipper = None
            pack = None
            packAnterior = None
            numberPack = None
            totalPacks = 0
            for linha in totalFicheiros:
                if linha.masterSerialID == masterSerial:
                    if linha.packItem == pack:
                        totalPacks = totalPacks + int(linha.numberOfPacks, base=10)
                    else:
                        novoElemento = finalFicheiroShippers(
                            None, masterSerial, preShipper, pack, totalPacks
                        )
                        novoElemento.save()
                        totalPacks = int(linha.numberOfPacks, base=10)
                else:
                    if linha.numberOfPacks != "Number of Packs":
                        if pack == packAnterior:
                            totalPacks = totalPacks
                            novoElemento = finalFicheiroShippers(
                                None, masterSerial, preShipper, pack, totalPacks
                            )
                            novoElemento.save()
                            totalPacks = int(linha.numberOfPacks, base=10)
                        else:
                            totalPacks = int(linha.numberOfPacks, base=10)
                            novoElemento = finalFicheiroShippers(
                                None, masterSerial, preShipper, pack, numberPack
                            )
                            novoElemento.save()

                masterSerial = linha.masterSerialID
                preShipper = linha.preShipperShipper
                packAnterior = pack
                pack = linha.packItem
                numberPack = linha.numberOfPacks

            # Para alcançar o ultimo elemento da lista
            for i in range(0, len(totalFicheiros)):
                if i == (len(totalFicheiros) - 1):
                    novoElemento = finalFicheiroShippers(
                        None, masterSerial, preShipper, pack, numberPack
                    )
                    novoElemento.save()

    return redirect("shippers:shippersConfirmation")


def resetVeiculoDB(request):
    if request.method == "POST":
        GatewayTipoViatura.objects.all().delete()
    return redirect("shippers:portaria")

def enviaMailDemora(request):
        #id_atraso = request["id"]

        dataHoraEntrada = ""
        abandono = request.POST["abandono"]
        comentariosEntrada = request.POST["comentariosEntrada"]
        dataHoraSaida = ""
        comentariosSaida = request.POST["comentariosSaida"]
        id = request.POST["id"]
        enviarEmailAbandonado = ""

        if request.POST["dataHoraEntrada"] != "":
            dataHoraEntrada = request.POST["dataHoraEntrada"]
        if request.POST["dataHoraSaida"] != "":
            dataHoraSaida = request.POST["dataHoraSaida"]

        elemento = Gateway.objects.get(id=id)
        if dataHoraSaida != "":
            elemento.dataHoraSaida = dataHoraSaida
        if dataHoraEntrada != "":
            elemento.dataHoraEntrada = dataHoraEntrada
            elemento.estado = "terminado"

        if elemento.abandono == "true":
            if abandono == "false":
                elemento.dataHoraEntrada = ""
                elemento.estado = ""
                enviarEmailAbandonado = "voltou"
        else:
            if abandono == "true":
                elemento.dataHoraEntrada = datetime.today().strftime("%Y-%m-%d %H:%M")
                elemento.dataHoraSaida = datetime.today().strftime("%Y-%m-%d %H:%M")
                elemento.estado = "terminado"
                enviarEmailAbandonado = "abandonou"
        elemento.abandono = abandono
        if comentariosEntrada != "":
            elemento.comentEntrada = comentariosEntrada
        if comentariosSaida != "":
            elemento.comentSaida = comentariosSaida
        if elemento.dataHoraSaida != "" or elemento.abandono == "true":
            elemento.estado = "feitoEsaiu"
        elemento.save()

        # alterei a primeira linha, e adicionei a cor diretamente no style das células
        if enviarEmailAbandonado == "abandonou":
            table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>ABANDONO</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'

            table += (
                '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.dataHoraChegada
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.condutor
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.ident
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.contacto
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.empresa
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.primeiraMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.segundaMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.cargaDescarga
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.doca
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.destinoCarga
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.tipoViatura
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + elemento.dataHoraEntrada
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + "X"
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + elemento.comentEntrada
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.dataHoraSaida
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.comentEntrada
                + "</td></tr>"
            )
            table += (
                '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
            )
            table += "</body></table>"

            subject, from_email, to = (
                "Alteração em Shipping - Portaria",
                "noreply@visteon.com",
                ["pmarti30@visteon.com", 'npires2@visteon.com'],
            )

            # [ 'aroque1@visteon.com', 'npires2@visteon.com', 'pmarti30@visteon.com']

            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()

      



def submitGatewayRowUpdate(request):
    if (
        request.method == "POST"
        and request.user.groups.filter(name="shippingPortaria").exists()
    ):
        print("REQUEST->", request.POST)
        dataHoraEntrada = ""
        abandono = request.POST["abandono"]
        comentariosEntrada = request.POST["comentariosEntrada"]
        dataHoraSaida = ""
        comentariosSaida = request.POST["comentariosSaida"]
        id = request.POST["id"]
        enviarEmailAbandonado = ""

        if request.POST["dataHoraEntrada"] != "":
            dataHoraEntrada = request.POST["dataHoraEntrada"]
        if request.POST["dataHoraSaida"] != "":
            dataHoraSaida = request.POST["dataHoraSaida"]

        elemento = Gateway.objects.get(id=id)
        if dataHoraSaida != "":
            elemento.dataHoraSaida = dataHoraSaida
        if dataHoraEntrada != "":
            elemento.dataHoraEntrada = dataHoraEntrada
            elemento.estado = "terminado"

        if elemento.abandono == "true":
            if abandono == "false":
                elemento.dataHoraEntrada = ""
                elemento.estado = ""
                enviarEmailAbandonado = "voltou"
        else:
            if abandono == "true":
                elemento.dataHoraEntrada = datetime.today().strftime("%Y-%m-%d %H:%M")
                elemento.dataHoraSaida = datetime.today().strftime("%Y-%m-%d %H:%M")
                elemento.estado = "terminado"
                enviarEmailAbandonado = "abandonou"
        elemento.abandono = abandono
        if comentariosEntrada != "":
            elemento.comentEntrada = comentariosEntrada
        if comentariosSaida != "":
            elemento.comentSaida = comentariosSaida
        if elemento.dataHoraSaida != "" or elemento.abandono == "true":
            elemento.estado = "feitoEsaiu"
        print("A apagar todos os dados ta tabela")
        LastUpdatePortaria.objects.all().delete()
        updatePortaria = LastUpdatePortaria(
            dataUpdate =  datetime.today().strftime("%Y-%m-%d %H:%M:%S.%f")
        ).save()
        print("A guardar dados na lastUpdatePOrtaria")

        elemento.save()

        # alterei a primeira linha, e adicionei a cor diretamente no style das células
        if enviarEmailAbandonado == "abandonou":
            table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>ABANDONO</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'

            table += (
                '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.dataHoraChegada
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.condutor
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.ident
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.contacto
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.empresa
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.primeiraMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.segundaMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.cargaDescarga
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.doca
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.destinoCarga
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.tipoViatura
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + elemento.dataHoraEntrada
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + "X"
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: red">'
                + elemento.comentEntrada
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.dataHoraSaida
                + '</td><td style="padding:0 15px 0 15px;background-color: red">'
                + elemento.comentEntrada
                + "</td></tr>"
            )
            table += (
                '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
            )
            table += "</body></table>"

            subject, from_email, to = (
                "Alteração em Shipping - Portaria",
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )

            # [ 'aroque1@visteon.com', 'npires2@visteon.com', 'pmarti30@visteon.com']

            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()
            # alterei a primeira linha, e adicionei a cor diretamente no style das células
        if enviarEmailAbandonado == "voltou":
            table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>ABANDONO</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'
            #'<tr style="background-color: whitesmoke">' + \
            table += (
                '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.dataHoraChegada
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.condutor
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.ident
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.contacto
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.empresa
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.primeiraMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.segundaMatricula
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.cargaDescarga
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.doca
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.destinoCarga
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.tipoViatura
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: whitesmoke">'
                + elemento.dataHoraEntrada
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: whitesmoke">'
                + ""
                + '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: whitesmoke">'
                + elemento.comentEntrada
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.dataHoraSaida
                + '</td><td style="padding:0 15px 0 15px;background-color: whitesmoke">'
                + elemento.comentEntrada
                + "</td></tr>"
            )
            table += (
                '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
            )
            table += "</body></table>"

            subject, from_email, to = (
                "Alteração em Shipping - Portaria",
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )

            #'aroque1@visteon.com', 'npires2@visteon.com','pmarti30@visteon.com']

            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()
        return redirect("shippers:portaria")

    elif not request.user.groups.filter(name="shippingPortaria").exists():
        empresas = GatewayEmpresa.objects
        condutores = GatewayCondutor.objects
        primeiraMatricula = GatewayPrimeiraMatricula.objects
        segundaMtricula = GatewaySegundaMatricula.objects
        cargaDescarga = GatewayCargaDescarga.objects
        docas = GatewayDoca.objects
        destinosCarga = GatewayDestinoCarga.objects
        elementosGateway = Gateway.objects
        viatura = GatewayTipoViatura.objects.first()
        return render(
            request,
            "shippers/portaria.html",
            {
                "empresas": empresas,
                "condutores": condutores,
                "primeiraMatricula": primeiraMatricula,
                "segundaMtricula": segundaMtricula,
                "cargaDescarga": cargaDescarga,
                "docas": docas,
                "destinosCarga": destinosCarga,
                "elementosGateway": elementosGateway,
                "erro": "Não tem permissão para fazer alterações",
                "viatura": viatura,
            },
        )
    else:
        return redirect("shippers:portaria")


def submitPortaria(request):
    import threading


    
    if (
        request.method == "POST"
        and request.user.groups.filter(name="shippingPortaria").exists()
    ):
        empresa = request.POST["empresa"]
        if not GatewayEmpresa.objects.filter(nome=empresa).exists():
            GatewayEmpresa(None, empresa).save()
        condutor = request.POST["condutor"]
        if not GatewayCondutor.objects.filter(nome=condutor).exists():
            GatewayCondutor(None, condutor).save()
        condutorID = request.POST["condutorID"]
        if not GatewayCondutorID.objects.filter(nome=condutorID).exists():
            GatewayCondutorID(None, condutorID).save()
        matricula1 = request.POST["1matricula"]
        if not GatewayPrimeiraMatricula.objects.filter(nome=matricula1).exists():
            GatewayPrimeiraMatricula(None, matricula1).save()
        matricula2 = request.POST["2matricula"]
        if not GatewaySegundaMatricula.objects.filter(nome=matricula2).exists():
            GatewaySegundaMatricula(None, matricula2).save()
        cargaDescarga = request.POST["cargaDescarga"]
        if not GatewayCargaDescarga.objects.filter(nome=cargaDescarga).exists():
            GatewayCargaDescarga(None, cargaDescarga).save()
        doca = request.POST["doca"]
        if not GatewayDoca.objects.filter(nome=doca).exists():
            GatewayDoca(None, doca).save()
        destinoCarga = request.POST["destinoCarga"]
        if not GatewayDestinoCarga.objects.filter(nome=destinoCarga).exists():
            GatewayDestinoCarga(None, destinoCarga).save()
        contacto = request.POST["contacto"]
        if not GatewayContactoCondutor.objects.filter(nome=contacto).exists():
            GatewayContactoCondutor(None, contacto).save()
        dataHoraChegada = datetime.today().strftime("%Y-%m-%d %H:%M")
        viatura = GatewayTipoViatura.objects.first()

        novoElemento = Gateway(
            None,
            dataHoraChegada,
            empresa,
            condutor,
            condutorID,
            contacto,
            matricula1,
            matricula2,
            cargaDescarga,
            doca,
            destinoCarga,
            viatura,
            "",
            "",
            "",
            "",
            "",
            "",
        )

        """ now = datetime.datetime.now()
        formatted_timestamp = now.strftime("%Y-%m-%d %H:%M:%S.%f") """

        #print(formatted_timestamp)
        LastUpdatePortaria.objects.all().delete()
        updatePortaria = LastUpdatePortaria(
            dataUpdate =  datetime.today().strftime("%Y-%m-%d %H:%M:%S.%f")
        ).save()
        
        novoElemento.save()

      
         
        #-------------------///////////--Thred que serve para enviar um mail caso esta entrada não mude a hora ed entrada/saida-----///////

        print("ID NOV-->-->", novoElemento.id)
        t = threading.Thread(target=temporizadorEntradasPortaria,args=[novoElemento.id, novoElemento.dataHoraChegada,
            novoElemento.empresa, novoElemento.condutor,condutorID,contacto,
            novoElemento.primeiraMatricula, novoElemento.segundaMatricula, novoElemento.cargaDescarga, 
            novoElemento.doca, novoElemento.destinoCarga, novoElemento.tipoViatura, novoElemento.comentEntrada,
            novoElemento.comentSaida])
        t.start()
        #-------------------///////////------------------/////////////--------------------/////

        condutorP = ""
        condutorIDP = ""
        contactoP = ""
        empresaP = ""
        if GatewayCondutor.objects.filter(nome=condutor).exists():
            condutorP = GatewayCondutor.objects.get(nome=condutor)
        if GatewayCondutorID.objects.filter(nome=condutorID).exists():
            condutorIDP = GatewayCondutorID.objects.get(nome=condutorID)
        if GatewayContactoCondutor.objects.filter(nome=contacto).exists():
            contactoP = GatewayContactoCondutor.objects.get(nome=contacto)
        if GatewayEmpresa.objects.filter(nome=empresa).exists():
            empresaP = GatewayEmpresa.objects.get(nome=empresa)
        if GatewayInfoCondutor.objects.filter(
            Q(condutor=condutorP) | Q(condutorID=condutorIDP) | Q(contacto=contactoP)
        ).exists():
            GatewayInfoCondutor.objects.filter(
                Q(condutor=condutorP)
                | Q(condutorID=condutorIDP)
                | Q(contacto=contactoP)
            ).delete()
        novoCondutor = GatewayInfoCondutor(
            None, condutorP.id, condutorIDP.id, contactoP.id, empresaP.id
        )
        novoCondutor.save()
        GatewayTipoViatura.objects.all().delete()
        return redirect("shippers:portaria")
    else:
        empresas = GatewayEmpresa.objects
        condutores = GatewayCondutor.objects
        primeiraMatricula = GatewayPrimeiraMatricula.objects
        segundaMtricula = GatewaySegundaMatricula.objects
        cargaDescarga = GatewayCargaDescarga.objects
        docas = GatewayDoca.objects
        destinosCarga = GatewayDestinoCarga.objects
        elementosGateway = Gateway.objects
        #print("Novo Elemento Portaria->",novoElemento.id)


        viatura = GatewayTipoViatura.objects.first()
        novoCondutor.save()
        GatewayTipoViatura.objects.all().delete()
        return redirect("shippers:portaria")
    
          

        return render(
            request,
            "shippers/portaria.html",
            {
                "empresas": empresas,
                "condutores": condutores,
                "primeiraMatricula": primeiraMatricula,
                "segundaMtricula": segundaMtricula,
                "cargaDescarga": cargaDescarga,
                "docas": docas,
                "destinosCarga": destinosCarga,
                "elementosGateway": elementosGateway,
                "erro": "Não tem permissão para fazer alterações",
                "viatura": viatura,
            },
        )


def submitViatura(request):
    if request.method == "POST":
        posicao = request.POST["data"]

        if posicao == "":
            GatewayTipoViatura.objects.all().delete()
            return redirect("shippers:portaria")

        GatewayTipoViatura.objects.all().delete()
        viatura = GatewayTipoViatura()
        viatura.nome = posicao
        viatura.save()
        return redirect("shippers:portaria")


def comparaMatriculas(request):
    if request.method == "POST":
        matricula = request.POST["matricula"]
        actualDay = datetime.today().strftime("%Y-%m-%d")
        gateway = (
            Gateway.objects.filter(
                Q(primeiraMatricula__contains=matricula)
                | Q(segundaMatricula__contains=matricula)
            )
            .filter(abandono="", dataHoraSaida="")
            .distinct()
        )
        dadosQADabs = (
            AbsMstr.objects.filter(abs_shp_date=actualDay)
            .filter(Q(abs_id__startswith="s") | Q(abs_id__startswith="p"))
            .filter((Q(abs_type="s")) | (Q(abs_type="p")))
            .filter(abs_shipfrom="3515")
            .values(
                "abs_id",
                "abs_shp_date",
                "abs_shp_time",
                "abs_qad01",
                "oid_abs_mstr",
                "abs_status",
                "abs_shipto",
                "abs_item",
                "abs_domain",
            )
        )
        for elem in gateway.all():
            elem.estado = "verde"
            elem.save()
        return redirect("shippers:tracking2")


def changeUserGroups(request):
    if request.method == "POST":
        user = User.objects.get(username=request.POST["username"])
        grupo = request.POST["paginas"]
        if User.objects.filter(
            username=request.POST["username"], groups__name="shippingPortaria"
        ):
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="shippingPortariaReadOnly"
        ):
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="TrackingAdmin"
        ):
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="shippingSecurity"
        ):
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="shippingConfirmation"
        ):
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.remove(user)
        if grupo == "none":
            # message = 'User ' + user.username + ' perdeu acesso às páginas de Shipping.'
            # subject, from_email, to = 'Alteração em Receiving - Line Request - Configurations', 'noreply@visteon.com', [
            #     'aroque1@visteon.com']
            # msg = EmailMultiAlternatives(subject, message, from_email, to)
            # msg.attach_alternative(message, "text/html")
            # msg.send()
            return redirect("shippers:configurations")
        if grupo == "portaria":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
        if grupo == "tracking":
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
        if grupo == "security":
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "confirmation":
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)

        if grupo == "portaria/tracking":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
        if grupo == "portaria/security":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "portaria/confirmation":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/tracking":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/security":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/confirmation":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "tracking/security":
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "tracking/confirmation":
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "security/confirmation":
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)

        if grupo == "portaria/tracking/security":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "portaria/tracking/confirmation":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "portaria/security/confirmation":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/tracking/security":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/tracking/confirmation":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/security/confirmation":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "tracking/security/confirmation":
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)

        if grupo == "portaria/tracking/security/confirmation":
            my_group = Group.objects.using("default").get(name="shippingPortaria")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        if grupo == "portariaReadOnly/tracking/security/confirmation":
            my_group = Group.objects.using("default").get(
                name="shippingPortariaReadOnly"
            )
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="TrackingAdmin")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user)
        
        if grupo == "TrackingAdmin":
            my_group = Group.objects.using("default").get(
                name="ShippersTracking"
            )
            my_group.user_set.add(user)
            """ my_group = Group.objects.using("default").get(name="ShippersTracking")
            my_group.user_set.add(user) """
            """ my_group = Group.objects.using("default").get(name="shippingSecurity")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="shippingConfirmation")
            my_group.user_set.add(user) """
        # message = 'User ' + user.username + ' com páginas acessiveis - ' + grupo.replace("/", " , ")
        # subject, from_email, to = 'Alteração em Shipping - Configurations', 'noreply@visteon.com', [
        #     'aroque1@visteon.com']
        # msg = EmailMultiAlternatives(subject, message, from_email, to)
        # msg.attach_alternative(message, "text/html")
        # msg.send()
        return redirect("shippers:configurations")


def searchAllInfo(request):
    if request.method == "GET":
        resposta = [] 
        # Substituir este for por uma query raw que traga esses valoresporque pelos vistos a cada iteração do for é feita 
       
        todos = GatewayInfoCondutor.objects.raw("SELECT * FROM shippers_gatewayinfocondutor")


        
        for elem in todos:
            row = {
                "condutor": elem.condutor.nome,
                "condutorID": elem.condutorID.nome,
                "contacto": elem.contacto.nome,
                "empresa": elem.empresa.nome,
            }
            resposta.append(row)
        return JsonResponse({"resposta": resposta})


def searchEmpresa(request):
    with timer():
        if request.method == "GET":
            empresa = request.GET["empresa"]
            resposta = []
            if GatewayInfoCondutor.objects.filter(empresa__nome=empresa).exists():
                elem = GatewayInfoCondutor.objects.filter(empresa__nome=empresa)
                for elem1 in elem.all():
                    row = {
                        "condutor": elem1.condutor.nome,
                        "condutorID": elem1.condutorID.nome,
                        "contacto": elem1.contacto.nome,
                        "empresa": elem1.empresa.nome,
                    }
                    resposta.append(row)
            return JsonResponse({"resposta": resposta})

def searchCondutorExistente(request):
    with timer():
        if request.method == "GET":
            condutor = request.GET["condutor"]
            resposta = []
            if GatewayInfoCondutor.objects.filter(
                condutor__in=GatewayCondutor.objects.filter(nome=condutor)
            ).exists():
                elem1 = f"SELECT * FROM shippers_gatewaycondutor WHERE nome LIKE {condutor}"
                elem = GatewayInfoCondutor.objects.get(
                    condutor=GatewayCondutor.objects.get(nome=condutor)
                )
                row = {
                    "condutor": elem1.nome,
                    "condutorID": elem.condutorID.nome,
                    "contacto": elem.contacto.nome,
                    "empresa": elem.empresa.nome,
                }
                resposta.append(row)
            return JsonResponse({"resposta": resposta})

""" def searchCondutor(request):
    with timer():
        if request.method == "GET":
            condutor = request.GET["condutor"]
            resposta = []
            if GatewayInfoCondutor.objects.filter(
                condutor__in=GatewayCondutor.objects.filter(nome=condutor)
            ).exists():
                elem = GatewayInfoCondutor.objects.get(
                    condutor=GatewayCondutor.objects.get(nome=condutor)
                )
                row = {
                    "condutor": elem.condutor.nome,
                    "condutorID": elem.condutorID.nome,
                    "contacto": elem.contacto.nome,
                    "empresa": elem.empresa.nome,
                }
                resposta.append(row)
            return JsonResponse({"resposta": resposta}) """


def searchCondutor(request):
    print("ENTROu")
    with timer():
        if request.method == "POST":
            condutor = request.POST["condutor_id"]
            print("Condutor-->", condutor)
            condutor_id = GatewayCondutor.objects.raw(f"""SELECT id FROM shippers_gatewaycondutor WHERE nome = '{request.POST["condutor_id"]}'""")
            # for i in condutor_id:
            #     print("condutor ID --)",condutor_id, "I--)",i.id )
            resposta = []
            condutorRow = GatewayCondutor.objects.raw(f"""SELECT DISTINCT(shippers_gatewaycondutor.nome), shippers_gatewaycondutor.id
                                                    FROM shippers_gatewaycondutor
                                                    INNER join shippers_gatewayinfocondutor on shippers_gatewaycondutor.id = {1}
                                                    """)
            # for i in condutorRow:
            #     print("linha-->", i, i.nome)
            contactoRow = GatewayContactoCondutor.objects.raw(f""" SELECT DISTINCT(shippers_gatewaycontactocondutor.nome), shippers_gatewaycontactocondutor.id
                    FROM shippers_gatewaycontactocondutor
                    INNER JOIN shippers_gatewayinfocondutor on shippers_gatewayinfocondutor.contacto_id = shippers_gatewaycontactocondutor.id
                    INNER join shippers_gatewaycondutor on shippers_gatewayinfocondutor.condutor_id = {1} """)
            empresaCondutorRow = GatewayEmpresa.objects.raw(f""" SELECT DISTINCT(shippers_gatewayempresa.nome), shippers_gatewayempresa.id
                    FROM shippers_gatewayempresa
                    INNER JOIN shippers_gatewayinfocondutor on shippers_gatewayinfocondutor.empresa_id = shippers_gatewayempresa.id
                    INNER join shippers_gatewaycondutor on shippers_gatewayinfocondutor.condutor_id = {1} """)
            condutorIdRow = GatewayCondutorID.objects.raw(f""" SELECT DISTINCT(shippers_gatewaycondutorid.nome), shippers_gatewaycondutorid.id
                    FROM shippers_gatewaycondutorid
                    INNER JOIN shippers_gatewayinfocondutor on shippers_gatewayinfocondutor.condutorID_id = shippers_gatewaycondutorid.id
                    INNER join shippers_gatewaycondutor on shippers_gatewayinfocondutor.condutorID_id = {1} """)
            #for condutor_row, contacto_row, empresa_row, condutorId_Row in zip(condutorRow, contactoRow,empresaCondutorRow,condutorIdRow):

            #print("LInha-->", condutor_row,contacto_row, empresa_row, condutorId_Row )
                #print("LInha NOME-->", condutor_row.nome,contacto_row.nome, empresa_row.nome, condutorId_Row.nome )
            linha = {} #tens de fazer um dicionario de listas
            for i in contactoRow:
                linha = {"contacto":i}
                print("CONTACTO ->",i)
            for i in condutorIdRow:
                linha = {"condutorId":i}
                print("CONDUTORID->",i)
            for i in empresaCondutorRow:
                linha = {"empresa":i}
                print("EMPRESA ->",i)
            print("DICT COM LINHA -->", linha)
            """ row = {
                "condutor": elem.condutor.nome,
                "condutorID": elem.condutorID.nome,
                "contacto": elem.contacto.nome,
                "empresa": elem.empresa.nome,
            }
            resposta.append(row) """
            return JsonResponse({"resposta": resposta})


def searchCondutorID(request):
    with timer():
        if request.method == "GET":
            condutorID = request.GET["condutorID"]
            resposta = []
            if GatewayInfoCondutor.objects.filter(
                condutorID__in=GatewayCondutorID.objects.filter(nome=condutorID)
            ).exists():
                elem = GatewayInfoCondutor.objects.get(
                    condutorID=GatewayCondutorID.objects.get(nome=condutorID)
                )
                row = {
                    "condutor": elem.condutor.nome,
                    "condutorID": elem.condutorID.nome,
                    "contacto": elem.contacto.nome,
                    "empresa": elem.empresa.nome,
                }
                resposta.append(row)
            return JsonResponse({"resposta": resposta})


""" def searchContacto1(request):
    with timer():
        if request.method == "GET":
            contacto = request.GET["contacto"]
            resposta = []
            if GatewayInfoCondutor.objects.filter(
                contacto__in=GatewayContactoCondutor.objects.filter(nome=contacto)
            ).exists():
                elem = GatewayInfoCondutor.objects.get(
                    contacto=GatewayContactoCondutor.objects.get(nome=contacto)
                )
                row = {
                    "condutor": elem.condutor.nome,
                    "condutorID": elem.contacto.nome,
                    "contacto": elem.contacto.nome,
                    "empresa": elem.empresa.nome,
                }
                resposta.append(row)
            return JsonResponse({"resposta": resposta}) """

def searchContacto(request):
    with timer():
        if request.method == "POST":
            contacto = request.POST["contacto"]
            resposta = []
            condutor = GatewayCondutor.object.raw(f"""SELECT id FROM shippers_gatewaycondutor WHERE id = { request.POST["condutor"] }""")
            
            print("condutor.id----)", condutor.id)
            for i in condutor:
                print("condutor ID ---)",i)
            query = GatewayInfoCondutor.objects.raw(f""" select DISTINCT(shippers_gatewaycontactocondutor.nome) 
                    from shippers_gatewaycontactocondutor
                    INNER JOIN shippers_gatewayinfocondutor on shippers_gatewayinfocondutor.contacto_id = shippers_gatewaycontactocondutor.id
                    INNER join shippers_gatewaycondutor on shippers_gatewayinfocondutor.condutorID_id = {condutor.id} """)

            for i in query:
                row = {
                    "condutor": i.condutor.nome,
                    "condutorID": i.contacto.nome,
                    "contacto": i.contacto.nome,
                    "empresa": i.empresa.nome,
                }
                resposta.append(row)
            return JsonResponse({"resposta": resposta})



def setDataHoraSaida(request):
    if request.method == "POST":
        id = request.POST["id"]
        elem = Gateway.objects.get(id=id)
        if elem.dataHoraEntrada != "":
            elem.dataHoraSaida = datetime.today().strftime("%Y-%m-%d %H:%M")
        elem.save()
        print("A apagar todos os dados ta tabela")
        LastUpdatePortaria.objects.all().delete()
        updatePortaria = LastUpdatePortaria(
            dataUpdate =  datetime.today().strftime("%Y-%m-%d %H:%M:%S.%f")
        ).save()
        print("A guardar dados na lastUpdatePOrtaria")
        return redirect("shippers:portaria")


def setDataHoraEntrada(request):
    if request.method == "POST":
        id = request.POST["id"]
        elem = Gateway.objects.get(id=id)
        elem.dataHoraEntrada = datetime.today().strftime("%Y-%m-%d %H:%M")
        elem.save()
        print("A apagar todos os dados ta tabela")
        LastUpdatePortaria.objects.all().delete()
        updatePortaria = LastUpdatePortaria(
            dataUpdate =  datetime.today().strftime("%Y-%m-%d %H:%M:%S.%f")
        ).save()
        print("A guardar dados na lastUpdatePOrtaria")
        return redirect("shippers:portaria")


def mudaEstadoFiltro(request):
    with timer():
        print("MUDA Estado req->", request.POST)
        if request.method == "POST":
            posicao = request.POST["posicao"]
            estado = request.POST["estado"]

            if posicao == "entrada":
                if estado == "on":
                    request.session["entrada"] = "entrada"
                else:
                    del request.session["entrada"]
            #     botao.entrada = estado
            if posicao == "saida":
                if estado == "on":
                    request.session["saida"] = "saida"
                else:
                    del request.session["saida"]
            #     botao.saida = estado
            # botao.save()

            return redirect("shippers:portaria")


def downloadExcel(request):
    if request.method == "GET":
        elementos = Gateway_View.objects.all() # Gateway.objects

        caminho = "media/shippers/portaria"
        if os.path.exists(caminho):
            for entry in os.listdir(caminho):
                if os.path.isfile(os.path.join(caminho, entry)):
                    os.remove(caminho + "/" + entry)

        wbProduction = Workbook()
        sheetProduction = wbProduction.add_sheet("Sheet 1")

        row = 0
        col = 0
        sheetProduction.write(row, col, "Data/hora chegada")
        sheetProduction.write(row, col + 1, "Condutor")
        sheetProduction.write(row, col + 2, "ID Condutor")
        sheetProduction.write(row, col + 3, "Contacto")
        sheetProduction.write(row, col + 4, "Empresa")
        sheetProduction.write(row, col + 5, "1ª matricula")
        sheetProduction.write(row, col + 6, "2ª matricula")
        sheetProduction.write(row, col + 7, "Carga/descarga")
        sheetProduction.write(row, col + 8, "Doca")
        sheetProduction.write(row, col + 9, "Destino carga")
        sheetProduction.write(row, col + 10, "Tipo viatura")
        sheetProduction.write(row, col + 11, "Data/hora entrada")
        sheetProduction.write(row, col + 12, "Abandono")
        sheetProduction.write(row, col + 13, "Comentários entrada")
        sheetProduction.write(row, col + 14, "Data/hora saida")
        sheetProduction.write(row, col + 15, "Comentários saída")
        row += 1

        for elem in elementos.all():
            sheetProduction.write(row, col, elem.dataHoraChegada)
            sheetProduction.write(row, col + 1, elem.condutor)
            sheetProduction.write(row, col + 2, elem.ident)
            sheetProduction.write(row, col + 3, elem.contacto)
            sheetProduction.write(row, col + 4, elem.empresa)
            sheetProduction.write(row, col + 5, elem.primeiraMatricula)
            sheetProduction.write(row, col + 6, elem.segundaMatricula)
            sheetProduction.write(row, col + 7, elem.cargaDescarga)
            sheetProduction.write(row, col + 8, elem.doca)
            sheetProduction.write(row, col + 9, elem.destinoCarga)
            sheetProduction.write(row, col + 10, elem.tipoViatura)
            sheetProduction.write(row, col + 11, elem.dataHoraEntrada)
            sheetProduction.write(row, col + 12, elem.abandono)
            sheetProduction.write(row, col + 13, elem.comentEntrada)
            sheetProduction.write(row, col + 14, elem.dataHoraSaida)
            sheetProduction.write(row, col + 15, elem.comentSaida)
            row += 1
        wbProduction.save(
            "media\\shippers\\portaria\\workbookPortaria.xls"
        )

        return redirect("shippers:portaria")


def uploadDataPortaria(request):
    if request.method == "POST":
        if request.FILES.get("ficheiro"):

            ficheiro = [request.FILES["ficheiro"]]

            workbook = openpyxl.load_workbook(ficheiro[0])
            worksheet = workbook["Sheet 1"]

            for element in worksheet:
                if element[0].value != None and element[1].value != "Condutor":
                    novoElemento = Gateway(
                        None,
                        element[0].value,
                        element[4].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        "feitoEsaiu",
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                    )
                    novoElemento.save()

        return redirect("shippers:portaria")


def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    return "%d:%02d:%02d" % (hour, minutes, seconds)


def reportPortaria(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Shippers - Portaria",
            "noreply@visteon.com",
            ["pmarti30@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("shippers:portaria")


def reportTracking(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Shippers - Tracking",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("shippers:tracking")


def reportSecurity(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Shippers - Security",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("shippers:security")


def reportConfirmation(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Shippers - Confirmation",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("shippers:shippersConfirmation")

 

def definicaoTempoEmailAtraso(request):
    print("REQ----->",request.POST )
    

    numeroHoras = NumeroHorasEmailDiario.objects #FAZER O SAVE COMO ESTÁ NO CROSS DOCKING L1532
    numeroHoras.all().delete()
    numeroHoras = NumeroHorasEmailDiario(numHoras1 = float(request.POST["numHoras1"]), numHoras2 = float(request.POST["numHoras2"]))
    numeroHoras.save()
   
    return JsonResponse({"message": "OK"})


def temporizadorEntradasPortaria(*args):
    import time
    # tens de ir buscar o id da nova entrada
    # ir buscar o id dessa nova entrada na base de dados7
    numHoras = NumeroHorasEmailDiario.objects.first()
  

    #print("NUMERO DE HORAS-->", definicaoTempoEmailAtraso(var_tempo_email))
    #id_bd = Gateway.objects.raw(f"SELECT * FROM shippers_gateway WHERE dataHoraChegada = {args[0]} AND empresa = {args[1]} AND condutor = {args[2]} AND condutor = {args[4]} AND primeiraMatricula = {args[7]} AND segundaMatricula = {args[8]}")
    """ id_bd1 = Gateway.objects.filter(dataHoraChegada = args[0], empresa = args[1], condutor = args[2],
    contacto = args[4], primeiraMatricula = args[5] , segundaMatricula = args[6], cargaDescarga =  args[7], doca = args[8],
    destinoCarga =  args[9])  """
    id_bd = Gateway.objects.raw(f"SELECT * FROM shippers_gateway WHERE id = {args[0]}")
    num_horas = NumeroHorasEmailDiario.objects
    numHoras = NumeroHorasEmailDiario.objects.first()
    print("SEGUNDO CICLO",numHoras.numHoras2)
    time.sleep(int(numHoras.numHoras1) * 60)

    dados_atuais = {
        "ident": "",
        "dataHoraChegada": "",
        "empresa": "",
        "condutor": "",
        "contacto": "",
        "condutorID": "",
        "segundaMatricula": "",
        "cargaDescarga": "",
        "doca": "",
        "destinoCarga": "",
        "tipoViatura": "",
        "comentEntrada": "",
        "comentSaida" : "",
        "dataHoraSaida" : "",
        "abandono" : ""
    }
    print("Tamanho bd-> ",len(id_bd))
    for i in id_bd:
        print("Printando o i",i)
        print("dataHoraChegada -)",i.dataHoraChegada)
        dados_atuais["ident"] = i.ident
        dados_atuais["dataHoraChegada"] = i.dataHoraChegada
        dados_atuais["empresa"] = i.empresa
        dados_atuais["condutor"] = i.condutor
        dados_atuais["contacto"] = i.condutor
        dados_atuais["primeiraMatricula"] = i.primeiraMatricula
        dados_atuais["dataHoraEntrada"] = i.dataHoraEntrada
        #dados_atuais["condutorID"] = i.condutorID dataHoraEntrada
        dados_atuais["segundaMatricula"] = i.segundaMatricula
        dados_atuais["cargaDescarga"] = i.cargaDescarga
        dados_atuais["doca"] = i.doca
        dados_atuais["destinoCarga"] = i.destinoCarga
        dados_atuais["tipoViatura"] = i.tipoViatura
        dados_atuais["comentEntrada"] = i.comentEntrada
        dados_atuais["comentSaida"] = i.comentSaida
        dados_atuais["dataHoraSaida"] = i.dataHoraSaida
        dados_atuais["abandono"] = i.abandono
        print("-> ", dados_atuais["abandono"], " <-")
    estado = 0
    print("numHoras -->", numHoras.numHoras1, numHoras.numHoras2)
    while len(dados_atuais["dataHoraSaida"]) == 0 and int(numHoras.numHoras1) != 0 or int(numHoras.numHoras2) !=0:
        numHoras = NumeroHorasEmailDiario.objects.first()
        print("numHoras -->", int(numHoras.numHoras1), int(numHoras.numHoras2))
        
        id_bd = Gateway.objects.raw(f"SELECT * FROM shippers_gateway WHERE {args[0]} = id ")
        print("A Ddentro do while",dados_atuais["dataHoraChegada"])
        for i in id_bd:
            print("dataHoraChegada -)",i.dataHoraChegada)
            dados_atuais["ident"] = i.ident
            dados_atuais["dataHoraChegada"] = i.dataHoraChegada
            dados_atuais["empresa"] = i.empresa
            dados_atuais["condutor"] = i.condutor
            dados_atuais["contacto"] = i.condutor
            dados_atuais["primeiraMatricula"] = i.primeiraMatriculaguardaFicheiro
            dados_atuais["dataHoraEntrada"] = i.dataHoraEntrada
            #dados_atuais["condutorID"] = i.condutorID dataHoraEntrada
            dados_atuais["segundaMatricula"] = i.segundaMatricula
            dados_atuais["cargaDescarga"] = i.cargaDescarga
            dados_atuais["doca"] = i.doca
            dados_atuais["destinoCarga"] = i.destinoCarga
            dados_atuais["tipoViatura"] = i.tipoViatura
            dados_atuais["comentEntrada"] = i.comentEntrada
            dados_atuais["comentSaida"] = i.comentSaida
            dados_atuais["dataHoraSaida"] = i.dataHoraSaida
            dados_atuais["abandono"] = i.abandono
        print("-> ", dados_atuais["ident"], " <-")
        
        if len(dados_atuais["dataHoraEntrada"]) == 0 and estado == 0 :
            print("Dados não foram alterados, vai mandar email", id_bd)
            table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'
            table += '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["dataHoraChegada"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["condutor"] ) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["ident"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["contacto"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["empresa"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["primeiraMatricula"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["segundaMatricula"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["cargaDescarga"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["doca"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["destinoCarga"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["tipoViatura"]) + \
                    '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: yellow">' + str(dados_atuais["dataHoraEntrada"]) + \
                    '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: yellow">' + str(dados_atuais["comentEntrada"])+ \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["dataHoraSaida"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: yellow">' + str(dados_atuais["comentSaida"]) + '</td></tr>'
            table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
            table += '</body></table>'
            #------------//////////--------------------------------------//////-----------------
            

            subject, from_email, to = (
                "Atraso na entrada/saída - Portaria",
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )

            # [ 'aroque1@visteon.com', 'npires2@visteon.com', 'pmarti30@visteon.com']

            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()
            print("Enviou Email passadas 3 horas")
            estado = 1
        
        time.sleep(int(numHoras.numHoras2) * 60)
        id_bd = Gateway.objects.raw(f"SELECT * FROM shippers_gateway WHERE {args[0]} = id ")
        for i in id_bd:
            dados_atuais["ident"] = i.ident
            dados_atuais["dataHoraChegada"] = i.dataHoraChegada
            dados_atuais["empresa"] = i.empresa
            dados_atuais["condutor"] = i.condutor
            dados_atuais["contacto"] = i.condutor
            dados_atuais["primeiraMatricula"] = i.primeiraMatricula
            dados_atuais["dataHoraEntrada"] = i.dataHoraEntrada
            #dados_atuais["condutorID"] = i.condutorID dataHoraEntrada
            dados_atuais["segundaMatricula"] = i.segundaMatricula
            dados_atuais["cargaDescarga"] = i.cargaDescarga
            dados_atuais["doca"] = i.doca
            dados_atuais["destinoCarga"] = i.destinoCarga
            dados_atuais["tipoViatura"] = i.tipoViatura
            dados_atuais["comentEntrada"] = i.comentEntrada
            dados_atuais["comentSaida"] = i.comentSaida
            dados_atuais["dataHoraSaida"] = i.dataHoraSaida
        numHoras = NumeroHorasEmailDiario.objects.first()
        if len(dados_atuais["dataHoraSaida"])  == 0  and estado ==1 and int(numHoras.numHoras1) != 0 or int(numHoras.numHoras2) !=0 :
            print("Dados não foram alterados e passaram 1 horas", str(dados_atuais["dataHoraChegada"]))

            table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'
            table += '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["dataHoraChegada"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["condutor"] ) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["ident"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["contacto"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["empresa"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["primeiraMatricula"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["segundaMatricula"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["cargaDescarga"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["doca"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["destinoCarga"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["tipoViatura"]) + \
                    '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: orange">' + str(dados_atuais["dataHoraEntrada"]) + \
                    '</td><td style="padding:0 15px 0 15px; text-align: center;background-color: orange">' + str(dados_atuais["comentEntrada"])+ \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["dataHoraSaida"]) + \
                    '</td><td style="padding:0 15px 0 15px;background-color: orange">' + str(dados_atuais["comentSaida"]) + '</td></tr>'
            table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
            table += '</body></table>'

            #-----------------------//////------------------------------/////-------------------------
            

            subject, from_email, to = (
                "Atraso na entrada/saída - Portaria",
                "noreply@visteon.com",
                ["pmarti30@visteon.com"],
            )

            # [ 'aroque1@visteon.com', 'npires2@visteon.com', 'pmarti30@visteon.com']

            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()
            print("Enviou Email de hora em hora")
    print("Fim do loop")
    print("Thread concluida")
    
    # se dados não foram alterados
    
    #envia mail

   #para a thread e sai 



# Edita uma linha da table table(html) e faz edição na db
""" @login_required()
@user_passes_test(
    lambda u: u.groups.filter(Q(name="TrackingAdmin")).exists()) """
def trackingPage (request):
    import datetime
    from dateutil import parser
    with timer():
    
        #tracking = TrackingPage.objects.raw("SELECT * FROM shippers_trackingpage")
        tracking = TrackingPage.objects.all()
        actualDay_menos2 = datetime.datetime.now() - datetime.timedelta(days=2) #date
        print(type(actualDay_menos2))
        actualDay_menos2 = actualDay_menos2.strftime("%Y-%m-%d")  #str
        actualDay_menos2 = parser.parse(actualDay_menos2)
        actualDay_menos2 = actualDay_menos2.date()
        print("DATA 2-> ", actualDay_menos2, type(actualDay_menos2))


    #A função  demorou 0.0015433999999991954s
    #datetime_object = datetime.strptime(my_date_string, '%b %d %Y %I:%M%p')
    #actualDay_menos2 = time.strptime(actualDay_menos2, "%Y-%m-%d")
    """ print("DIA----> ",actualDay_menos2, type(actualDay_menos2))
    datetime_object = datetime .strptime(actualDay_menos2, "%Y-%m-%d")
    print("DATA 2-> ", datetime_object, type(datetime_object)) """
   

        #print("DATA->" ,i.inicioPrep, type(i.inicioPrep)," + ", actualDay_menos2, type(actualDay_menos2) )
    return render(
        request,
        "shippers/tracking.html",
        {
            "items": tracking,
            "actualDay_menos2": actualDay_menos2,
        },
    )

    
def trigger_error(request):
    division_by_zero = 1 / 0


def get_time(value):
    if not value:
        return None
    try:
        float(value)
    except ValueError:
        return datetime.strptime(value, "%Y-%m-%d %H:%S")
    else:
        return datetime.fromtimestamp(float(value))

def addNewRowTracking(request):
    if request.method == "POST":
        id = request.POST.get("id")
        if id:
            query = TrackingPage.objects.get(id=id)
        else:
            query = TrackingPage()
        if (nShipper := request.POST.get("nShipper")):
            query.nShipper = nShipper
        if (qtyCaixas := request.POST.get("qtyCaixas")):
            query.qtyCaixas = qtyCaixas
        if (inicioPrep := get_time(request.POST.get("inicioPrep"))):
            query.inicioPrep = inicioPrep
        if (fimPrep := get_time(request.POST.get("fimPrep"))):
            query.fimPrep = fimPrep

        if (comentarios := request.POST.get("comentarios")):
            query.comentarios = comentarios
        query.save()
    return JsonResponse({"message": "OK"})


def addLine(request):
        dataEmpty = TrackingPage()
        dataEmpty.save()
        return redirect("shippers:trackingPage")

 
def guardaFicheiroHistorico(): #guardaFicheiroHistorico
    import openpyxl as xl
    print("ENtrou no scheduler")
    #esta func vai correr todos os dias 1 do mes e vai guardar dados referentes aos 2 meses anteriores
    #VER COMO FUNCIONsA O SCHEDULER no DJANGO
    #1-> criar o excel com os dados certos
    #2_> agarrar no ficheiro e fguardar numa dererminada pasta
    #3-> por a cena do calendario a dar trigger e depois definir para o primeiro dia do mes(já está em principio) "C:\\Users\\PMARTI30\Desktop\\fileTeste1"
    elementos_historico = TrackingPage.objects.raw("SELECT  * FROM shippers_trackingpage WHERE MONTH(inicioPrep) <=  MONTH(DATEADD(MONTH, -2, CURRENT_TIMESTAMP))")
    wbProduction = Workbook()
    file = io.BytesIO()
    sheetProduction = wbProduction.get_sheet_by_name("Sheet")

    sheetProduction.append(
        [
            "Nº Shipper",
            "QT Cx",
            "Inicio Prep",
            "Fim Prep",
            "Ship Date",
            "Ship Time",
            "Carrier",
            "Comentários"
        ]
        )

    for elem in elementos_historico: # em principio não precisa do all()
        sheetProduction.append(
            [
                str(elem.nShipper),
                str(elem.qtyCaixas),
                str(elem.inicioPrep),
                str(elem.fimPrep),
                str(elem.comentarios),
                str(elem.ship_date),
                str(elem.ship_time),
                str(elem.ship_carrier),
            ]
        )

    #wbProduction.save(caminho)
    # Voltar a pocição inicial do IO file object
    file.seek(0)

    data_fmt = datetime.now().strftime("%d-%m-%Y")
    caminho =  "C:\\Users\\PMARTI30\\Desktop\\historicoTrackingPage"+data_fmt+".xlsx"
    caminho1 =  "\\\\pavpd002\\e_proj\sharedir\\MP&L\\PROCEDIMENTOS\\Packaging\\TMP_NP\\historicoTrackingPage"+data_fmt+".xlsx"

     
    wbProduction.save(caminho)
    wbProduction.save(file)
    print("FICHEIRO GUARDADO")
    #wb=xl.Workbook()
    #wb.save(f"C:\\Users\\PMARTI30\Desktop\\TrackingPage_Historico2Mes_{data_fmt}.xlsx")
    #caminho =  "C:\\Users\\PMARTI30\Desktop\\fileTeste1"
    #wb.save(file)
    #file.open(caminho+"teste", fresp)

    #return fresp

 
 

def botaoDadosQAD(request):
    from qad_ee.models import AbsMstr, AbscDet
    from datetime import timedelta
    from django.http import HttpResponseNotFound
    #aqui vão ser retornados mais valores do QAD, mas estes não são para serem msotrados na tablea, apenas serão mostrados quando é fieto o download do ficheiro para excel
    

    nShipper = request.POST["nShipper"]
    if not nShipper:
        return HttpResponseNotFound(
            "nShipper não pode estar vazio"
            )
    results_mstr = AbsMstr.objects.raw(f"SELECT * FROM abs_mstr WHERE abs_domain = 3511010 AND abs_id LIKE '_{nShipper}'")
    results_det = AbscDet.objects.raw(f"SELECT * FROM absc_det WHERE absc_domain = 3511010 AND absc_abs_id LIKE '_{nShipper}'")
    abs_mstr = list(results_mstr)
    absc_det = list(results_det)
    print("RESULTS MSTR-->", abs_mstr)
    print("RESULTS MSTR-->", abs_mstr)
    id_tracking = request.POST["id"]
    
    if not abs_mstr or not absc_det:
        return HttpResponseNotFound(
            "Não foi encontrado nenhum value "
            )
    else:
        ref = TrackingPage.objects.get(id=id_tracking)
        for i in abs_mstr:
            abs_shp_date = i.abs_shp_date
            abs_shp_time = i.abs_shp_time
            convert = str(timedelta(seconds=abs_shp_time))
            convertData= str(abs_shp_date).split(" ")
            ref.ship_date =  convertData[0]
            ref.ship_time = convert
            print("DATA->",  convertData[0])
        for i in absc_det:
            absc_carrier= i.absc_carrier
            ref.ship_carrier = absc_carrier
        ref.confirmacao = datetime.now()
        ref.save()
        print("GUARDOU!")
    return redirect("shippers:trackingPage")


def downloadExcelHistoricoTracking(request): # downloadExcelHistoricoTracking
    #aqui ele tem de sair pelo menos os dados do ultimo mês, até dois dias antes do dia atual
    #caso apanhes btns ou campos vazios/Empty, no ficheiro tem de aparecer vazio ou mete o texto dos btns
    import io
    if request.method == "GET":
        elementos_historico = TrackingPage.objects.raw("SELECT  * FROM shippers_trackingpage WHERE MONTH(inicioPrep) >=  MONTH(DATEADD(MONTH, -1, CURRENT_TIMESTAMP))")
        wbProduction = Workbook()
        file = io.BytesIO()
        # Não criar um sheet novo e ter um existente ;)
        sheetProduction = wbProduction.get_sheet_by_name("Sheet")
        sheetProduction.append(
            [
                "Nº Shipper",
                "QT Cx",
                "Inicio Prep",
                "Fim Prep",
                "Ship Date",
                "Ship Time",
                "Carrier",
                "Comentários"
            ]
        )
        for elem in elementos_historico: 
            print("RLEMS--->", elem.confirmacao)
            sheetProduction.append(
                [
                    str(elem.nShipper),
                    str(elem.qtyCaixas),
                    str(elem.inicioPrep),
                    str(elem.fimPrep),
                    str(elem.comentarios),
                    str(elem.ship_date),
                    str(elem.ship_time),
                    str(elem.ship_carrier),
                ]
            )

        wbProduction.save(file)
        # Voltar a pocição inicial do IO file object
        file.seek(0) 

        data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
        fresp = FileResponse(
            file, filename=f"TrackingPage_Historico1Mes_{data_fmt}.xlsx", as_attachment=True
        )
        return fresp
    


def deleteRowTracking(request):
    stockPackagePack = TrackingPage.objects
    if request.method == "POST":
        print("REQ",request.POST)
        row_id = request.POST["rowIdDelete"]
        print(stockPackagePack.all().get(id=row_id))
        stockPackagePack.all().get(id=row_id).delete()
    return redirect("shippers:trackingPage")


def mediaMensalPortaria(request):
    import calendar
    import datetime
    from datetime import date
    print("CHEGAS_TE á func da media total")
    now = datetime.datetime.now()
    media_total = MediaTotalPortaria()
    #media_empresas = Gateway.objects.raw("SELECT 1 id,COUNT(id) as contagem FROM shippers_gateway AS contagem WHERE dataHoraChegada BETWEEN '2022-10-11' AND '2022-11-11'")
    media_empresas1 = Gateway.objects.raw('SELECT 1 id, COUNT(id) as contagem FROM shippers_gateway WHERE dataHoraChegada BETWEEN DATEADD(d,1,EOMONTH(GETDATE(),-2)) AND CONVERT(DATE,GETDATE())')
    print("acabou a query",media_empresas1)
    for i in media_empresas1:
        print("i->",i.contagem)
        soma = i.contagem
    media_total.media = round(soma / calendar.monthrange(now.year, now.month-2)[1],2)
    media_total.data = date.today().strftime('%Y-%m-%d') # BUG está a stressar com as datas
    media_total.save()
    #copiar esta func, ver se dá para mudar o valor do dicionario logo co a media dessa empresa
    return JsonResponse({"message": "OK"})

 

def mediaMensalEmpresa(request):
    import calendar
    import datetime
    from datetime import date
    dict_empresas = Gateway.objects.raw("SELECT DISTINCT(empresa), 1 id, COUNT(id) as contagem FROM shippers_gateway  WHERE dataHoraChegada BETWEEN DATEADD(d,1,EOMONTH(GETDATE(),-2)) AND CONVERT(DATE,GETDATE())GROUP BY empresa")
    now = datetime.datetime.now()
    nDias_mesPassado = calendar.monthrange(now.year, now.month-2)[1]
    for i in dict_empresas: #se calhar podes guardar logo na bd
        print("Printando o I",i.empresa, "\n", i.contagem)
        nova_linha = MediaMensalEmpresa()
        nova_linha.empresa = i.empresa
        nova_linha.media = round(i.contagem / calendar.monthrange(now.year, now.month-2)[1],2)
        nova_linha.data = date.today().strftime('%Y-%m-%d')
        nova_linha.save()
        print("Nova linhaBD",nova_linha.empresa, nova_linha.media)
    #Fazer a parte da func para criar e atualizar o excel se já existir
    #Fazer scheduler para mensalmente ele correr esta func (NO CROSSDOCKING)
    return JsonResponse({"message": "OK"})
       
""" def createExcelPortaria(request):
    # não tas a apanhar os valores para fazer as comparações para os pores no DataValidation
    if request.method == "POST":
        print("REQUEST",request.POST["start_date"],request.POST["end_date"])
        #recebes as datas no post nas variaveis start_date AND end_date 
        # ir na bd receber os dados que estão nesse entrevalo de datas

        from datetime import datetime
        import io
        from openpyxl import load_workbook
        from openpyxl import Workbook
        elementos = Gateway.objects.raw(f""SELECT * FROM shippers_gateway WHERE dataHoraChegada BETWEEN '{request.POST["start_date"]}' AND '{request.POST["end_date"]}' "")
        wbProduction = Workbook()
        file = io.BytesIO()
        caminho = "C:/visteon/media"
        
        # Não criar um sheet novo e ter um existente ;)
        sheetProduction = wbProduction.worksheets[0]

        # Verificar se ele está a ir buscar os campos à view ou aos models
        sheetProduction.append(
            [
                "Data/Hora Chegada",
                "Condutor",
                "ID",
                "Contacto",
                "Empresa",
                "Primeira Matricula",
                "Segunda Matricula",
                "Doca",
                "Destino Carga",
                "Tipo Viatura",
                "Data/Hora Entrada",
                "Abandono",
                "Comentario Entrada",
                "Data/Hora Saida",
                "Comentario Saida",
                "Estado",
            ]
        )

        for i in elementos:
            print("I ->", i.condutor)
            sheetProduction.append(
                [
                    str(i.dataHoraChegada),
                    str(i.condutor),
                    str(i.ident),
                    str(i.contacto),
                    str(i.empresa),
                    str(i.primeiraMatricula),
                    str(i.segundaMatricula),
                    str(i.doca),
                    str(i.destinoCarga),
                    str(i.tipoViatura),
                    str(i.dataHoraEntrada),
                    str(i.abandono),
                    str(i.comentEntrada),
                    str(i.dataHoraSaida),
                    str(i.comentSaida),
                    str(i.estado),
                ]
            )
        wbProduction.save(
            "C:\\visteon\\media\\workbookPortariaRestricaoData.xls"
        )
        return redirect("shippers:portaria")
        #wbProduction.save(file)
        # Voltar a pocição inicial do IO file object
        file.seek(0)

        wbProduction.save(
            "C:\\visteon\\media\\shippers\\portaria\\workbookPortaria.xls"
        )

        data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
        print("FILETHING ---)", wbProduction)
        print("FILETHING ---)", elementos)
        fresp = FileResponse(
            file, content_type="application/vnd.ms-excel",
            filename=f"Portaria_Calendarizado_{data_fmt}.xlsx", as_attachment=True
        )
        print("DEVERIA TER FEITO O DOWNLOAD DO FICHEIRO", fresp.close)
   
        return fresp """

def createExcelPortaria(request):
    print("ENTROU")
    print("req->",request.POST)
    if request.method == "POST":
        data_str_start = request.POST["start_date"]
        data_str_end = request.POST["end_date"]
        
        data_str_start_parts = data_str_start.split(' GMT')[0]
        data_str_end_parts = data_str_end.split(' GMT')[0]

        data_objStart = datetime.strptime(data_str_start_parts[:25], '%a %b %d %Y %H:%M:%S')
        data_objEnd = datetime.strptime(data_str_end_parts[:25], '%a %b %d %Y %H:%M:%S')

        elementos = Gateway.objects.raw(f"""SELECT * FROM shippers_gateway WHERE dataHoraChegada BETWEEN '{data_objStart.strftime('%Y-%m-%d')} 00:00' AND '{data_objEnd.strftime('%Y-%m-%d')} 24:00' """)
        print(elementos)
        if len(elementos) == 0:
            print("não trouxe dados,tens de ir buscar à tabela do shippers_backup")
            elementos = Gateway.objects.raw(f"""SELECT * FROM shippers_gatewaybackup WHERE dataHoraChegada BETWEEN '{data_objStart.strftime('%Y-%m-%d %H:%M')} 00:00' AND '{data_objEnd.strftime('%Y-%m-%d %H:%M')} 24:00' """)

        """ caminho = "C:\\Users\\PMARTI30\\Desktop\\visteon\\media\\" """
        """ if os.path.exists(caminho):
            for entry in os.listdir(caminho):
                if os.path.isfile(os.path.join(caminho, entry)):
                    os.remove(caminho + "/" + entry) """

        wbProduction = Workbook()
        sheetProduction = wbProduction.add_sheet("Sheet 1")

        row = 0
        col = 0
        sheetProduction.write(row, col, "Data/hora chegada")
        sheetProduction.write(row, col + 1, "Condutor")
        sheetProduction.write(row, col + 2, "ID Condutor")
        sheetProduction.write(row, col + 3, "Contacto")
        sheetProduction.write(row, col + 4, "Empresa")
        sheetProduction.write(row, col + 5, "1ª matricula")
        sheetProduction.write(row, col + 6, "2ª matricula")
        sheetProduction.write(row, col + 7, "Carga/descarga")
        sheetProduction.write(row, col + 8, "Doca")
        sheetProduction.write(row, col + 9, "Destino carga")
        sheetProduction.write(row, col + 10, "Tipo viatura")
        sheetProduction.write(row, col + 11, "Data/hora entrada")
        sheetProduction.write(row, col + 12, "Abandono")
        sheetProduction.write(row, col + 13, "Comentários entrada")
        sheetProduction.write(row, col + 14, "Data/hora saida")
        sheetProduction.write(row, col + 15, "Comentários saída")
        row += 1

        for elem in elementos:
            print("ELEMENTOS-->",elem.condutor)
            sheetProduction.write(row, col, elem.dataHoraChegada)
            sheetProduction.write(row, col + 1, elem.condutor)
            sheetProduction.write(row, col + 2, elem.ident)
            sheetProduction.write(row, col + 3, elem.contacto)
            sheetProduction.write(row, col + 4, elem.empresa)
            sheetProduction.write(row, col + 5, elem.primeiraMatricula)
            sheetProduction.write(row, col + 6, elem.segundaMatricula)
            sheetProduction.write(row, col + 7, elem.cargaDescarga)
            sheetProduction.write(row, col + 8, elem.doca)
            sheetProduction.write(row, col + 9, elem.destinoCarga)
            sheetProduction.write(row, col + 10, elem.tipoViatura)
            sheetProduction.write(row, col + 11, elem.dataHoraEntrada)
            sheetProduction.write(row, col + 12, elem.abandono)
            sheetProduction.write(row, col + 13, elem.comentEntrada)
            sheetProduction.write(row, col + 14, elem.dataHoraSaida)
            sheetProduction.write(row, col + 15, elem.comentSaida)
            row += 1
            #dá para mudar a pasta onde queres guardar o ficheiro
        wbProduction.save(
            "media\\shippers\\portaria\\workbookPortariaEntreData.xls"
        )
        

        return redirect("shippers:portaria")
""" def createExcelPortaria(request):
    print("REQUEST",request.POST["start_date"],request.POST["end_date"])
    #recebes as datas no post nas variaveis start_date AND end_date 
    # ir na bd receber os dados que estão nesse entrevalo de datas

    from datetime import datetime
    import io
    import openpyxl
    from openpyxl import Workbook
    elementos = Gateway.objects.raw(f"SELECT * FROM shippers_gateway WHERE dataHoraChegada BETWEEN '{request.POST["start_date"]}' AND '{request.POST["end_date"]}' ")
    wbProduction = Workbook()
    file = io.BytesIO()
    # Não criar um sheet novo se já existente
    #tentar abrir o ficheiro
    #se não esistente
    workbook = load_workbook(file.open("rb"), data_only = True)

    sheetProduction = wbProduction.get_sheet_by_name("Sheet")
    print("a entrar no for")
    lista=[]
    for i in elementos:
        print("A printar o i",i)
        lista.append(i.dataHoraChegada)
        lista.append(i.condutor)
        lista.append(i.ident)
        lista.append(i.contacto)
        lista.append(i.empresa)
        lista.append(i.primeiraMatricula)
        lista.append(i.segundaMatricula)
        lista.append(i.cargaDescarga)
        lista.append(i.doca)
        lista.append(i.destinoCarga)
        lista.append(i.tipoViatura)
        lista.append(i.dataHoraEntrada)
        lista.append(i.abandono)
        lista.append(i.comentEntrada)
        lista.append(i.dataHoraSaida)
        lista.append(i.comentSaida)
        lista.append(i.estado)  

        lista.append("Dia")
    sheetProduction.append(
            lista
        )
    
    
    print("saiu do segundo for")
    #Guardar o ficheiro numa pasta especifica
    wbProduction.save(file)
    # Voltar a pocição inicial do IO file object
    file.seek(0)
    print("guardou")
    data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
    #seguinte codigo serve para fazer com que seja transferido o ficheiro
    fresp = FileResponse(
        file, filename=f"MediaEmpresasMensal_{data_fmt}.xlsx", as_attachment=True
    )
    return fresp
    #se já existir
    #book = openpyxl.load_workbook('sample.xlsx') 
    #dár append dos dados do mes novo ao excel já existente
 """


def deleteLinhaPortaria(request):
    # TENS DE MONTAR UM EMAIL COM TODOS OS CABEÇALHOS PARA SER ENVIADO PARA O EMAIL
    print("REQ DO FILTRO",request.POST)
    #stockPackagePack = StockPackage.objects
    linhaToDel = Gateway.objects
    if request.method == "POST":
        row_id = request.POST["rowIdDelete"]
        linha = linhaToDel.all().filter(id=request.POST["rowIdDelete"]) 
        condutorId = request.POST['condutorId']
        #condutorId
        print("REQ ID-->", request.POST["rowIdDelete"])
        
        print("A apagar todos os dados ta tabela")
        LastUpdatePortaria.objects.all().delete()
        updatePortaria = LastUpdatePortaria(
            dataUpdate =  datetime.today().strftime("%Y-%m-%d %H:%M:%S.%f")
        ).save()
        print("A guardar dados na lastUpdatePOrtaria")




        
        message = ""
        message += "<b>Entrada da portaria eliminada pelo User: </b>" + request.user.username + "</br>"

        message += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA/HORA CHEGADA</th><th>CONDUTOR</th><th>CONDUTOR ID</th><th>CONTACTO</th><th>EMPRESA</th><th>1ª MATRICULA</th><th>2ª MATRICULA</th><th>CARGA/DESCARGA</th><th>DOCA</th><th>DESTINO CARGA</th><th>TIPO VIATURA</th><th>DATA/HORA ENTRADA</th><th>COMENTARIOS ENTRADA</th><th>DATA/HORA SAIDA</th><th>COMENTARIOS SAIDA</th></tr></thead><tbody>'
        for prod in linha:
            message += (
            '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.dataHoraChegada
            +'</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.condutor
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + condutorId
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.contacto
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.empresa
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.primeiraMatricula
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.segundaMatricula
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.cargaDescarga
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.doca
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.destinoCarga
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.tipoViatura
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.dataHoraEntrada
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.comentEntrada
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.dataHoraSaida
            + '</td><td style="padding:0 15px 0 15px;background-color: gray">'
            + prod.comentSaida
            + "</td></tr>"
            )

        mensagem = request.POST["reportDelete"]
        message += "</table></br></br><b>Motivo: </b>" + mensagem
        subject, from_email, to = (
            "Delete na página Portaria - Shippers",
            "noreply@visteon.com",
            ["pmarti30@visteon.com"], 
        )  #"npires2@visteon.com"
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        #msg.send()
    linhaToDel.all().get(id=row_id).delete()

    return redirect("shippers:portaria")


def historicoTabelaShippersTracking(request):
    from django.http import JsonResponse

    #elementosGateway = Gateway.objects.all() 
    #print("POST---)",request.POST)
    print(request.POST)
    data_str_start = request.POST["startDate"]
    data_str_end = request.POST["endDate"]
    
    data_str_start_parts = data_str_start.split(' GMT')[0]
    data_str_end_parts = data_str_end.split(' GMT')[0]

    data_objStart = datetime.strptime(data_str_start_parts[:25], '%a %b %d %Y %H:%M:%S')
    data_objEnd = datetime.strptime(data_str_end_parts[:25], '%a %b %d %Y %H:%M:%S')
    #elementostrackingShippers = Tracking_ShippersPage.objects.filter(timestamp__contains=request.POST["data"])
    #elementos = Gateway.objects.raw(f"""SELECT * FROM shippers_gateway WHERE dataHoraChegada BETWEEN '{data_objStart.strftime('%Y-%m-%d %H:%M')} 00:00' AND '{data_objEnd.strftime('%Y-%m-%d %H:%M')} 24:00' """)
    try:
        elementostrackingShippers = Tracking_ShippersPage.objects.raw(f"""SELECT * FROM shippers_tracking_shipperspage WHERE ship_date BETWEEN '{data_objStart.strftime('%Y-%m-%d')}' AND '{data_objEnd.strftime('%Y-%m-%d')} 24:00' """)
    #elementostrackingShippers = Tracking_ShippersPage.objects.filter(timestamp=request.POST["data"]) #Tracking_ShippersPage.objects.all()
    #elementostrackingShippers = Tracking_ShippersPage.objects.raw(f""" SELECT * FROM shippers_tracking_shipperspage WHERE timestamp BETWEEN '{request.POST["data"]}' AND '{request.POST["data"]}' """) #Tracking_ShippersPage.objects.all()
        dados_list = [item.to_dict() for item in elementostrackingShippers]
        print("POST---)",elementostrackingShippers, "\n")
        return JsonResponse(dados_list, safe=False)
    except:
        elementostrackingShippersStart = Tracking_ShippersPage.objects.raw(f"""SELECT * FROM shippers_tracking_shipperspage WHERE ship_date BETWEEN '{data_objStart.strftime('%Y-%m-%d')}' AND '{data_objStart.strftime('%Y-%m-%d')} 24:00' """)
        elementostrackingShippersEnd = Tracking_ShippersPage.objects.raw(f"""SELECT * FROM shippers_tracking_shipperspage WHERE ship_date BETWEEN '{data_objEnd.strftime('%Y-%m-%d')}' AND '{data_objEnd.strftime('%Y-%m-%d')} 24:00' """)
        return JsonResponse({"contadorElementostrackingShippersStart": len(elementostrackingShippersStart) , "contadorElementostrackingShippersEnd" : len(elementostrackingShippersEnd)})
            



def hojeTabelaShippersTracking(request):
    from django.http import JsonResponse
    #elementosGateway = Gateway.objects.all() 
    #print("POST---)",request.POST)
    print(request.POST)
    elementostrackingShippers = Tracking_ShippersPage.objects.filter(ship_date__contains=request.POST["data"])#.strip()
    #elementostrackingShippers = Tracking_ShippersPage.objects.raw(f"""SELECT * FROM shippers_tracking_shipperspage WHERE timestamp BETWEEN '{request.POST["startDate"]}' AND '{request.POST["endDate"]}' """)
    #elementostrackingShippers = Tracking_ShippersPage.objects.filter(timestamp=request.POST["data"]) #Tracking_ShippersPage.objects.all()
    #elementostrackingShippers = Tracking_ShippersPage.objects.raw(f""" SELECT * FROM shippers_tracking_shipperspage WHERE timestamp BETWEEN '{request.POST["data"]}' AND '{request.POST["data"]}' """) #Tracking_ShippersPage.objects.all()
    dados_list = [item.to_dict() for item in elementostrackingShippers]
    print("POST---)",len(dados_list))
    return JsonResponse(dados_list, safe=False)




#func do portatil
def populate_shippers_trackingPage(request = None):
    from datetime import datetime
    from qad_ee.models import AbsMstr  
    trackers = AbsMstr.objects.raw(f""" 
            SELECT abs_mstr.prrowid, abs_id , abs_shp_date, abs_shp_time, ad_city, ad_country, absc_carrier ,    abs_mstr.abs__qad01 AS Fob, abs_mstr.abs__qad01 AS Fob, oid_abs_mstr, abs_status
                FROM abs_mstr , absc_det, ad_mstr
                WHERE ad_mstr.ad_addr = abs_mstr.abs_shipto AND absc_det.absc_abs_id = abs_mstr.abs_id  AND abs_shp_date >=   convert(datetime, convert(char(10), getdate(), 120)) 
                AND abs_mstr.abs_domain = absc_det.absc_domain   and abs_mstr.abs_domain = ad_mstr.ad_domain and abs_mstr.abs_domain =  ad_mstr.ad_domain
                ORDER BY abs_shp_date DESC

                """)
    tracking_shipperspage_refresh = Tracking_ShippersPage_Refresh.objects.all().first() 
    #tracking_shipperspage.delete()
    colecao = []
    for i in trackers:
        #print("CONTRY", i.ad_city)
        try:
            obj = Tracking_ShippersPage_View.objects.get(abs_id = i.abs_id[1:], estado__in=["novo","updated"])
            if (i.abs_id[1:] == obj.abs_id
                and str(obj.ship_time) == str(convert_to_time(i.abs_shp_time/3600))
                and obj.city == i.ad_city and obj.country ==  i.ad_country
                and  obj.carrier == i.absc_carrier
                and obj.fob == i.Fob[20:40] and obj.mode_of_transport == i.Fob[60:80] 
                and obj.vehicle_id == i.Fob[80:100] #and obj.confirmed == i.abs_status[1:2]
                and str(obj.ship_date) == str(i.abs_shp_date.strftime("%Y-%m-%d"))
            ):
                print("Duplicado")

            elif(i.abs_id[1:] == obj.abs_id or obj.estado!= "novo"
                or str(obj.ship_time) !=  str(convert_to_time(i.abs_shp_time/3600))
                or obj.city !=  i.ad_city or obj.country !=   i.ad_country
                or  obj.carrier !=  i.absc_carrier
                or obj.fob !=  i.Fob[20:40] or obj.mode_of_transport !=  i.Fob[60:80] 
                or obj.vehicle_id !=  i.Fob[80:100] #or obj.confirmed !=  i.abs_status[1:2]
                or str(obj.ship_date) !=  str(i.abs_shp_date.strftime("%Y-%m-%d"))
            ):
                print("valor com alteracoes")
                shippers_tracking_original = Tracking_ShippersPage_View.objects.get(abs_id = i.abs_id[1:], estado__in=["novo","updated"] )
                shippers_tracking_original.estado = "outdated"
                shippers_tracking_original.save()
                """ print("TESTE",shippers_tracking_original.estado) """
                id_to_refresh = Tracking_ShippersPage_Refresh.objects.all().delete()
                """ variaveis que controlam o refresh desta pagina """
                id_to_refresh = Tracking_ShippersPage_Refresh(lastUpdate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")).save()
                
                shippers_tracking =  Tracking_ShippersPage(
                        abs_id= i.abs_id[1:], ship_date = i.abs_shp_date.strftime("%Y-%m-%d"), ship_time = convert_to_time(i.abs_shp_time/3600),
                        city = i.ad_city, country =  i.ad_country, carrier = i.absc_carrier,
                        fob = i.Fob[20:40], mode_of_transport =i.Fob[60:80], vehicle_id = i.Fob[80:100],total_master_packs =i.abs_status,
                        confirmed =  "Yes" if i.abs_status[1:2] == "y" else "No",  estado ="updated", timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ).save()
                print("Guardou valor novo ALTERADO")

        except Tracking_ShippersPage.MultipleObjectsReturned: 
            print("Existem varios ids",i.abs_status)
        except:
            #valor não existe na view (ainda não apareceu hoje)
            """ print("ID NOVO",i.abs_id[1:]) """
            id_to_refresh = Tracking_ShippersPage_Refresh.objects.all().delete()
            id_to_refresh = Tracking_ShippersPage_Refresh(lastUpdate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")).save()
            shippers_tracking =  Tracking_ShippersPage(
                        abs_id= i.abs_id[1:], ship_date = i.abs_shp_date.strftime("%Y-%m-%d"), ship_time = convert_to_time(i.abs_shp_time/3600),
                        city = i.ad_city, country =  i.ad_country, carrier = i.absc_carrier,
                        fob = i.Fob[20:40], mode_of_transport =i.Fob[60:80], vehicle_id = i.Fob[80:100],total_master_packs =i.abs_status,  
                        confirmed =  "Yes" if i.abs_status[1:2] == "y" else "No", estado ="novo", timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ).save()
            print("Guardou valor NOVO",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
    Tracking_ShippersPage_Refresh.objects.all().delete()
    Tracking_ShippersPage_Refresh(lastUpdate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")).save()
    setDeletedShippersTracking()
    setDateValidationShippersTracking()
    return JsonResponse({"resposta": "OK"})
     
def setDateValidationShippersTracking():
    tracking_shipperspage = Tracking_ShippersPage_View.objects.all()
    for i in tracking_shipperspage:
        tracking_shipperspage = Tracking_ShippersPage.objects.filter(abs_id=i.abs_id).earliest('id')

        print("if = true, data maior que timestamp")
        if tracking_shipperspage.timestamp.split(" ")[0] >= i.ship_date:
                print("para por falso",tracking_shipperspage.timestamp.split(" ")[0])
                i.date_validation = "False"
                i.save()
        else:
            print("vai ficar a true",tracking_shipperspage.timestamp.split(" ")[0])
            i.date_validation = "True"
            i.save() 

            



            """ 
        data = i.timestamp.split(" ")[0]
        if data >= i.ship_date:
            print("para por falso",data)
            i.date_validation = "False"
            i.save()
        else:
            print("para por verdadeiro",data)
            tracking_shipperspage = Tracking_ShippersPage.objects.filter(abs_id = i.abs_id)
            for i in tracking_shipperspage:
                i.date_validation = "True"
                i.save() """

# def setDateValidationShippersTracking():
#     from datetime import datetime
#     tracking_shipperspage = Tracking_ShippersPage_View.objects.all()
#     for i in tracking_shipperspage:
#         data = i.timestamp.split(" ")[0]
#         if data >= datetime.now().strftime("%Y-%m-%d"):
#             print("para por falso",data)
#             i.date_validation = "False"
#             i.save()
#         else:
#             print("para por verdadeiro",data)
#             i.date_validation = "True"
#             i.save()



""" SELECT DISTINCT abs_id , abs_shp_date, abs_shp_time,  abs_mstr.abs__qad01 AS Fob, oid_abs_mstr, abs_status ,abs_shipfrom, abs_item, abs_qty, abs_ship_qty
    FROM abs_mstr , absc_det, ad_mstr
    WHERE abs_shp_date >=   '2023-04-26 00:00:00.000' AND abs_id LIKE '%s2330298%' """
def setSomatoriosShippersTracking():
    tracking_shipperspage = Tracking_ShippersPage_View.objects.all()
    for i in tracking_shipperspage: 
        soma_abs_qty = 0
        soma_abs_ship_qty = 0
        #tens de precorrer todos os valores relacionados com este abs_id que vem do for anterior
        sub_items = Tracking_ShippersPage_Sub_Items.objects.filter(tracking_shipperspage_sub_items = i) #(tracking_shipperspage_sub_items = i.abs_id)
        for x in sub_items:
            #precorre todos os sub_items para este abs_id e vai somando os valores dos campos certos
            soma_qty += x.abs_qty
            soma_abs_ship_qty += x.abs_ship_qty
        i.sum_abs_qty = soma_abs_qty
        i.sum_abs_ship_qty = soma_abs_ship_qty
        #i.save()
         



        
        




def setSubItemsShippersTracking():
    from qad_ee.models import AbsMstr
    print("Entrou na setSubItemsShippersTracking")
    with timer():
        for page in Tracking_ShippersPage_View.objects.all():
            sub_items = []
            querry = AbsMstr.objects.raw(f""" SELECT DISTINCT abs_id, abs_status ,abs_shipfrom, abs_item, abs_qty, abs_ship_qty, abs_mstr.prrowid
                FROM abs_mstr , absc_det, ad_mstr
                WHERE abs_id LIKE %s """, ['%'+page.abs_id+'%'])
            for item in querry:
                sub_item = Tracking_ShippersPage_Sub_Items(
                    abs_id=item.abs_id,
                    abs_status=item.abs_status,
                    abs_shipfrom=item.abs_shipfrom,
                    abs_item=item.abs_item,
                    abs_qty=item.abs_qty,
                    abs_ship_qty=item.abs_ship_qty,
                )
                sub_item.save()
                sub_items.append(sub_item)

            # busca o objeto Tracking_ShippersPage existente para adicionar os sub_items
            try:
                existing_page = Tracking_ShippersPage.objects.filter(
                    abs_id=page.abs_id).first()
                
                existing_page.tracking_shipperspage_sub_items.add(*sub_items)  

                existing_page.sub_item_flag = "True"
                existing_page.save()
            except:
                print("não deu para encontrar o valor para o id: ",page.abs_id)


def _model_to_dict(model) -> dict[str, t.Any]:
    ret = {}
    for k in model._meta.get_fields():
        if isinstance(k, (models.ManyToOneRel, models.ManyToManyField)):
            continue
        
        value = getattr(model, k.name, None)
        if isinstance(k, models.ForeignKey) and value is not None:
            ret[k.name] = _model_to_dict(value)
        elif isinstance(k, models.FileField):
            ret[k.name] = value.name
        else:
            ret[k.name] = value

    return ret

def queryset_to_dict(queryset):
    ret = []
    for row in queryset:    
        ret.append(_model_to_dict(row))
    return ret

def getSubItemsShippersTracking(request):
    print("REQUEST->", request.POST)
    s = Tracking_ShippersPage.tracking_shipperspage_sub_items.through.objects.filter(tracking_shipperspage_id=request.POST["tracking_shipperspage_sub_items_id"])
    #sub_items = Tracking_ShippersPage_Sub_Items.objects.filter(tracking_shipperspage_sub_items=request.POST.get('tracking_shipperspage_sub_items_id'))
    return JsonResponse({"data": queryset_to_dict(s)}, safe=False)


        

def setDeletedShippersTracking():
    from datetime import datetime
    from qad_ee.models import AbsMstr  
    trackers = AbsMstr.objects.raw(f""" 
            SELECT abs_mstr.prrowid, abs_id , abs_shp_date, abs_shp_time, ad_city, ad_country, absc_carrier ,    abs_mstr.abs__qad01 AS Fob, abs_mstr.abs__qad01 AS Fob, oid_abs_mstr, abs_status
                FROM abs_mstr , absc_det, ad_mstr   
                WHERE ad_mstr.ad_addr = abs_mstr.abs_shipto AND absc_det.absc_abs_id = abs_mstr.abs_id  AND abs_shp_date >=   convert(datetime, convert(char(10), getdate(), 120)) 
                AND abs_mstr.abs_domain = absc_det.absc_domain   and abs_mstr.abs_domain = ad_mstr.ad_domain and abs_mstr.abs_domain =  ad_mstr.ad_domain
                ORDER BY abs_shp_date DESC""")
    tracking_shipperspage = Tracking_ShippersPage_View.objects.all()
    if tracking_shipperspage:
        set2=set([])
        set1=set([])
        for x in trackers:
            set1.add(x.abs_id[1:])

        for x in tracking_shipperspage:
            set2.add(x.abs_id)
        set3 = set2 - set1

        for i in set3:
            valores_apagados = Tracking_ShippersPage_View.objects.filter(abs_id=i) 
            for i in valores_apagados:
                i.estado= "deleted"
                i.save()
    else:
        print("bd vazia")
        
         
    return None
#Em uso
def convert_to_time(value):
    from datetime import timedelta
    time = timedelta(hours=int(value), minutes=int((value - int(value)) * 60))
    return str(time)

def getIdToRefreshShippersTracking(*request):
    id_to_refresh = Tracking_ShippersPage_Refresh.objects.all().values('lastUpdate').first().get('lastUpdate')
    return JsonResponse({"refreshDate":id_to_refresh})

@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists() or u.groups.filter(name="ShippersTracking").exists())
#@user_passes_test(lambda u: u.groups.filter(name="ShippersTracking").exists())
def shippers_trackingPage(request):
    import datetime
 
    #populate_shippers_trackingPage()
    shippers_tracking = Tracking_ShippersPage_View.objects.all()
    id_to_refresh = Tracking_ShippersPage_Refresh.objects.all().values('lastUpdate').first().get('lastUpdate')
    Tracking_ShippersPage.objects.filter(comentarios=None).update(comentarios='')
    """ data_atual = datetime.date.today()
    data_formatada = data_atual.strftime('%Y-%m-%d') """
    #setSubItemsShippersTracking()
    #shippers_tracking.all().delete()   
    return render(
            request,
            "shippers/shippers_tracking_1.0.html",
            {"shippers_tracking" : shippers_tracking,
            "id_to_refresh" : id_to_refresh}
            
        )


def addCommentShippersTracking(request):
    print("req",request.POST)
    abs_id = request.POST["id"].strip()
    timestamp = request.POST["timestamp"].strip()
    shippers_tracking = Tracking_ShippersPage.objects.get(abs_id=abs_id, timestamp=timestamp)
    shippers_tracking1 = Tracking_ShippersPage.objects.filter(abs_id=abs_id, timestamp=timestamp)
    shippers_tracking.comentarios = request.POST["comentarios"]
    shippers_tracking.save()
    return JsonResponse({"message": "OK" ,"shippers_tracking": len(shippers_tracking1), "request": request.POST})


def downloadExcelShippersTracking(request):
    if request.method == "GET":
        elementos = Tracking_ShippersPage.objects.all() # Gateway.objects
        print("ENTROu")
         

        wbProduction = Workbook()
        sheetProduction = wbProduction.add_sheet("Sheet 1")

        row = 0
        col = 0
        sheetProduction.write(row, col, "ID")
        sheetProduction.write(row, col + 1, "Ship Date")
        sheetProduction.write(row, col + 2, "Ship Time")
        sheetProduction.write(row, col + 3, "City")
        sheetProduction.write(row, col + 4, "Country")
        sheetProduction.write(row, col + 5, "Carrier")
        sheetProduction.write(row, col + 6, "FOB")
        sheetProduction.write(row, col + 7, "Mode of Transport")
        sheetProduction.write(row, col + 8, "Vehicle ID")
        sheetProduction.write(row, col + 9, "Confirmed")
        sheetProduction.write(row, col + 10, "Comments")
        sheetProduction.write(row, col + 11, "Estado")
        sheetProduction.write(row, col + 12, "Timestamp")

        row += 1

        for elem in elementos.all():
            sheetProduction.write(row, col, elem.abs_id)
            sheetProduction.write(row, col + 1, elem.ship_date)
            sheetProduction.write(row, col + 2, elem.ship_time)
            sheetProduction.write(row, col + 3, elem.city)
            sheetProduction.write(row, col + 4, elem.country)
            sheetProduction.write(row, col + 5, elem.carrier)
            sheetProduction.write(row, col + 6, elem.fob)
            sheetProduction.write(row, col + 7, elem.mode_of_transport)
            sheetProduction.write(row, col + 8, elem.vehicle_id)
            sheetProduction.write(row, col + 9, elem.confirmed)
            sheetProduction.write(row, col + 10, elem.comentarios)
            sheetProduction.write(row, col + 11, elem.estado)
            sheetProduction.write(row, col + 12, elem.timestamp)
            row += 1
        wbProduction.save(
            "media\\shippers\\portaria\\workbookShipersTracking.xls"
        )

        return redirect("shippers:shippersTrackingPage")


def reportPortariaTracking(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Shippers - Portaria Tracking",
            "noreply@visteon.com",
            ["pmarti30@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
    return redirect("shippers:shippersTrackingPage")

