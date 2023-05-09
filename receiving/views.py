import json
import os
import datetime

import numpy as np
import openpyxl
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.models import Group

from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMultiAlternatives
from django.core.serializers import serialize
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.db.models import Q

from datetime import datetime, date, timedelta
from collections import Counter

from xlwt import Workbook, easyxf

from qad_ee.models import *
from receiving.models import *

#from shippers.models import FASupplyRedItems, BPDropinSupplyRedItems, BPSMDSupplyRedItems, FaSupplyReportRedItems, BPDropinSupplyReportRedItems, BPSMDSupplyReportRedItems

def testeSchedule():
    print("ENTROU NO TESTE SO SCHEDULE")
    #funcao de teste, esta apenas a guardar uma nova linha com o timestamp atual
    HistoricoErros(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pagina ="Tenda_Mecflow", erro = "Teste schedule").save()
    
def tendaMECFLOW(request):
    try:
        #print(2/0)
        ldDetWHOUT1 = LdDet.objects.filter(ld_site="3515",ld_loc__in=["WHOUT","MECFLOW"]).values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetWHOUT = LdDet.objects.filter(ld_site="3515",  ld_loc="WHOUT").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetMECFLOW = LdDet.objects.filter(ld_site="3515",  ld_loc ="MECFLOW").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetNotWHOUTNotMECFLOW = ( LdDet.objects.filter(ld_site="3515", ld_part__in=ldDetWHOUT1.values_list("ld_part"))
            .exclude(Q(ld_loc="MECFLOW") | Q(ld_loc="WHOUT") |  Q(ld_loc="QUALITY") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC") | Q(ld_loc__contains="WIP")) #para ficar apenas o WHOUT -) tenda
            .values("ld_part", "ld_loc")
        )

        ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT1.values_list("ld_part"), pt_site="3515"
        ).values("pt_part", "pt_desc1", "pt_desc2")
        
        serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515", ser_loc="WHOUT"
        ).values("ser_part")

        serMstrMECFLOW = SerMstr.objects.filter(ser_part__in=ldDetMECFLOW.values_list("ld_part"), ser_site="3515", ser_loc="MECFLOW"
        ).values("ser_part")

        serMstrNotWHOUTNotMECFLOW = (SerMstr.objects.filter(ser_part__in=ldDetNotWHOUTNotMECFLOW.values_list("ld_part"), ser_site="3515")
            .exclude(Q(ser_loc="WHOUT") | Q(ser_loc="MECFLOW") | Q(ser_loc="QUALITY") | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC") | Q(ser_loc__contains="WIP"))
            .values("ser_part", "ser_qty_avail")
        )
        elementos = []

        counterWHOUT = Counter()
        #counterNotWHOUT = Counter()
        counterSerWHOUT = Counter()
        #counterSerNotWHOUT = Counter() 

        counterMECFLOW = Counter()
        counterSerMECFLOW = Counter()
        counterSerNotWHOUTNotMECFLOW = Counter()
        counterNotWHOUTNotMECFLOW = Counter()

        for ld in ldDetWHOUT:
            counterWHOUT[ld["ld_part"]] += ld["ld_qty_oh"]

        for ld in ldDetMECFLOW:
            counterMECFLOW[ld["ld_part"]] += ld["ld_qty_oh"]

        
        for ser in serMstrWHOUT:
            counterSerWHOUT[ser["ser_part"].upper()] += 1

        for ser in serMstrMECFLOW:
            counterSerMECFLOW[ser["ser_part"].upper()] += 1

        for ser in serMstrNotWHOUTNotMECFLOW:   #contador de serials para nenhum dos casos
            counterSerNotWHOUTNotMECFLOW[ser["ser_part"].upper()] += 1
            counterNotWHOUTNotMECFLOW[ser["ser_part"].upper()] += ser["ser_qty_avail"]

        for ld in ldDetWHOUT1:
            for pt in ptMstr:
                if pt["pt_part"] == ld["ld_part"]:
                    if len(pt["pt_desc1"]) == 24:
                        pt["pt_desc1"] = pt["pt_desc1"] + "" + pt["pt_desc2"]
                    else:
                        pt["pt_desc1"] = pt["pt_desc1"] + " " + pt["pt_desc2"]
                    if str(counterWHOUT[ld["ld_part"]]) == "0E-10":
                        counterWHOUT[ld["ld_part"]] = "0.0000000000"
                    if not any(elem["itemNumber"] == ld["ld_part"] for elem in elementos):
                        row = {
                            "itemNumber": ld["ld_part"],
                            "descricao": pt["pt_desc1"],
                            "qtyOnHand": 0,
                            "serialsWHOUT": counterSerWHOUT[ld["ld_part"]],
                            "qtyOnHandMECFLOW": 0,
                            "serialsMECFLOW": counterSerMECFLOW[ld["ld_part"]],
                            "qtyOnHandNotWHOUTNotMECFLOW": 0,
                            "serialsNotWHOUTNotMECFLOW": counterSerNotWHOUTNotMECFLOW[ld["ld_part"]],
                            "unitCost": "-",
                            "min": "-",
                            "max": "-",
                            "consumptionValue": 0,
                            "consumptionValueDiaSeguinte": 0,
                            "consumptionValueDiaAposDiaSeguinte": 0,
                            "totalCost": "-",
                        }
                        elementos.append(row)
        for ldWHOUT in ldDetWHOUT:
            for elem in elementos:
                if elem["itemNumber"] == ldWHOUT["ld_part"]:
                    elem["qtyOnHand"] = str(counterWHOUT[ldWHOUT["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHand"] == "":
                        elem["qtyOnHand"] = "0"
                    break
        
        for ldMECFLOW in ldDetMECFLOW:
            for elem in elementos:
                if elem["itemNumber"] == ldMECFLOW["ld_part"]:
                    elem["qtyOnHandMECFLOW"] = str(counterMECFLOW[ldMECFLOW["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHandMECFLOW"] == "":
                        elem["qtyOnHandMECFLOW"] = "0"
                    break

        for ldMECFLOW in ldDetNotWHOUTNotMECFLOW:
            for elem in elementos:
                if elem["itemNumber"] == ldMECFLOW["ld_part"]:
                    elem["qtyOnHandNotWHOUTNotMECFLOW"] = str(counterNotWHOUTNotMECFLOW[ldMECFLOW["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHandNotWHOUTNotMECFLOW"] == "":
                        elem["qtyOnHandNotWHOUTNotMECFLOW"] = "0"
                    break
        return render(request, "tenda_Mecflow.html", {"elementos": elementos})
    except:
        HistoricoErros(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pagina ="Tenda_Mecflow", erro = "Funcao não conseguiu realizar querrys para Tenda_Mecflow").save()
    # for xx in xxusrw:
    #     for wo in woMstr:
    #         if wo['wo_part'] == xx['xxusrw_key4']:
    #             for elem in elementos:
    #                 if elem['itemNumber'] == xx['xxusrw_key6']:
    #                     elem['contas'] += xx['xxusrw_key4'] + " " + str(xx['xxusrw_decfld_1']) + "*" + str(
    #                         wo['wo_qty_exp_complete'])
    #                     elem['consumptionValue'] += round(
    #                         float(xx['xxusrw_decfld_1']) * float(wo['wo_qty_exp_complete']), 2)
    #                     break

    return render(request, "tenda_Mecflow.html")






# def tenda2Ponto0(request):
#     ldDetWHOUT = LdDet.objects.filter(ld_site='3515', ld_loc__in=["WHOUT", "whout"]).values('ld_part', 'ld_loc', 'ld_qty_oh')
#     ldDetNotWHOUT = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT.values_list('ld_part')).exclude(
#         Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY")).values(
#         'ld_part', 'ld_loc')
#     ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list('ld_part'), pt_site='3515').values('pt_part',
#                                                                                                          'pt_desc1',
#                                                                                                          'pt_desc2')
#     serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515",
#                                           ser_loc__in=["WHOUT", "whout"]).values('ser_part')
    
#     serMstrNotWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515").exclude(
#         Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY")).values('ser_part', 'ser_qty_avail')

#    # ///---------------/////----------------------/////
    
#     ldDetWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_loc__in=["WHOUT", "whout"]).values('ld_part', 'ld_loc', 'ld_qty_oh')
#     ldDetNotWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT.values_list('ld_part')).exclude(
#         Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC")).values(
#         'ld_part', 'ld_loc')
#     ptMstrNovo = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list('ld_part'), pt_site='3515').values('pt_part',
#                                                                                                          'pt_desc1',
#                                                                                                          'pt_desc2')
#     serMstrWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515",
#                                           ser_loc__in=["WHOUT", "whout"]).values('ser_part')
    
#     serMstrNotWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515").exclude(
#         Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY")  | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC")).values('ser_part', 'ser_qty_avail')
    

    
#     elementos = []

#     xxusrw = XxusrwWkfl.objects.filter(xxusrw_key6="KCB00602-AA", xxusrw_domain="3511010").values('xxusrw_decfld_1',
#                                                                                                   'xxusrw_key4',
#                                                                                                   'xxusrw_key6',
#                                                                                                   'xxusrw_decfld_2')
#     # woMstr = WoMstr.objects.filter(wo_part__in=xxusrw.values_list('xxusrw_key4'),
#     #                                wo_due_date="2022-04-20").values(
#     #     'wo_part', 'wo_qty_exp_complete')

#     counterWHOUT = Counter()
#     counterNotWHOUT = Counter()
#     counterSerWHOUT = Counter()
#     counterSerNotWHOUT = Counter()

#     #///-----------////---------------------///-----///

#     counterWHOUTNovo = Counter()
#     counterNotWHOUTNovo = Counter()
#     counterSerWHOUTNovo = Counter()
#     counterSerNotWHOUTNovo = Counter()



#     for ld in ldDetWHOUT:
        
#         counterWHOUT[ld['ld_part'].upper()] += ld['ld_qty_oh']

#     # for ld in ldDetNotWHOUT:
#     #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

#     for ser in serMstrWHOUT:
#         counterSerWHOUT[ser['ser_part'].upper()] += 1

#     for ser in serMstrNotWHOUT:
#         counterSerNotWHOUT[ser['ser_part'].upper()] += 1
#         counterNotWHOUT[ser['ser_part'].upper()] += ser['ser_qty_avail']

#     #//----------------------------//-----------------------------------////////---------------------------

#     for ld in ldDetWHOUTNovo:
#         counterWHOUTNovo[ld['ld_part'].upper()] += ld['ld_qty_oh']

#     # for ld in ldDetNotWHOUT:
#     #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

#     for ser in serMstrWHOUTNovo:
#         counterSerWHOUTNovo[ser['ser_part'].upper()] += 1

#     for ser in serMstrNotWHOUTNovo:
#         print("->",ser)
#         counterSerNotWHOUTNovo[ser['ser_part'].upper()] += 1
#         counterNotWHOUTNovo[ser['ser_part'].upper()] += ser['ser_qty_avail']

#     for ld in ldDetWHOUT:
#         for pt in ptMstr:
#             #print("List-->", pt['pt_part'], ld['ld_part'])
#             if pt['pt_part'].upper() == ld['ld_part'].upper():
#                 if len(pt['pt_desc1']) == 24:
#                     pt['pt_desc1'] = pt['pt_desc1'] + "" + pt['pt_desc2']
#                 else:
#                     pt['pt_desc1'] = pt['pt_desc1'] + " " + pt['pt_desc2']
#                 if str(counterWHOUT[ld['ld_part']]) == '0E-10':
#                     counterWHOUT[ld['ld_part']] = '0.0000000000'
#                 if not any(elem['itemNumber'].upper() == ld['ld_part'].upper() for elem in elementos):
#                     row = {
#                         'itemNumber': ld['ld_part'],
#                         'descricao': pt['pt_desc1'],
#                         'qtyOnHand': str(counterWHOUT[ld['ld_part']])[:-11],
#                         'serialsWHOUT': counterSerWHOUT[ld['ld_part']],
#                         'qtyOnHandNotWHOUT': 0,
#                         'serialsNotWHOUT': counterSerNotWHOUTNovo[ld['ld_part']],
#                         'unitCost': '-',
#                         'min': '-',
#                         'max': '-',
#                         'consumptionValue': 0,
#                         'consumptionValueDiaSeguinte':0,
#                         'consumptionValueDiaAposDiaSeguinte': 0,
#                         'totalCost': '-',
#                     }
#                     elementos.append(row)
    
    
#     for ldNotWHOUT in ldDetNotWHOUTNovo:
#         for elem in elementos:

#             if elem['itemNumber'].upper() == ldNotWHOUT['ld_part'].upper():
#                 elem['qtyOnHandNotWHOUT'] = str(counterNotWHOUTNovo[ldNotWHOUT['ld_part']])[:-11]
#                 if elem['qtyOnHandNotWHOUT'] == '':
#                     elem['qtyOnHandNotWHOUT'] = '0'
#                 break

#     return render(request, 'tenda.html', {'elementos': elementos})


def tenda(request):
    ldDetWHOUT = LdDet.objects.filter(ld_site='3515', ld_loc__in=["WHOUT"]).values('ld_part', 'ld_loc', 'ld_qty_oh')
    ldDetNotWHOUT = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT.values_list('ld_part')).exclude(
        Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY") | Q(ld_loc="whout") | Q(ld_loc="quality")).values(
        'ld_part', 'ld_loc')
    ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list('ld_part'), pt_site='3515').values('pt_part',
                                                                                                         'pt_desc1',
                                                                                                         'pt_desc2')
    serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515",
                                          ser_loc='WHOUT').values('ser_part')
    
    serMstrNotWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515").exclude(
        Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY") | Q(ser_loc="whout") | Q(ser_loc="quality")).values('ser_part', 'ser_qty_avail')

   # ///---------------/////----------------------/////
    
    ldDetWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_loc="WHOUT").values('ld_part', 'ld_loc', 'ld_qty_oh')
    ldDetNotWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT.values_list('ld_part')).exclude(
        Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC") |
         Q(ld_loc="whout") | Q(ld_loc="quality") | Q(ld_loc__contains="pwb") | Q(ld_loc__contains="plc")).values(
        'ld_part', 'ld_loc')
    ptMstrNovo = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list('ld_part'), pt_site='3515').values('pt_part',
                                                                                                         'pt_desc1',
                                                                                                         'pt_desc2')
    serMstrWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515",
                                          ser_loc='WHOUT').values('ser_part')
    
    serMstrNotWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515").exclude(
        Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY")  | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC")|
        Q(ser_loc="whout") | Q(ser_loc="quality")  | Q(ser_loc__contains="pwb") | Q(ser_loc__contains="plc")).values('ser_part', 'ser_qty_avail')
    

    
    elementos = []

    xxusrw = XxusrwWkfl.objects.filter(xxusrw_key6="KCB00602-AA", xxusrw_domain="3511010").values('xxusrw_decfld_1',
                                                                                                  'xxusrw_key4',
                                                                                                  'xxusrw_key6',
                                                                                                  'xxusrw_decfld_2')
    # woMstr = WoMstr.objects.filter(wo_part__in=xxusrw.values_list('xxusrw_key4'),
    #                                wo_due_date="2022-04-20").values(
    #     'wo_part', 'wo_qty_exp_complete')

    counterWHOUT = Counter()
    counterNotWHOUT = Counter()
    counterSerWHOUT = Counter()
    counterSerNotWHOUT = Counter()

    #///-----------////---------------------///-----///

    counterWHOUTNovo = Counter()
    counterNotWHOUTNovo = Counter()
    counterSerWHOUTNovo = Counter()
    counterSerNotWHOUTNovo = Counter()



    for ld in ldDetWHOUT:
        counterWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    # for ld in ldDetNotWHOUT:
    #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUT:
        counterSerWHOUT[ser['ser_part']] += 1

    for ser in serMstrNotWHOUT:
        counterSerNotWHOUT[ser['ser_part']] += 1
        counterNotWHOUT[ser['ser_part']] += ser['ser_qty_avail']

    #//----------------------------//-----------------------------------////////---------------------------

    for ld in ldDetWHOUTNovo:
        counterWHOUTNovo[ld['ld_part']] += ld['ld_qty_oh']

    # for ld in ldDetNotWHOUT:
    #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUTNovo:
        counterSerWHOUTNovo[ser['ser_part']] += 1

    for ser in serMstrNotWHOUTNovo:
        counterSerNotWHOUTNovo[ser['ser_part']] += 1
        counterNotWHOUTNovo[ser['ser_part']] += ser['ser_qty_avail']

    for ld in ldDetWHOUT:
        for pt in ptMstr:
            if pt['pt_part'] == ld['ld_part']:
                if len(pt['pt_desc1']) == 24:
                    pt['pt_desc1'] = pt['pt_desc1'] + "" + pt['pt_desc2']
                else:
                    pt['pt_desc1'] = pt['pt_desc1'] + " " + pt['pt_desc2']
                if str(counterWHOUT[ld['ld_part']]) == '0E-10':
                    counterWHOUT[ld['ld_part']] = '0.0000000000'
                if not any(elem['itemNumber'] == ld['ld_part'] for elem in elementos):
                    row = {
                        'itemNumber': ld['ld_part'],
                        'descricao': pt['pt_desc1'],
                        'qtyOnHand': str(counterWHOUT[ld['ld_part']])[:-11],
                        'serialsWHOUT': counterSerWHOUT[ld['ld_part']],
                        'qtyOnHandNotWHOUT': 0,
                        'serialsNotWHOUT': counterSerNotWHOUTNovo[ld['ld_part']],
                        'unitCost': '-',
                        'min': '-',
                        'max': '-',
                        'consumptionValue': 0,
                        'consumptionValueDiaSeguinte':0,
                        'consumptionValueDiaAposDiaSeguinte': 0,
                        'totalCost': '-',
                    }
                    elementos.append(row)
    
    #contadores
    for ldNotWHOUT in ldDetNotWHOUTNovo:
        for elem in elementos:
            print("->",ldNotWHOUT['ld_part'],"|", elem['itemNumber'])
            if elem['itemNumber'] == ldNotWHOUT['ld_part']:
                elem['qtyOnHandNotWHOUT'] = str(counterNotWHOUTNovo[ldNotWHOUT['ld_part']])[:-11]
                if elem['qtyOnHandNotWHOUT'] == '':
                    elem['qtyOnHandNotWHOUT'] = '0'
                break

    # for xx in xxusrw:
    #     for wo in woMstr:
    #         if wo['wo_part'] == xx['xxusrw_key4']:
    #             for elem in elementos:
    #                 if elem['itemNumber'] == xx['xxusrw_key6']:
    #                     elem['contas'] += xx['xxusrw_key4'] + " " + str(xx['xxusrw_decfld_1']) + "*" + str(
    #                         wo['wo_qty_exp_complete'])
    #                     elem['consumptionValue'] += round(
    #                         float(xx['xxusrw_decfld_1']) * float(wo['wo_qty_exp_complete']), 2)
    #                     break

    return render(request, 'tenda.html', {'elementos': elementos})



def tendaMeclux(request):
    # = Table.SelectRows(dbo_ser_mstr, each [ser_domain] = "3511010" and [ser_loc] = "WHOUT" or [ser_loc] = "MECLUX")
    # = QAD_EE1{[Schema="dbo",Item="ser_mstr"]}[Data]

    """ SELECT * FROM ser_mstr
        WHERE ser_domain = '3511010' AND (ser_loc = 'WHOUT' OR ser_loc = 'MECFLOW');
    """

    #APENAS PARA WHOUT
    ldDetWHOUT = LdDet.objects.filter(ld_site='3515', ld_loc__in=["WHOUT"]).values('ld_part', 'ld_loc', 'ld_qty_oh')
    ldDetNotWHOUT = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT.values_list('ld_part')).exclude(
        Q(ld_loc="MECFLOW") | Q(ld_loc="WIP") | Q(ld_loc="PWB") | Q(ld_loc="PLC")).values(
        'ld_part', 'ld_loc')
    ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list('ld_part'), pt_site='3515').values('pt_part',
                                                                                                         'pt_desc1',
                                                                                                         'pt_desc2')
    serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515",
                                          ser_loc__in=["WHOUT"]).values('ser_part')
    
    serMstrNotWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list('ld_part'), ser_site="3515").exclude(
        Q(ser_loc="MECFLOW") | Q(ser_loc="WIP")).values('ser_part', 'ser_qty_avail')

   # ///---------------/////----------------------/////
    # #APENAS PARA MECFLOW
    # ldDetWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_loc__in=["MECFLOW"]).values('ld_part', 'ld_loc', 'ld_qty_oh')
    # ldDetNotWHOUTNovo = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUTNovo.values_list('ld_part')).exclude(
    #     Q(ld_loc="WHOUT") | Q(ld_loc="WIP") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC")).values(
    #     'ld_part', 'ld_loc')
    # ptMstrNovo = PtMstr.objects.filter(pt_part__in=ldDetWHOUTNovo.values_list('ld_part'), pt_site='3515').values('pt_part',
    #                                                                                                      'pt_desc1',
    #                                                                                                      'pt_desc2')
    # serMstrWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUTNovo.values_list('ld_part'), ser_site="3515",
    #                                       ser_loc__in=["WHOUT"]).values('ser_part')
    
    # serMstrNotWHOUTNovo = SerMstr.objects.filter(ser_part__in=ldDetWHOUTNovo.values_list('ld_part'), ser_site="3515").exclude(
    #     Q(ser_loc="WHOUT")  |Q(ser_loc="WIP")  | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC")).values('ser_part', 'ser_qty_avail')
    
    # # ///---------------/////----------------------/////
    # # PARA NENHUM DOS DOIS
    # ldDetWHOUT_MECALUX = LdDet.objects.filter(ld_site='3515', ld_loc__in=["WHOUT", "MECFLOW"]).values('ld_part', 'ld_loc', 'ld_qty_oh')

    # ldDetNotWHOUT_MECALUX = LdDet.objects.filter(ld_site='3515', ld_part__in=ldDetWHOUT_MECALUX.values_list('ld_part')).exclude(
    #     Q(ld_loc="WHOUT") | Q(ld_loc="WIP") | Q(ld_loc="MECFLOW") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC")).values(
    #     'ld_part', 'ld_loc')
    # ptMstr_MECALUX = PtMstr.objects.filter(pt_part__in=ldDetWHOUT_MECALUX.values_list('ld_part'), pt_site='3515').values('pt_part',
    #                                                                                                      'pt_desc1',
    #                                                                                                      'pt_desc2')
    # serMstrWHOUT_MECALUX = SerMstr.objects.filter(ser_part__in=ldDetWHOUT_MECALUX.values_list('ld_part'), ser_site="3515",
    #                                       ser_loc__in=["WHOUT", "MECFLOW"]).values('ser_part')
    
    # serMstrNotWHOUT_MECALUX = SerMstr.objects.filter(ser_part__in=ldDetWHOUT_MECALUX.values_list('ld_part'), ser_site="3515").exclude(
    #     Q(ser_loc="WHOUT") | Q(ser_loc="WIP") | Q(ser_loc="MECFLOW")  | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC")).values('ser_part', 'ser_qty_avail')
    
    
    elementos = []

    xxusrw = XxusrwWkfl.objects.filter(xxusrw_key6="KCB00602-AA", xxusrw_domain="3511010").values('xxusrw_decfld_1',
                                                                                                  'xxusrw_key4',
                                                                                                  'xxusrw_key6',
                                                                                                  'xxusrw_decfld_2')
    # woMstr = WoMstr.objects.filter(wo_part__in=xxusrw.values_list('xxusrw_key4'),
    #                                wo_due_date="2022-04-20").values(
    #     'wo_part', 'wo_qty_exp_complete')

    counterWHOUT = Counter()
    counterNotWHOUT = Counter()
    counterSerWHOUT = Counter()
    counterSerNotWHOUT = Counter()

    #///-----------////---------------------///-----///

    counterWHOUTNovo = Counter()
    counterNotWHOUTNovo = Counter()
    counterSerWHOUTNovo = Counter()
    counterSerNotWHOUTNovo = Counter()

    #///-----------////---------------------///-----///

    counterWHOUT_MECLUX = Counter()
    counterNotWHOUT_MECLUX = Counter()
    counterSerWHOUT_MECLUX = Counter()
    counterSerNotWHOUT_MECLUX = Counter()



    for ld in ldDetWHOUT:
        
        counterWHOUT[ld['ld_part'].upper()] += ld['ld_qty_oh']

    # for ld in ldDetNotWHOUT:
    #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUT:
        counterSerWHOUT[ser['ser_part'].upper()] += 1

    for ser in serMstrNotWHOUT:
        counterSerNotWHOUT[ser['ser_part'].upper()] += 1
        counterNotWHOUT[ser['ser_part'].upper()] += ser['ser_qty_avail']

    #//----------------------------//-----------------------------------////////---------------------------

    for ld in ldDetWHOUTNovo:
        counterWHOUTNovo[ld['ld_part'].upper()] += ld['ld_qty_oh']

    # for ld in ldDetNotWHOUT:
    #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUTNovo:
        counterSerWHOUTNovo[ser['ser_part'].upper()] += 1

    for ser in serMstrNotWHOUTNovo:
        print("->",ser)
        counterSerNotWHOUTNovo[ser['ser_part'].upper()] += 1
        counterNotWHOUTNovo[ser['ser_part'].upper()] += ser['ser_qty_avail']

    #//----------------------------//-----------------------------------////////---------------------------

    for ld in ldDetWHOUT_MECALUX:
        counterWHOUT_MECLUX[ld['ld_part'].upper()] += ld['ld_qty_oh']

    # for ld in ldDetNotWHOUT:
    #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUT_MECALUX:
        counterSerWHOUT_MECLUX[ser['ser_part'].upper()] += 1

    for ser in serMstrNotWHOUT_MECALUX:
        print("->",ser)
        counterSerNotWHOUT_MECLUX[ser['ser_part'].upper()] += 1
        counterNotWHOUT_MECLUX[ser['ser_part'].upper()] += ser['ser_qty_avail'] #conta os que não são de whou nem de meclux

    #//----------------------------//-----------------------------------////////---------------------------

    for ld in ldDetWHOUT:
        for pt in ptMstr:
            if pt['pt_part'].upper() == ld['ld_part'].upper():
                if len(pt['pt_desc1']) == 24:
                    pt['pt_desc1'] = pt['pt_desc1'] + "" + pt['pt_desc2']
                else:
                    pt['pt_desc1'] = pt['pt_desc1'] + " " + pt['pt_desc2']
                if str(counterWHOUT[ld['ld_part']]) == '0E-10':
                    counterWHOUT[ld['ld_part']] = '0.0000000000'
                if not any(elem['itemNumber'].upper() == ld['ld_part'].upper() for elem in elementos):
                    row = {
                        'itemNumber': ld['ld_part'],
                        'descricao': pt['pt_desc1'],
                        'qtyOnHand': str(counterWHOUT[ld['ld_part']])[:-11],
                        'serialsWHOUT': counterSerWHOUT[ld['ld_part']],     
                        'qtyOnHandNotWHOUT': 0,
                        'serialsNotWHOUT': counterSerNotWHOUT[ld['ld_part']],
                        'unitCost': '-',
                        'min': '-',
                        'max': '-',
                        'consumptionValue': 0,
                        'consumptionValueDiaSeguinte':0,
                        'consumptionValueDiaAposDiaSeguinte': 0,
                        'totalCost': '-',
                        'qtyOnHand_mecalux':str(counterWHOUT[ld['ld_part']])[:-11] ,
                        'serialsMECALUX': counterSerNotWHOUTNovo[ld['ld_part']],

                    }
                    elementos.append(row)
    
    
    for ldNotWHOUT in ldDetNotWHOUTNovo:
        for elem in elementos:

            if elem['itemNumber'].upper() == ldNotWHOUT['ld_part'].upper():
                elem['qtyOnHandNotWHOUT'] = str(counterNotWHOUTNovo[ldNotWHOUT['ld_part']])[:-11]
                if elem['qtyOnHandNotWHOUT'] == '':
                    elem['qtyOnHandNotWHOUT'] = '0'
                break
    for ldNotWHOUT in ldDetNotWHOUT_MECALUX:
        for elem in elementos:

            if elem['itemNumber'].upper() == ldNotWHOUT['ld_part'].upper():
                elem['qtyOnHand_mecalux'] = str(counterNotWHOUT_MECLUX[ldNotWHOUT['ld_part']])[:-11]
                if elem['qtyOnHand_mecalux'] == '':
                    elem['qtyOnHand_mecalux'] = '0'
                break
    

    return render(request, 'tenda_Meclux.html', {'elementos': elementos})


def downladExcelTenda(): #guardaFicheiroHistorico
    import io
    import openpyxl as xl
    print("ENtrou no scheduler")
    ldDetWHOUT = LdDet.objects.filter(ld_site="3515", ld_loc="WHOUT").values(
        "ld_part", "ld_loc", "ld_qty_oh"
    )
    ldDetNotWHOUT = (LdDet.objects.filter(ld_site="3515", ld_part__in=ldDetWHOUT.values_list("ld_part"))
        .exclude(Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY"))
        .values("ld_part", "ld_loc")
    )

    ldDetNotWHOUTNovo = (LdDet.objects.filter(ld_site="3515", ld_part__in=ldDetWHOUT.values_list("ld_part"))
    .exclude(Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY")).values("ld_part", "ld_loc") ) 
   
    ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT.values_list("ld_part"), pt_site="3515").values("pt_part", "pt_desc1", "pt_desc2")
    
    serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515", ser_loc="WHOUT").values("ser_part")
    serMstrNotWHOUT = (
        SerMstr.objects.filter(
            ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515"
        )
        .exclude(Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY"))
        .values("ser_part", "ser_qty_avail")
    )


    elementos = []

    xxusrw = XxusrwWkfl.objects.filter(
        xxusrw_key6="KCB00602-AA", xxusrw_domain="3511010"
    ).values("xxusrw_decfld_1", "xxusrw_key4", "xxusrw_key6", "xxusrw_decfld_2")
    # woMstr = WoMstr.objects.filter(wo_part__in=xxusrw.values_list('xxusrw_key4'),
    #                                wo_due_date="2022-04-20").values(
    #     'wo_part', 'wo_qty_exp_complete')

    counterWHOUT = Counter()
    counterNotWHOUT = Counter()
    counterSerWHOUT = Counter()
    counterSerNotWHOUT = Counter()

    for ld in ldDetWHOUT:
        counterWHOUT[ld["ld_part"]] += ld["ld_qty_oh"]

    for ld in ldDetNotWHOUT:
        counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

    for ser in serMstrWHOUT:
        counterSerWHOUT[ser["ser_part"]] += 1

    for ser in serMstrNotWHOUT:
        counterSerNotWHOUT[ser["ser_part"]] += 1
        counterNotWHOUT[ser["ser_part"]] += ser["ser_qty_avail"]

    for ld in ldDetWHOUT:
        for pt in ptMstr:
            if pt["pt_part"] == ld["ld_part"]:
                if len(pt["pt_desc1"]) == 24:
                    pt["pt_desc1"] = pt["pt_desc1"] + "" + pt["pt_desc2"]
                else:
                    pt["pt_desc1"] = pt["pt_desc1"] + " " + pt["pt_desc2"]
                if str(counterWHOUT[ld["ld_part"]]) == "0E-10":
                    counterWHOUT[ld["ld_part"]] = "0.0000000000"
                if not any(elem["itemNumber"] == ld["ld_part"] for elem in elementos):
                    row = {
                        "itemNumber": ld["ld_part"],
                        "descricao": pt["pt_desc1"],
                        "qtyOnHand": str(counterWHOUT[ld["ld_part"]])[:-11],
                        "serialsWHOUT": counterSerWHOUT[ld["ld_part"]],
                        "qtyOnHandNotWHOUT": 0,
                        "serialsNotWHOUT": counterSerNotWHOUT[ld["ld_part"]],
                        "unitCost": "-",
                        "min": "-",
                        "max": counterNotWHOUT,
                        "consumptionValue": 0,
                        "consumptionValueDiaSeguinte": 0,
                        "consumptionValueDiaAposDiaSeguinte": 0,
                        "totalCost": "-",
                    }
                    elementos.append(row)
    for ldNotWHOUT in ldDetNotWHOUT:
        for elem in elementos:
            if elem["itemNumber"] == ldNotWHOUT["ld_part"]:
                elem["qtyOnHandNotWHOUT"] = str(counterNotWHOUT[ldNotWHOUT["ld_part"]])[
                    :-11
                ]
                if elem["qtyOnHandNotWHOUT"] == "":
                    elem["qtyOnHandNotWHOUT"] = "0"
                break

    
    wbProduction = Workbook()
    file = io.BytesIO()
    sheetProduction = wbProduction.get_sheet_by_name("Sheet")
    sheetProduction.append(
        [
            "Item Number",
            "Description",
            "Qty on Hand-WHOUT",
            "Qt",
            "Qty on Hand-not WHOUT",
            "Qt",
            "Unit Cost",
            "Total Cost (WHOUT)",
            "1 day Consuption",
            "2 day Consuption",
            "3 day Consuption"
        ])
    for elem in elementos: # em principio não precisa do all()
        sheetProduction.append(
            [
                str(elem.itemNumber),
                str(elem.descricao),
                str(elem.qtyOnHand),
                str(elem.serialsWHOUT),
                str(elem.qtyOnHandNotWHOUT),
                str(elem.serialsNotWHOUT),
                str(elem.unitCost),
                str(elem.totalCost),
                str(elem.consumptionValue),
                str(elem.consumptionValueDiaSeguinte),
                str(elem.consumptionValueDiaAposDiaSeguinte)
            ]
        )

    #wbProduction.save(caminho)
    # Voltar a pocição inicial do IO file object
    file.seek(0)

    data_fmt = datetime.now().strftime("%d-%m-%Y")
    #var caminho não vai dar a nenhum sitio
    caminho =  "C:\\Users\\PMARTI30\\Desktop\\historicoTrackingPage"+data_fmt+".xlsx"
    caminho1 =  "\\\\pavpd002\\e_proj\sharedir\\MP&L\\PROCEDIMENTOS\\Packaging\\TMP_NP\\historicoTrackingPage"+data_fmt+".xlsx"

     
    """ wbProduction.save(caminho)
    wbProduction.save(file) """
    print("FICHEIRO GUARDADO")
    wbProduction.save(file)
    # Voltar a pocição inicial do IO file object
    file.seek(0) 

    data_fmt = datetime.now().strftime("%d-%m-%Y %H:%M")
    fresp = FileResponse(
        file, filename=f"TrackingPage_Historico1Mes_{data_fmt}.xlsx", as_attachment=True
    )
    return fresp

    return fresp





def uploadTendaCosts(request):
    if request.method == "GET":
        ldDetWHOUT = LdDet.objects.filter(ld_site="3515", ld_loc="WHOUT").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )
        ldDetNotWHOUT = (
            LdDet.objects.filter(
                ld_site="3515", ld_part__in=ldDetWHOUT.values_list("ld_part")
            )
            .exclude(Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY"))
            .values("ld_part", "ld_loc")
        )
        ptMstr = PtMstr.objects.filter(
            pt_part__in=ldDetWHOUT.values_list("ld_part"), pt_site="3515"
        ).values("pt_part", "pt_desc1", "pt_desc2")
        serMstrWHOUT = SerMstr.objects.filter(
            ser_part__in=ldDetWHOUT.values_list("ld_part"),
            ser_site="3515",
            ser_loc="WHOUT",
        ).values("ser_part")
        serMstrNotWHOUT = (
            SerMstr.objects.filter(
                ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515"
            )
            .exclude(Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY"))
            .values("ser_part", "ser_qty_avail")
        )

        xx = []
        elementos = []

        partNumberList = ldDetWHOUT.values_list("ld_part").values("ld_part").distinct()

        for pn in partNumberList:
            if SctDet.objects.filter(
                sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
            ).exists():
                xx.append(
                    SctDet.objects.filter(
                        sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
                    ).values("sct_cst_tot", "sct_part")
                )
            # if XxusrwWkfl.objects.filter(xxusrw_key6=pn['ld_part'], xxusrw_domain="3511010").exists():
            #     xx.append(XxusrwWkfl.objects.filter(xxusrw_key6=pn['ld_part'], xxusrw_domain="3511010").values(
            #         'xxusrw_decfld_1',
            #         'xxusrw_key4',
            #         'xxusrw_key6',
            #         'xxusrw_decfld_2')[0])

        counterWHOUT = Counter()
        counterNotWHOUT = Counter()
        counterSerWHOUT = Counter()
        counterSerNotWHOUT = Counter()

        for ld in ldDetWHOUT:
            counterWHOUT[ld["ld_part"]] += ld["ld_qty_oh"]

        # for ld in ldDetNotWHOUT:
        #     counterNotWHOUT[ld['ld_part']] += ld['ld_qty_oh']

        for ser in serMstrWHOUT:
            counterSerWHOUT[ser["ser_part"]] += 1

        for ser in serMstrNotWHOUT:
            counterSerNotWHOUT[ser["ser_part"]] += 1
            counterNotWHOUT[ser["ser_part"]] += ser["ser_qty_avail"]

        for ld in ldDetWHOUT:
            for pt in ptMstr:
                if pt["pt_part"] == ld["ld_part"]:
                    if len(pt["pt_desc1"]) == 24:
                        pt["pt_desc1"] = pt["pt_desc1"] + "" + pt["pt_desc2"]
                    else:
                        pt["pt_desc1"] = pt["pt_desc1"] + " " + pt["pt_desc2"]
                    if not any(
                        elem["itemNumber"] == ld["ld_part"] for elem in elementos
                    ):
                        row = {
                            "itemNumber": ld["ld_part"],
                            "descricao": pt["pt_desc1"],
                            "qtyOnHand": str(counterWHOUT[ld["ld_part"]])[:-11],
                            "serialsWHOUT": counterSerWHOUT[ld["ld_part"]],
                            "qtyOnHandNotWHOUT": 0,
                            "serialsNotWHOUT": counterSerNotWHOUT[ld["ld_part"]],
                            "unitCost": "-",
                            "min": "-",
                            "max": "-",
                            "consumptionValue": 0,
                            "consumptionValueDiaSeguinte": 0,
                            "consumptionValueDiaAposDiaSeguinte": 0,
                            "totalCost": "-",
                        }
                        elementos.append(row)
        for ldNotWHOUT in ldDetNotWHOUT:
            for elem in elementos:
                if elem["itemNumber"] == ldNotWHOUT["ld_part"]:
                    elem["qtyOnHandNotWHOUT"] = str(
                        counterNotWHOUT[ldNotWHOUT["ld_part"]]
                    )[:-11]
                    if elem["qtyOnHandNotWHOUT"] == "":
                        elem["qtyOnHandNotWHOUT"] = "0"
                    break

        for elem in elementos:
            for elemXX in xx:
                for elemenXX in elemXX:
                    if elem["itemNumber"] == elemenXX["sct_part"]:
                        if elem["qtyOnHand"] == "":
                            elem["qtyOnHand"] = "0"
                        elem["unitCost"] = round(float(elemenXX["sct_cst_tot"]), 2)
                        elem["totalCost"] = round(
                            float(elem["qtyOnHand"]) * elem["unitCost"], 2
                        )
                        break

        return render(request, "tenda.html", {"elementos": elementos})



def uploadTendaCostsMECFLOW(request):
    if request.method == "GET":
        ldDetWHOUT1 = LdDet.objects.filter(ld_site="3515",ld_loc__in=["WHOUT","MECFLOW"]).values(
        "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetWHOUT = LdDet.objects.filter(ld_site="3515",  ld_loc="WHOUT").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetMECFLOW = LdDet.objects.filter(ld_site="3515",  ld_loc ="MECFLOW").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )

        ldDetNotWHOUTNotMECFLOW = ( LdDet.objects.filter(ld_site="3515", ld_part__in=ldDetWHOUT1.values_list("ld_part"))
            .exclude(Q(ld_loc__contains="MECFLOW") | Q(ld_loc__contains="WHOUT") |  Q(ld_loc__contains="QUALITY") | Q(ld_loc__contains="PWB") | Q(ld_loc__contains="PLC") | Q(ld_loc="WIP")) #para ficar apenas o WHOUT -) tenda
            .values("ld_part", "ld_loc")
        )

        ptMstr = PtMstr.objects.filter(pt_part__in=ldDetWHOUT1.values_list("ld_part"), pt_site="3515"
        ).values("pt_part", "pt_desc1", "pt_desc2")
        
        serMstrWHOUT = SerMstr.objects.filter(ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515", ser_loc="WHOUT"
        ).values("ser_part")

        serMstrMECFLOW = SerMstr.objects.filter(ser_part__in=ldDetMECFLOW.values_list("ld_part"), ser_site="3515", ser_loc="MECFLOW"
        ).values("ser_part")

        serMstrNotWHOUTNotMECFLOW = (SerMstr.objects.filter(ser_part__in=ldDetNotWHOUTNotMECFLOW.values_list("ld_part"), ser_site="3515")
            .exclude(Q(ser_loc__contains="WHOUT") | Q(ser_loc__contains="MECFLOW") | Q(ser_loc__contains="QUALITY") | Q(ser_loc__contains="PWB") | Q(ser_loc__contains="PLC") | Q(ser_loc="WIP"))
            .values("ser_part", "ser_qty_avail")
        )

        xx = []
        elementos = []

        partNumberList = ldDetWHOUT1.values_list("ld_part").values("ld_part").distinct()

        for pn in partNumberList:
            if SctDet.objects.filter(
                sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
            ).exists():
                xx.append(
                    SctDet.objects.filter(
                        sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
                    ).values("sct_cst_tot", "sct_part")
                )
            # if XxusrwWkfl.objects.filter(xxusrw_key6=pn['ld_part'], xxusrw_domain="3511010").exists():
            #     xx.append(XxusrwWkfl.objects.filter(xxusrw_key6=pn['ld_part'], xxusrw_domain="3511010").values(
            #         'xxusrw_decfld_1',
            #         'xxusrw_key4',
            #         'xxusrw_key6',
            #         'xxusrw_decfld_2')[0])

        counterWHOUT = Counter()
        #counterNotWHOUT = Counter()
        counterSerWHOUT = Counter()
        #counterSerNotWHOUT = Counter()

        counterMECFLOW = Counter()
        counterSerMECFLOW = Counter()
        counterSerNotWHOUTNotMECFLOW = Counter()
        counterNotWHOUTNotMECFLOW = Counter()

        for ld in ldDetWHOUT:
            counterWHOUT[ld["ld_part"]] += ld["ld_qty_oh"]

        for ld in ldDetMECFLOW:
            counterMECFLOW[ld["ld_part"]] += ld["ld_qty_oh"]

        for ser in serMstrWHOUT:
            counterSerWHOUT[ser["ser_part"].upper()] += 1

        for ser in serMstrMECFLOW:
            counterSerMECFLOW[ser["ser_part"].upper()] += 1

        for ser in serMstrNotWHOUTNotMECFLOW:   #contador de serials para nenhum dos casos
            counterSerNotWHOUTNotMECFLOW[ser["ser_part"].upper()] += 1
            counterNotWHOUTNotMECFLOW[ser["ser_part"].upper()] += ser["ser_qty_avail"]

        for ld in ldDetWHOUT:
            for pt in ptMstr:
                if pt["pt_part"] == ld["ld_part"]:
                    if len(pt["pt_desc1"]) == 24:
                        pt["pt_desc1"] = pt["pt_desc1"] + "" + pt["pt_desc2"]
                    else:
                        pt["pt_desc1"] = pt["pt_desc1"] + " " + pt["pt_desc2"]
                    if not any(
                        elem["itemNumber"] == ld["ld_part"] for elem in elementos
                    ):
                        row = {
                            "itemNumber": ld["ld_part"],
                            "descricao": pt["pt_desc1"],
                            "qtyOnHand": 0,
                            "serialsWHOUT": counterSerWHOUT[ld["ld_part"]],
                            "qtyOnHandMECFLOW": 0,
                            "serialsMECFLOW": counterSerMECFLOW[ld["ld_part"]],
                            "qtyOnHandNotWHOUTNotMECFLOW": 0,
                            "serialsNotWHOUTNotMECFLOW": counterSerNotWHOUTNotMECFLOW[ld["ld_part"]],
                            "unitCost": "-",
                            "min": "-",
                            "max": "-",
                            "consumptionValue": 0,
                            "consumptionValueDiaSeguinte": 0,
                            "consumptionValueDiaAposDiaSeguinte": 0,
                            "totalCost": "-",
                        }
                        elementos.append(row)
        for ldWHOUT in ldDetWHOUT:
            for elem in elementos:
                if elem["itemNumber"] == ldWHOUT["ld_part"]:
                    elem["qtyOnHand"] = str(counterWHOUT[ldWHOUT["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHand"] == "":
                        elem["qtyOnHand"] = "0"
                    break
    
        for ldMECFLOW in ldDetMECFLOW:
            for elem in elementos:
                if elem["itemNumber"] == ldMECFLOW["ld_part"]:
                    elem["qtyOnHandMECFLOW"] = str(counterMECFLOW[ldMECFLOW["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHandMECFLOW"] == "":
                        elem["qtyOnHandMECFLOW"] = "0"
                    break

        for ldMECFLOW in ldDetNotWHOUTNotMECFLOW:
            for elem in elementos:
                if elem["itemNumber"] == ldMECFLOW["ld_part"]:
                    elem["qtyOnHandNotWHOUTNotMECFLOW"] = str(counterNotWHOUTNotMECFLOW[ldMECFLOW["ld_part"]])[
                        :-11
                    ]
                    if elem["qtyOnHandNotWHOUTNotMECFLOW"] == "":
                        elem["qtyOnHandNotWHOUTNotMECFLOW"] = "0"
                    break

        for elem in elementos:
            for elemXX in xx:
                for elemenXX in elemXX:
                    if elem["itemNumber"] == elemenXX["sct_part"]:
                        if elem["qtyOnHand"] == "":
                            elem["qtyOnHand"] = "0"
                        elem["unitCost"] = round(float(elemenXX["sct_cst_tot"]), 2)
                        elem["totalCost"] = round(
                            float(elem["qtyOnHand"]) * elem["unitCost"], 2
                        )
                        break

        return render(request, "tenda_Mecflow.html", {"elementos": elementos})
    




def updateTenda(request):
    if request.method == "GET":
        diaCorrente = date.today()
        diaSeguinte = date.today() + timedelta(days=1)
        diaAposDiaSeguinte = date.today() + timedelta(days=2)

        # quinta
        if diaCorrente.weekday() == 3:
            diaSeguinte = date.today() + timedelta(days=1)
            diaAposDiaSeguinte = date.today() + timedelta(days=4)
        # sexta
        if diaCorrente.weekday() == 4:
            diaSeguinte = date.today() + timedelta(days=3)
            diaAposDiaSeguinte = date.today() + timedelta(days=4)
        # sabado
        if diaCorrente.weekday() == 5:
            diaCorrente = date.today() + timedelta(days=2)
            diaSeguinte = date.today() + timedelta(days=3)
            diaAposDiaSeguinte = date.today() + timedelta(days=4)
        # domingo
        if diaCorrente.weekday() == 6:
            diaCorrente = date.today() + timedelta(days=1)
            diaSeguinte = date.today() + timedelta(days=2)
            diaAposDiaSeguinte = date.today() + timedelta(days=3)

        ldDetWHOUT = LdDet.objects.filter(ld_site="3515", ld_loc="WHOUT").values(
            "ld_part", "ld_loc", "ld_qty_oh"
        )
        ldDetNotWHOUT = (
            LdDet.objects.filter(
                ld_site="3515", ld_part__in=ldDetWHOUT.values_list("ld_part")
            )
            .exclude(Q(ld_loc="WHOUT") | Q(ld_loc="QUALITY"))
            .values("ld_part", "ld_loc")
        )
        ptMstr = PtMstr.objects.filter(
            pt_part__in=ldDetWHOUT.values_list("ld_part"), pt_site="3515"
        ).values("pt_part", "pt_desc1", "pt_desc2")
        serMstrWHOUT = SerMstr.objects.filter(
            ser_part__in=ldDetWHOUT.values_list("ld_part"),
            ser_site="3515",
            ser_loc="WHOUT",
        ).values("ser_part")
        serMstrNotWHOUT = (
            SerMstr.objects.filter(
                ser_part__in=ldDetWHOUT.values_list("ld_part"), ser_site="3515"
            )
            .exclude(Q(ser_loc="WHOUT") | Q(ser_loc="QUALITY"))
            .values("ser_part", "ser_qty_avail")
        )
        xxusrw = XxusrwWkfl.objects.filter(
            xxusrw_key6__in=ldDetWHOUT.values_list("ld_part"), xxusrw_domain="3511010"
        ).values("xxusrw_decfld_1", "xxusrw_key4", "xxusrw_key6", "xxusrw_decfld_2")
        woMstr = WoMstr.objects.filter(
            wo_part__in=xxusrw.values_list("xxusrw_key4"),
            wo_due_date=diaCorrente.strftime("%Y-%m-%d"),
        ).values("wo_part", "wo_qty_exp_complete")
        woMstrDiaSeguinte = WoMstr.objects.filter(
            wo_part__in=xxusrw.values_list("xxusrw_key4"),
            wo_due_date=diaSeguinte.strftime("%Y-%m-%d"),
        ).values("wo_part", "wo_qty_exp_complete")
        woMstrDiaAposDiaSeguinte = WoMstr.objects.filter(
            wo_part__in=xxusrw.values_list("xxusrw_key4"),
            wo_due_date=diaAposDiaSeguinte.strftime("%Y-%m-%d"),
        ).values("wo_part", "wo_qty_exp_complete")

        elementos = []
        xxList = []

        partNumberList = ldDetWHOUT.values_list("ld_part").values("ld_part").distinct()

        for pn in partNumberList:
            if SctDet.objects.filter(
                sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
            ).exists():
                xxList.append(
                    SctDet.objects.filter(
                        sct_part=pn["ld_part"], sct_site="3515", sct_sim="Standard"
                    ).values("sct_cst_tot", "sct_part")
                )

        counterWHOUT = Counter()
        counterNotWHOUT = Counter()
        counterSerWHOUT = Counter()
        counterSerNotWHOUT = Counter()

        for ld in ldDetWHOUT:
            counterWHOUT[ld["ld_part"]] += ld["ld_qty_oh"]

        for ser in serMstrWHOUT:
            counterSerWHOUT[ser["ser_part"]] += 1

        for ser in serMstrNotWHOUT:
            counterSerNotWHOUT[ser["ser_part"]] += 1
            counterNotWHOUT[ser["ser_part"]] += ser["ser_qty_avail"]

        for ld in ldDetWHOUT:
            for pt in ptMstr:
                if pt["pt_part"] == ld["ld_part"]:
                    if len(pt["pt_desc1"]) == 24:
                        pt["pt_desc1"] = pt["pt_desc1"] + "" + pt["pt_desc2"]
                    else:
                        pt["pt_desc1"] = pt["pt_desc1"] + " " + pt["pt_desc2"]
                    if not any(
                        elem["itemNumber"] == ld["ld_part"] for elem in elementos
                    ):
                        row = {
                            "diaAtual": diaCorrente,
                            "diaSeguinte": diaSeguinte,
                            "diaAposDiaSeguinte": diaAposDiaSeguinte,
                            "itemNumber": ld["ld_part"],
                            "descricao": pt["pt_desc1"],
                            "qtyOnHand": str(counterWHOUT[ld["ld_part"]])[:-11],
                            "serialsWHOUT": counterSerWHOUT[ld["ld_part"]],
                            "qtyOnHandNotWHOUT": 0,
                            "serialsNotWHOUT": counterSerNotWHOUT[ld["ld_part"]],
                            "unitCost": "-",
                            "min": "-",
                            "max": "-",
                            "consumptionValue": 0,
                            "consumptionValueDiaSeguinte": 0,
                            "consumptionValueDiaAposDiaSeguinte": 0,
                            "totalCost": "-",
                        }
                        elementos.append(row)
        for ldNotWHOUT in ldDetNotWHOUT:
            for elem in elementos:
                if elem["itemNumber"] == ldNotWHOUT["ld_part"]:
                    elem["qtyOnHandNotWHOUT"] = str(
                        counterNotWHOUT[ldNotWHOUT["ld_part"]]
                    )[:-11]
                    if elem["qtyOnHandNotWHOUT"] == "":
                        elem["qtyOnHandNotWHOUT"] = "0"
                    break

        for xx in xxusrw:
            for wo in woMstr:
                if wo["wo_part"] == xx["xxusrw_key4"]:
                    for elem in elementos:
                        if elem["itemNumber"] == xx["xxusrw_key6"]:
                            elem["consumptionValue"] += round(
                                float(xx["xxusrw_decfld_1"])
                                * float(wo["wo_qty_exp_complete"]),
                                2,
                            )
                            break
            for woDiaSeguinte in woMstrDiaSeguinte:
                if woDiaSeguinte["wo_part"] == xx["xxusrw_key4"]:
                    for elem in elementos:
                        if elem["itemNumber"] == xx["xxusrw_key6"]:
                            elem["consumptionValueDiaSeguinte"] += round(
                                float(xx["xxusrw_decfld_1"])
                                * float(woDiaSeguinte["wo_qty_exp_complete"]),
                                2,
                            )
                            break
            for woDiaAposDiaSeguinte in woMstrDiaAposDiaSeguinte:
                if woDiaAposDiaSeguinte["wo_part"] == xx["xxusrw_key4"]:
                    for elem in elementos:
                        if elem["itemNumber"] == xx["xxusrw_key6"]:
                            elem["consumptionValueDiaAposDiaSeguinte"] += round(
                                float(xx["xxusrw_decfld_1"])
                                * float(woDiaAposDiaSeguinte["wo_qty_exp_complete"]),
                                2,
                            )
                            break
        for elem in elementos:
            for elemXX in xxList:
                for elemenXX in elemXX:
                    if elem["itemNumber"] == elemenXX["sct_part"]:
                        if elem["qtyOnHand"] == "":
                            elem["qtyOnHand"] = "0"
                        elem["unitCost"] = round(float(elemenXX["sct_cst_tot"]), 2)
                        elem["totalCost"] = round(
                            float(elem["qtyOnHand"]) * elem["unitCost"], 2
                        )
                        break
        return render(request, "tenda.html", {"elementos": elementos})


def icdr(request):
    # UserICDR.objects.all().delete()
    # workbook = openpyxl.load_workbook(
    #     "//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Receiving/ICDR/Copy of Copy of Copy of ICDR.xlsx")
    # worksheet = workbook['Users']
    #
    # for element in worksheet:
    #     if element[0].value != "":
    #         novoElemento = UserICDR(
    #             None,
    #             element[3].value,
    #             element[0].value,
    #             element[1].value,
    #             element[2].value
    #         )
    #         novoElemento.save()

    ultimoElemento = ICDRultimoValor.objects.first()
    motivos = ICDRmotivo.objects
    utilizadoresICDR = UserICDR.objects
    listaICDR = ICDR.objects
    filtroDesativado = request.session.get("filtroDesativado")
    tipo1 = request.user.groups.filter(name="ICDR - tipo1").exists()
    tipo2 = request.user.groups.filter(name="ICDR - tipo2").exists()
    tipo3 = request.user.groups.filter(name="ICDR - tipo3").exists()
    tipo1tipo2tipo3 = tipo1 and tipo2 and tipo3
    tipo1tipo2 = tipo1 and tipo2
    tipo1tipo3 = tipo1 and tipo3
    tipo2tipo3 = tipo2 and tipo3
    return render(
        request,
        "icdr.html",
        {
            "ultimoElemento": ultimoElemento,
            "motivos": motivos,
            "utilizadoresICDR": utilizadoresICDR,
            "listaICDR": listaICDR,
            "tipo1": tipo1,
            "tipo2": tipo2,
            "tipo3": tipo3,
            "tipo1tipo2": tipo1tipo2,
            "tipo2tipo3": tipo2tipo3,
            "tipo1tipo3": tipo1tipo3,
            "tipo1tipo2tipo3": tipo1tipo2tipo3,
            "filtroDesativado": filtroDesativado,
        },
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingLINES").exists())
def lineRequest(request):
    lineRequestProcessing = LineRequestProcessing.objects
    lineRequestFinished = LineRequestFinished.objects
    linhas = Line.objects
    justificacoes = Justification.objects
    return render(
        request,
        "lineRequest.html",
        {
            "lineRequestFinished": lineRequestFinished,
            "lineRequestProcessing": lineRequestProcessing,
            "linhas": linhas,
            "justificacoes": justificacoes,
        },
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="coordReceiving").exists())
def configurations(request):
    linhas = Line.objects
    justificacoes = Justification.objects
    # users = User.objects.filter(groups__name__in=['receivingLINES', 'receivingMNFG', 'receivingTPM']).distinct()
    users = User.objects
    return render(
        request,
        "configurations.html",
        {"linhas": linhas, "justificacoes": justificacoes, "users": users},
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="coordReceiving").exists())
def configurationsTPM(request):
    posicao1 = ReceivingPosicao1Items.objects
    posicao2 = ReceivingPosicao2Items.objects
    posicao3 = ReceivingPosicao3Items.objects
    posicao4 = ReceivingPosicao4Items.objects
    posicao5 = ReceivingPosicao5Items.objects
    posicao6 = ReceivingPosicao6Items.objects
    posicao7 = ReceivingPosicao7Items.objects
    posicao8 = ReceivingPosicao8Items.objects
    tipoSubItems = ReceivingSubItemsTipo.objects
    subItems = ReceivingSubItems.objects
    users = User.objects
    habilitados = HabilitarQuadrados.objects
    tempos = DefinirTempos.objects
    return render(
        request,
        "configurationsTPM.html",
        {
            "posicao1": posicao1,
            "posicao2": posicao2,
            "posicao3": posicao3,
            "posicao4": posicao4,
            "posicao5": posicao5,
            "posicao6": posicao6,
            "posicao7": posicao7,
            "posicao8": posicao8,
            "subItems": subItems,
            "habilitados": habilitados,
            "tempos": tempos,
            "users": users,
            "tipoSubItems": tipoSubItems,
        },
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingMNFG").exists())
def mnfgSupply(request):
    hoje = datetime.today()
    agora = datetime.now().time()
    data_formatada = hoje.strftime('%m/%d/%Y')
    hora_formatada = agora.strftime('%H:%M')
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')

    timestamp = MNFGSupplyItems.objects.first()
    print("Timesmtamp-->", timestamp.timestamp)
    
     
    
    taskBrowse = TaskBrowse.objects
    data = LastUpdate.objects
    areaA = AreaA.objects
    areaB = AreaB.objects
    kardex = Kardex.objects
    

    dia = date.today() - timedelta(days=10)
    """ dadosQAD = WtskMstr.objects.filter(
        wtsk_create_date__gt=dia.strftime("%Y-%m-%d")
    ).exclude(wtsk_to_loc="kardex") """

    wevDetFilterFA = WevdDet.objects.filter(
        wevd_create_date__gt=dia.strftime("%Y-%m-%d"),
        wevd_loc_to__startswith="F",
        wevd_status="PENDING",
        wevd_line="Repl02",
        wevd_domain="3511010",
    ).values("wevd_event_id", "wevd_from_part", "wevd_qty_remain", "wevd_create_date")

    wevDetFilterBP = WevdDet.objects.filter(
        wevd_create_date__gt=dia.strftime("%Y-%m-%d"),
        wevd_loc_to__startswith="S",
        wevd_status="PENDING",
        wevd_line="Repl02",
        wevd_domain="3511010",
    ).values("wevd_event_id", "wevd_from_part", "wevd_qty_remain", "wevd_create_date")

    wevMstrFA = WevMstr.objects.filter(
        wev_create_date__gt=dia.strftime("%Y-%m-%d"),
        wev_event_id__in=wevDetFilterFA.values_list("wevd_event_id"),
    ).values("wev_event_id", "wev_create_time")

    wevMstrBP = WevMstr.objects.filter(
        wev_create_date__gt=dia.strftime("%Y-%m-%d"),
        wev_event_id__in=wevDetFilterBP.values_list("wevd_event_id"),
    ).values("wev_event_id", "wev_create_time")

    elementosFA = []
    elementosBP = []
    for mstr in wevMstrFA:
        for det in wevDetFilterFA:
            if mstr["wev_event_id"] == det["wevd_event_id"]:
                row = {
                    "item": det["wevd_from_part"],
                    "qtyRemaining": str(det["wevd_qty_remain"])[:-11],
                    "date": str(det["wevd_create_date"])[:-9],
                    "time": mstr["wev_create_time"],
                }
                elementosFA.append(row)
                break
    for mstr in wevMstrBP:
        for det in wevDetFilterBP:
            if mstr["wev_event_id"] == det["wevd_event_id"]:
                row = {
                    "item": det["wevd_from_part"],
                    "qtyRemaining": str(det["wevd_qty_remain"])[:-11],
                    "date": str(det["wevd_create_date"])[:-9],
                    "time": mstr["wev_create_time"],
                }
                elementosBP.append(row)
                break
    contador =0
    
    """ for i in dadosQAD:
        if i.wtsk_from_stor_zone == "KARDEX": #assim estás a ver BPSMD SUPPLY
            contador +=1
            print("-",i.wtsk_create_time,contador) """
        # agora =  datetime.now()
        # data_atual = datetime.now().date()
        # data_atual1 = data_atual.strftime('%Y-%m-%d')
        # i_time = datetime.strptime(i["time"], "%H%M%S")
        # proxima_hora = agora + timedelta(hours=1) 
        # proxHora = proxima_hora.strftime("%H%M%S")
        # if data_atual1 == i["date"]:
        #     print("Data igual ao dia de hoje!", proxHora)
        #     proxHora_dt = datetime.strptime(proxHora, "%H%M%S")
        #     if proxHora_dt >= i_time:
        #         contador += 1
    dadosQAD = MNFGSupplyItems.objects.all()
    #dadosQAD = MNFGSupplyItems.objects.filter(wtsk_from_stor_zone__in=areaB.values_list("storageZone"))
    return render(
        request,
        "mNFGSupply.html",
        {
            "taskBrowse": taskBrowse,
            "data": data,
            "areaA": areaA,
            "areaB": areaB,
            "kardex": kardex,
            "dadosQAD": dadosQAD,
            "elementosFA": elementosFA,
            "elementosBP": elementosBP,
            "timestamp": timestamp.timestamp
            
            
        },
    )
def mostraContadores(request):
    print(request.POST)
    hoje = datetime.today()
    agora = datetime.now().time()
    data_formatada = hoje.strftime('%m/%d/%Y')
    hora_formatada = agora.strftime('%H:%M')
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
    #precisas de percorrer as tabelas todas para pegar o contador de cada uma ads 3 tabelas, sem if de tabela
    print("MOSTRA contadores")

    if request.POST["turno"] == "1":
        print("Primeiro Turno")
        contadorFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
        contadorBpsmd = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
        contadorBpDrop = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
    elif request.POST["turno"] == "2":
        print("Segundo turno")
        contadorFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
        contadorBpsmd = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
        contadorBpDrop = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  16:30' AND '{data_formatada} 01:00' """)
    else:
        print("Terceiro turno")
        contadorFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
        contadorBpsmd = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
        contadorBpDrop = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
    print(contadorBpsmd)

    return JsonResponse({"message": "OK", "contadorFA" : len(contadorFA), "contadorBpsmd" : len(contadorBpsmd), "contadorBpDrop" : len(contadorBpDrop)})


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingMNFG").exists())
def faSupply(request):
    from datetime import datetime
  
    taskBrowse = TaskBrowse.objects
    areaB = AreaB.objects
    data = LastUpdate.objects
    redItemsBd =  FASupplyRedItems.objects
    now = datetime.now()
    #adiciona_val_vermelhos_automatico_Supply()
    # hoje = datetime.today()
    # data_formatada = hoje.strftime('%m/%d/%Y')
    # agora = datetime.now().time()
    # hora_formatada = agora.strftime('%H:%M')
    # print("->", type(hora_formatada))
    # dia_seguinte = hoje + timedelta(days=1)
    # data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
    # hora_formatada = "08:40"
    # if hora_formatada > "08:00" and hora_formatada < "16:30": 
    #     print("Primeiro Turno")
    #     contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
    # elif hora_formatada > "16:30" and hora_formatada < "24:00":
    #     print("Segundo turno")
    #     contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
    # else:
    #     print("Terceiro turno")
    #     contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
    # print("QUERRY-->",data)
    #contador = FASupplyRedItems.objects.raw(f"""SELECT * from receiving_fasupplyreditems WHERE timestamp >= '{now.strftime("%m/%d/%Y")}'""")
    #dateLastUpdate = Receiving_SupplyPages_Refresh.objects.all().values('lastUpdate').last()#.get('lastUpdate')
    #print("dateLastUpdate---",dateLastUpdate)
    #bd_estado = Receiving_SupplyPages_Refresh.objects.all().values('estado').last()#.get('estado')
    dia = date.today() - timedelta(days=10)
    
    """ dadosQAD = WtskMstr.objects.filter(
        wtsk_create_date__gt=dia.strftime("%Y-%m-%d")
    ).exclude(wtsk_to_loc="kardex") """
    #if line.wtsk_from_stor_zone == areaB.storageZone
    dadosQAD = MNFGSupplyItems.objects.filter(wtsk_from_stor_zone__in=areaB.values_list("storageZone"))#.all()
    #ldDetNotWHOUTNotMECFLOW = ( LdDet.objects.filter(ld_site="3515", ld_part__in=ldDetWHOUT1.values_list("ld_part"))
                               
    print("TAMANHO DOS DADOS", len(dadosQAD.all()))
    return render(
        request,
        "fASupplyNovaFunc.html",
        {"taskBrowse": taskBrowse, "data": data, "areaB": areaB, "dadosQAD": dadosQAD, "redItemsBd" : redItemsBd},
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingMNFG").exists())
def bpDropinSupply(request):
    from datetime import datetime
    taskBrowse = TaskBrowse.objects
    data = LastUpdate.objects
    areaA = AreaA.objects
    #para trazer também
    dia = date.today() - timedelta(days=10)
    
    hoje = datetime.today()
    data_formatada = hoje.strftime('%m/%d/%Y')
    agora = datetime.now().time()
    hora_formatada = agora.strftime('%H:%M')
    print("->", type(hora_formatada))
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
    hora_formatada = "08:40"

    #dateLastUpdate = Receiving_SupplyPages_Refresh.objects.all().values('lastUpdate').last()#.get('lastUpdate')
    #print("dateLastUpdate---",dateLastUpdate)
    #bd_estado = Receiving_SupplyPages_Refresh.objects.all().values('estado').last()
    #contador = FASupplyRedItems.objects.raw(f"""SELECT * from receiving_fasupplyreditems WHERE timestamp >= '{now.strftime("%m/%d/%Y")}'""")
    #tens de trazer este codigo quando for para fazer para correr com o apScheduler
    """ dadosQAD = WtskMstr.objects.filter(
        wtsk_create_date__gt=dia.strftime("%Y-%m-%d")
    ).exclude(wtsk_to_loc="kardex") """
    
    dadosQAD = MNFGSupplyItems.objects.filter(wtsk_from_stor_zone__in=areaA.values_list("storageZone"))
    return render(
        request,
        "bPDropinSupply.html",
        { "taskBrowse": taskBrowse, "data": data, "areaA": areaA, "dadosQAD": dadosQAD},
    )

@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingMNFG").exists())
def bpSMDSupply(request):
    from datetime import datetime
    

    taskBrowse = TaskBrowse.objects
    data = LastUpdate.objects
    kardex = Kardex.objects
    dia = date.today() - timedelta(days=10)
    redItemsBd =  BPSMDSupplyRedItems.objects
    
    hoje = datetime.today()
    data_formatada = hoje.strftime('%m/%d/%Y')
    agora = datetime.now().time()
    hora_formatada = agora.strftime('%H:%M')
    dia_seguinte = hoje + timedelta(days=1)

    #dateLastUpdate = Receiving_SupplyPages_Refresh.objects.all().values('lastUpdate').last()#.get('lastUpdate')
    #bd_estado = Receiving_SupplyPages_Refresh.objects.all().values('estado').last()
    dadosQAD = WtskMstr.objects.filter(
        wtsk_create_date__gt=dia.strftime("%Y-%m-%d")
    ).exclude(wtsk_to_loc="kardex")

    dadosQAD = MNFGSupplyItems.objects.filter(wtsk_from_stor_zone__in=kardex.values_list("storageZone"))
    return render(
        request,
        "bPSMDSupply.html",
        {
            "taskBrowse": taskBrowse,
            "data": data,
            "kardex": kardex,
            "dadosQAD": dadosQAD,
            "redItemsBd" : redItemsBd,
           

        },
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists())
def errorLog(request):
    return render(request, "errorLog.html")


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingMNFG").exists())
def pending(request):
    dia = date.today() - timedelta(days=10)

    wevDetFilterFA = WevdDet.objects.filter(
        wevd_create_date__gt=dia.strftime("%Y-%m-%d"),
        wevd_loc_to__startswith="F",
        wevd_status="PENDING",
        wevd_line="Repl02",
        wevd_domain="3511010",
    ).values(
        "wevd_event_id",
        "wevd_ev_sub_id",
        "wevd_from_part",
        "wevd_loc_to",
        "wevd_qty_req",
        "wevd_qty_remain",
        "wevd_create_date",
    )

    wevDetFilterBP = WevdDet.objects.filter(
        wevd_create_date__gt=dia.strftime("%Y-%m-%d"),
        wevd_loc_to__startswith="S",
        wevd_status="PENDING",
        wevd_line="Repl02",
        wevd_domain="3511010",
    ).values(
        "wevd_event_id",
        "wevd_ev_sub_id",
        "wevd_from_part",
        "wevd_loc_to",
        "wevd_qty_req",
        "wevd_qty_remain",
        "wevd_create_date",
    )

    wevMstrFA = WevMstr.objects.filter(
        wev_create_date__gt=dia.strftime("%Y-%m-%d"),
        wev_event_id__in=wevDetFilterFA.values_list("wevd_event_id"),
    ).values("wev_event_id", "wev_create_time")

    wevMstrBP = WevMstr.objects.filter(
        wev_create_date__gt=dia.strftime("%Y-%m-%d"),
        wev_event_id__in=wevDetFilterBP.values_list("wevd_event_id"),
    ).values("wev_event_id", "wev_create_time")

    data = LastUpdate.objects

    elementosFA = []
    elementosBP = []
    for mstr in wevMstrFA:
        for det in wevDetFilterFA:
            if mstr["wev_event_id"] == det["wevd_event_id"]:
                row = {
                    "id": det["wevd_event_id"],
                    "subID": det["wevd_ev_sub_id"],
                    "item": det["wevd_from_part"],
                    "locTo": det["wevd_loc_to"],
                    "qtyRequired": str(det["wevd_qty_req"])[:-11],
                    "qtyRemaining": str(det["wevd_qty_remain"])[:-11],
                    "date": str(det["wevd_create_date"])[:-9],
                    "time": mstr["wev_create_time"],
                }
                elementosFA.append(row)
                break
    for mstr in wevMstrBP:
        for det in wevDetFilterBP:
            if mstr["wev_event_id"] == det["wevd_event_id"]:
                row = {
                    "id": det["wevd_event_id"],
                    "subID": det["wevd_ev_sub_id"],
                    "item": det["wevd_from_part"],
                    "locTo": det["wevd_loc_to"],
                    "qtyRequired": str(det["wevd_qty_req"])[:-11],
                    "qtyRemaining": str(det["wevd_qty_remain"])[:-11],
                    "date": str(det["wevd_create_date"])[:-9],
                    "time": mstr["wev_create_time"],
                }
                elementosBP.append(row)
                break

    return render(
        request,
        "pending.html",
        {"data": data, "elementosFA": elementosFA, "elementosBP": elementosBP},
    )

def getNokSmart(request):
    #buscar os noks tens de ver se o val é 1 ou null
    print("entrou",request.POST["data_report"])
    smartNoksHistInicioT = ReceivingPosicao2Historico.objects.filter(dataPosicao__contains =request.POST["data_report"], inicioTurnoNok = "1")
    for x in smartNoksHistInicioT:
        print("VALOR ENCONTRADO para NOKs",x)
    smartNoksHistFimT = ReceivingPosicao2Historico.objects.filter(dataPosicao__contains =request.POST["data_report"], fimTurnoNok = "1")
    smartSubItemsNok = ReceivingPosicao2SubItems.objects.filter(nok ="1")
    wbProduction = Workbook()
    sheetProduction = wbProduction.add_sheet("Sheet 1")
    row = 0
    col = 0
    sheetProduction.write(row, col, "Data Posicao")
    sheetProduction.write(row, col + 1, "Item")
    sheetProduction.write(row, col + 2, "Comentario Final Turno")
    sheetProduction.write(row, col + 3, "Comentario Inicio Turno")
    sheetProduction.write(row, col + 4, "Responsavel Inicio Turno")
    sheetProduction.write(row, col + 5, "Responsavel Final Turno")
    sheetProduction.write(row, col + 6, "Data Final Turno")
    sheetProduction.write(row, col + 7, "Data Inicio Turno")
    sheetProduction.write(row, col + 8, "Hora Final Turno")
    sheetProduction.write(row, col + 9, "Hora Inicio Turno")
    sheetProduction.write(row, col + 10, "Fim Turno Ok")
    sheetProduction.write(row, col + 11, "Fim Turno Nok")
    sheetProduction.write(row, col + 12, "Inicio Turno Ok")
    sheetProduction.write(row, col + 13, "Inicio Turno Nok")
    sheetProduction.write(row, col + 14, "Tipo")
    row += 1

    for elem in smartNoksHistInicioT:
        sheetProduction.write(row, col, elem.dataPosicao)
        sheetProduction.write(row, col + 1, elem.item)
        sheetProduction.write(row, col + 2, elem.comentarioFinalTurno)
        sheetProduction.write(row, col + 3, elem.comentarioInicioTurno)
        sheetProduction.write(row, col + 4, elem.responsavelInicioTurno)
        sheetProduction.write(row, col + 5, elem.responsavelFinalTurno)
        sheetProduction.write(row, col + 6, elem.dataFinalTurno)
        sheetProduction.write(row, col + 7, elem.dataInicioTurno)
        sheetProduction.write(row, col + 8, elem.horaFinalTurno)
        sheetProduction.write(row, col + 9, elem.horaInicioTurno)
        sheetProduction.write(row, col + 10, elem.fimTurnoOk)
        sheetProduction.write(row, col + 11, elem.fimTurnoNok)
        sheetProduction.write(row, col + 12, elem.inicioTurnoOk)
        sheetProduction.write(row, col + 13, elem.inicioTurnoNok)
        sheetProduction.write(row, col + 14, elem.tipo)
        row += 1
        
    for elem in smartNoksHistFimT:
        sheetProduction.write(row, col , elem.dataPosicao)
        sheetProduction.write(row, col + 1, elem.item)
        sheetProduction.write(row, col + 2, elem.comentarioFinalTurno)
        sheetProduction.write(row, col + 3, elem.comentarioInicioTurno)
        sheetProduction.write(row, col + 4, elem.responsavelInicioTurno)
        sheetProduction.write(row, col + 5, elem.responsavelFinalTurno)
        sheetProduction.write(row, col + 6, elem.dataFinalTurno)
        sheetProduction.write(row, col + 7, elem.dataInicioTurno)
        sheetProduction.write(row, col + 8, elem.horaFinalTurno)
        sheetProduction.write(row, col + 9, elem.horaInicioTurno)
        sheetProduction.write(row, col + 10, elem.fimTurnoOk)
        sheetProduction.write(row, col + 11, elem.fimTurnoNok)
        sheetProduction.write(row, col + 12, elem.inicioTurnoOk)
        sheetProduction.write(row, col + 13, elem.inicioTurnoNok)
        sheetProduction.write(row, col + 14, elem.tipo)
        row += 1
    
    row += 1
    sheetProduction.write(row, col+2, "SUB ITEMS")
    row += 1

    sheetProduction.write(row, col, "Id Dia")
    sheetProduction.write(row, col + 1, "Item")
    sheetProduction.write(row, col + 2, "Ok")
    sheetProduction.write(row, col + 3, "Nok")
    sheetProduction.write(row, col + 4, "Comentario")
    row += 1

    for elem in smartSubItemsNok:
        sheetProduction.write(row, col, elem.idDia)
        sheetProduction.write(row, col + 1, elem.item)
        sheetProduction.write(row, col + 2, elem.ok)
        sheetProduction.write(row, col + 3, elem.nok)
        sheetProduction.write(row, col + 4, elem.comentario)
        row += 1
    
    wbProduction.save("..\\visteon\\media\\receiving\\tpmReports\\workbookTPMNoks.xls")
    return JsonResponse({"message": "OK"})
@login_required()
@user_passes_test(lambda u: u.groups.filter(name="receivingTPM").exists())
def tpm(request):
    # import keyboard
    # import threading

 

    # def on_press(event):
    #     global text
    #     print(event.name)
    #     """  if event.name == "space":
    #         print(text, end=" ")
    #         text = ""
    #     elif event.name == "esc":
    #         print(text)
    #         text = ""
    #     else:
    #         text += event.name """

    # def start_listener():
    #     # Adiciona um listener para todas as teclas pressionadas
    #     keyboard.on_press(on_press)

    #     # Mantém o script em execução
    #     keyboard.wait()

    # # Cria uma nova thread para executar o listener
    # listener_thread = threading.Thread(target=start_listener)
    # listener_thread.start()

    # # Aguarda a thread terminar (o que nunca acontecerá)
    # listener_thread.join()



    posicao1 = ReceivingPosicao1Items.objects
    posicao2 = ReceivingPosicao2Items.objects
    posicao3 = ReceivingPosicao3Items.objects
    posicao4 = ReceivingPosicao4Items.objects
    posicao5 = ReceivingPosicao5Items.objects
    posicao6 = ReceivingPosicao6Items.objects
    posicao7 = ReceivingPosicao7Items.objects
    posicao8 = ReceivingPosicao8Items.objects
    tipoSubItems = ReceivingSubItemsTipo.objects
    subItems = ReceivingSubItems.objects
    habilitados = HabilitarQuadrados.objects
    tempos = DefinirTempos.objects
    return render(
        request,
        "tpm.html",
        {
            "posicao1": posicao1,
            "posicao2": posicao2,
            "posicao3": posicao3,
            "posicao4": posicao4,
            "posicao5": posicao5,
            "posicao6": posicao6,
            "posicao7": posicao7,
            "posicao8": posicao8,
            "subItems": subItems,
            "habilitados": habilitados,
            "tempos": tempos,
            "tipoSubItems": tipoSubItems,
        },
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists())
def operationAnalyse(request):
    return render(request, "operationAnalyse.html")


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="coordReceiving").exists())
def configurationMNFG(request):
    areaA = AreaA.objects
    areaB = AreaB.objects
    kardex = Kardex.objects
    data = LastUpdate.objects
    return render(
        request,
        "configurationsMNFG.html",
        {"areaA": areaA, "areaB": areaB, "kardex": kardex, "data": data},
    )


@login_required()
@user_passes_test(lambda u: u.groups.filter(name="admin").exists())
def configurationsICDR(request):
    users = User.objects
    utilizadoresICDR = UserICDR.objects
    triggers = TriggerEnvioEmailsICDR.objects
    listasUsers = ListaUsersICDR.objects
    return render(
        request,
        "configurationsICDR.html",
        {
            "users": users,
            "utilizadoresICDR": utilizadoresICDR,
            "triggers": triggers,
            "listasUsers": listasUsers,
        },
    )


def addRequest(request):
    if request.method == "POST":
        partNumber = request.POST["partNumber"]
        horaPedido = request.POST["horaPedidoUsar"]
        requisitante = request.POST["requisitante"]
        justification = request.POST["justification"]
        comentario = request.POST["comentario"]
        receiver = request.POST["receiver"]

        if request.POST["lineSelect"]:
            line = request.POST["lineSelect"]
        if request.POST["lineInput"]:
            line = request.POST["lineInput"]

        if "_pending" in request.POST:
            novoRequest = LineRequestProcessing(
                None,
                horaPedido,
                partNumber,
                line,
                requisitante,
                receiver,
                justification,
                comentario,
            )
            novoRequest.save()
            return redirect("receiving:lineRequest")

        if "_finished" in request.POST:
            novoRequest = LineRequestFinished(
                None,
                horaPedido,
                partNumber,
                line,
                requisitante,
                receiver,
                justification,
                comentario,
            )
            novoRequest.save()
            return redirect("receiving:lineRequest")


def updatePending(request):
    if request.method == "POST":
        newPartNumber = request.POST["newPartNumberPending"]
        newLine = request.POST["newLinePending"]
        newRequisitante = request.POST["newRequirePending"]
        newReceiver = request.POST["newReceiverPending"]
        newJustification = request.POST["newJustificationPending"]
        newComentario = request.POST["newCommentPending"]

        oldPartNumber = request.POST["partNumberPendingVal"]
        partNumber = request.POST["partNumberPendingVal"]
        horaPedido = request.POST["horaPedidoPendingVal"]
        line = request.POST["linePendingVal"]
        requisitante = request.POST["requirePendingVal"]
        receiver = request.POST["receiverPendingVal"]
        justification = request.POST["justificationPendingVal"]
        comentario = request.POST["commentPendingVal"]

        if newPartNumber != "":
            partNumber = newPartNumber
        if newLine != "":
            line = newLine
        if newRequisitante != "":
            requisitante = newRequisitante
        if newReceiver != "":
            receiver = newReceiver
        if newJustification != "":
            justification = newJustification
        if newComentario != "":
            comentario = newComentario

        if "_deletePending" in request.POST:
            oldLine = LineRequestProcessing.objects.get(
                horaPedido=horaPedido, partNumber=oldPartNumber
            )
            oldLine.delete()
            return redirect("receiving:lineRequest")

        if "_pending" in request.POST:
            oldLine = LineRequestProcessing.objects.get(
                horaPedido=horaPedido, partNumber=oldPartNumber
            )
            oldLine.delete()
            novoRequest = LineRequestProcessing(
                None,
                horaPedido,
                partNumber,
                line,
                requisitante,
                receiver,
                justification,
                comentario,
            )
            novoRequest.save()
            return redirect("receiving:lineRequest")

        if justification == "":
            lineRequestProcessing = LineRequestProcessing.objects
            lineRequestFinished = LineRequestFinished.objects
            linhas = Line.objects
            justificacoes = Justification.objects
            return render(
                request,
                "lineRequest.html",
                {
                    "lineRequestFinished": lineRequestFinished,
                    "lineRequestProcessing": lineRequestProcessing,
                    "linhas": linhas,
                    "justificacoes": justificacoes,
                    "alert": "Not succeed! To proceed to finish need Justification",
                },
            )

        if "_finished" in request.POST:
            oldLine = LineRequestProcessing.objects.get(
                horaPedido=horaPedido,
                partNumber=partNumber,
                linha=line,
                receiver=receiver,
            )
            oldLine.delete()
            novoRequest = LineRequestFinished(
                None,
                horaPedido,
                partNumber,
                line,
                requisitante,
                receiver,
                justification,
                comentario,
            )
        novoRequest.save()
        return redirect("receiving:lineRequest")

    return redirect("receiving:lineRequest")


def createLine(request):
    if request.method == "POST":
        try:
            linha = request.POST["newLine"]
            linha = Line.objects.get(linha=linha)
        except ObjectDoesNotExist:
            novaLinha = Line()
            novaLinha.linha = request.POST["newLine"]
            novaLinha.save()

            message = "Adicionada linha " + novaLinha.linha
            subject, from_email, to = (
                "Alteração em Receiving - Line Request - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()

            return redirect("receiving:configurations")
    linhas = Line.objects
    justificacoes = Justification.objects
    return render(
        request,
        "configurations.html",
        {
            "linhas": linhas,
            "justificacoes": justificacoes,
            "erro": "Line already exists",
        },
    )


def createJustification(request):
    if request.method == "POST":
        try:
            justificacao = request.POST["newJustification"]
            justificacao = Justification.objects.get(justificacao=justificacao)
        except ObjectDoesNotExist:
            novaJustificacao = Justification()
            novaJustificacao.justificacao = request.POST["newJustification"]
            novaJustificacao.save()

            message = "Adicionada justificação - " + novaJustificacao.justificacao
            subject, from_email, to = (
                "Alteração em Receiving - Line Request - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect("receiving:configurations")
    linhas = Line.objects
    justificacoes = Justification.objects
    return render(
        request,
        "configurations.html",
        {
            "linhas": linhas,
            "justificacoes": justificacoes,
            "erro": "Justification already exists",
        },
    )


def deleteLine(request):
    if request.method == "POST":
        linha = request.POST["nome2"]

        nome = Line.objects.get(linha=linha)

        message = "Eliminada linha " + nome.linha
        subject, from_email, to = (
            "Alteração em Receiving - Line Request - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("receiving:configurations")


def deleteJustification(request):
    if request.method == "POST":
        justificacao = request.POST["nome3"]

        nome = Justification.objects.get(justificacao=justificacao)

        message = "Eliminada justificação - " + nome.justificacao
        subject, from_email, to = (
            "Alteração em Receiving - Line Request - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("receiving:configurations")


def uploadTaskBrowse(request):
    if request.method == "POST":

        # TaskBrowse.objects.all().delete()
        # workbook = openpyxl.load_workbook("//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Receiving/Task Browse.xlsx")
        # worksheet = workbook['Data']
        #
        # for element in worksheet:
        #     if element[23].value != "KARDEX" and element[0].value != "Task ID":
        #         novoElemento = TaskBrowse(
        #             None,
        #             element[0].value,
        #             element[3].value,
        #             element[5].value,
        #             element[13].value,
        #             element[18].value,
        #             element[23].value,
        #             element[39].value,
        #             element[40].value
        #         )
        #         novoElemento.save()
        #
        # worksheet = workbook['Info']
        diaCorrente = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        oldLastUpdate = LastUpdate.objects.get()
        oldLastUpdate1 = LastUpdate.objects.all()#.raw("SELECT * FROM receiving_lastupdate")
        novaData = LastUpdate(
            None,
            diaCorrente,
            oldLastUpdate.redHour,
            oldLastUpdate.yellowHour,
            oldLastUpdate.lastUpdateRed,
        )
        oldLastUpdate.delete()
        novaData.save()


        if "_updateFA" in request.POST:
            return redirect("receiving:faSupply")
        elif "_updateSMD" in request.POST:
            return redirect("receiving:bpSMDSupply")
        elif "_updateDropin" in request.POST:
            return redirect("receiving:bpDropinSupply")
        elif "_updatePending" in request.POST:
            return redirect("receiving:pending")
        else:
            return redirect("receiving:mnfgSupply")


# def getTaskBrowse(request):
#     if request.method == 'GET':
#         TaskBrowse.objects.all().delete()
#         LastUpdate.objects.all().delete()
#         workbook = openpyxl.load_workbook("//PAVPD002/E_Proj/sharedir/MP&L/Warehouse/PWMS/Receiving/Task Browse.xlsx")
#         worksheet = workbook['Data']
#
#         for element in worksheet:
#             if element[23].value != "KARDEX" and element[0].value != "Task ID":
#                 novoElemento = TaskBrowse(
#                     None,
#                     element[0].value,
#                     element[3].value,
#                     element[5].value,
#                     element[13].value,
#                     element[18].value,
#                     element[23].value,
#                     element[39].value,
#                     element[40].value
#                 )
#                 novoElemento.save()
#
#         worksheet = workbook['Info']
#         novaData = LastUpdate(
#             None,
#             worksheet["B3"].value
#         )
#         novaData.save()
#
#         return redirect('receiving:faSupply')


def createAreaA(request):
    if request.method == "POST":
        try:
            newAreaA = request.POST["newAreaA"]
            newAreaA = AreaA.objects.get(storageZone=newAreaA)
        except ObjectDoesNotExist:
            novaAreaA = AreaA()
            novaAreaA.storageZone = request.POST["newAreaA"]
            novaAreaA.save()

            message = (
                "<b>BP Dropin Supply </b></br>Adicionado " + request.POST["newAreaA"]
            )
            subject, from_email, to = (
                "Alteração em Receiving - MNFG Supply - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect("receiving:configurationMNFG")
    return redirect("receiving:configurationMNFG")


def createAreaB(request):
    if request.method == "POST":
        try:
            newAreaB = request.POST["newAreaB"]
            newAreaB = AreaB.objects.get(storageZone=newAreaB)
        except ObjectDoesNotExist:
            novaAreaB = AreaB()
            novaAreaB.storageZone = request.POST["newAreaB"]
            novaAreaB.save()

            message = "<b>FA Supply </b></br>Adicionado " + request.POST["newAreaB"]
            subject, from_email, to = (
                "Alteração em Receiving - MNFG Supply - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()

            return redirect("receiving:configurationMNFG")
    return redirect("receiving:configurationMNFG")


def createKardex(request):
    if request.method == "POST":
        try:
            newKardex = request.POST["newKardex"]
            newKardex = Kardex.objects.get(storageZone=newKardex)
        except ObjectDoesNotExist:
            novaKardex = Kardex()
            novaKardex.storageZone = request.POST["newKardex"]
            novaKardex.save()

            message = (
                "<b>BP SMD Supply </b></br>Adicionado " + request.POST["newKardex"]
            )
            subject, from_email, to = (
                "Alteração em Receiving - MNFG Supply - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()

            return redirect("receiving:configurationMNFG")
    return redirect("receiving:configurationMNFG")

def alteraEstadoAlertaSupply(request):
    bd_item = Receiving_SupplyPages_Refresh.objects.all()
    for i in bd_item:
        i.estado= "false"
        i.save()

  
    return redirect("receiving:configurationMNFG")

def getEstadoAlertaSupply(request):
    bd_item = Receiving_SupplyPages_Refresh.objects.all().values('estado').last().get('estado')
    return JsonResponse({"message": "OK","estado" : bd_item})

  


def deleteAreaA(request):
    if request.method == "POST":
        storageZone = request.POST["nomeAreaA"]

        nome = AreaA.objects.get(storageZone=storageZone)

        message = "<b>BP Dropin Supply </b></br>Removido " + nome.storageZone
        subject, from_email, to = (
            "Alteração em Receiving - MNFG Supply - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("receiving:configurationMNFG")


def deleteAreaB(request):
    if request.method == "POST":
        storageZone = request.POST["nomeAreaB"]

        nome = AreaB.objects.get(storageZone=storageZone)

        message = "<b>FA Supply </b></br>Removido " + nome.storageZone
        subject, from_email, to = (
            "Alteração em Receiving - MNFG Supply - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("receiving:configurationMNFG")


def deleteKardex(request):
    if request.method == "POST":
        storageZone = request.POST["nomeKardex"]

        nome = Kardex.objects.get(storageZone=storageZone)

        message = "<b>BP SMD Supply </b></br>Removido " + nome.storageZone
        subject, from_email, to = (
            "Alteração em Receiving - MNFG Supply - Configurations",
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        nome.delete()
        return redirect("receiving:configurationMNFG")


def updateTimers(request):
    if request.method == "POST":
        for x in LastUpdate.objects.all():
            print("var-->",x)
        oldLastUpdate = LastUpdate.objects.first()
        lastUpdate = LastUpdate()

        if request.POST["redLine"] != "":
            lastUpdate.redHour = request.POST["redLine"]
        else:
            lastUpdate.redHour = oldLastUpdate.redHour
        if request.POST["yellowLine"] != "":
            lastUpdate.yellowHour = request.POST["yellowLine"]
        else:
            lastUpdate.yellowHour = oldLastUpdate.yellowHour
        if request.POST["redTime"] != "":
            lastUpdate.lastUpdateRed = request.POST["redTime"]
        else:
            lastUpdate.lastUpdateRed = oldLastUpdate.lastUpdateRed
        print("Teste->")
        lastUpdate.data = oldLastUpdate.data
        print("lastUpdate->",lastUpdate.data)
        lastUpdate.save()

        message = (
            "<b>Alterados Timers </b></br></br><b>Now: </b></br>Red line: "
            + lastUpdate.redHour
            + "</br>Yellow line: "
            + lastUpdate.yellowHour
            + "</br>Red last update Time: "
            + lastUpdate.lastUpdateRed
            + "</br></br><b>Before: </b></br>Red Line: "
            + str(oldLastUpdate.redHour)
            + "</br>Yellow line: "
            + str(oldLastUpdate.yellowHour)
            + "</br>Red last update Time: "
            + str(oldLastUpdate.lastUpdateRed)
        )
        subject, from_email, to = (
            "Alteração em Receiving - MNFG Supply - Configurations",
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()

        oldLastUpdate.delete()

        return redirect("receiving:configurationMNFG")


def addNewItemReceivingPos(request):
    if request.method == "POST":
        itemName = request.POST["itemAdd"]
        posicao = request.POST["posicaoModal"]
        tipoSubItem = request.POST["subItemTipo"]

        if tipoSubItem == "Novo":
            novoTipo = ReceivingSubItemsTipo()
            novoTipo.tipo = itemName
            novoTipo.save()
            return redirect("receiving:configurationsTPM")

        if posicao == "1":
            novoItem = ReceivingPosicao1Items()
            if not ReceivingPosicao1Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "2":
            novoItem = ReceivingPosicao2Items()
            if not ReceivingPosicao2Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "3":
            novoItem = ReceivingPosicao3Items()
            if not ReceivingPosicao3Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "4":
            novoItem = ReceivingPosicao4Items()
            if not ReceivingPosicao4Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "5":
            novoItem = ReceivingPosicao5Items()
            if not ReceivingPosicao5Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "6":
            novoItem = ReceivingPosicao6Items()
            if not ReceivingPosicao6Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "7":
            novoItem = ReceivingPosicao7Items()
            if not ReceivingPosicao7Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        elif posicao == "8":
            novoItem = ReceivingPosicao8Items()
            if not ReceivingPosicao8Items.objects.filter(item=itemName).exists():
                novoItem.item = itemName
                novoItem.tipo = tipoSubItem
                novoItem.posicao = posicao
                novoItem.save()
        return redirect("receiving:configurationsTPM")


def getBodyInfo(request):
    if request.method == "GET":
        tipo = request.GET["tipo"]

        subItems = ReceivingSubItems.objects

        body = []

        for subItem in subItems.all():
            if subItem.tipo == tipo:
                if subItem.usar == "Stop":
                    elemento = {
                        "linha": '<tr style="background-color: rgb(255, 0, 0, .2)">'
                        "<td>"
                        + subItem.item
                        + '</td><td><input type="checkbox" id="turnoOk" required>&nbsp;&nbsp;&nbsp'
                        ';&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<input type="checkbox" id="turnoNok" required></td><td>'
                        '<input style="width: 370px; margin-left: 8px" class="" type="text" '
                        'name="inicio{{ pos.id }}" id="inicio{{ pos.id }}"></td><td '
                        'style="display:none" id="posicaoGenerico"></td><td style="display:none"'
                        ' id="turnoGenerico"></td><td style="display:none" id="idDiaGenerico"></td></tr>'
                    }
                    body.append(elemento)
                else:
                    elemento = {
                        "linha": "<tr><td>"
                        + subItem.item
                        + '</td><td><input type="checkbox" id="turnoOk" required>&nbsp;&nbsp;&nbsp'
                        ';&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<input type="checkbox" id="turnoNok" required></td><td>'
                        '<input style="width: 370px; margin-left: 8px" class="" type="text" '
                        'name="inicio{{ pos.id }}" id="inicio{{ pos.id }}"></td><td '
                        'style="display:none" id="posicaoGenerico"></td><td style="display:none"'
                        ' id="turnoGenerico"></td><td style="display:none" id="idDiaGenerico"></td></tr>'
                    }
                    body.append(elemento)
        return JsonResponse({"body": body})


# onde é feito o delete de cada linha individual
# def deleteItemReceivingPos1(request):
#     if request.method == "POST":
#         itemID = request.POST['idDelete']
#
#         ReceivingPosicao1Items.objects.get(id=itemID).delete()
#
#         return redirect('receiving:tpm')

# onde é feito o update de cada linha individual
# def editItemReceivingPos1(request):
#     if request.method == "POST":
#         itemID = request.POST['idEdit']
#         itemName = request.POST['itemEdit']
#         itemComentarioFinal = request.POST['comentarioFinalEdit']
#         itemComentarioInicio = request.POST['comentarioInicioEdit']
#
#         item = ReceivingPosicao1Items.objects.get(id=itemID)
#
#         if (itemName):
#             item.item = itemName
#         if (itemComentarioFinal):
#             item.comentarioFinalTurno = itemComentarioFinal
#         if (itemComentarioInicio):
#             item.comentarioInicioTurno = itemComentarioInicio
#
#         item.save()
#
#         return redirect('receiving:tpm')
def submitSubItems(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        idLinhaBotao = request.POST["idLinhaBotao"]
        for elem in datatable:
            posicao = elem.get("posicao", None)
            idDiaAtual = elem.get("idDia", None)
            idDia = idLinhaBotao + " " + idDiaAtual
            item = elem.get("item", None)
            if posicao == "1":
                if ReceivingPosicao1SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao1SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao1SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "2":
                if ReceivingPosicao2SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao2SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao2SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "3":
                if ReceivingPosicao3SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao3SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao3SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "4":
                if ReceivingPosicao4SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao4SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao4SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "5":
                if ReceivingPosicao5SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao5SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao5SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "6":
                if ReceivingPosicao6SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao6SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao6SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "7":
                if ReceivingPosicao7SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao7SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao7SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
            elif posicao == "8":
                if ReceivingPosicao8SubItems.objects.filter(
                    idDia=idDia, item=item
                ).exists():
                    elemento = ReceivingPosicao8SubItems.objects.get(
                        idDia=idDia, item=item
                    )
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
                else:
                    elemento = ReceivingPosicao8SubItems()
                    elemento.idDia = idDia
                    elemento.item = elem.get("item", None)
                    elemento.comentario = elem.get("comentario", None)
                    elemento.ok = elem.get("ok", None)
                    elemento.nok = elem.get("nok", None)
                    elemento.save()
    return redirect("receiving:tpm")


# def submitSubItems1(request):
#     if request.method == "POST":
#         datatable = json.loads(request.POST['datatable'])
#         for elem in datatable:
#             posicao = elem.get('posicao', None)
#             item = elem.get('item', None)
#             if ReceivingPosicao1SubItems.objects.filter(idDia=idDia, item=item).exists():
#                 elemento = ReceivingPosicao1SubItems.objects.get(idDia=idDia, item=item)
#                 elemento.idDia = elem.get('idDia', None)
#                 elemento.item = elem.get('item', None)
#                 elemento.comentario = elem.get('comentario', None)
#                 elemento.ok = elem.get('ok', None)
#                 elemento.nok = elem.get('nok', None)
#                 elemento.save()
#             else:
#                 elemento = ReceivingPosicao1SubItems()
#                 elemento.idDia = elem.get('idDia', None)
#                 elemento.item = elem.get('item', None)
#                 elemento.comentario = elem.get('comentario', None)
#                 elemento.ok = elem.get('ok', None)
#                 elemento.nok = elem.get('nok', None)
#                 elemento.save()
#     return redirect('receiving:tpm')


def submitDataFinalTurno(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao1Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao1Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao1Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao1Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao1Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao1Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno2(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao2Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao2Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao2Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno2(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao2Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao2Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao2Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno3(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao3Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao3Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao3Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno3(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao3Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao3Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao3Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno4(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao4Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao4Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao4Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno4(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao4Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao4Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao4Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno5(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao5Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao5Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao5Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno5(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao5Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao5Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao5Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno6(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao6Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao6Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao6Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno6(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao6Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao6Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao6Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno7(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao7Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao7Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao7Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno7(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao7Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao7Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao7Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataFinalTurno8(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao8Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao8Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioFinalTurno = elem.get("comentario", None)
                    fimDoTurno = elem.get("fimDoTurno", None)
                    if fimDoTurno == False:
                        fimDoTurno = None
                    elemento.fimTurnoOk = fimDoTurno
                    fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                    if fimDoTurnoNok == False:
                        fimDoTurnoNok = None
                    elemento.fimTurnoNok = fimDoTurnoNok
                    elemento.dataFinalTurno = elem.get("data", None)
                    elemento.horaFinalTurno = elem.get("hora", None)
                    elemento.responsavelFinalTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao8Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioFinalTurno = elem.get("comentario", None)
                fimDoTurno = elem.get("fimDoTurno", None)
                if fimDoTurno == False:
                    fimDoTurno = None
                elemento.fimTurnoOk = fimDoTurno
                fimDoTurnoNok = elem.get("fimDoTurnoNok", None)
                if fimDoTurnoNok == False:
                    fimDoTurnoNok = None
                elemento.fimTurnoNok = fimDoTurnoNok
                elemento.dataFinalTurno = elem.get("data", None)
                elemento.horaFinalTurno = elem.get("hora", None)
                elemento.responsavelFinalTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def submitDataInicioTurno8(request):
    if request.method == "POST":
        datatable = json.loads(request.POST["datatable"])
        for elem in datatable:
            dataPosicao = elem.get("posicao", None)
            item = elem.get("item", None)
            if ReceivingPosicao8Historico.objects.filter(
                dataPosicao=dataPosicao, item=item
            ).exists():
                elemento = ReceivingPosicao8Historico.objects.get(
                    dataPosicao=dataPosicao, item=item
                )
                if (elemento.inicioTurnoOk or elemento.inicioTurnoNok) and (
                    elemento.fimTurnoOk or elemento.fimTurnoNok
                ):
                    return redirect("receiving:tpm")
                else:
                    elemento.dataPosicao = elem.get("posicao", None)
                    elemento.item = elem.get("item", None)
                    elemento.comentarioInicioTurno = elem.get("comentario", None)
                    inicioDoTurno = elem.get("inicioDoTurno", None)
                    if inicioDoTurno == False:
                        inicioDoTurno = None
                    elemento.inicioTurnoOk = inicioDoTurno
                    inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                    if inicioDoTurnoNok == False:
                        inicioDoTurnoNok = None
                    elemento.inicioTurnoNok = inicioDoTurnoNok
                    elemento.dataInicioTurno = elem.get("data", None)
                    elemento.horaInicioTurno = elem.get("hora", None)
                    elemento.responsavelInicioTurno = elem.get("responsavel", None)
                    elemento.tipo = elem.get("tipo", None)
                    elemento.save()
            else:
                elemento = ReceivingPosicao8Historico()
                elemento.dataPosicao = elem.get("posicao", None)
                elemento.item = elem.get("item", None)
                elemento.comentarioInicioTurno = elem.get("comentario", None)
                inicioDoTurno = elem.get("inicioDoTurno", None)
                if inicioDoTurno == False:
                    inicioDoTurno = None
                elemento.inicioTurnoOk = inicioDoTurno
                inicioDoTurnoNok = elem.get("inicioDoTurnoNok", None)
                if inicioDoTurnoNok == False:
                    inicioDoTurnoNok = None
                elemento.inicioTurnoNok = inicioDoTurnoNok
                elemento.dataInicioTurno = elem.get("data", None)
                elemento.horaInicioTurno = elem.get("hora", None)
                elemento.responsavelInicioTurno = elem.get("responsavel", None)
                elemento.tipo = elem.get("tipo", None)
                elemento.save()
    return redirect("receiving:tpm")


def getPosicao1Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao1Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao2Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao2Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao3Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao3Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao4Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao4Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao5Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao5Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao6Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao6Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao7Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao7Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def getPosicao8Data(request):
    if request.method == "GET":
        dataPosicao = request.GET["posicaoData"]
        elementosFiltrados = ReceivingPosicao8Historico.objects.filter(
            dataPosicao=dataPosicao
        )
        listElementosFiltrados = serialize("json", elementosFiltrados)
        return JsonResponse(listElementosFiltrados, safe=False)


def deleteSelectedItems(request):
    if request.method == "POST":
        items = json.loads(request.POST["items[]"])
        posicao = request.POST["posicao"]

        if posicao == "1":
            for elem in items:
                ReceivingPosicao1Items.objects.get(item=elem).delete()
        if posicao == "2":
            for elem in items:
                ReceivingPosicao2Items.objects.get(item=elem).delete()
        if posicao == "3":
            for elem in items:
                ReceivingPosicao3Items.objects.get(item=elem).delete()
        if posicao == "4":
            for elem in items:
                ReceivingPosicao4Items.objects.get(item=elem).delete()
        if posicao == "5":
            for elem in items:
                ReceivingPosicao5Items.objects.get(item=elem).delete()
        if posicao == "6":
            for elem in items:
                ReceivingPosicao6Items.objects.get(item=elem).delete()
        if posicao == "7":
            for elem in items:
                ReceivingPosicao7Items.objects.get(item=elem).delete()
        if posicao == "8":
            for elem in items:
                ReceivingPosicao8Items.objects.get(item=elem).delete()
    return redirect("receiving:configurationsTPM")

def editSelectedItems(request):
    if request.method == "POST":
        id_selected_item = json.loads(request.POST["items[]"])#json.loads(request.POST["items[]"])
        posicao = request.POST["posicao"]

        if posicao == "1":
            item = ReceivingPosicao1Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "2":
            item = ReceivingPosicao2Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()            
        if posicao == "3":
            item = ReceivingPosicao3Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "4":
            item = ReceivingPosicao4Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "5":
            item = ReceivingPosicao5Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "6":
            item = ReceivingPosicao6Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "7":
            item = ReceivingPosicao7Items.objects.get(item=id_selected_item)
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()
        if posicao == "8":
            item = ReceivingPosicao8Items.objects.get(item=id_selected_item)  
            item.item = request.POST["itemName"]
            item.tipo = request.POST["tipoSubItem"]
            item.posicao = request.POST["posicao"]
            item.save()  

    return redirect("receiving:configurationsTPM")


def deleteSelectedSubItems(request):
    items = json.loads(request.POST["items[]"])
    for elem in items:
        ReceivingSubItems.objects.get(item=elem).delete()
    return redirect("receiving:configurationsTPM")


# def updateComentario1(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao1Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao1Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario2(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao2Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao2Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario3(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao3Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao3Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario4(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao4Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao4Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario5(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao5Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao5Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario6(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao6Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao6Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario7(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao7Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao7Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')
#
#
# def updateComentario8(request):
#     if request.method == "POST":
#         id = request.POST['nome']
#         if id.startswith('inicio'):
#             id = id.removeprefix('inicio')
#             item = ReceivingPosicao8Items.objects.get(id=id)
#             item.comentarioInicioTurno = request.POST['comentario']
#             item.save()
#
#         elif id.startswith('fim'):
#             id = int(id.removeprefix('fim'), base=10)
#             item = ReceivingPosicao8Items.objects.get(id=id)
#             item.comentarioFinalTurno = request.POST['comentario']
#             item.save()
#         return redirect('receiving:tpm')


def getSubItems(request):
    if request.method == "GET":
        dataPosicao = request.GET["dataPosicao"]
        posicao = request.GET["posicao"]
        nomeLinhaBotao = request.GET["nomeLinhaBotao"]
        listaObjetos = []

        if "inicio" in dataPosicao:
            dataPosicao = dataPosicao[-30:]
        elif "final" in dataPosicao:
            dataPosicao = dataPosicao[-29:]

        idDia = nomeLinhaBotao + " " + dataPosicao

        if posicao == "1":
            listaObjetos = ReceivingPosicao1SubItems.objects.filter(idDia=idDia)
        elif posicao == "2":
            listaObjetos = ReceivingPosicao2SubItems.objects.filter(idDia=idDia)
        elif posicao == "3":
            listaObjetos = ReceivingPosicao3SubItems.objects.filter(idDia=idDia)
        elif posicao == "4":
            listaObjetos = ReceivingPosicao4SubItems.objects.filter(idDia=idDia)
        elif posicao == "5":
            listaObjetos = ReceivingPosicao5SubItems.objects.filter(idDia=idDia)
        elif posicao == "6":
            listaObjetos = ReceivingPosicao6SubItems.objects.filter(idDia=idDia)
        elif posicao == "7":
            listaObjetos = ReceivingPosicao7SubItems.objects.filter(idDia=idDia)
        elif posicao == "8":
            listaObjetos = ReceivingPosicao8SubItems.objects.filter(idDia=idDia)

        listElementosFiltrados = serialize("json", listaObjetos)
        return JsonResponse(listElementosFiltrados, safe=False)
    return redirect("receiving:tpm")


def addNewSubItem(request):
    nome = request.POST["itemAdd"]
    tipo = request.POST.getlist("subItemTipo")
    usar = request.POST["usar"]
    stringTipo = ""

    for elem in tipo:
        stringTipo += ", " + elem

    novoItem = ReceivingSubItems()
    if not ReceivingSubItems.objects.filter(item=nome).exists():
        novoItem.item = nome
        novoItem.tipo = stringTipo[2:]
        novoItem.usar = usar
        novoItem.save()

    return redirect("receiving:configurationsTPM")


def changeUserGroups(request):
    if request.method == "POST":
        user = User.objects.get(username=request.POST["username"])
        grupo = request.POST["paginas"]
        if User.objects.filter(
            username=request.POST["username"], groups__name="receivingMNFG"
        ):
            my_group = Group.objects.using("default").get(name="receivingMNFG")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="receivingLINES"
        ):
            my_group = Group.objects.using("default").get(name="receivingLINES")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="receivingTPM"
        ):
            my_group = Group.objects.using("default").get(name="receivingTPM")
            my_group.user_set.remove(user)
        if grupo == "none":
            message = (
                "User " + user.username + " perdeu acesso às páginas de Receiving."
            )
            subject, from_email, to = (
                "Alteração em Receiving - Line Request - Configurations",
                "noreply@visteon.com",
                ["aroque1@visteon.com"],
            )
            msg = EmailMultiAlternatives(subject, message, from_email, to)
            msg.attach_alternative(message, "text/html")
            msg.send()
            return redirect("receiving:configurations")
        if grupo == "tpm":
            my_group = Group.objects.using("default").get(name="receivingTPM")
            my_group.user_set.add(user)
        if grupo == "lineRequest":
            my_group = Group.objects.using("default").get(name="receivingLINES")
            my_group.user_set.add(user)
        if grupo == "mnfgSupply":
            my_group = Group.objects.using("default").get(name="receivingMNFG")
            my_group.user_set.add(user)
        if grupo == "lineRequest/mnfgSupply":
            my_group = Group.objects.using("default").get(name="receivingMNFG")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="receivingLINES")
            my_group.user_set.add(user)
        if grupo == "lineRequest/tpm":
            my_group = Group.objects.using("default").get(name="receivingLINES")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="receivingTPM")
            my_group.user_set.add(user)
        if grupo == "mnfgSupply/tpm":
            my_group = Group.objects.using("default").get(name="receivingMNFG")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="receivingTPM")
            my_group.user_set.add(user)
        if grupo == "lineRequest/mnfgSupply/tpm/pending":
            my_group = Group.objects.using("default").get(name="receivingMNFG")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="receivingLINES")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="receivingTPM")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="pending")
            my_group.user_set.add(user)
        message = (
            "User "
            + user.username
            + " com páginas acessiveis - "
            + grupo.replace("/", " , ")
        )
        subject, from_email, to = (
            "Alteração em Receiving - Line Request - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:configurations")


def habilitarQuadrados(request):
    if request.method == "POST":
        message = "</br><b>Alterações de visibilidade </b></br></br>"
        if request.POST["1"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="1")
            pos.valor = request.POST["1"]
            pos.save()
            message += (
                "Alterado estado visivel de Receções Físicas. </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["2"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="2")
            pos.valor = request.POST["2"]
            pos.save()
            message += (
                "Alterado estado visivel de Smart. </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["3"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="3")
            pos.valor = request.POST["3"]
            pos.save()
            message += (
                "Alterado estado visivel de Kardex.  </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["4"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="4")
            pos.valor = request.POST["4"]
            pos.save()
            message += (
                "Alterado estado visivel de Área A.  </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["5"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="5")
            pos.valor = request.POST["5"]
            pos.save()
            message += (
                "Alterado estado visivel de Área B.  </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["6"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="6")
            pos.valor = request.POST["6"]
            pos.save()
            message += (
                "Alterado estado visivel de Área B embalagens/Mecalux.  </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["7"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="7")
            pos.valor = request.POST["7"]
            pos.save()
            message += (
                "Alterado estado visivel de Empilhadores.  </br>Visível? <b>"
                + pos.valor
                + "</b></br></br>"
            )
        if request.POST["8"] != "":
            pos = HabilitarQuadrados.objects.get(posicao="8")
            pos.valor = request.POST["8"]
            pos.save()
            message += (
                "Alterado estado visivel de Doca.  </br>Visível? <b>"
                + pos.valor
                + "</b>"
            )
        subject, from_email, to = (
            "Alteração em Receiving - TPM - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
    return redirect("receiving:configurationsTPM")


def definirTempos(request):
    if request.method == "POST":
        elementos = DefinirTempos.objects
        message = "</br><b>Alterações de tempos </b></br></br>"
        elem1 = elementos.get(turno="1")
        elem2 = elementos.get(turno="2")
        elem4 = elementos.get(turno="4")
        if request.POST["inicio1"] != elem1.inicio:
            message += (
                "Alterado tempo inicio turno 1.</br>Antigo: <b>"
                + elem1.inicio
                + "h</b>. Novo: <b>"
                + request.POST["inicio1"]
                + "h</b></br></br>"
            )
            elem1.inicio = request.POST["inicio1"]
            elem1.save()
        if request.POST["final1"] != elem1.final:
            message += (
                "Alterado tempo final turno 1.</br>Antigo: <b>"
                + elem1.final
                + "h</b>. Novo: <b>"
                + request.POST["final1"]
                + "h</b></br></br>"
            )
            elem1.final = request.POST["final1"]
            elem1.save()
        if request.POST["inicio2"] != elem2.inicio:
            message += (
                "Alterado tempo inicio turno 2.</br>Antigo: <b>"
                + elem2.inicio
                + "h</b>. Novo: <b>"
                + request.POST["inicio2"]
                + "h</b></br></br>"
            )
            elem2.inicio = request.POST["inicio2"]
            elem2.save()
        if request.POST["final2"] != elem2.final:
            message += (
                "Alterado tempo final turno 2.</br>Antigo: <b>"
                + elem2.final
                + "h</b>. Novo: <b>"
                + request.POST["final2"]
                + "h</b></br></br>"
            )
            elem2.final = request.POST["final2"]
            elem2.save()
        if request.POST["inicio4"] != elem4.inicio:
            message += (
                "Alterado tempo inicio turno 4.</br>Antigo: <b>"
                + elem4.inicio
                + "h</b>. Novo: <b>"
                + request.POST["inicio4"]
                + "h</b></br></br>"
            )
            elem4.inicio = request.POST["inicio4"]
            elem4.save()
        if request.POST["final4"] != elem4.final:
            message += (
                "Alterado tempo final turno 4.</br>Antigo: <b>"
                + elem4.final
                + "h</b>. Novo: <b>"
                + request.POST["final4"]
                + "h</b>"
            )
            elem4.final = request.POST["final4"]
            elem4.save()

        subject, from_email, to = (
            "Alteração em Receiving - TPM - Configurations",
            "noreply@visteon.com",
            ["aroque1@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:configurationsTPM")



def definirTempos(request):
    if request.method == "POST":
        print(request.POST)
        elementos = DefinirTempos.objects.all()
        elementos.delete()
        DefinirTempos(turno = "1", p1Inicio = request.POST["inicioT1P1"],  p1Fim = request.POST["fimT1P1"],  p2Inicio = request.POST["inicioT1P2"] , p2Fim = request.POST["fimT1P2"]).save()
        DefinirTempos(turno = "2", p1Inicio = request.POST["inicioT2P1"],  p1Fim = request.POST["fimT2P1"],  p2Inicio = request.POST["inicioT2P2"] , p2Fim = request.POST["fimT2P2"]).save()
        DefinirTempos(turno = "4", p1Inicio = request.POST["inicioT4P1"],  p1Fim = request.POST["fimT4P1"],  p2Inicio = request.POST["inicioT4P2"] , p2Fim = request.POST["fimT4P2"]).save()
        return redirect("receiving:configurationsTPM")


def reAbrirICDR(request):
    if request.method == "POST":
        actualDay = datetime.today()
        actualDayString = actualDay.strftime("%Y-%m-%d")

        id = request.POST["id"]

        diaCorrente = datetime.now().strftime("%d/%m/%Y")
        horaCorrente = datetime.now().strftime("%H:%M:%S")

        icdr = ICDR.objects.get(id=id)
        elemDayString = icdr.aberturaICDR[0:10]
        contador = np.busday_count(elemDayString, actualDayString)
        icdr.ageing = contador
        icdr.date = ""
        comentarios = icdr.comentarioFechoICDR
        comentarios += (
            "\n\rICDR reaberto por - "
            + request.user.username
            + " - em "
            + diaCorrente
            + " às "
            + horaCorrente
        )
        icdr.comentarioFechoICDR = comentarios
        icdr.save()

        rctUnpCheck = ""
        cycleCountCheck = ""

        trigger = TriggerEnvioEmailsICDR.objects.get(nome="Reabertura")
        listTo = []

        for user in trigger.users.split(";"):
            if not user in listTo:
                if ListaUsersICDR.objects.filter(nome=user).exists():
                    lista = ListaUsersICDR.objects.get(nome=user)
                    for userListaBD in lista.user.split(";"):
                        if not userListaBD in listTo and not userListaBD == "":
                            listTo.append(userListaBD + "@visteon.com")
                else:
                    if not user == "":
                        listTo.append(user + "@visteon.com")

        if trigger.estado == "ativo":
            if icdr.rctUnpCheck == "false" or icdr.rctUnpCheck == "":
                rctUnpCheck = ""
            elif icdr.rctUnpCheck == "true":
                rctUnpCheck = "X"

            if icdr.cycleCountCheck == "false" or icdr.cycleCountCheck == "":
                cycleCountCheck = ""
            elif icdr.cycleCountCheck == "true":
                cycleCountCheck = "X"

            table = "<h3>User: <b>" + request.user.username + " </b></h3> "
            table += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
            table += (
                '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                + icdr.aberturaICDR
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(icdr.ageing)
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.nAno
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.fornecedor
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.partnumber
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.quantidade
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.tipo
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.simNao
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.responsavel
                + "/"
                + icdr.departamento
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.comentarioFecho
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.unCost
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.totalCost
                + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                + rctUnpCheck
                + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                + cycleCountCheck
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.date
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.cycleCount
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.dataCycleCount
                + '</td><td style="padding:0 15px 0 15px;">'
                + icdr.auditCheck
                + "</td></tr>"
            )
            table += "</body></table>"

            subject, from_email, to = (
                "Reabertura de ICDR",
                "noreply@visteon.com",
                listTo,
            )
            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()

        return redirect("receiving:icdr")


def criarICDR(request):
    if request.method == "POST":
        numeroAno = request.POST["numeroAno"]
        fornecedor = request.POST["fornecedor"]
        partNumber = request.POST["partnumber"]
        quantidade = request.POST["quantidade"]
        tipo = request.POST["tipo"]
        alfandegaSimNao = request.POST["alfandegaSimNao"]
        responsavel = request.POST["responsavel"]
        departamento = request.POST["departamento"]
        comentario = request.POST["comentarios"]

        numeroICDR = numeroAno.split("/")
        numeroICDR = int(numeroICDR[0][1:], base=10) + 1

        ultimoICDR = ICDRultimoValor.objects.filter().first()
        if ultimoICDR.valor == numeroICDR - 1:
            ultimoICDR.valor = numeroICDR
            ultimoICDR.nome = (
                "D" + str(numeroICDR) + "/" + str(datetime.now().strftime("%Y"))
            )
            ultimoICDR.save()

        novoICDR = ICDR()
        novoICDR.aberturaICDR = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        novoICDR.ageing = "0"
        novoICDR.nAno = numeroAno
        novoICDR.fornecedor = fornecedor
        novoICDR.partnumber = partNumber
        novoICDR.quantidade = quantidade
        novoICDR.tipo = tipo
        novoICDR.simNao = alfandegaSimNao
        novoICDR.responsavel = responsavel
        novoICDR.departamento = departamento
        novoICDR.comentarioFecho = comentario
        novoICDR.unCost = ""
        novoICDR.totalCost = ""
        novoICDR.rctUnpCheck = ""
        novoICDR.cycleCountCheck = ""
        novoICDR.consumption = ""
        novoICDR.po = ""
        novoICDR.comentarioFechoICDR = ""
        novoICDR.date = ""
        novoICDR.cycleCount = ""
        novoICDR.dataCycleCount = ""
        novoICDR.auditCheck = ""
        novoICDR.save()

        rctUnpCheck = ""
        cycleCountCheck = ""

        trigger = TriggerEnvioEmailsICDR.objects.get(nome="Criação")
        listTo = []

        for user in trigger.users.split(";"):
            if not user in listTo:
                if ListaUsersICDR.objects.filter(nome=user).exists():
                    lista = ListaUsersICDR.objects.get(nome=user)
                    for userListaBD in lista.user.split(";"):
                        if not userListaBD in listTo and not userListaBD == "":
                            listTo.append(userListaBD + "@visteon.com")
                else:
                    if not user == "":
                        listTo.append(user + "@visteon.com")

        if trigger.estado == "ativo":
            if novoICDR.rctUnpCheck == "false" or novoICDR.rctUnpCheck == "":
                rctUnpCheck = ""
            elif novoICDR.rctUnpCheck == "true":
                rctUnpCheck = "X"

            if novoICDR.cycleCountCheck == "false" or novoICDR.cycleCountCheck == "":
                cycleCountCheck = ""
            elif novoICDR.cycleCountCheck == "true":
                cycleCountCheck = "X"

            table = "<h3>User: <b>" + request.user.username + " </b></h3> "
            table += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
            table += (
                '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                + novoICDR.aberturaICDR
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(novoICDR.ageing)
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.nAno
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.fornecedor
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.partnumber
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.quantidade
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.tipo
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.simNao
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.responsavel
                + "/"
                + novoICDR.departamento
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.comentarioFecho
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.unCost
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.totalCost
                + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                + rctUnpCheck
                + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                + cycleCountCheck
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.date
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.cycleCount
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.dataCycleCount
                + '</td><td style="padding:0 15px 0 15px;">'
                + novoICDR.auditCheck
                + "</td></tr>"
            )
            table += "</body></table>"

            subject, from_email, to = "Novo ICDR", "noreply@visteon.com", listTo
            msg = EmailMultiAlternatives(subject, table, from_email, to)
            msg.attach_alternative(table, "text/html")
            msg.send()

    return redirect("receiving:icdr")


def getDepartamento(request):
    if request.method == "GET":
        responsavel = request.GET["responsavel"]
        elemento = list(UserICDR.objects.filter(username=responsavel).values())
        return JsonResponse({"elemento": elemento})


def getUtilizadoresPorDepartamento(request):
    if request.method == "GET":
        departamento = request.GET["departamento"]
        elementos = list(UserICDR.objects.filter(area=departamento).values())
        return JsonResponse({"elementos": elementos})


def searchAllInfoUtilizadores(request):
    if request.method == "GET":
        resposta = []
        todos = UserICDR.objects.all()
        for elem in todos:
            row = {
                "username": elem.username,
                "nome": elem.nome,
                "area": elem.area,
                "email": elem.email,
            }
            resposta.append(row)
        return JsonResponse({"resposta": resposta})


def changeUserGroupsICDR(request):
    if request.method == "POST":
        user = User.objects.get(username=request.POST["username"])
        grupo = request.POST["paginas"]
        if User.objects.filter(
            username=request.POST["username"], groups__name="ICDR - read only"
        ):
            my_group = Group.objects.using("default").get(name="ICDR - read only")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="ICDR - tipo1"
        ):
            my_group = Group.objects.using("default").get(name="ICDR - tipo1")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="ICDR - tipo2"
        ):
            my_group = Group.objects.using("default").get(name="ICDR - tipo2")
            my_group.user_set.remove(user)
        if User.objects.filter(
            username=request.POST["username"], groups__name="ICDR - tipo3"
        ):
            my_group = Group.objects.using("default").get(name="ICDR - tipo3")
            my_group.user_set.remove(user)
        if grupo == "none":
            # message = 'User ' + user.username + ' perdeu acesso às páginas de Shipping.'
            # subject, from_email, to = 'Alteração em Receiving - Line Request - Configurations', 'noreply@visteon.com', [
            #     'aroque1@visteon.com']
            # msg = EmailMultiAlternatives(subject, message, from_email, to)
            # msg.attach_alternative(message, "text/html")
            # msg.send()
            return redirect("shippers:configurations")
        if grupo == "readOnly":
            my_group = Group.objects.using("default").get(name="ICDR - read only")
            my_group.user_set.add(user)
        if grupo == "tipo1":
            my_group = Group.objects.using("default").get(name="ICDR - tipo1")
            my_group.user_set.add(user)
        if grupo == "tipo2":
            my_group = Group.objects.using("default").get(name="ICDR - tipo2")
            my_group.user_set.add(user)
        if grupo == "tipo3":
            my_group = Group.objects.using("default").get(name="ICDR - tipo3")
            my_group.user_set.add(user)

        if grupo == "tipo1/tipo2":
            my_group = Group.objects.using("default").get(name="ICDR - tipo1")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="ICDR - tipo2")
            my_group.user_set.add(user)
        if grupo == "tipo1/tipo3":
            my_group = Group.objects.using("default").get(name="ICDR - tipo1")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="ICDR - tipo3")
            my_group.user_set.add(user)
        if grupo == "tipo2/tipo3":
            my_group = Group.objects.using("default").get(name="ICDR - tipo2")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="ICDR - tipo3")
            my_group.user_set.add(user)
        if grupo == "tipo1/tipo2/tipo3":
            my_group = Group.objects.using("default").get(name="ICDR - tipo1")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="ICDR - tipo2")
            my_group.user_set.add(user)
            my_group = Group.objects.using("default").get(name="ICDR - tipo3")
            my_group.user_set.add(user)
        return redirect("receiving:configurationsICDR")
        # message = 'User ' + user.username + ' com páginas acessiveis - ' + grupo.replace("/", " , ")
        # subject, from_email, to = 'Alteração em Shipping - Configurations', 'noreply@visteon.com', [
        #     'aroque1@visteon.com']
        # msg = EmailMultiAlternatives(subject, message, from_email, to)
        # msg.attach_alternative(message, "text/html")
        # msg.send()


def setDataHoraFechoICDR(request):
    if request.method == "POST":
        id = request.POST["id"]
        elem = ICDR.objects.get(id=id)
        elem.date = datetime.today().strftime("%Y-%m-%d %H:%M")
        elem.ageing = "-"
        elem.save()

        rctUnpCheck = ""
        cycleCountCheck = ""

        trigger = TriggerEnvioEmailsICDR.objects.get(nome="Fecho")
        listTo = []

        for user in trigger.users.split(";"):
            if not user in listTo:
                if ListaUsersICDR.objects.filter(nome=user).exists():
                    lista = ListaUsersICDR.objects.get(nome=user)
                    for userListaBD in lista.user.split(";"):
                        if not userListaBD in listTo and not userListaBD == "":
                            listTo.append(userListaBD + "@visteon.com")
                else:
                    if not user == "":
                        listTo.append(user + "@visteon.com")

        if trigger.estado == "ativo":

            if elem.rctUnpCheck == "false" or elem.rctUnpCheck == "":
                rctUnpCheck = ""
            elif elem.rctUnpCheck == "true":
                rctUnpCheck = "X"

            if elem.cycleCountCheck == "false" or elem.cycleCountCheck == "":
                cycleCountCheck = ""
            elif elem.cycleCountCheck == "true":
                cycleCountCheck = "X"

            trigger = TriggerEnvioEmailsICDR.objects.get(nome="Fecho")
            if trigger.estado == "ativo":
                table = "<h3>User: <b>" + request.user.username + " </b></h3> "
                table += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
                table += (
                    '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                    + elem.aberturaICDR
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem.ageing)
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.nAno
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.fornecedor
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.partnumber
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.quantidade
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.tipo
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.simNao
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.responsavel
                    + "/"
                    + elem.departamento
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.comentarioFecho
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.unCost
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.totalCost
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + rctUnpCheck
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + cycleCountCheck
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.date
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.cycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.dataCycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.auditCheck
                    + "</td></tr>"
                )
                table += "</body></table>"

                subject, from_email, to = "Fecho de ICDR", "noreply@visteon.com", listTo
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                msg.send()
        return redirect("receiving:icdr")


def updateICDR(request):
    if request.method == "POST":
        id = request.POST["id"]
        poNumber = ""
        cycleCount = ""
        dataCycleCount = ""
        auditCheck = ""
        checkedCycleCount = request.POST["checkedCycleCount"]
        checkedRctUnp = request.POST["checkedRctUnp"]

        if request.POST["poNumber"] != "":
            poNumber = request.POST["poNumber"]
        if request.POST["cycleCount"] != "":
            cycleCount = request.POST["cycleCount"]
        if request.POST["dataCycleCount"] != "":
            dataCycleCount = request.POST["dataCycleCount"]
        if request.POST["auditCheck"] != "":
            auditCheck = request.POST["auditCheck"]

        elem = ICDR.objects.get(id=id)
        if poNumber != "":
            elem.po = poNumber
        if cycleCount != "":
            elem.cycleCount = cycleCount
        if dataCycleCount != "":
            elem.dataCycleCount = dataCycleCount
        if auditCheck != "":
            elem.auditCheck = auditCheck
        elem.cycleCountCheck = checkedCycleCount
        elem.rctUnpCheck = checkedRctUnp
        elem.save()
        return redirect("receiving:icdr")


def getElem(request):
    if request.method == "GET":
        id = request.GET["id"]
        atributo = request.GET["atributo"]
        elem = ICDR.objects.get(id=id)

        if atributo == "consumption":
            return JsonResponse({"response": elem.consumption})
        if atributo == "poNumber":
            return JsonResponse({"response": elem.po})
        if atributo == "comentariosFecho":
            return JsonResponse({"response": elem.comentarioFechoICDR})


def updateConsumption(request):
    if request.method == "POST":
        id = request.POST["openICDRElemID"]
        comentario = request.POST["novoOpenICDR"]
        diaCorrente = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")

        elem = ICDR.objects.get(id=id)
        comentarioAnterior = elem.consumption
        comentarioNovo = (
            comentarioAnterior
            + "\n\r- "
            + str(request.user.username)
            + ", dia "
            + diaCorrente
            + ": "
            + "\n"
            + comentario
        )
        elem.consumption = comentarioNovo
        elem.save()
        return redirect("receiving:icdr")


def updatePONumber(request):
    if request.method == "POST":
        id = request.POST["poNumberElemID"]
        comentario = request.POST["novoQADEntry"]
        diaCorrente = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")

        elem = ICDR.objects.get(id=id)
        comentarioAnterior = elem.po
        comentarioNovo = (
            comentarioAnterior
            + "\n\r- "
            + str(request.user.username)
            + ", dia "
            + diaCorrente
            + ": "
            + ": "
            + "\n"
            + comentario
        )
        elem.po = comentarioNovo
        elem.save()
        return redirect("receiving:icdr")


def updateComentarioFechoICDR(request):
    if request.method == "POST":
        id = request.POST["comentarioFechoElemID"]
        comentario = request.POST["novoComentarioFechoICDR"]
        diaCorrente = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")

        elem = ICDR.objects.get(id=id)
        comentarioAnterior = elem.comentarioFechoICDR
        comentarioNovo = (
            comentarioAnterior
            + "\n\r- "
            + str(request.user.username)
            + ", dia "
            + diaCorrente
            + ": "
            + "\n"
            + comentario
        )
        elem.comentarioFechoICDR = comentarioNovo
        elem.save()
        return redirect("receiving:icdr")


def mudaEstadoFiltroICDR(request):
    if request.method == "POST":
        id = ""
        estado = ""
        elemento = ""

        if "idConfirmarCC" in request.POST:
            id = request.POST["idConfirmarCC"]
        elif "idRetirarCC" in request.POST:
            id = request.POST["idRetirarCC"]
        else:
            id = request.POST["id"]

        if "estadoConfirmarCC" in request.POST:
            estado = request.POST["estadoConfirmarCC"]
        elif "estadoRetirarCC" in request.POST:
            estado = request.POST["estadoRetirarCC"]
        else:
            estado = request.POST["estado"]

        if "elementoConfirmarCC" in request.POST:
            elemento = request.POST["elementoConfirmarCC"]
        elif "elementoRetirarCC" in request.POST:
            elemento = request.POST["elementoRetirarCC"]
        else:
            elemento = request.POST["elemento"]

        elem = ICDR.objects.get(id=id)
        if elemento == "cycle":
            elem.cycleCountCheck = estado
            if elem.cycleCountCheck == "true":
                rctUnpCheck = ""
                cycleCountCheck = ""

                trigger = TriggerEnvioEmailsICDR.objects.get(nome="Check_Cycle_Count")
                listTo = []

                for user in trigger.users.split(";"):
                    if not user in listTo:
                        if ListaUsersICDR.objects.filter(nome=user).exists():
                            lista = ListaUsersICDR.objects.get(nome=user)
                            for userListaBD in lista.user.split(";"):
                                if not userListaBD in listTo and not userListaBD == "":
                                    listTo.append(userListaBD + "@visteon.com")
                        else:
                            if not user == "":
                                listTo.append(user + "@visteon.com")

                if trigger.estado == "ativo":
                    if elem.rctUnpCheck == "false" or elem.rctUnpCheck == "":
                        rctUnpCheck = ""
                    elif elem.rctUnpCheck == "true":
                        rctUnpCheck = "X"

                    if elem.cycleCountCheck == "false" or elem.cycleCountCheck == "":
                        cycleCountCheck = ""
                    elif elem.cycleCountCheck == "true":
                        cycleCountCheck = "X"

                    table = "<h3>User: <b>" + request.user.username + " </b></h3> "
                    table += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
                    table += (
                        '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                        + elem.aberturaICDR
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem.ageing)
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.nAno
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.fornecedor
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.partnumber
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.quantidade
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.tipo
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.simNao
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.responsavel
                        + "/"
                        + elem.departamento
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.comentarioFecho
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.unCost
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.totalCost
                        + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                        + rctUnpCheck
                        + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                        + cycleCountCheck
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.date
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.cycleCount
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.dataCycleCount
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.auditCheck
                        + "</td></tr>"
                    )
                    table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
                    table += "</body></table>"

                    subject, from_email, to = (
                        "Alteração de Cycle Count - ICDR",
                        "noreply@visteon.com",
                        listTo,
                    )
                    msg = EmailMultiAlternatives(subject, table, from_email, to)
                    msg.attach_alternative(table, "text/html")
                    msg.send()
            elif elem.cycleCountCheck == "false":
                rctUnpCheck = ""
                cycleCountCheck = ""

                trigger = TriggerEnvioEmailsICDR.objects.get(nome="Check_Cycle_Count")
                listTo = []

                for user in trigger.users.split(";"):
                    if not user in listTo:
                        if ListaUsersICDR.objects.filter(nome=user).exists():
                            lista = ListaUsersICDR.objects.get(nome=user)
                            for userListaBD in lista.user.split(";"):
                                if not userListaBD in listTo and not userListaBD == "":
                                    listTo.append(userListaBD + "@visteon.com")
                        else:
                            if not user == "":
                                listTo.append(user + "@visteon.com")

                if trigger.estado == "ativo":
                    if elem.rctUnpCheck == "false" or elem.rctUnpCheck == "":
                        rctUnpCheck = ""
                    elif elem.rctUnpCheck == "true":
                        rctUnpCheck = "X"

                    if elem.cycleCountCheck == "false" or elem.cycleCountCheck == "":
                        cycleCountCheck = ""
                    elif elem.cycleCountCheck == "true":
                        cycleCountCheck = "X"

                    table = "<h3>User: <b>" + request.user.username + " </b></h3> "
                    table += '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
                    table += (
                        '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                        + elem.aberturaICDR
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + str(elem.ageing)
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.nAno
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.fornecedor
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.partnumber
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.quantidade
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.tipo
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.simNao
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.responsavel
                        + "/"
                        + elem.departamento
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.comentarioFecho
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.unCost
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.totalCost
                        + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                        + rctUnpCheck
                        + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                        + cycleCountCheck
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.date
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.cycleCount
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.dataCycleCount
                        + '</td><td style="padding:0 15px 0 15px;">'
                        + elem.auditCheck
                        + "</td></tr>"
                    )
                    table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
                    table += "</body></table>"

                    subject, from_email, to = (
                        "Alteração de Cycle Count - ICDR",
                        "noreply@visteon.com",
                        listTo,
                    )
                    msg = EmailMultiAlternatives(subject, table, from_email, to)
                    msg.attach_alternative(table, "text/html")
                    msg.send()
        if elemento == "rct":
            elem.rctUnpCheck = estado
        elem.save()
        return redirect("receiving:icdr")


def downloadExcel(request):
    if request.method == "GET":
        elementos = ICDR.objects

        caminho = "C:/visteon/media/receiving/icdr"
        if os.path.exists(caminho):
            for entry in os.listdir(caminho):
                if os.path.isfile(os.path.join(caminho, entry)):
                    os.remove(caminho + "/" + entry)

        wbProduction = Workbook()
        sheetProduction = wbProduction.add_sheet("Sheet 1")

        row = 0
        col = 0
        sheetProduction.write(row, col, "Data abertura")
        sheetProduction.write(row, col + 1, "Ageing (dias)")
        sheetProduction.write(row, col + 2, "Nº/Ano")
        sheetProduction.write(row, col + 3, "Fornecedor")
        sheetProduction.write(row, col + 4, "Partnumber")
        sheetProduction.write(row, col + 5, "Quantidade")
        sheetProduction.write(row, col + 6, "Motivo")
        sheetProduction.write(row, col + 7, "Alfandega")
        sheetProduction.write(row, col + 8, "Responsável")
        sheetProduction.write(row, col + 9, "Departamento")
        sheetProduction.write(row, col + 10, "Comentários (abertura)")
        sheetProduction.write(row, col + 11, "Un cost")
        sheetProduction.write(row, col + 12, "Total cost")
        sheetProduction.write(row, col + 13, "RCT UNP")
        sheetProduction.write(row, col + 14, "Cycle count")
        sheetProduction.write(row, col + 15, "Consumo")
        sheetProduction.write(row, col + 16, "QAD (PO nº)")
        sheetProduction.write(row, col + 17, "Comentários")
        sheetProduction.write(row, col + 18, "Data fecho")
        sheetProduction.write(row, col + 19, "Cycle count")
        sheetProduction.write(row, col + 20, "Data cc")
        sheetProduction.write(row, col + 21, "Audit check")
        row += 1

        for elem in elementos.all():
            sheetProduction.write(row, col, elem.aberturaICDR)
            sheetProduction.write(row, col + 1, elem.ageing)
            sheetProduction.write(row, col + 2, elem.nAno)
            sheetProduction.write(row, col + 3, elem.fornecedor)
            sheetProduction.write(row, col + 4, elem.partnumber)
            sheetProduction.write(row, col + 5, elem.quantidade)
            sheetProduction.write(row, col + 6, elem.tipo)
            sheetProduction.write(row, col + 7, elem.simNao)
            sheetProduction.write(row, col + 8, elem.responsavel)
            sheetProduction.write(row, col + 9, elem.departamento)
            sheetProduction.write(row, col + 10, elem.comentarioFecho)
            sheetProduction.write(row, col + 11, elem.unCost)
            sheetProduction.write(row, col + 12, elem.totalCost)
            sheetProduction.write(row, col + 13, elem.rctUnpCheck)
            sheetProduction.write(row, col + 14, elem.cycleCountCheck)
            sheetProduction.write(row, col + 15, elem.consumption)
            sheetProduction.write(row, col + 16, elem.po)
            sheetProduction.write(row, col + 17, elem.comentarioFechoICDR)
            sheetProduction.write(row, col + 18, elem.date)
            sheetProduction.write(row, col + 19, elem.cycleCount)
            sheetProduction.write(row, col + 20, elem.dataCycleCount)
            sheetProduction.write(row, col + 21, elem.auditCheck)
            row += 1
        wbProduction.save("..\\visteon\\media\\receiving\\icdr\\workbookICDR.xls")

        return redirect("receiving:icdr")


def mudaEstadoSession(request):
    if request.method == "POST":
        estado = request.POST["estado"]

        if estado == "on":
            request.session["filtroDesativado"] = "filtroDesativado"
        else:
            del request.session["filtroDesativado"]

        return redirect("receiving:icdr")


def uploadDataICDR(request):
    if request.method == "POST":
        if request.FILES.get("ficheiro"):

            ficheiro = [request.FILES["ficheiro"]]

            workbook = openpyxl.load_workbook(ficheiro[0])
            worksheet = workbook["Sheet 1"]

            for element in worksheet:
                if element[0].value != None and element[0].value != "Data Abertura":
                    novoElemento = ICDR(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                        element[7].value,
                        element[8].value,
                        element[9].value,
                        element[10].value,
                        element[11].value,
                        element[12].value,
                        element[13].value,
                        element[14].value,
                        element[15].value,
                        element[16].value,
                        element[17].value,
                        element[18].value,
                        element[19].value,
                        element[20].value,
                        element[21].value,
                    )
                    novoElemento.save()

        return redirect("receiving:icdr")


def uploadDataLineRequest(request):
    if request.method == "POST":
        if request.FILES.get("ficheiro"):

            ficheiro = [request.FILES["ficheiro"]]

            workbook = openpyxl.load_workbook(ficheiro[0])
            worksheet = workbook["Sheet1"]

            for element in worksheet:
                if element[0].value != None and element[0].value != "Request Date":
                    novoElemento = LineRequestFinished(
                        None,
                        element[0].value,
                        element[1].value,
                        element[2].value,
                        element[3].value,
                        element[4].value,
                        element[5].value,
                        element[6].value,
                    )
                    novoElemento.save()

        return redirect("receiving:lineRequest")


def criarUserICDR(request):
    if request.method == "POST":
        nome = request.POST["nomeUser"]
        area = request.POST["areaUser"]
        email = request.POST["emailUser"]
        username = email.split("@")[0]

        novoUser = UserICDR(None, username, nome, area, email)
        novoUser.save()

        if "_configurations" in request.POST:
            return redirect("receiving:configurationsICDR")
        return redirect("receiving:icdr")


def criarListaUsersICDR(request):
    if request.method == "POST":
        nomeLista = request.POST["nomeListaUser"]
        users = request.POST.getlist("listaDeUsers")
        novoUserLista = ""

        novaLista = ListaUsersICDR()
        novaLista.nome = nomeLista
        for user in users:
            novoUserLista += ";" + user
        novaLista.user = novoUserLista
        novaLista.save()
        return redirect("receiving:configurationsICDR")


def removerUtilizadorICDR(request):
    if request.method == "POST":
        username = request.POST["utilizador"]

        UserICDR.objects.get(username=username).delete()

        return redirect("receiving:configurationsICDR")


def adicionarUtilizadorListaICDR(request):
    if request.method == "POST":
        lista = ListaUsersICDR.objects.get(nome=request.POST["lista_adicionar"])
        users = request.POST.getlist("listaDeUsersDisponiveis")

        novoUserLista = ""
        if lista.user != None:
            novoUserLista = lista.user

        for user in users:
            novoUserLista += ";" + user

        lista.user = novoUserLista
        lista.save()
        return redirect("receiving:configurationsICDR")


def removerUtilizadorListaICDR(request):
    if request.method == "POST":
        nomeLista = ListaUsersICDR.objects.get(nome=request.POST["lista_remover"])
        users = request.POST.getlist("listaDeUsersRemover")

        novaStringUsers = ""

        lista = ListaUsersICDR.objects.get(nome=nomeLista)
        for elem in lista.user.split(";"):
            if not elem in users:
                novaStringUsers += ";" + elem
        novaStringUsers = novaStringUsers.replace(";;", "")
        lista.user = novaStringUsers
        lista.save()
        return redirect("receiving:configurationsICDR")


def alterarEstadoTriggerICDR(request):
    if request.method == "POST":
        trigger = request.POST["triggers"]
        nome = trigger.split("-")[0]
        estado = trigger.split("-")[1]

        triggerBD = TriggerEnvioEmailsICDR.objects.get(nome=nome)
        if estado == "ativo":
            triggerBD.estado = "inativo"
        elif estado == "inativo":
            triggerBD.estado = "ativo"
        triggerBD.save()

        return redirect("receiving:configurationsICDR")


def adicionarListaAlertaICDR(request):
    if request.method == "POST":
        trigger = request.POST["listaAlerta_adicionar"]
        listas = request.POST.getlist("listaEscolhidaAlerta")

        users = ""
        alerta = TriggerEnvioEmailsICDR.objects.get(nome=trigger)
        if alerta.users != None:
            users = alerta.users

        for user in listas:
            if not user in users:
                users += ";" + user
        users = users.replace(";;", "")
        alerta.users = users
        alerta.save()
        return redirect("receiving:configurationsICDR")


def removerListaAlertaICDR(request):
    if request.method == "POST":
        trigger = request.POST["listaAlerta_remover"]
        listas = request.POST.getlist("listaEscolhidaAlertaRemover")

        alerta = TriggerEnvioEmailsICDR.objects.get(nome=trigger)
        novaLista = ""

        for listaBD in alerta.users.split(";"):
            if not listaBD in listas:
                novaLista += ";" + listaBD
        novaLista = novaLista.replace(";;", "")
        alerta.users = novaLista
        alerta.save()
        return redirect("receiving:configurationsICDR")


def searchUsersPorLista(request):
    if request.method == "GET":
        lista = request.GET["lista"]
        listaResposta = []
        elementoLista = ""
        elementosRestantesLista = ""
        usersBD = UserICDR.objects

        if lista == "":
            for utilizador in usersBD.all():
                if utilizador.username != None:
                    elementoLista += ";" + utilizador.username
        else:
            listaBD = ListaUsersICDR.objects.get(nome=lista)
            elementoLista = listaBD.user
        listaResposta.append(elementoLista)

        if "utilizadoresRestantes" in request.GET:
            if lista == "":
                for utilizador in usersBD.all():
                    if utilizador.username != None:
                        elementosRestantesLista += ";" + utilizador.username
            else:
                listaBD = ListaUsersICDR.objects.filter(nome=lista).values_list(
                    "user", flat=True
                )
                if listaBD[0] != None:
                    elementosRestantes = listaBD[0].split(";")
                    for utilizador in usersBD.all():
                        if utilizador.username != None:
                            if not (utilizador.username in elementoLista) and not (
                                utilizador.username in elementosRestantesLista
                            ):
                                elementosRestantesLista += ";" + utilizador.username
                else:
                    for utilizador in usersBD.all():
                        elementosRestantesLista += ";" + utilizador.username
                listaResposta.append(elementosRestantesLista)

        return JsonResponse({"listaResposta": listaResposta}, safe=False)


def searchListasPorAlerta(request):
    if request.method == "GET":
        usersBD = UserICDR.objects
        listasBD = ListaUsersICDR.objects

        alerta = request.GET["alerta"]
        listaResposta = []
        elementoLista = ""

        if alerta == "":
            for lista in listasBD.all():
                elementoLista += ";" + lista.nome
            for utilizador in usersBD.all():
                elementoLista += ";" + utilizador.username
        else:
            alerta = TriggerEnvioEmailsICDR.objects.get(nome=alerta)
            elementoLista = alerta.users
        listaResposta.append(elementoLista)
        return JsonResponse({"listaResposta": listaResposta}, safe=False)


def reportICDR(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Receiving -  ICDR",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:icdr")


def reportLineRequest(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Receiving - Line Request",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:lineRequest")


def reportMNFGSupply(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Receiving - MNFG Supply",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:mnfgSupply")


def reportTPM(request):
    if request.method == "POST":
        texto = request.POST["reportICDR"].replace("\n", "</br>")

        message = "<b>Report criado pelo User: " + request.user.username + "</b>"
        message += "</br></br>" + texto
        subject, from_email, to = (
            "Novo report na página Receiving - TPM",
            "noreply@visteon.com",
            ["aroque1@visteon.com", "npires2@visteon.com"],
        )
        msg = EmailMultiAlternatives(subject, message, from_email, to)
        msg.attach_alternative(message, "text/html")
        msg.send()
        return redirect("receiving:tpm")


def atualizarAgeing():
    elementosICDR = ICDR.objects
    actualDay = datetime.today()
    actualDayString = actualDay.strftime("%Y-%m-%d")

    if actualDay.strftime("%A") != "Saturday" and actualDay.strftime("%A") != "Sunday":
        for elem in elementosICDR.all():
            if elem.ageing != "-":
                elemDayString = elem.aberturaICDR[0:10]
                contador = np.busday_count(elemDayString, actualDayString)
                elem.ageing = contador
            elem.save()


def enviarEmailsICDR():
    print("entrou")
    elementosICDR = ICDR.objects

    for elem in elementosICDR.all():
        trigger = TriggerEnvioEmailsICDR.objects.get(nome="24H")
        listTo = []

        for user in trigger.users.split(";"):
            if not user in listTo:
                if ListaUsersICDR.objects.filter(nome=user).exists():
                    lista = ListaUsersICDR.objects.get(nome=user)
                    for userListaBD in lista.user.split(";"):
                        if not userListaBD in listTo and not userListaBD == "":
                            listTo.append(userListaBD + "@visteon.com")
                else:
                    if not user == "":
                        listTo.append(user + "@visteon.com")

        if trigger.estado == "ativo":
            if elem.ageing == "1":
                rctUnpCheck = ""
                cycleCountCheck = ""
                if elem.rctUnpCheck == "false" or elem.rctUnpCheck == "":
                    rctUnpCheck = ""
                elif elem.rctUnpCheck == "true":
                    rctUnpCheck = "X"

                if elem.cycleCountCheck == "false" or elem.cycleCountCheck == "":
                    cycleCountCheck = ""
                elif elem.cycleCountCheck == "true":
                    cycleCountCheck = "X"

                table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
                table += (
                    '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                    + elem.aberturaICDR
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem.ageing)
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.nAno
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.fornecedor
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.partnumber
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.quantidade
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.tipo
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.simNao
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.responsavel
                    + "/"
                    + elem.departamento
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.comentarioFecho
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.unCost
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.totalCost
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + rctUnpCheck
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + cycleCountCheck
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.date
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.cycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.dataCycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.auditCheck
                    + "</td></tr>"
                )
                table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
                table += "</body></table>"

                subject, from_email, to = "Report ICDR", "noreply@visteon.com", listTo
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                msg.send()

        trigger2 = TriggerEnvioEmailsICDR.objects.get(nome="+48H")
        listTo2 = []

        for user in trigger2.users.split(";"):
            if not user in listTo2:
                if ListaUsersICDR.objects.filter(nome=user).exists():
                    lista = ListaUsersICDR.objects.get(nome=user)
                    for userListaBD in lista.user.split(";"):
                        if not userListaBD in listTo2 and not userListaBD == "":
                            listTo2.append(userListaBD + "@visteon.com")
                else:
                    if not user == "":
                        listTo2.append(user + "@visteon.com")

        if trigger2.estado == "ativo":
            if elem.ageing != "0" and elem.ageing != "-" and elem.ageing != "1":
                rctUnpCheck = ""
                cycleCountCheck = ""
                if elem.rctUnpCheck == "false" or elem.rctUnpCheck == "":
                    rctUnpCheck = ""
                elif elem.rctUnpCheck == "true":
                    rctUnpCheck = "X"

                if elem.cycleCountCheck == "false" or elem.cycleCountCheck == "":
                    cycleCountCheck = ""
                elif elem.cycleCountCheck == "true":
                    cycleCountCheck = "X"

                table = '</br><table class="display"><thead style="background-color: lightgray"><tr><th>DATA ABERTURA</th><th>AGEING (DIAS)</th><th>Nº/ANO</th><th>FORNECEDOR</th><th>PARTNUMBER</th><th>QT</th><th>MOTIVO</th><th>ALFANDEGA (S/N)</th><th>RESPONSÁVEL (NOME/DEPARTAMENTO)</th><th>COMENTÁRIOS (ABERTURA)</th><th>UN COST</th><th>TOTAL COST</th><th>RCT UNP</th><th>CYCLE COUNT</th><th>DATA FECHO</th><th>CYCLE COUNT</th><th>DATA CC</th><th>AUDIT CHECK</th></tr></thead><tbody>'
                table += (
                    '<tr style="background-color: whitesmoke"><td style="padding:0 15px 0 15px;">'
                    + elem.aberturaICDR
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + str(elem.ageing)
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.nAno
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.fornecedor
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.partnumber
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.quantidade
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.tipo
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.simNao
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.responsavel
                    + "/"
                    + elem.departamento
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.comentarioFecho
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.unCost
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.totalCost
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + rctUnpCheck
                    + '</td><td style="padding:0 15px 0 15px; text-align: center">'
                    + cycleCountCheck
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.date
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.cycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.dataCycleCount
                    + '</td><td style="padding:0 15px 0 15px;">'
                    + elem.auditCheck
                    + "</td></tr>"
                )
                table += '<tr style="background-color: lightgray"><td colspan="18"> </td></tr>'
                table += "</body></table>"

                subject, from_email, to = "Report ICDR", "noreply@visteon.com", listTo2
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                msg.send()


def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    return "%02d:%02d:%02d" % (hour, minutes, seconds)

def getVermelhosSupply(request):
    hoje = datetime.today()
    data_formatada = hoje.strftime('%m/%d/%Y')
    agora = datetime.now().time()
    hora_formatada = agora.strftime('%H:%M')
    print("->", type(hora_formatada))
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
     

    if request.POST["table"] == "receiving_fasupplyreditems":
        #if request.POST["turno"] == "1"
        if hora_formatada > "08:00" and hora_formatada < "16:30": 
            contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            print("Primeiro Turno",len(contador),contador)
        #elif request.POST["turno"] == "2"
        elif hora_formatada > "16:30" and hora_formatada <= "24:00":
            print("Segundo turno")
            contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
        else:
            print("Terceiro turno")
            contador = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
        return JsonResponse({"message": "OK", "contador": len(contador)})
    
    elif request.POST["table"] == "receiving_bpsmdsupplyreditems":
        if hora_formatada > "08:00" and hora_formatada < "16:30": 
            contador = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            print("Primeiro Turno",len(contador))

        elif hora_formatada > "16:30" and hora_formatada < "24:00":
            print("Segundo turno")
            contador = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
        else:
            print("Terceiro turno")
            contador = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
        return JsonResponse({"message": "OK", "contador": len(contador)})
    
    elif request.POST["table"] == "all":
        """-------------------------------- TODOS OS TOTAIS -------------------------------------------- """
        print("queres ver os totais para todas as tabelas")
        if hora_formatada > "08:00" and hora_formatada < "16:30": 
            contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            return JsonResponse({"message": "OK", "contadorBP": len(contadorBPT1), "contadorBPSMD": len(contadorBPSMDT1), "contadorFA": len(contadorFAT1)})

        elif hora_formatada > "16:30" and hora_formatada < "24:00":
            print("Segundo turno")
            contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
            contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
            contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
            return JsonResponse({"message": "OK", "contadorBP": len(contadorBPT2), "contadorBPSMD": len(contadorBPSMDT2), "contadorFA": len(contadorFAT2)})
        
        else:
            print("Terceiro turno")
            contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
            contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
            contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
            return JsonResponse({"message": "OK", "contadorBP": len(contadorBPT3), "contadorBPSMD": len(contadorBPSMDT3), "contadorFA": len(contadorFAT3)})
    else:
        if hora_formatada > "08:00" and hora_formatada < "16:30": 
            contador = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 08:00' AND '{data_formatada} 16:30' """)
            print("Primeiro Turno",len(contador))

        elif hora_formatada > "16:30" and hora_formatada < "24:00":
            print("Segundo turno")
            contador = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada} 16:30' AND '{data_seguinte_formatada} 01:00"' """)
        else:
            print("Terceiro turno")
            contador = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada}  01:00' AND '{data_formatada} 08:00' """)
        return JsonResponse({"message": "OK", "contador": len(contador)})


    return JsonResponse({"message": "OK"})

def getErros(request):
    print("REQUESt->",request.POST)
    erro = HistoricoErros_View.objects.filter(pagina =request.POST["pagina"])
    timestamp = HistoricoErros_View.objects.filter(pagina =request.POST["pagina"]).values("timestamp").first()
    print("tamanho ", len(erro))
    return JsonResponse({"message": "OK", "erro" : len(erro), "timestamp" : timestamp})


def sendRelatorioDiarioSupply(request = None): 
    print("entrou no sendRelatorioDiario")
    import datetime
    """  ["pmarti30@visteon.com",'npires2@visteon.com',
                     'abrandao@visteon.com','sanasta1@visteon.com',
                    'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                    'abilro1@visteon.com', 'jrodri80@visteon.com',
                    'evenanc1@visteon.com'] """
    #timeNow = datetime.datetime.now()
    print("FUNC REPORT CONTADORES MNFG SUPPLY EMAIL")
    #print("REQUEST para tirar excel", request.POST)

    #acabar formatacao
    table = """<h1>REPORT MNFG PAGES:</h1> <h2><br></br> Contadores </h2></br><table class="display"><thead style="background-color: lightgray">
    <tr><th></th>
    <th> 1 Turno </th>
    <th> 2 Turno </th>
    <th> 3 Turno </th>
    </tr></thead>
    <tbody>"""
    hoje = datetime.datetime.now()
    """ hoje = datetime.datetime.now()
    data_formatada = hoje.strftime('%m/%d/%Y')
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
    dia_anterior = hoje - timedelta(days=1)
    data_anterior_formatada = dia_anterior.strftime('%m/%d/%Y') """
    try:
       
        if request.POST["data_report"]:
        #if len(request) > 0:
            data_report = request.POST["data_report"]
            data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
            data_str = data_obj.strftime("%m/%d/%Y")
            dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
            dia_seguinte_str = dia_seguinte_obj.strftime("%m/%d/%Y")
            dia_anterior = data_obj - datetime.timedelta(days=1)
            data_anterior_formatada = dia_anterior.strftime("%m/%d/%Y")

            print("Primeiro turno")
            contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
            contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
            contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)

            print("Segundo turno")
            contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)
            contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)
            contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)

            contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            print("Terceiro turno",request.POST["data_report"], "|", contadorFAT3)
        
        else:
            #Se nao houver request
            print("não ha request")
            data_report = request.POST["data_report"]
            data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
            data_str = data_obj.strftime("%m/%d/%Y")
            dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
            dia_seguinte_str = dia_seguinte_obj.strftime("%m/%d/%Y")
            dia_anterior = data_obj - datetime.timedelta(days=1)
            data_anterior_formatada = dia_anterior.strftime("%m/%d/%Y")

            data_atual = datetime.datetime.now()
            data_atual_str = data_atual.strftime("%m/%d/%Y")
            dia_seguinte_obj_novo = data_atual + datetime.timedelta(days=1)
            dia_seguinte_str_novo = dia_seguinte_obj_novo.strftime("%m/%d/%Y")

            print("Primeiro turno")
            contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)
            contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)
            contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)

            print("Segundo turno")
            contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)
            contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)
            contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)

            print("Terceiro turno")
            contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)
            contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)
            contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)


        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "FA Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "BPDropin Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "BPSMD Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += "</tbody></table>"

        
        subject, from_email, to = (
            "Receiving Supply Report: " + hoje.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        ) 
        
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        #print("Mensagem a ser enviada no mail", nextDay.strftime("%d-%m-%Y"))
        print(2/0)
        msg.send()
        return JsonResponse({"message": "OK"})
    except:
        #se nao for possivel enviar a informacao por email ele cria um excel e faz download
        # data_report = request.POST["data_report"]
        # data_obj = datetime.strptime(data_report, "%d/%m/%Y")
        # dia_seguinte_obj = data_obj + timedelta(days=1)
        # dia_seguinte_str = dia_seguinte_obj.strftime("%d/%m/%Y")
            try:
                data_report = request.POST["data_report"]
                data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
                dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
                dia_seguinte_str = dia_seguinte_obj.strftime("%d/%m/%Y")

                data_report = request.POST["data_report"]
                data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
                data_str = data_obj.strftime("%d/%m/%Y")

                print("Primeiro turno")
                contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
                contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
                contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)

                print("Segundo turno")
                contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)
                contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)
                contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)

                print("Terceiro turno", contadorBPT3, "|", data_obj)
                contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                # nome do file workbookReportMNFG
                style = easyxf('pattern: pattern solid, fore_colour red;')
                wbProduction = Workbook()
                sheetProduction = wbProduction.add_sheet("Sheet 1")
                row = 0
                col = 0
                
                sheetProduction.write(row, col, "Vermelhos MNFG",style)
                sheetProduction.write(row, col + 1, "")
                sheetProduction.write(row, col + 2, "")
                sheetProduction.write(row, col + 3, "")
                
                row += 1

                sheetProduction.write(row, col, "")
                sheetProduction.write(row, col + 1, "Turno 1")
                sheetProduction.write(row, col + 2, "Turno 2")
                sheetProduction.write(row, col + 3, "Turno 4")

                row += 1

                sheetProduction.write(row, col,"FA")
                sheetProduction.write(row, col + 1,len(contadorFAT1))
                sheetProduction.write(row, col + 2,len(contadorFAT2))
                sheetProduction.write(row, col + 3,len(contadorFAT3))

                row += 1

                sheetProduction.write(row, col , "BPSMD")
                sheetProduction.write(row, col + 1,len(contadorBPSMDT1))
                sheetProduction.write(row, col + 2,len(contadorBPSMDT2))
                sheetProduction.write(row, col + 3,len(contadorBPSMDT3))

                row += 1

                sheetProduction.write(row, col , "BPDropin")
                sheetProduction.write(row, col + 1,len(contadorBPT1))
                sheetProduction.write(row, col + 2,len(contadorBPT2))
                sheetProduction.write(row, col + 3,len(contadorBPT3))

                wbProduction.save("..\\visteon\\media\\receiving\\MNFG\\workbookReportMNFG.xls")
                return JsonResponse({"message": "Excel"})
            except:
                data_atual = datetime.datetime.now()
                data_formatada_atual = data_atual.strftime("%m/%d/%Y")

                data_ontem = data_atual - datetime.timedelta(days=1)
                data_formatada_ontem = data_ontem.strftime("%m/%d/%Y")

                print("Primeiro turno",data_formatada_atual,data_formatada_ontem)

                contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)
                contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)
                contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)

                print("Segundo turno",contadorBPT1)
                contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)
                contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)
                contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)

                print("Terceiro turno",contadorBPT2)
                contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "FA Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "BPDropin Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "BPSMD Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += "</tbody></table>"

                
                subject, from_email, to = (
                    "Receiving Supply Report: " + hoje.strftime("%d-%m-%Y"),
                    "noreply@visteon.com",
                    ["pmarti30@visteon.com"],
                ) 
                
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                print("Mensagem a ser enviada no mail")
                #print(2/0)
                msg.send()
                return JsonResponse({"message": "OK"})
        
    return JsonResponse({"message": "OK"})
    return redirect("crossdocking:configurationsEmailDiario")

#verrsão sem excel
def sendRelatorioDiarioSupply(request = None): 
    print("entrou no sendRelatorioDiario")
    import datetime
    """  ["pmarti30@visteon.com",'npires2@visteon.com',
                     'abrandao@visteon.com','sanasta1@visteon.com',
                    'rsalgue2@visteon.com', 'nlopes8@visteon.com',
                    'abilro1@visteon.com', 'jrodri80@visteon.com',
                    'evenanc1@visteon.com'] """
    #timeNow = datetime.datetime.now()
    print("FUNC REPORT CONTADORES MNFG SUPPLY EMAIL")
    #print("REQUEST para tirar excel", request.POST)

    #acabar formatacao
    table = """<h1>REPORT MNFG PAGES:</h1> <h2><br></br> Contadores </h2></br><table class="display"><thead style="background-color: lightgray">
    <tr><th></th>
    <th> 1 Turno </th>
    <th> 2 Turno </th>
    <th> 3 Turno </th>
    </tr></thead>
    <tbody>"""
    hoje = datetime.datetime.now()
    """ hoje = datetime.datetime.now()
    data_formatada = hoje.strftime('%m/%d/%Y')
    dia_seguinte = hoje + timedelta(days=1)
    data_seguinte_formatada = dia_seguinte.strftime('%m/%d/%Y')
    dia_anterior = hoje - timedelta(days=1)
    data_anterior_formatada = dia_anterior.strftime('%m/%d/%Y') """
    try:
       
        if request.POST["data_report"]:
        #if len(request) > 0:
            data_report = request.POST["data_report"]
            data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
            data_str = data_obj.strftime("%m/%d/%Y")
            dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
            dia_seguinte_str = dia_seguinte_obj.strftime("%m/%d/%Y")
            dia_anterior = data_obj - datetime.timedelta(days=1)
            data_anterior_formatada = dia_anterior.strftime("%m/%d/%Y")

            print("Primeiro turno")
            contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
            contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
            contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)

            print("Segundo turno")
            contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)
            contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)
            contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00"' """)

            contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
            print("Terceiro turno",request.POST["data_report"], "|", contadorFAT3)
        
        else:
            #Se nao houver request
            print("não ha request")
            data_report = request.POST["data_report"]
            data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
            data_str = data_obj.strftime("%m/%d/%Y")
            dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
            dia_seguinte_str = dia_seguinte_obj.strftime("%m/%d/%Y")
            dia_anterior = data_obj - datetime.timedelta(days=1)
            data_anterior_formatada = dia_anterior.strftime("%m/%d/%Y")

            data_atual = datetime.datetime.now()
            data_atual_str = data_atual.strftime("%m/%d/%Y")
            dia_seguinte_obj_novo = data_atual + datetime.timedelta(days=1)
            dia_seguinte_str_novo = dia_seguinte_obj_novo.strftime("%m/%d/%Y")

            print("Primeiro turno")
            contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)
            contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)
            contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 08:00' AND '{data_atual_str} 16:30' """)

            print("Segundo turno")
            contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)
            contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)
            contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str} 16:30' AND '{dia_seguinte_str_novo} 01:00' """)

            print("Terceiro turno")
            contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)
            contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)
            contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_atual_str}  01:00' AND '{data_atual_str} 08:00' """)


        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "FA Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorFAT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "BPDropin Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += (
            '<tr><td style="padding:0 15px 0 15px;">'
        + "BPSMD Supply"
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT1)) # site
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT2)) # due_date
        + '</td><td style="padding:0 15px 0 15px;">'
        + str(len(contadorBPSMDT3)) # item_number
        + '</td>'
        + "</td></tr>"
        )

        table += "</tbody></table>"

        
        subject, from_email, to = (
            "Receiving Supply Report: " + hoje.strftime("%d-%m-%Y"),
            "noreply@visteon.com",
            ["pmarti30@visteon.com"],
        ) 
        
        msg = EmailMultiAlternatives(subject, table, from_email, to)
        msg.attach_alternative(table, "text/html")
        #print("Mensagem a ser enviada no mail", nextDay.strftime("%d-%m-%Y"))
        print(2/0)
        msg.send()
        return JsonResponse({"message": "OK"})
    except:
        #se nao for possivel enviar a informacao por email ele cria um excel e faz download
        # data_report = request.POST["data_report"]
        # data_obj = datetime.strptime(data_report, "%d/%m/%Y")
        # dia_seguinte_obj = data_obj + timedelta(days=1)
        # dia_seguinte_str = dia_seguinte_obj.strftime("%d/%m/%Y")
            try:
                data_report = request.POST["data_report"]
                data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
                dia_seguinte_obj = data_obj + datetime.timedelta(days=1)
                dia_seguinte_str = dia_seguinte_obj.strftime("%d/%m/%Y")

                data_report = request.POST["data_report"]
                data_obj = datetime.datetime.strptime(data_report, "%d/%m/%Y")
                data_str = data_obj.strftime("%d/%m/%Y")

                print("Primeiro turno")
                contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
                contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)
                contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 08:00' AND '{data_str} 16:30' """)

                print("Segundo turno")
                contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)
                contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)
                contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str} 16:30' AND '{dia_seguinte_str} 01:00' """)

                print("Terceiro turno", contadorBPT3, "|", data_obj)
                contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_str}  01:00' AND '{data_str} 08:00' """)
                # nome do file workbookReportMNFG
                style = easyxf('pattern: pattern solid, fore_colour red;')
                wbProduction = Workbook()
                sheetProduction = wbProduction.add_sheet("Sheet 1")
                row = 0
                col = 0
                
                sheetProduction.write(row, col, "Vermelhos MNFG",style)
                sheetProduction.write(row, col + 1, "")
                sheetProduction.write(row, col + 2, "")
                sheetProduction.write(row, col + 3, "")
                
                row += 1

                sheetProduction.write(row, col, "")
                sheetProduction.write(row, col + 1, "Turno 1")
                sheetProduction.write(row, col + 2, "Turno 2")
                sheetProduction.write(row, col + 3, "Turno 4")

                row += 1

                sheetProduction.write(row, col,"FA")
                sheetProduction.write(row, col + 1,len(contadorFAT1))
                sheetProduction.write(row, col + 2,len(contadorFAT2))
                sheetProduction.write(row, col + 3,len(contadorFAT3))

                row += 1

                sheetProduction.write(row, col , "BPSMD")
                sheetProduction.write(row, col + 1,len(contadorBPSMDT1))
                sheetProduction.write(row, col + 2,len(contadorBPSMDT2))
                sheetProduction.write(row, col + 3,len(contadorBPSMDT3))

                row += 1

                sheetProduction.write(row, col , "BPDropin")
                sheetProduction.write(row, col + 1,len(contadorBPT1))
                sheetProduction.write(row, col + 2,len(contadorBPT2))
                sheetProduction.write(row, col + 3,len(contadorBPT3))

                wbProduction.save("..\\visteon\\media\\receiving\\MNFG\\workbookReportMNFG.xls")
                return JsonResponse({"message": "Excel"})
            except:
                data_atual = datetime.datetime.now()
                data_formatada_atual = data_atual.strftime("%m/%d/%Y")

                data_ontem = data_atual - datetime.timedelta(days=1)
                data_formatada_ontem = data_ontem.strftime("%m/%d/%Y")

                print("Primeiro turno",data_formatada_atual,data_formatada_ontem)

                contadorBPT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)
                contadorBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)
                contadorFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 08:00' AND '{data_formatada_ontem} 16:30' """)

                print("Segundo turno",contadorBPT1)
                contadorBPT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)
                contadorBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)
                contadorFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem} 16:30' AND '{data_formatada_atual} 01:00' """)

                print("Terceiro turno",contadorBPT2)
                contadorBPT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                contadorBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                contadorFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{data_formatada_ontem}  01:00' AND '{data_formatada_ontem} 08:00' """)
                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "FA Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorFAT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "BPDropin Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += (
                    '<tr><td style="padding:0 15px 0 15px;">'
                + "BPSMD Supply"
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT1)) # site
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT2)) # due_date
                + '</td><td style="padding:0 15px 0 15px;">'
                + str(len(contadorBPSMDT3)) # item_number
                + '</td>'
                + "</td></tr>"
                )

                table += "</tbody></table>"

                
                subject, from_email, to = (
                    "Receiving Supply Report: " + hoje.strftime("%d-%m-%Y"),
                    "noreply@visteon.com",
                    ["pmarti30@visteon.com"],
                ) 
                
                msg = EmailMultiAlternatives(subject, table, from_email, to)
                msg.attach_alternative(table, "text/html")
                print("Mensagem a ser enviada no mail")
                #print(2/0)
                msg.send()
                return JsonResponse({"message": "OK"})
        
    return JsonResponse({"message": "OK"})
    return redirect("crossdocking:configurationsEmailDiario")



def adiciona_val_vermelhos_automatico_Supply(*var):
    from datetime import datetime, timedelta
    dia = date.today() - timedelta(days=10)
    dadosQAD = WtskMstr.objects.filter(
        wtsk_create_date__gt=dia.strftime("%Y-%m-%d")
    ).exclude(wtsk_to_loc="kardex")
    area_b = AreaB.objects.all()
    area_a = AreaA.objects.all()
    kardex = Kardex.objects.all()
     
    MNFGSupplyItems.objects.all().delete()
    for i in dadosQAD:
        now = datetime.now()
        hora_str = i.wtsk_create_time
        hora_atual = now.strftime("%H%M%S")
        data_str = i.wtsk_create_date
        dataQad = datetime.strptime(str(data_str), '%Y-%m-%d %H:%M:%S')

        data_atual = datetime.now()
        data_atual_formatada = data_atual.strftime("%Y-%m-%d 00:00:00")

        fa_supply_red_items = FASupplyRedItems.objects.filter(taskID = i.wtsk_event_id)
        bpDropin_supply_red_items = BPDropinSupplyRedItems.objects.filter(taskID = i.wtsk_event_id)
        bpSmd_supply_red_items = BPSMDSupplyRedItems.objects.filter(taskID = i.wtsk_event_id)
        
        MNFGSupplyItems(
            wtsk_id = i.wtsk_id,
            wtsk_event_id = i.wtsk_event_id,
            wtsk_wave_id = i.wtsk_wave_id,
            wtsk_task_type = i.wtsk_task_type,
            wtsk_from_part = i.wtsk_from_part,
            wtsk_from_stor_zone = i.wtsk_from_stor_zone,
            wtsk_to_loc = i.wtsk_to_loc,
            wtsk_qty_exp = i.wtsk_qty_exp,
            wtsk_create_date = i.wtsk_create_date,
            wtsk_create_time = i.wtsk_create_time,
            timestamp = now.strftime("%m/%d/%Y %H:%M"),  #%Y/%m/%d %H:%M"
        ).save()
        print("GUARDOU NOVO MNFG")
    
        """ AREA A """
        try:
            #print(2/0)
            for faItems in area_b:

                if i.wtsk_from_stor_zone == faItems.storageZone:
                    try: 
                        FASupplyRedItems.objects.get(taskID = i.wtsk_id)
                        
                    except:
                        hora_atual = now.strftime("%H%M%S")
                        a = i.wtsk_create_time
                        b = str(now.strftime("%H%M%S"))
                        time_a = datetime.strptime(a, "%H%M%S").time()
                        time_b = datetime.strptime(b, "%H%M%S").time()
                        if (datetime.combine(datetime.min, time_b) - datetime.combine(datetime.min, time_a)).total_seconds() > 7200 or  str(dataQad) != str(data_atual_formatada):
                            print("é para adicionar faItems")
                            FASupplyRedItems(
                        taskID = i.wtsk_id,
                        waveID = i.wtsk_wave_id,
                        taskType = i.wtsk_task_type,
                        fromPart = i.wtsk_from_part,
                        fromStorageZone = i.wtsk_from_stor_zone,
                        toLocation = i.wtsk_to_loc,
                        wtsk_create_time = i.wtsk_from_stor_zone,
                        createdDate = i.wtsk_create_date,
                        createdTime = i.wtsk_create_time,
                        timestamp = now.strftime("%m/%d/%Y %H:%M"),  #%Y/%m/%d %H:%M"
                    ).save()
                    print("NOVO RED FA")
                            
        except Exception as e:
            print("DEU ERRO NO FA",str(e))
            
            #HistoricoErros.objects.all().delete()
            val = HistoricoErros(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pagina ="MNFG", erro = "Funcao não conseguiu guardar nenhum valor FA").save()
            break
            #Esta tabela serve apenas para guardar historico de erros, esta pensada para servir todas as paginas do projeto
            """ AREA B """

        try :
            for bpDropItems in area_a:
                if i.wtsk_from_stor_zone == bpDropItems.storageZone:
                    try: 
                        BPDropinSupplyRedItems.objects.get(taskID = i.wtsk_id) 
                        
                    except:
                        hora_atual = now.strftime("%H%M%S")
                        a = i.wtsk_create_time
                        b = str(now.strftime("%H%M%S"))
                        time_a = datetime.strptime(a, "%H%M%S").time()
                        time_b = datetime.strptime(b, "%H%M%S").time()  
                        if (datetime.combine(datetime.min, time_b) - datetime.combine(datetime.min, time_a)).total_seconds() > 7200 or  str(dataQad) != str(data_atual_formatada):
                            bpDropin_supply_red_items = BPDropinSupplyRedItems(
                        taskID = i.wtsk_id,
                        waveID = i.wtsk_wave_id,
                        taskType = i.wtsk_task_type,
                        fromPart = i.wtsk_from_part,
                        fromStorageZone = i.wtsk_from_stor_zone,
                        toLocation = i.wtsk_to_loc,
                        createdDate = i.wtsk_create_date,
                        createdTime = i.wtsk_create_time,
                        timestamp = now.strftime("%m/%d/%Y %H:%M"), 
                    ).save()
                    print("NOVO RED BPDROPIN")
        except:
            val = HistoricoErros(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pagina ="MNFG", erro = "Funcao não conseguiu guardar nenhum valor MNFG").save()
            break
        
            """ AREA KARDEX """
        try:
            for kardexItems in kardex:
                if i.wtsk_from_stor_zone == kardexItems.storageZone:

                    try: 
                         

                        BPSMDSupplyRedItems.objects.get(taskID = i.wtsk_id)
                        #print("entrou no BPSMDSupplyRedItems")
                    except:
                        hora_atual = now.strftime("%H%M%S")
                        a = i.wtsk_create_time
                        b = str(now.strftime("%H%M%S"))
                        #print("entrou no try2")
                        time_a = datetime.strptime(a, "%H%M%S").time()
                        time_b = datetime.strptime(b, "%H%M%S").time()
                        if (datetime.combine(datetime.min, time_b) - datetime.combine(datetime.min, time_a)).total_seconds() > 7200 or  str(dataQad) != str(data_atual_formatada):
                            print("Guardou no BPSMDSupplyRedItems")
                            BPSMDSupplyRedItems(
                        taskID = i.wtsk_id,
                        waveID = i.wtsk_wave_id,
                        taskType = i.wtsk_task_type,
                        fromPart = i.wtsk_from_part,
                        fromStorageZone = i.wtsk_from_stor_zone,
                        toLocation = i.wtsk_to_loc,
                        createdDate = i.wtsk_create_date,
                        createdTime = i.wtsk_create_time,
                        timestamp = now.strftime("%m/%d/%Y %H:%M")
                    ).save()
                    print("NOVO RED BPSMD")
        except:
            val = HistoricoErros(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pagina ="MNFG", erro = "Funcao não conseguiu guardar nenhum valor BPSMD").save()
            break
    return JsonResponse({"message": "OK"})



#deixa de ser usado
def adiciona_val_para_mediaSupply(request):
    import datetime
    #receiving_fasupplyreditems / receiving_bpsmdsupplyreditems / receiving_bpdropinsupplyreditems
    #print("ENTROU--> ",request.POST)

    if request.POST["table"] == "receiving_fasupplyreditems":
        fa_supply_red_items = FASupplyRedItems.objects.filter(taskID = request.POST["task_id"])
        if  fa_supply_red_items:
            print("FaSupply REPETIDO")
        else:
            print("emNFG_supply_red_items Não está empty")
            fa_supply_red_items = FASupplyRedItems(
                taskID = request.POST["task_id"],
                waveID = request.POST["wave_id"],
                taskType = request.POST["task_type"],
                fromPart = request.POST["from_part"],
                fromStorageZone = request.POST["from_storage_zone"],
                toLocation = request.POST["to_location"],
                #quantity_expected = request.POST["qty_expected"],
                createdDate = request.POST["created_date"],
                createdTime = request.POST["created_time"],
                timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M"),  #%Y/%m/%d %H:%M"
            ).save()
            print("VAI GAURDAR NOVO mNFG")
    elif request.POST["table"] == "receiving_bpsmdsupplyreditems":
        bpsmd_supply_red_items = BPSMDSupplyRedItems.objects.filter(taskID = request.POST["task_id"])

        if  bpsmd_supply_red_items:
            print("BPSMDSupply REPETIDO")
        else:
            bpsmd_supply_red_items = BPSMDSupplyRedItems(
                taskID = request.POST["task_id"],
                waveID = request.POST["wave_id"],
                taskType = request.POST["task_type"],
                fromPart = request.POST["from_part"],
                fromStorageZone = request.POST["from_storage_zone"],
                toLocation = request.POST["to_location"],
                #quantity_expected = request.POST["qty_expected"], 
                createdDate = request.POST["created_date"],
                createdTime = request.POST["created_time"], 
                timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M")
            ).save()
            print("Guardou BPSMDSupply")
    else:
        bpDropin_supply_items = BPDropinSupplyRedItems.objects.filter(taskID = request.POST["task_id"])
        if  bpDropin_supply_items:
            print("BPDropinSupplyRedItems Repetido")
        
        else:
            print("emNFG_supply_red_items Não está empty")
            bpDropin_supply_red_items = BPDropinSupplyRedItems(
                taskID = request.POST["task_id"],
                waveID = request.POST["wave_id"],
                taskType = request.POST["task_type"],
                fromPart = request.POST["from_part"],
                fromStorageZone = request.POST["from_storage_zone"],
                toLocation = request.POST["to_location"],
                #quantity_expected = request.POST["qty_expected"],
                createdDate = request.POST["created_date"],
                createdTime = request.POST["created_time"],
                timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M"), 
            ).save()
            print("VAI GAURDAR NOVO BPDropinSupplyRedItems")
    return JsonResponse({"message": "OK"})



#não vai ser usado
def adiciona_val_mediaBPSMDSupply(request):
    import datetime
    contador = 0
    contadorNovo = 0
    
    
    bpsmd_supply_red_items = BPSMDSupplyRedItems.objects.filter(taskID = request.POST["task_id"])

    data_de_hoje = f"{datetime.datetime.now().date()}" 

    if  bpsmd_supply_red_items:
        print("Já tem esta linha")
        #se tiver algo lá dentro corre este bloc        
    else:
        BPSMDSupplyRedItems(
            taskID = request.POST["task_id"],
            waveID = request.POST["wave_id"],
            taskType = request.POST["task_type"],
            fromPart = request.POST["from_part"],
            fromStorageZone = request.POST["from_storage_zone"],
            toLocation = request.POST["to_location"],
            #quantity_expected = request.POST["qty_expected"], 
            createdDate = request.POST["created_date"],
            createdTime = request.POST["created_time"], 
            timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M")
        ).save()
        
       
    print("Fim da itera")
    return JsonResponse({"message": "OK"})



#não vai ser usado
def adiciona_val_mediaBPDropinSupply(request):
    import datetime
    mNFG_supply_red_items = BPDropinSupplyRedItems.objects.filter(taskID = request.POST["task_id"])

      
    if  mNFG_supply_red_items:
        print("Já tem esta linha lá dentro, não vai fazer nada")
      
    else:
        print("emNFG_supply_red_items Não está empty")
        mNFG_supply_red_items = BPDropinSupplyRedItems(
            taskID = request.POST["task_id"],
            waveID = request.POST["wave_id"],
            taskType = request.POST["task_type"],
            fromPart = request.POST["from_part"],
            fromStorageZone = request.POST["from_storage_zone"],
            toLocation = request.POST["to_location"],
            #quantity_expected = request.POST["qty_expected"],
            createdDate = request.POST["created_date"],
            createdTime = request.POST["created_time"],
            timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M"), 
        )
        print("VAI GAURDAR NOVO mNFG")

        mNFG_supply_red_items.save()

    return JsonResponse({"message": "OK"})


#não vai ser usado
def media_Supply_Fa(request):
    #aqui tens de receber duas datas 
    import calendar
    import datetime
    #mNFG_supply_red_items = MNFGSupplyRedItems.objects.raw(f"""SELECT COUNT(*) FROM shippers_MNFGSupplyRedItems WHERE timestamp BETWEEN  '{request.POST["month"]}' AND '{request.POST["month"]}'""")
    #em cima são retornados o numero de vermelhos ocorridos num determinado mês
    print("REQ -->", request.POST)
 
    start_date = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    start_date_formatted = start_date.strftime('%m/%d/%Y').replace('/', '/')

    end_date = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    end_date_formatted = end_date.strftime('%m/%d/%Y').replace('/', '/')
 
    print("Datas -->",request.POST["start_date"] ,start_date )
    resultados = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE receiving_fasupplyreditems.timestamp BETWEEN '{start_date_formatted}' AND '{end_date_formatted}' """)
   


    contador = len(list(resultados)) #SELECT * FROM receiving_fasupplyreditems WHERE receiving_fasupplyreditems.timestamp BETWEEN '2/1/2023, 12:00:00 PM' AND '2/28/2023, 12:00:00 PM'
    print("RESULT QUERY-->", len(resultados),resultados)
    
    difference = end_date - start_date
    number_of_days = difference.days
    media = contador / number_of_days 
    print("Num Dias -->", number_of_days, "|MEDIA->", round(media,2), "|Contador->",contador)

    return JsonResponse({"message": "OK", "media" : round(media,2)})

#não vai ser usado
def media_bp_Dropin_Supply(request):
    #aqui tens de receber duas datas 
    import calendar
    import datetime
    #mNFG_supply_red_items = MNFGSupplyRedItems.objects.raw(f"""SELECT COUNT(*) FROM shippers_MNFGSupplyRedItems WHERE timestamp BETWEEN  '{request.POST["month"]}' AND '{request.POST["month"]}'""")
    #em cima são retornados o numero de vermelhos ocorridos num determinado mês
    start_date = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    start_date_formatted = start_date.strftime('%m/%d/%Y').replace('/', '/')

    end_date = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    end_date_formatted = end_date.strftime('%m/%d/%Y').replace('/', '/')
 

    resultados = BPDropinSupplyRedItems.objects.raw(f"""SELECT id FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{start_date_formatted}' AND '{end_date_formatted}' """)
    contador = len(list(resultados))
    
    difference = end_date - start_date
    number_of_days = difference.days
    media = contador / number_of_days 
    print("Num Dias -->", number_of_days, "|MEDIA->", round(media,2), "|Contador->",contador)

    return JsonResponse({"message": "OK", "media" : round(media,2)})


#não vai ser usado
def media_bpsmdSupply(request):
    #aqui tens de receber duas datas 
    import calendar
    import datetime
     
    print("REQ -->", request.POST)

    start_date = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    start_date_formatted = start_date.strftime('%m/%d/%Y').replace('/', '/')

    end_date = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %I:%M:%S %p')
    end_date_formatted = end_date.strftime('%m/%d/%Y').replace('/', '/')

    resultados = BPSMDSupplyRedItems.objects.raw(f"""SELECT * from receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{start_date_formatted}' AND '{end_date_formatted}' """)

    contador = len(list(resultados))
        
    difference = end_date - start_date
    number_of_days = difference.days
    media = contador / number_of_days 
    print("Num Dias -->", number_of_days, "|MEDIA->", media, "|Contador->",contador)

    return JsonResponse({"message": "OK", "media" : round(media,2)})
    
    
def count_days_in_range(start_date, end_date):
    start = date(*[int(x) for x in start_date.split('-')])
    end = date(*[int(x) for x in end_date.split('-')])
    return (end - start).days + 1

def mediaTurnosEcontadores(request):
    import datetime
    print(request.POST," (/&/) ", datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p'))#ver o que dá quando selecionas apenas um dia
    start_date = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    start_date2 = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    start_date_formatted = start_date.strftime('%m/%d/%Y').replace('/', '/')
    end_date = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    end_date2 = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    end_date_formatted = end_date.strftime('%m/%d/%Y').replace('/', '/')
    end_date_nextDay = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p') +  datetime.timedelta(days=1) 
    days = []

    if request.POST["start_date"] == request.POST["end_date"] == "":
        days.append(start_date.date().strftime('%m/%d/%Y')) #precisas do dia seguinte para o segundo turno
    else:
        while start_date <= end_date:
            days.append(start_date.date().strftime('%m/%d/%Y'))
            start_date += datetime.timedelta(days=1)
    days.append(end_date_nextDay.strftime("%m/%d/%Y"))

    listaFaT1 = 0
    listaFaT2 = 0
    listaFaT3 = 0

    listaBPDropinT1 = 0
    listaBPDropinT2 = 0
    listaBPDropinT3 = 0

    listaBPSMDT1 = 0
    listaBPSMDT2 = 0
    listaBPSMDT3 = 0
    for i in range(len(days)-1):
        
      
        resultadosBPSMDT1 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} 08:00' AND '{days[i]} 16:30' """)
        resultadosBPSMDT2 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} 16:30' AND '{days[i+1]} 01:00' """)
        resultadosBPSMDT3 = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} 01:00' AND '{days[i]} 08:00' """)

        resultadosFAT1 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{days[i]} 08:00' AND '{ days[i]} 16:30' """)
        resultadosFAT2 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{days[i]} 16:30' AND '{days[i+1]} 01:00' """)
        resultadosFAT3 = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} 01:00' AND '{ days[i]} 08:00' """)

        resultadosBPDropinT1 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} 08:00' AND '{days[i]} 16:30' """)
        resultadosBPDropinT2 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} 16:30' AND '{days[i+1]} 01:00' """)
        resultadosBPDropinT3 = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} 01:00' AND '{days[i]} 08:00' """)

        listaFaT1 += len(list(resultadosFAT1)) 
        listaFaT2 += len(list(resultadosFAT2)) 
        listaFaT3 += len(list(resultadosFAT3)) 

        listaBPDropinT1 += len(list(resultadosBPDropinT1)) 
        listaBPDropinT2 += len(list(resultadosBPDropinT2)) 
        listaBPDropinT3 += len(list(resultadosBPDropinT3)) 

        listaBPSMDT1 += len(list(resultadosBPSMDT1))
        listaBPSMDT2 += len(list(resultadosBPSMDT2))
        listaBPSMDT3 += len(list(resultadosBPSMDT3))

    number_of_days=1
    difference = end_date2 - start_date2 + datetime.timedelta(days=1)
    if difference.days <= 0:
        number_of_days==1
    else:
        number_of_days = difference.days
        
    mediaFAT1 = listaFaT1 / number_of_days
    mediaFAT2 = listaFaT2 / number_of_days
    mediaFAT3 = listaFaT3 / number_of_days 

    mediaBPSMDT1 = listaBPSMDT1 / number_of_days 
    mediaBPSMDT2 = listaBPSMDT2 / number_of_days 
    mediaBPSMDT3 = listaBPSMDT3 / number_of_days 

    mediaBPDropinT1 = listaBPDropinT1 / number_of_days 
    mediaBPDropinT2 = listaBPDropinT2 / number_of_days 
    mediaBPDropinT3 = listaBPDropinT3 / number_of_days 
    print("Num Dias -->", number_of_days, "|MEDIA->", mediaFAT1, "|Contador->",listaFaT1, "DIF->",difference,"|" , end_date2 , start_date2)

    
    return JsonResponse({
        "message": "OK", 
        #Medias
        "mediaFAT1" : round(mediaFAT1,2),
        "mediaFAT2" : round(mediaFAT2,2),
        "mediaFAT3" : round(mediaFAT3,2),

        "mediaBPSMDT1" : round(mediaBPSMDT1,2),
        "mediaBPSMDT2" : round(mediaBPSMDT2,2),
        "mediaBPSMDT3" : round(mediaBPSMDT3,2),

        "mediaBPDropinT1" : round(mediaBPDropinT1,2),
        "mediaBPDropinT2" : round(mediaBPDropinT2,2),
        "mediaBPDropinT3" : round(mediaBPDropinT3,2),

        #Contadores
        "contadorFAT1" : listaFaT1,
        "contadorFAT2" : listaFaT2,
        "contadorFAT3" : listaFaT3,

        "contadorBpsmdT1" : listaBPSMDT1,
        "contadorBpsmdT2" : listaBPSMDT2,
        "contadorBpsmdT3" : listaBPSMDT3,

        "contadorBpDropT1" : listaBPDropinT1,
        "contadorBpDropT2" : listaBPDropinT2,
        "contadorBpDropT3" : listaBPDropinT3,
            })
    
    return JsonResponse({"message": "OK", "contadorFA" : len(contadorFA), "contadorBpsmd" : len(contadorBpsmd), "contadorBpDrop" : len(contadorBpDrop)})



#não vai ser usado
def mediaTurnos(request):
    import datetime
    #BD tables ----> receiving_fasupplyreditems / receiving_bpsmdsupplyreditems / receiving_bpdropinsupplyreditems
    print(request.POST)#ver o que dá quando selecionas apenas um dia
    start_date = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    start_date2 = datetime.datetime.strptime(request.POST["start_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    start_date_formatted = start_date.strftime('%m/%d/%Y').replace('/', '/')
    end_date = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    end_date2 = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p')
    end_date_formatted = end_date.strftime('%m/%d/%Y').replace('/', '/')
    end_date_nextDay = datetime.datetime.strptime(request.POST["end_date"].replace("-", "/"), '%m/%d/%Y, %H:%M:%S %p') +  datetime.timedelta(days=1) 
    days = []

    if request.POST["start_date"] == request.POST["end_date"] == "":
        days.append(start_date.date().strftime('%m/%d/%Y')) #precisas do dia seguinte para o segundo turno
    else:
        while start_date <= end_date:
            days.append(start_date.date().strftime('%m/%d/%Y'))
            start_date += datetime.timedelta(days=1)
    days.append(end_date_nextDay.strftime("%m/%d/%Y"))

        
    hora_inicio=""
    hora_fim=""
    #este codigo não está a funcionar bem
    listaFa = 0
    listaBPDropin = 0
    listaBPSMD = 0
    print("DIAS-->",days)
    for i in range(len(days)-1):
        
      
        #print("entrou no BPSMDSupplyRedItems",hora_inicio, hora_fim)
        if request.POST["turno"] == "1":
            hora_inicio="08:00"
            hora_fim="16:30"
            resultadosBPSMD = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i]} {hora_fim}' """)
            resultadosFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{ days[i]} {hora_fim}' """)
            resultadosBPDropin = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i]} {hora_fim}' """)

        elif request.POST["turno"] == "2":
            hora_inicio="16:30"
            hora_fim="01:00"
            resultadosBPSMD = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i+1]} {hora_fim}' """)
            resultadosBPDropin = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i+1]} {hora_fim}' """)
            resultadosFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{days[i+1]} {hora_fim}' """)
        
        else:
            hora_inicio="01:00"
            hora_fim="08:00"
            resultadosBPSMD = BPSMDSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpsmdsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i]} {hora_fim}' """)
            resultadosBPDropin = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i]} {hora_fim}' """)
            resultadosFA = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{ days[i]} {hora_fim}' """)

        listaFa += len(list(resultadosFA)) 
        listaBPDropin += len(list(resultadosBPDropin)) 
        listaBPSMD += len(list(resultadosBPSMD))
    print("listas->",listaFa,listaBPDropin,listaBPSMD) 

    contadorFA = listaFa
    contadorBPSMD = listaBPSMD
    contadorBPDropin = listaBPDropin
    number_of_days=1
    difference = end_date2 - start_date2 + datetime.timedelta(days=1)
    if difference.days <= 0:
        number_of_days==1
    else:
        number_of_days = difference.days
    mediaFA = contadorFA / number_of_days 
    mediaBPSMD = contadorBPSMD / number_of_days 
    mediaBPDropin = contadorBPDropin / number_of_days 
    print("Num Dias -->", number_of_days, "|MEDIA->", mediaBPDropin, "|Contador->",contadorBPDropin, "DIF->",difference,"|" ,difference , end_date2 , start_date2)

        # elif request.POST["table"] == "receiving_bpdropinsupplyreditems":
        #     print("entrou no BPDropinSupplyRedItems",hora_inicio, hora_fim)
        #     if request.POST["turno"] == "1":
        #         hora_inicio="08:00"
        #         hora_fim="16:30"
        #         resultados = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i]} {hora_fim}' """)
        #     elif request.POST["turno"] == "2":
        #         hora_inicio="16:30"
        #         hora_fim="01:00"
        #         resultados = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i+1]} {hora_fim}' """)
        #     else:
        #         hora_inicio="01:00"
        #         hora_fim="08:00"
        #         resultados = BPDropinSupplyRedItems.objects.raw(f"""SELECT * FROM receiving_bpdropinsupplyreditems WHERE timestamp BETWEEN '{days[i]} {hora_inicio}' AND '{days[i+1]} {hora_fim}' """)

        #     teste += len(list(resultados)) 
        
        # else:
        #     print("entrou no FASupplyRedItems",hora_inicio, hora_fim)
        #     if request.POST["turno"] == "1":
        #         hora_inicio="08:00"
        #         hora_fim="16:30"
        #         resultados = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{ days[i]} {hora_fim}' """)
        #     elif request.POST["turno"] == "2":
        #         hora_inicio="16:30"
        #         hora_fim="01:00"
        #         resultados = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{ days[i+1]} {hora_fim}' """)
        #     else:
        #         hora_inicio="01:00"
        #         hora_fim="08:00"
        #         resultados = FASupplyRedItems.objects.raw(f"""SELECT * FROM receiving_fasupplyreditems WHERE timestamp BETWEEN '{ days[i]} {hora_inicio}' AND '{ days[i]} {hora_fim}' """)

        #     teste += len(list(resultados)) 
        

    return JsonResponse({"message": "OK", "mediaFA" : round(mediaFA,2), "mediaBPSMD" : round(mediaBPSMD,2), "mediaBPDropin" : round(mediaBPDropin,2)})






#vai ter de ser refeita para futuras alterações, apenas guarda o resto que vem do request fas paginas fa/ bpDroping/ bpsmd
def addCommentsToReport(request):
    # receiving_fasupplyreditems / receiving_bpsmdsupplyreditems / receiving_bpdropinsupplyreditems
    if request.POST["table"] == "receiving_fasupplyreditems":
        FaSupplyReportRedItems(report = request.POST["report"], timestamp = datetime.now().strftime("%Y-%m-%d"), turno=request.POST["turno"]).save()
    
    elif request.POST["table"] == "receiving_bpdropinsupplyreditems":
        BPDropinSupplyReportRedItems(report = request.POST["report"], timestamp = datetime.now().strftime("%Y-%m-%d"), turno=request.POST["turno"]).save()
    else:
        BPSMDSupplyReportRedItems(report = request.POST["report"], timestamp = datetime.now().strftime("%Y-%m-%d"), turno=request.POST["turno"]).save()

    return JsonResponse({"message": "OK"})


def getNOK(request):
    #apanhar todas as pos
    ReceivingPosicao1SubItems.items.all()
    ReceivingPosicao2SubItems.items.all()
    ReceivingPosicao3SubItems.items.all()
    ReceivingPosicao4SubItems.items.all()
    ReceivingPosicao5SubItems.items.all()
    ReceivingPosicao6SubItems.items.all()
    ReceivingPosicao7SubItems.items.all()
    ReceivingPosicao8SubItems.items.all()
    
    #apanhar todas as pos dos subitems

